"""Ad-hoc: cross-check an XLSX exam timetable vs. live student enrolments.

Reads the XLSX (one row per course, columns: row_no, day, hijri_date,
greg_date, period, time, course_code), builds a slot map
{(greg_date, period): {course_codes}}, and intersects against the canonical
enrolled-sets the scheduler would use today.

Reports per-student conflicts (any student with 2+ exams scheduled in the
same slot), aggregated counts, and a per-slot breakdown.
"""

from __future__ import annotations

import os
import sys
from collections import Counter, defaultdict

import django

sys.stdout.reconfigure(encoding="utf-8")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from openpyxl import load_workbook  # noqa: E402

from core.models import Course, Student  # noqa: E402
from core.services.exam_timetable import build_enrolled_sets  # noqa: E402

XLSX_PATH = r"C:\Users\user\Downloads\final_exam_timetable_codes_only_1447H_V21_checked.xlsx"


# Multi-sitting filters: course_code → {slot_key: predicate(student_id)->bool}.
# When a course has multiple rows in the schedule (e.g. split by ID-prefix
# cohort), only students matching the slot's predicate are considered scheduled
# in that slot. Predicates are total — every enrolled student must match
# exactly one slot for the same course, otherwise that student has no exam.
# Code aliases — XLSX uses code on the left, DB uses code on the right.
CODE_ALIASES: dict[str, str] = {
    "ENGL102": "ENG102",  # confirmed by registrar
}

SITTING_FILTERS: dict[str, dict[tuple, callable]] = {
    "CS112": {
        # P1: student IDs starting with 46 or 47 (newer cohorts)
        ("08/06/2026", 1): lambda sid: str(sid)[:2] in {"46", "47"},
        # P2: everyone else (older cohorts: 42, 43, 44, 45, …)
        ("08/06/2026", 2): lambda sid: str(sid)[:2] not in {"46", "47"},
    },
}


def parse_schedule(path: str) -> tuple[dict[tuple, set[str]], dict[str, list[tuple]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    slot_to_courses: dict[tuple, set[str]] = defaultdict(set)
    course_to_slots: dict[str, list[tuple]] = defaultdict(list)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for r in rows:
        if not r or r[6] is None:
            continue
        _no, day, hijri, greg, period, time_str, code = r[:7]
        code = str(code).strip().upper()
        code = CODE_ALIASES.get(code, code)
        slot_key = (str(greg), int(period))
        slot_to_courses[slot_key].add(code)
        if slot_key in course_to_slots[code]:
            continue  # same row repeated, not a split
        prior = list(course_to_slots[code])
        course_to_slots[code].append(slot_key)
        if prior and code not in SITTING_FILTERS:
            print(f"  ! UNDECLARED multi-sitting: {code} in {prior + [slot_key]}")
    return slot_to_courses, dict(course_to_slots)


def main() -> int:
    print("Parsing XLSX schedule …")
    slot_to_courses, course_to_slots = parse_schedule(XLSX_PATH)
    print(f"  slots: {len(slot_to_courses)}")
    print(
        f"  courses scheduled: {len(course_to_slots)} "
        f"({sum(len(v) for v in course_to_slots.values())} slot-assignments incl. multi-sittings)"
    )

    print("Building enrolled sets from DB …")
    enrolled = build_enrolled_sets()
    print(f"  courses with enrolment data: {len(enrolled)}")
    total_students = len({s for sids in enrolled.values() for s in sids})
    print(f"  unique students enrolled: {total_students}")

    # Verify multi-sitting predicates partition the enrolled set
    print()
    print("=== Multi-sitting integrity ===")
    for code, by_slot in SITTING_FILTERS.items():
        if code not in enrolled:
            print(f"  {code}: declared multi-sitting but no enrolment data")
            continue
        sids = enrolled[code]
        coverage: dict[int, list[tuple]] = defaultdict(list)
        for sid in sids:
            for slot, pred in by_slot.items():
                if pred(sid):
                    coverage[sid].append(slot)
        unassigned = [sid for sid in sids if sid not in coverage]
        multi_assigned = {sid: slots for sid, slots in coverage.items() if len(slots) > 1}
        per_slot_counts = Counter(slots[0] for sid, slots in coverage.items() if len(slots) == 1)
        for slot, n in sorted(per_slot_counts.items()):
            print(f"  {code} {slot}: {n} students")
        if unassigned:
            print(
                f"  ! {code}: {len(unassigned)} enrolled students match NO sitting "
                f"(sample: {sorted(unassigned)[:5]})"
            )
        if multi_assigned:
            print(
                f"  ! {code}: {len(multi_assigned)} students match multiple sittings "
                f"(sample: {dict(list(multi_assigned.items())[:5])})"
            )

    scheduled_courses = set(course_to_slots.keys())
    enrolled_courses = set(enrolled.keys())

    missing_from_db = scheduled_courses - enrolled_courses
    missing_from_xlsx = enrolled_courses - scheduled_courses

    print()
    print("=== Coverage ===")
    print(f"Courses in XLSX but not in enrolment data: {len(missing_from_db)}")
    if missing_from_db:
        for c in sorted(missing_from_db):
            ex = Course.objects.filter(course_code=c).first()
            print(f"  {c}  ({'exists in Course table' if ex else 'NOT in Course table'})")

    print(f"Courses with enrolments but NOT in XLSX: {len(missing_from_xlsx)}")
    if missing_from_xlsx:
        # show counts so registrar can see size of unscheduled cohorts
        rows = sorted(
            ((c, len(enrolled[c])) for c in missing_from_xlsx),
            key=lambda t: -t[1],
        )
        for c, n in rows[:25]:
            print(f"  {c:10s}  {n} students")
        if len(rows) > 25:
            print(f"  … +{len(rows) - 25} more")

    print()
    print("=== Conflict scan ===")
    # student → {slot: [courses]}
    student_slots: dict[int, dict[tuple, list[str]]] = defaultdict(lambda: defaultdict(list))
    for code, slots in course_to_slots.items():
        for slot in slots:
            pred = None
            if code in SITTING_FILTERS and slot in SITTING_FILTERS[code]:
                pred = SITTING_FILTERS[code][slot]
            for sid in enrolled.get(code, ()):
                if pred is not None and not pred(sid):
                    continue
                student_slots[sid][slot].append(code)

    students_with_conflicts: list[tuple[int, list[tuple[tuple, list[str]]]]] = []
    for sid, by_slot in student_slots.items():
        clashes = [(slot, sorted(cs)) for slot, cs in by_slot.items() if len(cs) >= 2]
        if clashes:
            students_with_conflicts.append((sid, clashes))

    students_with_conflicts.sort(key=lambda t: -len(t[1]))

    total_clash_events = sum(len(c) for _, c in students_with_conflicts)
    print(
        f"Students with ≥1 clash: {len(students_with_conflicts)} / {len(student_slots)} scheduled"
    )
    print(f"Total clash events (student × slot pairs): {total_clash_events}")

    # per-slot pair counts
    pair_counter: Counter[tuple] = Counter()
    for _sid, clashes in students_with_conflicts:
        for _slot, cs in clashes:
            for i in range(len(cs)):
                for j in range(i + 1, len(cs)):
                    pair_counter[(cs[i], cs[j])] += 1

    if pair_counter:
        print()
        print("=== Top conflicting course pairs (by # students affected) ===")
        for (a, b), n in pair_counter.most_common(25):
            slots_a = course_to_slots.get(a, [])
            print(f"  {a:10s}  ⇄  {b:10s}   {n} students   slot(s)={slots_a}")

    if students_with_conflicts:
        print()
        print("=== Sample affected students (top 20 by # clashes) ===")
        sid_to_meta = {
            s.student_id: (s.program, s.section)
            for s in Student.objects.filter(
                student_id__in=[sid for sid, _ in students_with_conflicts[:20]]
            )
        }
        for sid, clashes in students_with_conflicts[:20]:
            prog, sec = sid_to_meta.get(sid, ("?", "?"))
            print(f"  student_id={sid}  prog={prog}  sec={sec}  clashes={len(clashes)}")
            for slot, cs in clashes:
                print(f"      slot {slot} → {', '.join(cs)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
