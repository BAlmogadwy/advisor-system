"""Persist the XLSX-imported exam timetable as ExamTimetableRun(label=...).

Reads the XLSX, builds a v3-shaped payload (slots + schedule + qa with
same_slot_conflicts), normalises with stamp_schema_version, and writes one
ExamTimetableRun row.
"""

from __future__ import annotations

import json
import os
import sys
from collections import defaultdict
from datetime import datetime

import django

sys.stdout.reconfigure(encoding="utf-8")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from openpyxl import load_workbook  # noqa: E402

from core.models import ExamTimetableRun  # noqa: E402
from core.services.exam_run_schema import (  # noqa: E402
    STATUS_DERIVATION_VERSION,
    derive_status_surface,
    load_normalised_run,
    stamp_schema_version,
)
from core.services.exam_timetable import build_enrolled_sets  # noqa: E402

XLSX_PATH = r"C:\Users\user\Downloads\final_exam_timetable_codes_only_1447H_V21_checked.xlsx"
LABEL = "final1stdraft"

# Mirror the cross-check script: known multi-sitting splits + code aliases.
CODE_ALIASES: dict[str, str] = {
    "ENGL102": "ENG102",  # confirmed by registrar
}

SITTING_FILTERS: dict[str, dict[tuple, callable]] = {
    "CS112": {
        ("08/06/2026", 1): lambda sid: str(sid)[:2] in {"46", "47"},
        ("08/06/2026", 2): lambda sid: str(sid)[:2] not in {"46", "47"},
    },
}


def parse_xlsx(path: str):
    """Return (slot_to_courses, course_to_slots, ordered_days)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    slot_to_courses: dict[tuple, set[str]] = defaultdict(set)
    course_to_slots: dict[str, list[tuple]] = defaultdict(list)
    day_seen_order: list[str] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[6] is None:
            continue
        _no, _dayname, _hijri, greg, period, _time, code = r[:7]
        code = str(code).strip().upper()
        code = CODE_ALIASES.get(code, code)
        slot_key = (str(greg), int(period))
        slot_to_courses[slot_key].add(code)
        if slot_key not in course_to_slots[code]:
            course_to_slots[code].append(slot_key)
        if greg not in day_seen_order:
            day_seen_order.append(str(greg))

    # chronological ordering of days
    def _sort_key(d: str) -> datetime:
        # XLSX uses dd/mm/yyyy
        return datetime.strptime(d, "%d/%m/%Y")

    ordered_days = sorted(day_seen_order, key=_sort_key)
    return dict(slot_to_courses), dict(course_to_slots), ordered_days


def build_payload() -> dict:
    slot_to_courses, course_to_slots, ordered_days = parse_xlsx(XLSX_PATH)
    periods = [1, 2, 3, 4]

    # Slot pool — Cartesian product, ordered by day then period.
    # Only include (day, period) pairs that actually appear in the XLSX
    # so empty slots don't pollute the UI grid; index is dense.
    slots: list[dict] = []
    slot_index_of: dict[tuple, int] = {}
    idx = 0
    for d in ordered_days:
        for p in periods:
            if (d, p) in slot_to_courses:
                slots.append({"index": idx, "day": d, "period": p})
                slot_index_of[(d, p)] = idx
                idx += 1

    # Schedule entries — one per (course, slot).
    schedule_entries: list[dict] = []
    for code, slots_for_course in course_to_slots.items():
        for slot_key in slots_for_course:
            si = slot_index_of[slot_key]
            day, period = slot_key
            schedule_entries.append(
                {
                    "course_code": code,
                    "slot_index": si,
                    "day": day,
                    "period": period,
                    "rooms": [],
                }
            )

    # Compute same_slot_conflicts against live enrolment.
    enrolled = build_enrolled_sets()

    student_slots: dict[int, dict[int, list[str]]] = defaultdict(lambda: defaultdict(list))
    for code, slots_for_course in course_to_slots.items():
        for slot_key in slots_for_course:
            pred = SITTING_FILTERS.get(code, {}).get(slot_key)
            si = slot_index_of[slot_key]
            for sid in enrolled.get(code, ()):
                if pred is not None and not pred(sid):
                    continue
                student_slots[sid][si].append(code)

    same_slot_conflicts: list[dict] = []
    for sid, by_slot in student_slots.items():
        for si, ccs in by_slot.items():
            if len(ccs) >= 2:
                same_slot_conflicts.append(
                    {
                        "student_id": sid,
                        "slot_index": si,
                        "courses": sorted(ccs),
                    }
                )
    same_slot_conflicts.sort(key=lambda c: (c["slot_index"], c["student_id"]))

    courses_list = sorted(course_to_slots.keys())
    all_students = {sid for sids in enrolled.values() for sid in sids}

    qa: dict = {
        "same_slot_conflicts": same_slot_conflicts,
        "conflict_count": len(same_slot_conflicts),
        "manual_override_count": len(same_slot_conflicts),
        "manual_override_details": list(same_slot_conflicts),
        "imported_from": "final_exam_timetable_codes_only_1447H_V21_checked.xlsx",
        "import_notes": [
            "ENGL102 row treated as ENG102 (registrar-confirmed).",
            "CS112 split into two sittings on 08/06/2026: P1 = IDs 46/47, P2 = older cohorts.",
        ],
    }

    payload: dict = {
        "status": "ok",
        "label": LABEL,
        "students_count": len(all_students),
        "courses": courses_list,
        "courses_count": len(courses_list),
        "conflicts": [],  # course-level conflict graph not recomputed for an imported run
        "conflicts_count": 0,
        "slots": slots,
        "schedule": schedule_entries,
        "qa": qa,
        "buckets_summary": {},
        "bucket_count": 0,
        "credit_map": {},
        "seed": None,
        "section_enrollment": {},
        "rooms_count": 0,
        "assign_rooms": False,
    }
    # Mirror build_exam_timetable: set the status surface BEFORE stamping so
    # the read path doesn't fall through to default "clean".
    primary_status, status_flags = derive_status_surface(payload)
    payload["primary_status"] = primary_status
    payload["status_flags"] = status_flags
    payload["status_derivation_version"] = STATUS_DERIVATION_VERSION
    return stamp_schema_version(payload)


def main() -> int:
    payload = build_payload()
    # Replace any pre-existing row with this label so re-runs are idempotent.
    deleted, _ = ExamTimetableRun.objects.filter(label=LABEL).delete()
    if deleted:
        print(f"Removed {deleted} prior row(s) with label={LABEL!r}")
    run = ExamTimetableRun.objects.create(
        label=LABEL,
        result_json=json.dumps(payload, ensure_ascii=False),
    )
    print(f"Saved ExamTimetableRun id={run.id} label={run.label!r}")

    # Round-trip verification — make sure load_normalised_run is happy
    loaded = load_normalised_run(run)
    print(f"  status                 = {loaded.get('status')}")
    print(f"  primary_status         = {loaded.get('primary_status')}")
    print(f"  schema_version         = {loaded.get('schema_version')}")
    print(f"  courses_count          = {loaded.get('courses_count')}")
    print(f"  slots                  = {len(loaded.get('slots', []))}")
    print(f"  schedule entries       = {len(loaded.get('schedule', []))}")
    print(f"  same_slot conflicts    = {loaded.get('qa', {}).get('conflict_count')}")
    print(f"  status_flags           = {loaded.get('status_flags')}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
