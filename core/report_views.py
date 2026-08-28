import csv
import re
from io import StringIO

from django.http import FileResponse, HttpRequest, HttpResponse, HttpResponseBase, JsonResponse
from django.views.decorators.http import require_GET

from core.authz import role_required
from core.models import Course, Prerequisite, ProgrammeRequirement, Student
from core.services.advisors import list_students_by_advisor, resolve_roster_scope
from core.services.conflict_matrix import build_conflict_matrix_report, export_conflict_matrix_xlsx
from core.services.course_priority import program_downstream_importance_scores
from core.services.debug_reporting import build_recommendation_debug_report
from core.services.eligibility import (
    build_course_eligibility_report,
    hour_gate,
    split_hour_prereqs,
)
from core.services.high_priority_missing import (
    export_missing_high_priority_xlsx,
    run_missing_high_priority_report,
)
from core.services.policy import require_program_scope, require_student_scope
from core.services.rbac import ROLE_ADVISOR, ROLE_GENERAL_ADVISOR
from core.services.recommender import recommend_next_courses
from core.services.reporting import build_aggregate_counts
from core.services.student_helpers import (
    get_program_prerequisites,
    get_student_course_status_sets,
    get_student_program,
    normalize_code,
)
from core.settings_views import load_defaults


def _safe_int(value: str | None, default: int) -> int:
    try:
        return int(value) if value else default
    except (ValueError, TypeError):
        return default


def _safe_float(value: str | None, default: float) -> float:
    try:
        return float(value) if value else default
    except (ValueError, TypeError):
        return default


def _parse_int(value: str | None, field: str) -> tuple[int | None, JsonResponse | None]:
    if value is None:
        return None, JsonResponse(
            {"error": f"Missing required query parameter: {field}"}, status=400
        )
    try:
        return int(value), None
    except ValueError:
        return None, JsonResponse({"error": f"Invalid integer for {field}: {value}"}, status=400)


def _parse_recommendation_mode(
    value: str | None,
    *,
    default: str = "strict",
) -> tuple[str | None, JsonResponse | None]:
    """Parse the only supported prerequisite modes without a relaxed fallback."""
    mode = (value or "").strip().casefold() or default
    if mode not in {"strict", "relaxed"}:
        return None, JsonResponse({"error": "mode must be strict or relaxed"}, status=400)
    return mode, None


def _excel_csv_response(filename: str, csv_text: str) -> HttpResponse:
    body = "\ufeff" + csv_text
    response = HttpResponse(body, content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _parse_programs(program: str | None) -> list[str] | None:
    if not program:
        return None
    return [p.strip() for p in program.split(",") if p.strip()]


def _student_programs_for_filter(program: str | None, section: str | None) -> list[str]:
    parsed = _parse_programs(program)
    if parsed is not None:
        return parsed

    qs = Student.objects.exclude(program__isnull=True).exclude(program="")
    if section:
        qs = qs.filter(section=section)
    return sorted({str(p).strip() for p in qs.values_list("program", flat=True) if str(p).strip()})


def _course_name_fallbacks(course_codes: list[str]) -> dict[str, str]:
    return {
        normalize_code(code): description or ""
        for code, description in Course.objects.filter(course_code__in=course_codes).values_list(
            "course_code", "description"
        )
    }


def _programme_course_names(program: str, course_codes: list[str]) -> dict[str, str]:
    return {
        normalize_code(code): course_name or ""
        for code, course_name in ProgrammeRequirement.objects.filter(
            program=program,
            course_code__in=course_codes,
        ).values_list("course_code", "course_name")
    }


def _build_batch_course_rows(
    *,
    year: int,
    semester: int,
    program: str | None,
    section: str | None,
    strict_passed_only: bool = True,
    limit: int | None = None,
) -> tuple[int, list[dict[str, object]]]:
    """Return batch recommender rows grouped by plan-specific course identity."""
    student_count, aggregate = build_aggregate_counts(
        year=year,
        semester=semester,
        program=program,
        section=section,
        strict_passed_only=strict_passed_only,
    )
    program_list = _student_programs_for_filter(program, section)
    show_programs = len(program_list) != 1

    course_codes = [normalize_code(code) for code in aggregate.keys()]
    fallback_names = _course_name_fallbacks(course_codes)
    merged: dict[tuple[str, str], dict[str, object]] = {}

    if program_list:
        for prog in program_list:
            _prog_student_count, prog_aggregate = build_aggregate_counts(
                year=year,
                semester=semester,
                program=prog,
                section=section,
                strict_passed_only=strict_passed_only,
            )
            prog_codes = [normalize_code(code) for code in prog_aggregate.keys()]
            prog_names = _programme_course_names(prog, prog_codes)
            for raw_code, count in prog_aggregate.items():
                code = normalize_code(raw_code)
                course_name = prog_names.get(code) or fallback_names.get(code, "")
                key = (code, course_name)
                if key not in merged:
                    merged[key] = {
                        "course_code": code,
                        "course_name": course_name,
                        "count": 0,
                        "programs": [],
                        "show_programs": show_programs,
                    }
                merged[key]["count"] = int(merged[key]["count"]) + int(count)
                programs = merged[key]["programs"]
                if isinstance(programs, list) and prog not in programs:
                    programs.append(prog)

    if not merged:
        for raw_code, count in aggregate.items():
            code = normalize_code(raw_code)
            course_name = fallback_names.get(code, "")
            merged[(code, course_name)] = {
                "course_code": code,
                "course_name": course_name,
                "count": int(count),
                "programs": program_list,
                "show_programs": show_programs,
            }

    rows = list(merged.values())
    for row in rows:
        programs = row.get("programs")
        if isinstance(programs, list):
            row["programs"] = sorted(programs)
    rows.sort(
        key=lambda row: (
            -int(row.get("count", 0)),
            str(row.get("course_code", "")),
            str(row.get("course_name", "")),
        )
    )
    if limit is not None:
        rows = rows[:limit]
    return student_count, rows


def _program_importance_scores(program: str) -> dict[str, float]:
    """Compatibility wrapper retaining the report's six-decimal output."""

    return {
        code: round(score, 6)
        for code, score in program_downstream_importance_scores(program).items()
    }


def _build_student_plan_payload(
    student_id: int,
    *,
    prerequisite_map: dict[str, list[str]] | None = None,
    additional_studying_codes: set[str] | None = None,
    strict_passed_only: bool = False,
) -> tuple[dict | None, JsonResponse | None]:
    program = get_student_program(student_id)
    if not program:
        return None, JsonResponse(
            {"error": f"Student not found or has no program: {student_id}"}, status=404
        )

    passed, studying, failed = get_student_course_status_sets(student_id)
    studying |= {
        normalize_code(code)
        for code in (additional_studying_codes or set())
        if normalize_code(code)
    }
    # A completed requirement stays completed if its course is being retaken.
    # A failed requirement with current registrar evidence is now being studied;
    # do not count it simultaneously in both plan-status buckets.
    studying -= passed
    failed -= passed | studying
    # Keep the shared service backward-compatible for non-screen callers, while
    # request handlers explicitly choose their screen policy.  Strict eligibility
    # treats only completed passes/earned hours as prerequisite evidence.
    satisfied_pool = passed if strict_passed_only else passed | studying
    importance_scores = _program_importance_scores(program)
    prerequisites_by_course = (
        prerequisite_map if prerequisite_map is not None else get_program_prerequisites(program)
    )

    pr_rows = (
        ProgrammeRequirement.objects.filter(
            program=program,
        )
        .order_by("programme_term", "course_code")
        .values_list(
            "course_code",
            "type",
            "programme_term",
            "credit_hours",
        )
    )

    terms: dict[int, list[dict[str, object]]] = {t: [] for t in range(1, 11)}

    for code_raw, ctype, term_raw, credits_raw in pr_rows:
        code = normalize_code(code_raw)
        term = int(term_raw) if term_raw is not None else 0

        if code in passed:
            status = "passed"
        elif code in studying:
            status = "studying"
        elif code in failed:
            status = "failed"
        else:
            status = "not_taken"

        prereqs = prerequisites_by_course.get(code, [])
        # A "146(HOURS)" prerequisite is a credit-hour gate, not a course. Tested as a
        # course code it can never be satisfied, which locked every capstone forever.
        course_prereqs, required_hours = split_hour_prereqs(prereqs)
        missing_prereqs = [p for p in course_prereqs if p not in satisfied_pool]
        gate = (
            hour_gate(
                student_id,
                required_hours,
                strict_passed_only=strict_passed_only,
            )
            if required_hours
            else None
        )
        if gate is not None and not gate["met"]:
            missing_prereqs = [*missing_prereqs, f"{required_hours}(HOURS)"]
        prereqs_ok = len(missing_prereqs) == 0
        can_register = status in {"not_taken", "failed"} and prereqs_ok

        item = {
            "course_code": code,
            "type": str(ctype) if ctype is not None else "",
            "programme_term": term,
            "credit_hours": int(credits_raw) if credits_raw is not None else None,
            "status": status,
            "can_register": can_register,
            "prerequisites": prereqs,
            "missing_prereqs": missing_prereqs,
            "importance_score": float(importance_scores.get(code, 0.0)),
        }

        if 1 <= term <= 10:
            terms[term].append(item)

    blocker_stats: dict[str, dict[str, float | int]] = {}
    for courses in terms.values():
        for c in courses:
            status_val = str(c.get("status", ""))
            can_register_val = bool(c.get("can_register", False))
            if status_val not in {"not_taken", "failed"} or can_register_val:
                continue

            missing_raw = c.get("missing_prereqs", [])
            missing_list = missing_raw if isinstance(missing_raw, list) else []
            for m in missing_list:
                key = str(m)
                if key not in blocker_stats:
                    blocker_stats[key] = {
                        "blocks": 0,
                        "unlock_score": float(importance_scores.get(key, 0.0)),
                    }
                blocker_stats[key]["blocks"] = int(blocker_stats[key]["blocks"]) + 1

    blocker_hints_unsorted: list[dict[str, str | int | float]] = [
        {
            "course_code": k,
            "blocks": int(v["blocks"]),
            "unlock_score": float(v["unlock_score"]),
        }
        for k, v in blocker_stats.items()
    ]

    blocker_hints = sorted(
        blocker_hints_unsorted,
        key=lambda x: (
            -int(x["blocks"]),
            -float(x["unlock_score"]),
            str(x["course_code"]),
        ),
    )[:10]

    payload = {
        "student_id": student_id,
        "program": program,
        "eligibility_mode": "strict" if strict_passed_only else "relaxed",
        "strict_passed_only": strict_passed_only,
        "summary": {
            "passed": len(passed),
            "studying": len(studying),
            "failed": len(failed),
            "not_taken_can_register": sum(
                1
                for t in terms.values()
                for c in t
                if c["status"] == "not_taken" and c["can_register"]
            ),
            "not_taken_locked": sum(
                1
                for t in terms.values()
                for c in t
                if c["status"] == "not_taken" and not c["can_register"]
            ),
            "failed_can_register": sum(
                1
                for t in terms.values()
                for c in t
                if c["status"] == "failed" and c["can_register"]
            ),
            "failed_locked": sum(
                1
                for t in terms.values()
                for c in t
                if c["status"] == "failed" and not c["can_register"]
            ),
        },
        "blocker_hints": blocker_hints,
        "terms": [{"term": t, "courses": terms[t]} for t in range(1, 11)],
    }
    return payload, None


@role_required(ROLE_GENERAL_ADVISOR)
@require_GET
def report_summary_view(request: HttpRequest) -> JsonResponse:
    year, err = _parse_int(request.GET.get("year"), "year")
    if err:
        return err
    semester, err = _parse_int(request.GET.get("semester"), "semester")
    if err:
        return err

    if year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    program = request.GET.get("program") or None
    section = request.GET.get("section") or None
    mode, err = _parse_recommendation_mode(request.GET.get("mode"))
    if err:
        return err
    if mode is None:
        return JsonResponse({"error": "Invalid mode"}, status=400)

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    student_count, rows = _build_batch_course_rows(
        year=year,
        semester=semester,
        program=program,
        section=section,
        strict_passed_only=mode == "strict",
        limit=20,
    )

    return JsonResponse(
        {
            "year": year,
            "semester": semester,
            "program": program,
            "section": section,
            "mode": mode,
            "strict_passed_only": mode == "strict",
            "student_count": student_count,
            "top_recommended_courses": rows,
        }
    )


@role_required(ROLE_ADVISOR)
@require_GET
def export_student_csv_view(request: HttpRequest) -> HttpResponse:
    student_id, err = _parse_int(request.GET.get("student_id"), "student_id")
    if err:
        return err
    year, err = _parse_int(request.GET.get("year"), "year")
    if err:
        return err
    semester, err = _parse_int(request.GET.get("semester"), "semester")
    if err:
        return err

    if student_id is None or year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    scope_err = require_student_scope(request, student_id)
    if scope_err:
        return scope_err

    mode = "relaxed" if request.GET.get("mode", "").strip().lower() == "relaxed" else "strict"
    recommendations = recommend_next_courses(
        student_id=student_id,
        current_academic_year=year,
        current_semester=semester,
        strict_passed_only=mode == "strict",
    )

    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(["student_id", "year", "semester", "course_code"])
    for code in recommendations:
        writer.writerow([student_id, year, semester, code])

    response = _excel_csv_response(f"student_{student_id}_{year}_{semester}.csv", out.getvalue())
    response["X-Recommendation-Mode"] = mode
    return response


@role_required(ROLE_GENERAL_ADVISOR)
@require_GET
def export_aggregate_csv_view(request: HttpRequest) -> HttpResponse:
    year, err = _parse_int(request.GET.get("year"), "year")
    if err:
        return err
    semester, err = _parse_int(request.GET.get("semester"), "semester")
    if err:
        return err

    if year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    program = request.GET.get("program") or None
    section = request.GET.get("section") or None
    mode, err = _parse_recommendation_mode(request.GET.get("mode"))
    if err:
        return err
    if mode is None:
        return JsonResponse({"error": "Invalid mode"}, status=400)

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    student_count, rows = _build_batch_course_rows(
        year=year,
        semester=semester,
        program=program,
        section=section,
        strict_passed_only=mode == "strict",
    )

    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(
        [
            "year",
            "semester",
            "mode",
            "program",
            "section",
            "student_count",
            "programs",
            "course_code",
            "course_name",
            "count",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                year,
                semester,
                mode,
                program or "",
                section or "",
                student_count,
                ",".join(row.get("programs", [])) if isinstance(row.get("programs"), list) else "",
                row.get("course_code", ""),
                row.get("course_name", ""),
                row.get("count", 0),
            ]
        )

    response = _excel_csv_response(
        f"aggregate_{year}_{semester}_{mode}.csv",
        out.getvalue(),
    )
    response["X-Recommendation-Mode"] = mode
    return response


@role_required(ROLE_ADVISOR)
@require_GET
def export_aggregate_xlsx_view(request: HttpRequest) -> HttpResponse:
    """Export batch recommender results as styled XLSX."""
    year, err = _parse_int(request.GET.get("year"), "year")
    if err:
        return err
    semester, err = _parse_int(request.GET.get("semester"), "semester")
    if err:
        return err
    if year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    program = request.GET.get("program") or None
    section = request.GET.get("section") or None
    mode, err = _parse_recommendation_mode(request.GET.get("mode"))
    if err:
        return err
    if mode is None:
        return JsonResponse({"error": "Invalid mode"}, status=400)

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    student_count, rows = _build_batch_course_rows(
        year=year,
        semester=semester,
        program=program,
        section=section,
        strict_passed_only=mode == "strict",
    )

    from core.services.batch_export import export_batch_recommender_xlsx

    path = export_batch_recommender_xlsx(
        year,
        semester,
        program,
        section,
        student_count,
        {str(row.get("course_code", "")): int(row.get("count", 0)) for row in rows},
        course_rows=rows,
        strict_passed_only=mode == "strict",
    )
    prog_label = program or "all"
    filename = f"batch_recommender_{prog_label}_{year}_T{semester}_{mode}.xlsx"
    response = FileResponse(
        open(path, "rb"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["X-Recommendation-Mode"] = mode
    return response


@role_required(ROLE_ADVISOR)
@require_GET
def student_plan_view(request: HttpRequest) -> JsonResponse:
    student_id, err = _parse_int(request.GET.get("student_id"), "student_id")
    if err:
        return err
    if student_id is None:
        return JsonResponse({"error": "Invalid student_id"}, status=400)

    scope_err = require_student_scope(request, student_id)
    if scope_err:
        return scope_err

    mode, mode_err = _parse_recommendation_mode(request.GET.get("eligibility_mode"))
    if mode_err:
        return mode_err
    assert mode is not None

    payload, payload_err = _build_student_plan_payload(
        student_id,
        strict_passed_only=mode == "strict",
    )
    if payload_err:
        return payload_err
    if payload is None:
        return JsonResponse({"error": "Failed to build student plan"}, status=500)

    return JsonResponse(payload)


@role_required(ROLE_ADVISOR)
@require_GET
def export_student_plan_csv_view(request: HttpRequest) -> HttpResponse:
    student_id, err = _parse_int(request.GET.get("student_id"), "student_id")
    if err:
        return err
    if student_id is None:
        return JsonResponse({"error": "Invalid student_id"}, status=400)

    scope_err = require_student_scope(request, student_id)
    if scope_err:
        return scope_err

    mode, mode_err = _parse_recommendation_mode(request.GET.get("eligibility_mode"))
    if mode_err:
        return mode_err
    assert mode is not None

    payload, payload_err = _build_student_plan_payload(
        student_id,
        strict_passed_only=mode == "strict",
    )
    if payload_err:
        return payload_err
    if payload is None:
        return JsonResponse({"error": "Failed to build student plan"}, status=500)

    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(
        [
            "student_id",
            "program",
            "term",
            "course_code",
            "credit_hours",
            "status",
            "can_register",
            "prerequisites",
        ]
    )

    program = str(payload["program"])
    terms = payload["terms"]
    for term_obj in terms:
        term = term_obj["term"]
        for course in term_obj["courses"]:
            writer.writerow(
                [
                    student_id,
                    program,
                    term,
                    course["course_code"],
                    course["credit_hours"],
                    course["status"],
                    course["can_register"],
                    ",".join(course["prerequisites"]),
                ]
            )

    return _excel_csv_response(f"student_plan_{student_id}.csv", out.getvalue())


@role_required(ROLE_ADVISOR)
@require_GET
def prerequisites_view(request: HttpRequest) -> JsonResponse:
    program = (request.GET.get("program") or "").strip()
    course_code = (request.GET.get("course_code") or "").strip().upper().replace(" ", "")

    if not program:
        return JsonResponse({"error": "Missing required query parameter: program"}, status=400)

    scope_err = require_program_scope(request, program, require_program_for_scoped=False)
    if scope_err:
        return scope_err

    qs = Prerequisite.objects.filter(program=program)
    if course_code:
        # Filter by normalized course code
        matching_codes = [p.course_code for p in qs if normalize_code(p.course_code) == course_code]
        qs = qs.filter(course_code__in=matching_codes) if matching_codes else qs.none()

    rows = qs.order_by("course_code", "prerequisite_course_code").values_list(
        "course_code",
        "prerequisite_course_code",
    )
    data = [{"course_code": str(r[0]), "prerequisite_course_code": str(r[1])} for r in rows]
    return JsonResponse({"program": program, "count": len(data), "items": data})


@role_required(ROLE_ADVISOR)
@require_GET
def export_prerequisites_xlsx_view(request: HttpRequest) -> HttpResponse:
    """Export course prerequisites as a styled XLSX with dependency graph."""
    import tempfile
    from collections import defaultdict
    from pathlib import Path

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    program = (request.GET.get("program") or "").strip()
    course_code = (request.GET.get("course_code") or "").strip().upper().replace(" ", "")

    if not program:
        return JsonResponse({"error": "Missing required query parameter: program"}, status=400)

    scope_err = require_program_scope(request, program, require_program_for_scoped=False)
    if scope_err:
        return scope_err

    qs = Prerequisite.objects.filter(program=program)
    if course_code:
        matching_codes = [p.course_code for p in qs if normalize_code(p.course_code) == course_code]
        qs = qs.filter(course_code__in=matching_codes) if matching_codes else qs.none()

    rows = list(
        qs.order_by("course_code", "prerequisite_course_code").values_list(
            "course_code",
            "prerequisite_course_code",
        )
    )

    # Declared programme terms drive the graph's vertical axis.
    term_of: dict[str, int] = {
        str(code): int(term)
        for code, term in ProgrammeRequirement.objects.filter(program=program).values_list(
            "course_code", "programme_term"
        )
        if term is not None
    }

    # ── Shared styles ───────────────────────────────────────────
    thin = Side(style="thin", color="D5D8DC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    hdr_fill = PatternFill(start_color="0A8E6E", end_color="0A8E6E", fill_type="solid")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill(start_color="1B2631", end_color="1B2631", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    row_alt = PatternFill(start_color="F4F6F7", end_color="F4F6F7", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left_a = Alignment(horizontal="left", vertical="center")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Dependency Graph ───────────────────────────────
    graph_img = _render_prereq_graph(rows, program, term_of)
    ws_graph = wb.create_sheet(title="Dependency Graph")
    ws_graph.sheet_properties.tabColor = "0A8E6E"

    if graph_img is not None:
        img = XlImage(graph_img)
        # Scale to fit nicely — max ~1200px wide in the image
        # Excel column default ~64px, so ~18 columns ≈ 1150px
        ws_graph.add_image(img, "A1")
        # Set row heights and column widths so the image area is clean
        img_w, img_h = img.width, img.height
        # Each Excel column ≈ 8.43 chars ≈ 64px, each row ≈ 15px
        cols_needed = max(1, int(img_w / 64) + 2)
        rows_needed = max(1, int(img_h / 15) + 2)
        for c in range(1, cols_needed + 1):
            ws_graph.column_dimensions[get_column_letter(c)].width = 10
        for r in range(1, rows_needed + 1):
            ws_graph.row_dimensions[r].height = 15
    elif not rows:
        ws_graph.cell(row=1, column=1, value="No prerequisite data to graph.")
    else:
        # rows exist but the image could not be produced — currently only when
        # Pillow (the optional graph renderer) is unavailable.
        ws_graph.cell(row=1, column=1, value="Graph image unavailable (Pillow not installed).")

    # ── Sheet 2: Grouped by Course ──────────────────────────────
    ws2 = wb.create_sheet(title="Grouped by Course")
    ws2.sheet_properties.tabColor = "2E86C1"

    ws2.merge_cells("A1:C1")
    c = ws2.cell(row=1, column=1, value=f"Prerequisites Grouped — {program}")
    c.fill = title_fill
    c.font = title_font
    c.alignment = center
    for col in range(2, 4):
        ws2.cell(row=1, column=col).fill = title_fill

    for col, h in enumerate(["Course Code", "Prerequisites", "Count"], 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.border = border
        c.alignment = center

    grouped: dict[str, list[str]] = defaultdict(list)
    for cc, pc in rows:
        grouped[cc].append(pc)

    r_idx = 3
    for cc in sorted(grouped):
        prereqs = grouped[cc]
        ws2.cell(row=r_idx, column=1, value=cc).alignment = left_a
        ws2.cell(row=r_idx, column=2, value=", ".join(sorted(prereqs))).alignment = left_a
        ws2.cell(row=r_idx, column=3, value=len(prereqs)).alignment = center
        if r_idx % 2 == 1:
            for col in range(1, 4):
                ws2.cell(row=r_idx, column=col).fill = row_alt
        for col in range(1, 4):
            ws2.cell(row=r_idx, column=col).border = border
        r_idx += 1

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 10
    ws2.freeze_panes = "A3"

    # ── Sheet 3: Summary ────────────────────────────────────────
    ws3 = wb.create_sheet(title="Summary")
    ws3.sheet_properties.tabColor = "F39C12"

    ws3.merge_cells("A1:B1")
    c = ws3.cell(row=1, column=1, value=f"Summary — {program}")
    c.fill = title_fill
    c.font = title_font
    c.alignment = center
    ws3.cell(row=1, column=2).fill = title_fill

    summary_data = [
        ("Program", program),
        ("Total Prerequisite Links", len(rows)),
        ("Courses with Prerequisites", len(grouped)),
        ("Unique Prerequisite Courses", len({pc for _, pc in rows})),
    ]
    for r_idx, (label, val) in enumerate(summary_data, 2):
        ws3.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=r_idx, column=1).border = border
        ws3.cell(row=r_idx, column=2, value=val).border = border

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    # ── Save and return ─────────────────────────────────────────
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    course_label = f"_{course_code}" if course_code else ""
    filename = f"prerequisites_{program}{course_label}.xlsx"
    response = FileResponse(
        open(Path(tmp.name), "rb"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ── Prereq-graph layout helpers ─────────────────────────────────────────────
# These mirror the frontend renderer in static/js/page-dashboard.js so the XLSX
# export and the on-screen graph agree.  Keep the two in step when either moves.
_PG_GATE_RE = re.compile(r"^\s*(\d+)\s*\(\s*HOURS?\s*\)\s*$", re.IGNORECASE)


def _pg_gate_hours(code: str) -> str | None:
    """Digit string for a credit-hour gate like ``144(HOURS)``, else ``None``."""
    m = _PG_GATE_RE.match(str(code))
    return m.group(1) if m else None


def _pg_term_rows(all_courses, dependents, term_of):
    """Assign every node a row = its declared programme term.

    Nodes with no declared term (credit-hour gates, and courses that gate this
    plan without belonging to it) are placed one row before the earliest course
    that depends on them, and recorded in ``inferred`` so the guess stays
    visible.  Mirrors ``pgTermRows``.
    """
    row: dict[str, int] = {}
    inferred: set[str] = set()
    for c in all_courses:
        t = term_of.get(c)
        if isinstance(t, int):
            row[c] = t
    floor = min(row.values()) if row else 1
    changed = True
    guard = 0
    while changed and guard <= len(all_courses):
        guard += 1
        changed = False
        for c in all_courses:
            if c in row:
                continue
            known = [row[d] for d in dependents.get(c, []) if d in row]
            if not known:
                continue
            row[c] = min(known) - 1
            inferred.add(c)
            changed = True
    for c in all_courses:
        if c not in row:
            row[c] = floor - 1
            inferred.add(c)
    return row, inferred


def _pg_build_edges(rows, row, inferred):
    """Build edge dicts ``{f, t, warn}`` from ``(course, prereq)`` rows.

    An edge is a "warn" (unsatisfiable as declared) only when BOTH endpoints have
    a DECLARED term and the prerequisite's term is at or after its dependent's.
    An edge touching an inferred-term node never warns — the "declared same/later
    term" message would be false there.  Mirrors the frontend edge build.
    """
    out = []
    for cc, pc in rows:
        warn = pc not in inferred and cc not in inferred and row[pc] >= row[cc]
        out.append({"f": pc, "t": cc, "warn": warn})
    return out


def _pg_build_slots(edges, row, min_r, max_r):
    """Give every multi-row edge a routing point in each band it crosses.

    Returns ``(slots, up, dn, chain)``: ``slots[r]`` is the ordered list of slot
    dicts in band ``r``; ``up``/``dn`` are adjacency over slot ids; ``chain[i]``
    is the id path for edge ``i`` (``None`` for warning edges).  Mirrors
    ``pgBuildSlots``.
    """
    slots = {r: [] for r in range(min_r, max_r + 1)}
    seen = set()
    for e in edges:
        for node_id in (e["f"], e["t"]):
            if node_id in seen:
                continue
            seen.add(node_id)
            slots[row[node_id]].append({"id": node_id, "kind": "node", "key": node_id})
    up: dict[str, list[str]] = {}
    dn: dict[str, list[str]] = {}

    def link(a, b):
        dn.setdefault(a, []).append(b)
        up.setdefault(b, []).append(a)

    chain: dict[int, list[str] | None] = {}
    for i, e in enumerate(edges):
        if e["warn"]:
            chain[i] = None
            continue
        r1, r2 = row[e["f"]], row[e["t"]]
        path = [e["f"]]
        for r in range(r1 + 1, r2):
            rid = f" d{i}@{r}"
            slots[r].append({"id": rid, "kind": "route", "key": f"{e['f']}>{e['t']}"})
            path.append(rid)
        path.append(e["t"])
        for j in range(len(path) - 1):
            link(path[j], path[j + 1])
        chain[i] = path
    return slots, up, dn, chain


def _pg_order_slots(slots, up, dn, edges, chain):
    """Barycentre sweeps that keep the least-crossing arrangement.

    Deterministic: alphabetical seed and tie-break.  Mirrors ``pgOrderSlots``.
    Mutates ``slots`` in place and returns the sorted band keys.
    """
    keys = sorted(slots.keys())
    for k in keys:
        slots[k].sort(key=lambda s: s["key"])
    row_of = {s["id"]: k for k in keys for s in slots[k]}
    idx: dict[str, int] = {}
    span: dict[int, int] = {}

    def reindex():
        for k in keys:
            arr = slots[k]
            span[k] = len(arr)
            for i, s in enumerate(arr):
                idx[s["id"]] = i

    reindex()

    def frac(node_id):
        return (idx[node_id] + 0.5) / max(1, span[row_of[node_id]])

    def segments():
        out = []
        for i in range(len(edges)):
            path = chain.get(i)
            if not path:
                continue
            for j in range(len(path) - 1):
                out.append((path[j], path[j + 1]))
        return out

    def crossings():
        sg = segments()
        n = 0
        for a in range(len(sg)):
            a1, b1 = sg[a]
            for b in range(a + 1, len(sg)):
                a2, b2 = sg[b]
                if row_of[a1] != row_of[a2] or row_of[b1] != row_of[b2]:
                    continue
                if a1 == a2 or b1 == b2:
                    continue
                if (idx[a1] - idx[a2]) * (idx[b1] - idx[b2]) < 0:
                    n += 1
        return n

    def snapshot():
        return {k: [s["id"] for s in slots[k]] for k in keys}

    def restore(snap):
        for k in keys:
            by_id = {s["id"]: s for s in slots[k]}
            slots[k] = [by_id[i] for i in snap[k]]

    best = snapshot()
    best_x = crossings()
    for p in range(8):
        down = p % 2 == 0
        for k in keys if down else list(reversed(keys)):
            arr = slots[k]
            bc = {}
            for s in arr:
                nb = [frac(x) for x in (up if down else dn).get(s["id"], []) if x in idx]
                bc[s["id"]] = sum(nb) / len(nb) if nb else frac(s["id"])
            arr.sort(key=lambda s: (bc[s["id"]], s["key"]))
            reindex()
        x = crossings()
        if x < best_x:
            best_x = x
            best = snapshot()
    restore(best)
    reindex()
    return keys


def _pg_dash_segment(draw, x1, y1, x2, y2, color, width, dash, gap):
    """Draw a dashed straight line from (x1,y1) to (x2,y2)."""
    from math import hypot

    dist = hypot(x2 - x1, y2 - y1)
    if dist == 0:
        return
    ux, uy = (x2 - x1) / dist, (y2 - y1) / dist
    n = 0.0
    while n < dist:
        a, b = n, min(n + dash, dist)
        draw.line([(x1 + ux * a, y1 + uy * a), (x1 + ux * b, y1 + uy * b)], fill=color, width=width)
        n += dash + gap


def _pg_dashed_rrect(draw, box, radius, color, width, dash, gap):
    """Dashed rounded-rectangle border (straight edges dashed, corners arced)."""
    x0, y0, x1, y1 = box
    r = radius
    _pg_dash_segment(draw, x0 + r, y0, x1 - r, y0, color, width, dash, gap)
    _pg_dash_segment(draw, x0 + r, y1, x1 - r, y1, color, width, dash, gap)
    _pg_dash_segment(draw, x0, y0 + r, x0, y1 - r, color, width, dash, gap)
    _pg_dash_segment(draw, x1, y0 + r, x1, y1 - r, color, width, dash, gap)
    draw.arc([x0, y0, x0 + 2 * r, y0 + 2 * r], 180, 270, fill=color, width=width)
    draw.arc([x1 - 2 * r, y0, x1, y0 + 2 * r], 270, 360, fill=color, width=width)
    draw.arc([x0, y1 - 2 * r, x0 + 2 * r, y1], 90, 180, fill=color, width=width)
    draw.arc([x1 - 2 * r, y1 - 2 * r, x1, y1], 0, 90, fill=color, width=width)


def _render_prereq_graph(
    rows: list[tuple[str, str]],
    program: str,
    term_of: dict[str, int] | None = None,
):
    """Render the prerequisite dependency graph as a high-quality PNG.

    The vertical axis is the declared programme term (falling back to longest
    prerequisite chain when the plan carries no terms), matching the on-screen
    ``renderPrereqGraph``: term-labelled bands, credit-hour gates as amber
    dashed pills, inferred-term nodes dashed, and unsatisfiable edges bowed
    out in amber.

    Returns a ``BytesIO`` ready for openpyxl ``Image()``, or ``None``.
    """
    from collections import defaultdict
    from io import BytesIO
    from math import atan2, cos, sin

    try:
        from PIL import Image, ImageDraw, ImageFilter, ImageFont
    except ImportError:
        # Pillow is optional; without it the export drops the graph image
        # rather than failing the whole download.
        return None

    if not rows:
        return None
    term_of = term_of or {}

    # ── adjacency ───────────────────────────────────────────────
    prereqs: dict[str, list[str]] = defaultdict(list)
    dependents: dict[str, list[str]] = defaultdict(list)
    # insertion-ordered set: term inference freezes on first assignment, so a
    # deterministic iteration order keeps the export reproducible and matched to
    # the on-screen graph (a plain set's order is PYTHONHASHSEED-salted).
    all_courses: dict[str, None] = {}
    for cc, pc in rows:
        all_courses.setdefault(cc, None)
        all_courses.setdefault(pc, None)
        if pc not in prereqs[cc]:
            prereqs[cc].append(pc)
        if cc not in dependents[pc]:
            dependents[pc].append(cc)

    is_pre_of: set[str] = {pc for _cc, pc in rows}
    gate_of = {c: _pg_gate_hours(c) for c in all_courses}

    # ── rows = declared programme term (else longest-chain depth) ─
    declared = [c for c in all_courses if isinstance(term_of.get(c), int)]
    by_term = bool(declared)
    if by_term:
        row, inferred = _pg_term_rows(all_courses, dependents, term_of)
    else:
        layers: dict[str, int] = {}

        def _depth(c, vis=None):
            if c in layers:
                return layers[c]
            vis = set() if vis is None else vis
            if c in vis:
                return 0
            vis.add(c)
            ps = prereqs.get(c, [])
            layers[c] = (max(_depth(p, set(vis)) for p in ps) + 1) if ps else 0
            return layers[c]

        for c in all_courses:
            _depth(c)
        row, inferred = layers, set()

    min_r, max_r = min(row.values()), max(row.values())

    # ── edges; unsatisfiable ones (declared prereq at/after its dependent) bypass routing ─
    edges = _pg_build_edges(rows, row, inferred)
    slots, up, dn, chain = _pg_build_slots(edges, row, min_r, max_r)
    keys = _pg_order_slots(slots, up, dn, edges, chain)

    # ── geometry (render at 3x for HiDPI) ───────────────────────
    S = 3  # supersampling factor
    node_h = 34 * S
    node_r = 8 * S
    gap_x = 18 * S
    route_w = 14 * S
    pad_x = 40 * S
    pad_top = 64 * S
    pad_bot = 74 * S
    band_pad_y = 20 * S
    empty_h = 30 * S
    font_size = 11 * S
    gutter = (104 * S) if by_term else 0

    max_chars = max(len(c) for c in all_courses)
    node_w = max(90 * S, (max_chars * 8 + 28) * S)

    def slot_w(s):
        return node_w if s["kind"] == "node" else route_w

    def band_w(k):
        arr = slots[k]
        return sum(slot_w(s) for s in arr) + max(0, len(arr) - 1) * gap_x

    # band heights: a band with a course is tall, an empty term is a thin rule
    band_y: dict[int, int] = {}
    band_h: dict[int, int] = {}
    y = pad_top
    for k in keys:
        has_node = any(s["kind"] == "node" for s in slots[k])
        h = (node_h + 2 * band_pad_y) if has_node else empty_h
        band_y[k] = y
        band_h[k] = h
        y += h
    img_h = y + pad_bot

    content_w = max((band_w(k) for k in keys), default=node_w)
    img_w = max(600 * S, gutter + pad_x * 2 + content_w)
    content_x0 = gutter + pad_x
    content_avail = img_w - gutter - pad_x * 2

    # ── slot positions (nodes and routing points share the band centre) ─
    pos: dict[str, tuple[int, int]] = {}
    for k in keys:
        x = content_x0 + max(0, (content_avail - band_w(k)) // 2)
        cy = band_y[k] + band_h[k] // 2
        for s in slots[k]:
            w = slot_w(s)
            pos[s["id"]] = (x + w // 2, cy)
            x += w + gap_x

    # ── Node styling by role (mirrors the on-screen palette) ────
    def node_style(c: str):
        """Return (fill, stroke, text_colour, shadow_colour) for course ``c``."""
        if gate_of.get(c):
            # Credit-hour gate — amber
            return (250, 240, 228), (180, 83, 9), (180, 83, 9), (180, 83, 9, 26)
        has_p = bool(prereqs.get(c))
        is_p = c in is_pre_of
        if not has_p and is_p:
            # Foundation — teal tint
            return (228, 244, 239), (10, 142, 110), (6, 100, 80), (10, 142, 110, 30)
        if has_p and not is_p:
            # Terminal — indigo tint
            return (232, 235, 252), (86, 104, 220), (48, 64, 180), (86, 104, 220, 30)
        # Intermediate — neutral white
        return (255, 255, 255), (195, 202, 212), (35, 45, 60), (0, 0, 0, 18)

    # ── Background: subtle vertical gradient ────────────────────
    img = Image.new("RGBA", (img_w, img_h), (0, 0, 0, 0))
    bg = Image.new("RGBA", (img_w, img_h))
    bg_draw = ImageDraw.Draw(bg)
    # Top: (240, 247, 250)  Bottom: (232, 240, 245)
    for y in range(img_h):
        t = y / max(1, img_h - 1)
        r = int(240 + (232 - 240) * t)
        g_c = int(247 + (240 - 247) * t)
        b = int(250 + (245 - 250) * t)
        bg_draw.line([(0, y), (img_w, y)], fill=(r, g_c, b, 255))
    img = Image.alpha_composite(img, bg)

    # ── Fonts ───────────────────────────────────────────────────
    def _try_font(names, size):
        for name in names:
            try:
                return ImageFont.truetype(name, size)
            except OSError:
                continue
        return ImageFont.load_default()

    font_mono = _try_font(
        ["consolab.ttf", "consola.ttf", "cour.ttf", "DejaVuSansMono-Bold.ttf"], font_size
    )
    font_title = _try_font(
        ["calibrib.ttf", "calibri.ttf", "arialbd.ttf", "arial.ttf"], int(font_size * 1.5)
    )
    font_legend = _try_font(["calibri.ttf", "arial.ttf", "segoeui.ttf"], int(font_size * 0.9))

    # ── Term bands: alternating tints + gutter labels behind everything ─
    band_draw = ImageDraw.Draw(img)
    if by_term and keys:
        for k in keys:
            has_node = any(s["kind"] == "node" for s in slots[k])
            if k % 2:
                band_draw.rectangle(
                    [0, band_y[k], img_w, band_y[k] + band_h[k]], fill=(17, 17, 68, 8)
                )
            band_draw.line(
                [(0, band_y[k]), (img_w, band_y[k])],
                fill=(17, 17, 68, 20),
                width=max(1, S // 2),
            )
            ly = band_y[k] + band_h[k] // 2
            lbl = f"TERM {k}"
            bb = band_draw.textbbox((0, 0), lbl, font=font_mono)
            tw, th = bb[2] - bb[0], bb[3] - bb[1]
            band_draw.text(
                (gutter - 16 * S - tw, ly - th // 2),
                lbl,
                fill=(120, 128, 150) if has_node else (150, 156, 172),
                font=font_mono,
            )
            if not has_node:
                band_draw.text(
                    (gutter + 14 * S, ly - th // 2),
                    "no linked courses",
                    fill=(160, 166, 180),
                    font=font_legend,
                )
        band_draw.line(
            [(gutter, band_y[keys[0]]), (gutter, band_y[keys[-1]] + band_h[keys[-1]])],
            fill=(17, 17, 68, 28),
            width=max(1, S // 2),
        )

    # ── Draw edges on a separate RGBA layer (crisp, no blur) ────
    edge_layer = Image.new("RGBA", (img_w, img_h), (0, 0, 0, 0))
    edge_draw = ImageDraw.Draw(edge_layer)

    def _bezier(x1, y1, x2, y2, steps=26):
        """Cubic bezier with vertical control points (prereq → course)."""
        cy = (y1 + y2) / 2
        return [
            (
                (1 - t) ** 3 * x1 + 3 * (1 - t) ** 2 * t * x1 + 3 * (1 - t) * t**2 * x2 + t**3 * x2,
                (1 - t) ** 3 * y1 + 3 * (1 - t) ** 2 * t * cy + 3 * (1 - t) * t**2 * cy + t**3 * y2,
            )
            for t in (s / steps for s in range(steps + 1))
        ]

    def _cubic(p0, p1, p2, p3, steps=26):
        """General cubic bezier through four control points."""
        out = []
        for s in range(steps + 1):
            t = s / steps
            u = 1 - t
            out.append(
                (
                    u**3 * p0[0] + 3 * u**2 * t * p1[0] + 3 * u * t**2 * p2[0] + t**3 * p3[0],
                    u**3 * p0[1] + 3 * u**2 * t * p1[1] + 3 * u * t**2 * p2[1] + t**3 * p3[1],
                )
            )
        return out

    line_w = max(3, S + S // 2)
    teal = (10, 142, 110)
    amber = (180, 83, 9)

    def _arrow(pts, color):
        if len(pts) < 4:
            return
        px, py = pts[-4]
        ax, ay = pts[-1]
        angle = atan2(ay - py, ax - px)
        sz = 5.5 * S
        edge_draw.polygon(
            [
                (ax, ay),
                (ax - sz * cos(angle - 0.4), ay - sz * sin(angle - 0.4)),
                (ax - sz * cos(angle + 0.4), ay - sz * sin(angle + 0.4)),
            ],
            fill=color,
        )

    for i, e in enumerate(edges):
        fc, tc = e["f"], e["t"]
        if fc not in pos or tc not in pos:
            continue
        if e["warn"]:
            # prereq at/after its own dependent — bow sideways, amber dashed
            fx, fy = pos[fc]
            tx, ty = pos[tc]
            d = 1 if tx >= fx else -1
            x1 = fx + d * (node_w // 2)
            x2 = tx - d * (node_w // 2)
            yb = max(fy, ty) + node_h // 2 + 14 * S
            pts = _cubic((x1, fy), (x1 + d * 26 * S, yb), (x2 - d * 26 * S, yb), (x2, ty))
            for j in range(len(pts) - 1):
                if j % 2 == 0:
                    edge_draw.line([pts[j], pts[j + 1]], fill=(*amber, 190), width=line_w)
            _arrow(pts, (*amber, 230))
            continue
        # route through the chain: node bottom → routing points → node top
        path = chain.get(i) or [fc, tc]
        anchors = []
        for j, nid in enumerate(path):
            qx, qy = pos[nid]
            if j == 0:
                anchors.append((qx, qy + node_h // 2 + 2 * S))
            elif j == len(path) - 1:
                anchors.append((qx, qy - node_h // 2 - 2 * S))
            else:
                anchors.append((qx, qy))
        curve = []
        for j in range(len(anchors) - 1):
            seg = _bezier(anchors[j][0], anchors[j][1], anchors[j + 1][0], anchors[j + 1][1])
            curve.extend(seg if j == 0 else seg[1:])
        for j in range(len(curve) - 1):
            edge_draw.line([curve[j], curve[j + 1]], fill=(*teal, 150), width=line_w)
        _arrow(curve, (*teal, 220))

    img = Image.alpha_composite(img, edge_layer)

    # ── Draw nodes ──────────────────────────────────────────────
    draw = ImageDraw.Draw(img)

    # Shadow pass first (all nodes)
    shadow_layer = Image.new("RGBA", (img_w, img_h), (0, 0, 0, 0))
    shadow_draw = ImageDraw.Draw(shadow_layer)
    shadow_offset = 3 * S
    for c in all_courses:
        p = pos.get(c)
        if not p:
            continue
        cx, cy = p
        _, _, _, sh_col = node_style(c)
        rad = node_h // 2 if gate_of.get(c) else node_r
        shadow_draw.rounded_rectangle(
            [
                cx - node_w // 2,
                cy - node_h // 2 + shadow_offset,
                cx + node_w // 2,
                cy + node_h // 2 + shadow_offset,
            ],
            radius=rad,
            fill=sh_col,
        )
    shadow_layer = shadow_layer.filter(ImageFilter.GaussianBlur(radius=4 * S))
    img = Image.alpha_composite(img, shadow_layer)

    # Node bodies
    draw = ImageDraw.Draw(img)
    for c in all_courses:
        p = pos.get(c)
        if not p:
            continue
        cx, cy = p
        fill_c, stroke_c, text_c, _ = node_style(c)
        is_gate = bool(gate_of.get(c))
        is_inf = c in inferred and not is_gate
        rad = node_h // 2 if is_gate else node_r
        box = [cx - node_w // 2, cy - node_h // 2, cx + node_w // 2, cy + node_h // 2]

        draw.rounded_rectangle(box, radius=rad, fill=(*fill_c, 240))
        if is_gate or is_inf:
            # gates and inferred-term nodes carry a dashed border (matches screen)
            _pg_dashed_rrect(draw, box, rad, (*stroke_c, 220), max(1, S), 5 * S, 3 * S)
        else:
            draw.rounded_rectangle(box, radius=rad, outline=(*stroke_c, 180), width=max(1, S))

        # Centred label — gates show their hours, courses their code
        label = f"{gate_of[c]} hrs" if is_gate else c
        bbox = draw.textbbox((0, 0), label, font=font_mono)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        draw.text((cx - tw // 2, cy - th // 2), label, fill=(*text_c, 255), font=font_mono)

    # ── Title ───────────────────────────────────────────────────
    title = f"Dependency Graph — {program}"
    bbox = draw.textbbox((0, 0), title, font=font_title)
    tw = bbox[2] - bbox[0]
    draw.text(
        ((img_w - tw) // 2, (pad_top - int(font_size * 1.5)) // 2),
        title,
        fill=(27, 38, 49, 255),
        font=font_title,
    )

    # ── Legend: node roles present in this graph (centred at bottom) ─
    legend_items: list[tuple[tuple[int, int, int], str]] = [
        ((10, 142, 110), "Foundation"),
        ((86, 104, 220), "Terminal"),
        ((120, 128, 150), "Intermediate"),
    ]
    if any(gate_of.values()):
        legend_items.append(((180, 83, 9), "Credit-hour gate"))
    if inferred:
        legend_items.append(((150, 158, 175), "Term inferred"))

    dot_r = 5 * S
    item_gap = 18 * S
    row_w = 0
    for _color, label in legend_items:
        bbox = draw.textbbox((0, 0), label, font=font_legend)
        row_w += dot_r * 2 + 6 * S + (bbox[2] - bbox[0]) + item_gap
    row_w -= item_gap  # no trailing gap

    lx = (img_w - row_w) // 2
    ly = img_h - pad_bot + 14 * S
    for color, label in legend_items:
        draw.ellipse([lx, ly, lx + dot_r * 2, ly + dot_r * 2], fill=(*color, 255))
        draw.text(
            (lx + dot_r * 2 + 6 * S, ly - 2 * S),
            label,
            fill=(80, 90, 100, 255),
            font=font_legend,
        )
        bbox = draw.textbbox((0, 0), label, font=font_legend)
        lx += dot_r * 2 + 6 * S + (bbox[2] - bbox[0]) + item_gap

    # ── Finalise: flatten to RGB, downscale for crisp output ────
    flat = Image.new("RGB", img.size, (240, 247, 250))
    flat.paste(img, mask=img.split()[3])
    final = flat.resize((img_w // S, img_h // S), Image.LANCZOS)
    buf = BytesIO()
    final.save(buf, format="PNG", dpi=(150, 150))
    buf.seek(0)
    return buf


@role_required(ROLE_ADVISOR)
@require_GET
def program_plan_view(request: HttpRequest) -> JsonResponse:
    program = (request.GET.get("program") or "").strip()
    if not program:
        return JsonResponse({"error": "Missing required query parameter: program"}, status=400)

    scope_err = require_program_scope(request, program, require_program_for_scoped=False)
    if scope_err:
        return scope_err

    pp_rows = (
        ProgrammeRequirement.objects.filter(
            program=program,
        )
        .order_by("programme_term", "course_code")
        .values_list(
            "course_code",
            "course_name",
            "programme_term",
            "credit_hours",
        )
    )

    items = [
        {
            "course_code": str(r[0]),
            "course_name": str(r[1] or ""),
            "programme_term": int(r[2]) if r[2] is not None else None,
            "credit_hours": int(r[3]) if r[3] is not None else None,
        }
        for r in pp_rows
    ]
    return JsonResponse({"program": program, "count": len(items), "items": items})


@role_required(ROLE_ADVISOR)
@require_GET
def recommendation_debug_view(request: HttpRequest) -> JsonResponse:
    _defaults = load_defaults()
    year_raw = request.GET.get("year", "").strip() or str(_defaults["academic_year"])
    semester_raw = request.GET.get("semester", "").strip() or str(_defaults["term"])

    year, err = _parse_int(year_raw, "year")
    if err:
        return err
    semester, err = _parse_int(semester_raw, "semester")
    if err:
        return err

    if year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    section = (request.GET.get("section") or "").strip().upper() or None
    program = (request.GET.get("program") or "").strip().upper() or None
    join_years_raw = (request.GET.get("join_years") or "").strip()
    join_years = (
        [x.strip() for x in join_years_raw.split(",") if x.strip()] if join_years_raw else None
    )
    limit = _safe_int(request.GET.get("limit"), 150)
    # Same vocabulary as course_eligibility_view: mode=strict, anything else relaxed.
    strict_mode = (request.GET.get("mode") or "").strip().lower() == "strict"

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    payload = build_recommendation_debug_report(
        current_academic_year=year,
        current_semester=semester,
        section=section,
        program=program,
        join_year_prefixes=join_years,
        limit=limit,
        strict_passed_only=strict_mode,
    )
    return JsonResponse(payload)


@role_required(ROLE_ADVISOR)
@require_GET
def conflict_matrix_view(request: HttpRequest) -> JsonResponse:
    _defaults = load_defaults()
    year_raw = request.GET.get("year", "").strip() or str(_defaults["academic_year"])
    semester_raw = request.GET.get("semester", "").strip() or str(_defaults["term"])

    year, err = _parse_int(year_raw, "year")
    if err:
        return err
    semester, err = _parse_int(semester_raw, "semester")
    if err:
        return err

    if year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    section = (request.GET.get("section") or "").strip().upper() or None
    program = (request.GET.get("program") or "").strip().upper() or None
    join_years_raw = (request.GET.get("join_years") or "").strip()
    join_years = (
        [x.strip() for x in join_years_raw.split(",") if x.strip()] if join_years_raw else None
    )
    limit = _safe_int(request.GET.get("limit"), 150)

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    payload = build_conflict_matrix_report(
        current_academic_year=year,
        current_semester=semester,
        section=section,
        program=program,
        join_year_prefixes=join_years,
        limit=limit,
    )
    return JsonResponse(payload)


@role_required(ROLE_ADVISOR)
@require_GET
def export_conflict_matrix_xlsx_view(request: HttpRequest) -> HttpResponseBase:
    _defaults = load_defaults()
    year_raw = request.GET.get("year", "").strip() or str(_defaults["academic_year"])
    semester_raw = request.GET.get("semester", "").strip() or str(_defaults["term"])

    year, err = _parse_int(year_raw, "year")
    if err:
        return err
    semester, err = _parse_int(semester_raw, "semester")
    if err:
        return err

    if year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    section = (request.GET.get("section") or "").strip().upper() or None
    program = (request.GET.get("program") or "").strip().upper() or None
    join_years_raw = (request.GET.get("join_years") or "").strip()
    join_years = (
        [x.strip() for x in join_years_raw.split(",") if x.strip()] if join_years_raw else None
    )
    limit = _safe_int(request.GET.get("limit"), 150)

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    try:
        path = export_conflict_matrix_xlsx(
            current_academic_year=year,
            current_semester=semester,
            section=section,
            program=program,
            join_year_prefixes=join_years,
            limit=limit,
        )
        return FileResponse(path.open("rb"), as_attachment=True, filename="conflict_matrix.xlsx")
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


@role_required(ROLE_ADVISOR)
@require_GET
def course_eligibility_view(request: HttpRequest) -> JsonResponse:
    course_code = (request.GET.get("course_code") or "").strip().upper()
    if not course_code:
        return JsonResponse({"error": "course_code is required"}, status=400)

    section = (request.GET.get("section") or "").strip().upper() or None
    program = (request.GET.get("program") or "").strip().upper() or None
    join_years_raw = (request.GET.get("join_years") or "").strip()
    join_years = (
        [x.strip() for x in join_years_raw.split(",") if x.strip()] if join_years_raw else None
    )
    mode, err = _parse_recommendation_mode(request.GET.get("mode"))
    if err:
        return err
    if mode is None:
        return JsonResponse({"error": "Invalid mode"}, status=400)
    strict_mode = mode == "strict"

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    payload = build_course_eligibility_report(
        course_code=course_code,
        section=section,
        program=program,
        join_year_prefixes=join_years,
        strict_passed_only=strict_mode,
    )
    return JsonResponse(payload)


@role_required(ROLE_ADVISOR)
@require_GET
def export_recommendation_debug_csv_view(request: HttpRequest) -> HttpResponse:
    _defaults = load_defaults()
    year_raw = request.GET.get("year", "").strip() or str(_defaults["academic_year"])
    semester_raw = request.GET.get("semester", "").strip() or str(_defaults["term"])

    year, err = _parse_int(year_raw, "year")
    if err:
        return err
    semester, err = _parse_int(semester_raw, "semester")
    if err:
        return err
    if year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    section = (request.GET.get("section") or "").strip().upper() or None
    program = (request.GET.get("program") or "").strip().upper() or None
    join_years_raw = (request.GET.get("join_years") or "").strip()
    join_years = (
        [x.strip() for x in join_years_raw.split(",") if x.strip()] if join_years_raw else None
    )
    limit = _safe_int(request.GET.get("limit"), 150)
    strict_mode = (request.GET.get("mode") or "").strip().lower() == "strict"

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    payload = build_recommendation_debug_report(
        year, semester, section, program, join_years, limit, strict_passed_only=strict_mode
    )

    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(
        [
            "student_id",
            "program",
            "real_term",
            "next_term",
            "passed",
            "studying",
            "recommended_courses",
        ]
    )
    for item in payload.get("items", []):
        writer.writerow(
            [
                item.get("student_id"),
                item.get("program"),
                item.get("real_term"),
                item.get("next_term"),
                ",".join(item.get("passed", [])),
                ",".join(item.get("studying", [])),
                ",".join(item.get("recommended_courses", [])),
            ]
        )

    # Name the assumption in the file itself: a strict export must not be
    # mistakable for a relaxed one once it leaves the browser.
    csv_name = "recommendation_debug_strict.csv" if strict_mode else "recommendation_debug.csv"
    return _excel_csv_response(csv_name, out.getvalue())


@role_required(ROLE_ADVISOR)
@require_GET
def export_recommendation_debug_xlsx_view(request: HttpRequest) -> HttpResponse:
    """Export recommendation debug report as styled XLSX workbook."""
    _defaults = load_defaults()
    year_raw = request.GET.get("year", "").strip() or str(_defaults["academic_year"])
    semester_raw = request.GET.get("semester", "").strip() or str(_defaults["term"])

    year, err = _parse_int(year_raw, "year")
    if err:
        return err
    semester, err = _parse_int(semester_raw, "semester")
    if err:
        return err
    if year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    section = (request.GET.get("section") or "").strip().upper() or None
    program = (request.GET.get("program") or "").strip().upper() or None
    join_years_raw = (request.GET.get("join_years") or "").strip()
    join_years = (
        [x.strip() for x in join_years_raw.split(",") if x.strip()] if join_years_raw else None
    )
    limit = _safe_int(request.GET.get("limit"), 150)
    strict_mode = (request.GET.get("mode") or "").strip().lower() == "strict"

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    payload = build_recommendation_debug_report(
        year, semester, section, program, join_years, limit, strict_passed_only=strict_mode
    )

    from core.services.debug_export import export_recommendation_debug_xlsx

    path = export_recommendation_debug_xlsx(payload)
    prog_label = program or "all"
    mode_suffix = "_strict" if strict_mode else ""
    filename = f"recommendation_debug_{prog_label}_{year}_T{semester}{mode_suffix}.xlsx"
    response = FileResponse(
        open(path, "rb"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@role_required(ROLE_ADVISOR)
@require_GET
def export_course_eligibility_csv_view(request: HttpRequest) -> HttpResponseBase:
    """Export course eligibility as a styled XLSX with multiple sheets."""
    course_code = (request.GET.get("course_code") or "").strip().upper()
    if not course_code:
        return JsonResponse({"error": "course_code is required"}, status=400)

    section = (request.GET.get("section") or "").strip().upper() or None
    program = (request.GET.get("program") or "").strip().upper() or None
    join_years_raw = (request.GET.get("join_years") or "").strip()
    join_years = (
        [x.strip() for x in join_years_raw.split(",") if x.strip()] if join_years_raw else None
    )
    mode, err = _parse_recommendation_mode(request.GET.get("mode"))
    if err:
        return err
    if mode is None:
        return JsonResponse({"error": "Invalid mode"}, status=400)
    strict_mode = mode == "strict"

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    payload = build_course_eligibility_report(
        course_code, section, program, join_years, strict_mode
    )

    from core.services.eligibility_export import export_eligibility_xlsx

    path = export_eligibility_xlsx(payload)
    filename = f"eligibility_{course_code}_{mode}.xlsx"
    response = FileResponse(path.open("rb"), as_attachment=True, filename=filename)
    response["X-Recommendation-Mode"] = mode
    return response


@role_required(ROLE_ADVISOR)
@require_GET
def missing_high_priority_view(request: HttpRequest) -> JsonResponse:
    year, err = _parse_int(request.GET.get("year"), "year")
    if err:
        return err
    semester, err = _parse_int(request.GET.get("semester"), "semester")
    if err:
        return err
    if year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)

    if semester not in {1, 2}:
        return JsonResponse({"error": "semester must be 1 or 2"}, status=400)

    section = (request.GET.get("section") or "").strip().upper() or None
    program = (request.GET.get("program") or "").strip().upper() or None
    join_years_raw = (request.GET.get("join_years") or "").strip()
    join_years = (
        [x.strip() for x in join_years_raw.split(",") if x.strip()] if join_years_raw else None
    )

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    # Main semesters are the source of truth for study-plan parity: semester
    # 1 serves odd plan terms and semester 2 serves even plan terms.  Keeping a
    # second, independently editable value allowed the HP report to disagree
    # with the global term selected by the operator.
    term_parity = semester - 1
    discount = (request.GET.get("discount") or "1_over_d").strip()
    min_score = _safe_float(request.GET.get("min_score"), 2.0)
    top_k = _safe_int(request.GET.get("top_k"), 10)
    studying_counts = (request.GET.get("studying_counts_as_passed") or "false").strip().lower() in {
        "1",
        "true",
        "yes",
        "y",
    }

    payload = run_missing_high_priority_report(
        year=year,
        semester=semester,
        section=section,
        program=program,
        join_year_prefixes=join_years,
        term_parity=term_parity,
        discount=discount,
        min_score=min_score,
        top_k_per_student=top_k,
        studying_counts_as_passed=studying_counts,
    )
    return JsonResponse(payload)


# ROLE_ADVISOR, matching the JSON endpoint this screen already reads. An adviser
# could see every row on screen and was refused the same rows as a file — and
# because #apCsvLink is an <a href>, clicking it navigated the whole tab to a raw
# 403 JSON blob.
#
# Only safe alongside `resolve_roster_scope`: before it, this view's sole control
# was this decorator plus a falsy-guarded scope check that a blank advisor_id
# skipped entirely. Relaxing the guard on top of that would have turned a 403 into
# a data leak. The resolver refuses an unscoped caller and answers a cross-advisor
# request with 403 rather than an empty file.
@role_required(ROLE_ADVISOR)
@require_GET
def export_students_by_advisor_csv_view(request: HttpRequest) -> HttpResponse:
    advisor_id = (request.GET.get("advisor_id") or "").strip()
    if not advisor_id:
        return JsonResponse({"error": "advisor_id is required"}, status=400)

    search = (request.GET.get("search") or "").strip() or None
    focus = (request.GET.get("focus") or "").strip() or None
    program_filter = (request.GET.get("program_filter") or "").strip() or None

    forced_advisor_id, allowed_departments, scope_error = resolve_roster_scope(
        request.user, advisor_id
    )
    if scope_error:
        # An explicit refusal. This view used to answer a scope mismatch with a 200
        # carrying nothing but the BOM and a header row, because the check below
        # reads only `mapping_ready` and discards `payload["error"]` — so "you may
        # not" arrived looking exactly like "this adviser has no students".
        return JsonResponse({"error": scope_error}, status=403)

    payload = list_students_by_advisor(
        advisor_id,
        search=search,
        focus=focus,
        program_filter=program_filter,
        forced_advisor_id=forced_advisor_id,
        allowed_departments=allowed_departments,
    )
    if payload.get("mapping_ready") is False:
        return JsonResponse(
            {"error": payload.get("message", "students.advisor_id column is not added yet.")},
            status=400,
        )

    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(
        [
            "advisor_id",
            "student_id",
            "registration_no",
            "name",
            "program",
            "section",
            "status",
            "gpa",
            "total_earned_credits",
            "total_registered_credits",
            "current_term_registered_hours",
            "has_high_priority_missing",
            "needs_attention",
            "risk_score",
            "attention_reasons",
            "missing_courses_compact",
        ]
    )
    for row in payload.get("items", []):
        writer.writerow(
            [
                advisor_id,
                row.get("student_id"),
                row.get("registration_no", ""),
                row.get("name", ""),
                row.get("program", ""),
                row.get("section", ""),
                row.get("status", ""),
                row.get("gpa", ""),
                row.get("total_earned_credits", ""),
                row.get("total_registered_credits", ""),
                row.get("current_term_registered_hours", ""),
                row.get("has_high_priority_missing", ""),
                row.get("needs_attention", ""),
                row.get("risk_score", ""),
                ",".join(row.get("attention_reasons", [])),
                row.get("missing_courses_compact", ""),
            ]
        )

    return _excel_csv_response(f"students_by_advisor_{advisor_id}.csv", out.getvalue())


@role_required(ROLE_ADVISOR)
@require_GET
def export_missing_high_priority_xlsx_view(request: HttpRequest) -> HttpResponseBase:
    year, err = _parse_int(request.GET.get("year"), "year")
    if err:
        return err
    semester, err = _parse_int(request.GET.get("semester"), "semester")
    if err:
        return err
    if year is None or semester is None:
        return JsonResponse({"error": "Invalid parameters"}, status=400)
    if semester not in {1, 2}:
        return JsonResponse({"error": "semester must be 1 or 2"}, status=400)

    section = (request.GET.get("section") or "").strip().upper() or None
    program = (request.GET.get("program") or "").strip().upper() or None
    join_years_raw = (request.GET.get("join_years") or "").strip()
    join_years = (
        [x.strip() for x in join_years_raw.split(",") if x.strip()] if join_years_raw else None
    )

    scope_err = require_program_scope(request, program)
    if scope_err:
        return scope_err

    term_parity = semester - 1
    discount = (request.GET.get("discount") or "1_over_d").strip()
    min_score = _safe_float(request.GET.get("min_score"), 2.0)
    top_k = _safe_int(request.GET.get("top_k"), 10)
    studying_counts = (request.GET.get("studying_counts_as_passed") or "false").strip().lower() in {
        "1",
        "true",
        "yes",
        "y",
    }

    try:
        path = export_missing_high_priority_xlsx(
            year=year,
            semester=semester,
            section=section,
            program=program,
            join_year_prefixes=join_years,
            term_parity=term_parity,
            discount=discount,
            min_score=min_score,
            top_k_per_student=top_k,
            studying_counts_as_passed=studying_counts,
        )
    except RuntimeError as exc:
        return JsonResponse({"error": str(exc)}, status=500)

    return FileResponse(
        path.open("rb"), as_attachment=True, filename="flagged_students_missing_high_priority.xlsx"
    )
