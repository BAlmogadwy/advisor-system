"""
core/services/timetable_per_plan_export.py
Per-plan XLSX export for the Timetable Workspace.

Where ``timetable_export.export_scenario_xlsx`` produces a single workbook
organised by **board nominal_term** (the "as-planned" sittings shared across
programs), this module produces **one workbook per program** organised by
**plan term** (``ProgrammeRequirement.programme_term``).

A scenario whose boards target ``program="DS,AI,AI2,DS2"`` therefore yields
four files (DS / AI / AI2 / DS2).  Each file shows that program's plan
courses laid out under their plan terms, regardless of which board the
planner actually placed them on — so a course like ``AI113`` (DS2 plan
term=5) appearing on the scenario's "Term 3" board still lands in the DS2
file's "Term 5" sheet.

The two outputs are complementary: the planner-facing combined export keeps
the board-as-sitting view, and this per-plan export gives each cohort an
auditable view of their own term-by-term week.

Returns a path to a ``.zip`` archive when the scenario spans multiple
programs, or a single ``.xlsx`` when it spans only one.

Public API
----------
* :func:`export_scenario_per_plan` — main entry point used by the view.
"""

from __future__ import annotations

import re
import tempfile
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from core.models import (
    DeliveryBoard,
    ProgrammeRequirement,
    ScenarioSectionBudget,
    SectionPlacement,
    Student,
    TermSectionMeeting,
    TimetableScenario,
)
from core.services.course_identity import display_course_label, planner_course_key
from core.services.timetable_autoplace import DEFAULT_LAB_SLOTS, DEFAULT_SLOTS
from core.services.timetable_demand import load_scenario_course_demands
from core.services.timetable_plan_lens import build_scenario_plan_lens

DAY_LABELS = ["SUN", "MON", "TUE", "WED", "THU"]
DAY_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]

COURSE_COLORS = [
    "D4E6F1",
    "D5F5E3",
    "FADBD8",
    "FCF3CF",
    "D7BDE2",
    "A9DFBF",
    "F9E79F",
    "AED6F1",
    "F5CBA7",
    "A3E4D7",
    "E8DAEF",
    "FDEBD0",
    "ABB2B9",
    "A2D9CE",
    "F5B7B1",
    "D6DBDF",
    "ABEBC6",
    "FAD7A0",
    "D2B4DE",
    "AEB6BF",
]


def _slugify(value: str) -> str:
    """Filesystem-safe slug for filenames (preserves commas-as-underscores)."""
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_")
    return s or "scenario"


def get_scenario_programs(scenario: TimetableScenario) -> list[str]:
    """Return the unique list of programs targeted by a scenario's boards.

    Boards may carry comma-separated program lists (e.g. ``"DS,AI,AI2,DS2"``).
    Programs are returned in board ``display_order`` then alphabetical order
    of first appearance, so DS,AI,AI2,DS2 stays in the registrar's intended
    sequence rather than shuffling alphabetically.
    """
    seen: list[str] = []
    for b in DeliveryBoard.objects.filter(scenario=scenario).order_by("display_order"):
        for token in (b.program or "").split(","):
            p = token.strip()
            if p and p not in seen:
                seen.append(p)
    return seen


@dataclass(frozen=True)
class ExportFilter:
    """Normalized filter passed from the timetable workspace search box."""

    original: str = ""
    program_tokens: frozenset[str] = frozenset()
    text_tokens: tuple[str, ...] = ()
    shared_only: bool = False


def _norm(value: object | None) -> str:
    return str(value or "").replace("\u00a0", " ").strip()


def _upper(value: object | None) -> str:
    return _norm(value).upper()


def _course_key_from_values(course_code: object | None, course_name: object | None = None) -> str:
    return planner_course_key(course_code, course_name)


def _requirement_course_key(req: ProgrammeRequirement) -> str:
    return _course_key_from_values(req.course_code, req.course_name)


def _budget_course_key(budget: ScenarioSectionBudget) -> str:
    return _norm(budget.course_key) or _course_key_from_values(
        budget.course_code,
        budget.course_name,
    )


def _term_section_course_key(section) -> str:
    if not section:
        return ""
    return _norm(getattr(section, "course_key", "")) or _course_key_from_values(
        getattr(section, "course_code", ""),
        getattr(section, "course_name", ""),
    )


def _course_code_from_key(course_key: str) -> str:
    return _upper(course_key.split("::", 1)[0])


def _name_from_key(course_key: str) -> str:
    if "::" not in course_key:
        return _course_code_from_key(course_key)
    return course_key.split("::", 1)[1].replace("_", " ").strip()


def _course_display_name(course_key: str, course_names: dict[str, str]) -> str:
    return (
        course_names.get(course_key)
        or _name_from_key(course_key)
        or _course_code_from_key(course_key)
    )


def _course_search_text(course_key: str, course_names: dict[str, str]) -> str:
    name = _course_display_name(course_key, course_names)
    return " ".join(
        part
        for part in [
            course_key,
            course_key.replace("::", " "),
            _course_code_from_key(course_key),
            name,
            display_course_label(_course_code_from_key(course_key), name),
        ]
        if part
    ).upper()


def _matches_text_tokens(tokens: tuple[str, ...], text: str) -> bool:
    if not tokens:
        return True
    haystack = _upper(text)
    return all(token in haystack for token in tokens)


def _build_export_filter(search: object | None, programs: list[str]) -> ExportFilter:
    raw = _norm(search)
    if not raw:
        return ExportFilter()
    program_set = {_upper(program) for program in programs}
    program_tokens: set[str] = set()
    text_tokens: list[str] = []
    shared_only = False
    for token in re.split(r"\s+", raw.upper()):
        token = token.strip()
        if not token:
            continue
        if token == "SHARED":  # nosec B105 — search-filter sentinel, not a secret
            shared_only = True
        elif token in program_set:
            program_tokens.add(token)
        else:
            text_tokens.append(token)
    return ExportFilter(
        original=raw,
        program_tokens=frozenset(program_tokens),
        text_tokens=tuple(text_tokens),
        shared_only=shared_only,
    )


def _section_matches_plan_filter(
    term_section_id: int | None,
    program: str,
    plan_lens: dict[str, Any],
    export_filter: ExportFilter,
) -> bool:
    section_lens = (plan_lens.get("sections") or {}).get(str(term_section_id or ""))
    if not section_lens:
        return True
    if export_filter.shared_only and not section_lens.get("shared"):
        return False
    filter_plans = {_upper(plan) for plan in section_lens.get("filter_plans") or []}
    return _upper(program) in filter_plans if filter_plans else True


def _plan_lens_course_payload(plan_lens: dict[str, Any], course_key: str) -> dict[str, Any] | None:
    """Return the Plan Lens course payload without collapsing duplicate codes."""
    courses = plan_lens.get("courses") or {}
    if course_key in courses and isinstance(courses[course_key], dict):
        return courses[course_key]

    # Legacy fallback only when the visible code is unambiguous. This keeps
    # CS111::PROGRAMMING_I distinct from other CS111 identities.
    code = _course_code_from_key(course_key)
    matches = [
        payload
        for key, payload in courses.items()
        if _course_code_from_key(str(key)) == code and isinstance(payload, dict)
    ]
    return matches[0] if len(matches) == 1 else None


def _plan_lens_program_demand(
    *,
    plan_lens: dict[str, Any],
    course_key: str,
    program: str,
    fallback: int,
) -> int:
    course_lens = _plan_lens_course_payload(plan_lens, course_key)
    if not course_lens:
        return int(fallback or 0)
    plans = course_lens.get("plans") or {}
    program_key = _upper(program)
    for plan, value in plans.items():
        if _upper(plan) == program_key:
            try:
                return int(value or 0)
            except (TypeError, ValueError):
                return 0
    return 0


def _plan_lens_visible_section_count(
    *,
    plan_lens: dict[str, Any],
    course_key: str,
    program: str,
    export_filter: ExportFilter,
    fallback: int,
) -> int:
    """Count sections allocated to this program in the Plan Lens.

    Shared sections count for a program only when that program is one of the
    shared contributors. Unallocated/extra sections are deliberately excluded
    from the per-plan count, because the workbook is meant to read as if the
    scenario were filtered to that plan.
    """
    sections = plan_lens.get("sections") or {}
    if not sections:
        return int(fallback or 0)

    program_key = _upper(program)
    count = 0
    saw_course = False
    for _raw_id, section_lens in sections.items():
        if not isinstance(section_lens, dict):
            continue
        if _norm(section_lens.get("course_key")) != course_key:
            continue
        saw_course = True
        if export_filter.shared_only and not section_lens.get("shared"):
            continue
        filter_plans = {_upper(plan) for plan in section_lens.get("filter_plans") or []}
        if program_key in filter_plans:
            count += 1

    if saw_course:
        return count
    return int(fallback or 0)


def _placement_search_text(
    placement: SectionPlacement,
    course_key: str,
    course_names: dict[str, str],
) -> str:
    section = placement.term_section
    parts = [
        _course_search_text(course_key, course_names),
        getattr(section, "section", ""),
        placement.day,
        placement.start_time,
        placement.end_time,
        placement.room,
    ]
    return " ".join(str(part) for part in parts if part).upper()


def _section_cell_label(section, duplicate_course_codes: set[str]) -> str:
    if not section:
        return "?"
    text = f"{section.course_code} {section.section}"
    if _upper(section.course_code) in duplicate_course_codes and _norm(section.course_name):
        text += f"\n{section.course_name}"
    return text


def get_plan_term_map(program: str) -> dict[str, int]:
    """Return ``{course_code -> programme_term}`` for one program's plan.

    Course codes are upper-cased so that comparisons against
    ``term_section.course_code`` are case-insensitive.  Duplicate rows in the
    requirement table (the same course appearing more than once for the same
    program) keep the smallest term — a defensive choice that lands the
    course earliest in the plan view.
    """
    out: dict[str, int] = {}
    for r in ProgrammeRequirement.objects.filter(program=program):
        if not r.course_code or r.programme_term is None:
            continue
        course_key = _requirement_course_key(r)
        if course_key not in out or (r.programme_term < out[course_key]):
            out[course_key] = r.programme_term
    return out


def _classify_placement_term(
    course_key: str, board_term: int | None, plan_map: dict[str, int]
) -> tuple[int | None, str]:
    """Decide which sheet a placement belongs to and tag its plan-vs-board status.

    Returns ``(plan_term, status)`` where ``status`` is one of:

    * ``"matched"`` — the course is in this program's plan AND the board's
      ``nominal_term`` matches the plan term (correctly placed).
    * ``"misplaced"`` — the course is in the plan but the board term differs
      (e.g. AI113 in DS2's plan at term 5 but on the scenario's term-3 board).
    * ``"out-of-plan"`` — the course is not in this program's plan at all
      (will be skipped when filtering, but kept for the audit sheet).
    """
    key = _norm(course_key)
    plan_term = plan_map.get(key)
    if plan_term is None:
        return None, "out-of-plan"
    if board_term is None or board_term == plan_term:
        return plan_term, "matched"
    return plan_term, "misplaced"


def export_scenario_per_plan(
    scenario_id: int,
    *,
    search: object | None = None,
) -> tuple[Path, str, bool]:
    """Build per-program XLSX files for a scenario.

    Parameters
    ----------
    scenario_id : int
        The :class:`TimetableScenario` to export.

    Returns
    -------
    (Path, str, bool)
        ``(output_path, suggested_filename, is_zip)``.  ``is_zip`` is
        ``True`` when more than one program was found and files were bundled
        into a zip archive; ``False`` when a single ``.xlsx`` is returned
        directly.

    Raises
    ------
    RuntimeError
        If ``openpyxl`` is not installed or no programs are found.
    TimetableScenario.DoesNotExist
        If the scenario id is invalid.
    """
    try:
        from openpyxl import Workbook  # noqa: F401  (import probe)
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    scenario = TimetableScenario.objects.get(id=scenario_id)
    programs = get_scenario_programs(scenario)
    if not programs:
        raise RuntimeError("Scenario has no boards with a configured program")

    export_filter = _build_export_filter(search, programs)
    selected_programs = [
        program
        for program in programs
        if not export_filter.program_tokens or _upper(program) in export_filter.program_tokens
    ]
    if not selected_programs:
        raise RuntimeError("The export filter did not match any scenario program")

    plan_lens = build_scenario_plan_lens(scenario_id)
    safe_name = _slugify(scenario.name)
    filter_suffix = (
        f"__filtered_{_slugify(export_filter.original)}" if export_filter.original else ""
    )

    workbook_paths: list[tuple[str, Path]] = []  # (filename, on-disk path)
    for prog in selected_programs:
        plan_map = get_plan_term_map(prog)
        wb_path = _build_program_workbook(
            scenario,
            prog,
            plan_map,
            plan_lens=plan_lens,
            export_filter=export_filter,
        )
        filename = f"{_slugify(prog)}.xlsx"
        workbook_paths.append((filename, wb_path))

    if len(workbook_paths) == 1:
        only_filename, only_path = workbook_paths[0]
        return only_path, only_filename, False

    # Bundle into a zip
    tmp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
    tmp.close()
    zip_path = Path(tmp.name)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, path in workbook_paths:
            zf.write(path, arcname=filename)

    # Clean up individual workbook tempfiles after they're inside the zip
    for _, path in workbook_paths:
        try:
            path.unlink()
        except OSError:
            pass

    suggested = f"{safe_name}{filter_suffix}__per_plan.zip"
    return zip_path, suggested, True


# ── workbook builder ─────────────────────────────────────────────────────────


def _build_program_workbook(
    scenario: TimetableScenario,
    program: str,
    plan_map: dict[str, int],
    *,
    plan_lens: dict[str, Any],
    export_filter: ExportFilter,
) -> Path:
    """Produce one XLSX file for a single program's plan view."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # ── Styles ───────────────────────────────────────────────────────────
    hdr_fill = PatternFill(start_color="0A8E6E", end_color="0A8E6E", fill_type="solid")
    hdr_font = Font(name="Consolas", size=9, bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="CCCCCC")
    thick_side = Side(style="medium", color="333333")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    bold_font = Font(bold=True, size=9)
    normal_font = Font(name="Consolas", size=9)
    center_align = Alignment(horizontal="center", vertical="center")
    conflict_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    board_hdr_fill = PatternFill(start_color="111144", end_color="111144", fill_type="solid")
    board_hdr_font = Font(name="Consolas", size=10, bold=True, color="FFFFFF")
    slot_fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
    info_hdr_fill = PatternFill(start_color="4056E3", end_color="4056E3", fill_type="solid")
    info_hdr_font = Font(name="Consolas", size=9, bold=True, color="FFFFFF")
    misplaced_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    missing_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    matched_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

    course_color_map: dict[str, str] = {}

    def _course_fill(code: str) -> PatternFill:
        if code not in course_color_map:
            idx = len(course_color_map) % len(COURSE_COLORS)
            course_color_map[code] = COURSE_COLORS[idx]
        hex_c = course_color_map[code]
        return PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")

    def _apply_outer_border(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                ex = cell.border
                cell.border = Border(
                    left=thick_side if c == min_col else ex.left,
                    right=thick_side if c == max_col else ex.right,
                    top=thick_side if r == min_row else ex.top,
                    bottom=thick_side if r == max_row else ex.bottom,
                )

    # ── Data prep ─────────────────────────────────────────────────────────
    slot_config = scenario.slot_config or DEFAULT_SLOTS
    lab_slot_config = scenario.lab_slot_config or DEFAULT_LAB_SLOTS  # noqa: F841

    boards = list(
        DeliveryBoard.objects.filter(scenario=scenario).order_by("display_order", "label")
    )
    board_term_by_id = {b.id: (b.nominal_term or 0) for b in boards}

    placements = list(
        SectionPlacement.objects.filter(board__scenario=scenario)
        .select_related("term_section")
        .order_by("day", "start_time")
    )
    budget_rows = list(ScenarioSectionBudget.objects.filter(scenario=scenario))

    # Build the budget lookup keyed by the same course identity used by the
    # planner, not the display code. This prevents CS111 variants collapsing.
    budget_by_key: dict[str, ScenarioSectionBudget] = {}
    for b in budget_rows:
        key = _budget_course_key(b)
        existing = budget_by_key.get(key)
        plan_term = plan_map.get(key)
        if existing is None:
            budget_by_key[key] = b
        elif plan_term is not None and (b.programme_term or 0) == plan_term:
            budget_by_key[key] = b

    # Course descriptions by planner identity, with code-only catalogue names
    # retained only as a fallback.
    from core.models import Course, TermSection

    course_names_by_code: dict[str, str] = {
        _upper(c.course_code): c.description or c.course_code for c in Course.objects.all()
    }
    course_names: dict[str, str] = {}
    for r in ProgrammeRequirement.objects.filter(program=program):
        key = _requirement_course_key(r)
        course_names[key] = r.course_name or course_names_by_code.get(
            _upper(r.course_code), r.course_code
        )
    for b in budget_rows:
        key = _budget_course_key(b)
        course_names.setdefault(
            key,
            b.course_name or course_names_by_code.get(_upper(b.course_code), b.course_code),
        )
    for ts in TermSection.objects.filter(scenario=scenario):
        key = _term_section_course_key(ts)
        if key:
            course_names.setdefault(
                key,
                ts.course_name or course_names_by_code.get(_upper(ts.course_code), ts.course_code),
            )

    code_keys: dict[str, set[str]] = defaultdict(set)
    for key in set(plan_map) | set(budget_by_key) | set(course_names):
        code_keys[_course_code_from_key(key)].add(key)
    duplicate_course_codes = {code for code, keys in code_keys.items() if len(keys) > 1}

    text_matched_keys = {
        key
        for key in plan_map
        if _matches_text_tokens(export_filter.text_tokens, _course_search_text(key, course_names))
    }

    # Filter to placements whose identity is in this program's plan and whose
    # section belongs to the same Plan Lens used on the screen.
    in_plan_placements: list[SectionPlacement] = []
    out_of_plan_codes: set[str] = set()
    course_status: dict[str, str] = {}
    course_board_terms: dict[str, set[int]] = defaultdict(set)
    ranking = {"matched": 0, "misplaced": 1, "missing": 2}
    for p in placements:
        ts = p.term_section
        key = _term_section_course_key(ts)
        board_term = board_term_by_id.get(p.board_id)
        placement_matches_text = _matches_text_tokens(
            export_filter.text_tokens,
            _placement_search_text(p, key, course_names),
        )
        plan_term, status = _classify_placement_term(key, board_term, plan_map)
        if plan_term is None:
            if not _section_matches_plan_filter(
                p.term_section_id, program, plan_lens, export_filter
            ):
                continue
            if not export_filter.text_tokens or placement_matches_text:
                out_of_plan_codes.add(key)
                course_board_terms[key].add(board_term or 0)
            continue
        if not _section_matches_plan_filter(p.term_section_id, program, plan_lens, export_filter):
            continue
        if export_filter.text_tokens and not (placement_matches_text or key in text_matched_keys):
            continue
        text_matched_keys.add(key)
        in_plan_placements.append(p)
        course_board_terms[key].add(board_term or 0)
        prev = course_status.get(key)
        if prev is None or ranking.get(status, 0) > ranking.get(prev, 0):
            course_status[key] = status

    visible_plan_map = (
        {key: term for key, term in plan_map.items() if key in text_matched_keys}
        if export_filter.text_tokens
        else dict(plan_map)
    )

    # Group filtered placements by plan_term
    by_plan_term: dict[int, list[SectionPlacement]] = defaultdict(list)
    for p in in_plan_placements:
        key = _term_section_course_key(p.term_section)
        plan_term = visible_plan_map.get(key)
        if plan_term is not None:
            by_plan_term[plan_term].append(p)

    # ── Build workbook ────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # Cover/audit sheet first
    _write_plan_coverage_sheet(
        wb=wb,
        scenario=scenario,
        program=program,
        plan_map=visible_plan_map,
        course_status=course_status,
        course_board_terms=course_board_terms,
        course_names=course_names,
        budget_by_key=budget_by_key,
        out_of_plan_codes=out_of_plan_codes,
        styles=dict(
            hdr_fill=hdr_fill,
            hdr_font=hdr_font,
            hdr_align=hdr_align,
            thin_border=thin_border,
            normal_font=normal_font,
            bold_font=bold_font,
            center_align=center_align,
            matched_fill=matched_fill,
            misplaced_fill=misplaced_fill,
            missing_fill=missing_fill,
            board_hdr_fill=board_hdr_fill,
            board_hdr_font=board_hdr_font,
        ),
        apply_outer_border=_apply_outer_border,
    )

    # One sheet per plan term (only terms that have *any* plan course, even
    # if the course is missing — so the registrar can see empty term sheets
    # as a visual gap)
    plan_terms_present = sorted(
        term_num for term_num, term_placements in by_plan_term.items() if term_placements
    )
    for term_num in plan_terms_present:
        ws = wb.create_sheet(title=f"Term {term_num}")
        term_placements = by_plan_term.get(term_num, [])
        _render_plan_term_sheet(
            ws=ws,
            scenario=scenario,
            program=program,
            term_num=term_num,
            plan_map=visible_plan_map,
            placements=term_placements,
            slot_config=slot_config,
            plan_lens=plan_lens,
            budget_by_key=budget_by_key,
            course_names=course_names,
            course_status=course_status,
            course_board_terms=course_board_terms,
            board_term_by_id=board_term_by_id,
            duplicate_course_codes=duplicate_course_codes,
            export_filter=export_filter,
            styles=dict(
                hdr_fill=hdr_fill,
                hdr_font=hdr_font,
                hdr_align=hdr_align,
                thin_border=thin_border,
                bold_font=bold_font,
                normal_font=normal_font,
                center_align=center_align,
                conflict_fill=conflict_fill,
                board_hdr_fill=board_hdr_fill,
                board_hdr_font=board_hdr_font,
                slot_fill=slot_fill,
                info_hdr_fill=info_hdr_fill,
                info_hdr_font=info_hdr_font,
                misplaced_fill=misplaced_fill,
                missing_fill=missing_fill,
            ),
            course_fill=_course_fill,
            apply_outer_border=_apply_outer_border,
        )

    _render_plan_rooms_sheet(
        wb=wb,
        placements=placements,
        lecture_slots=slot_config,
        lab_slots=scenario.lab_slot_config or DEFAULT_LAB_SLOTS,
        course_fill=_course_fill,
    )

    # Instructor timetables (each assigned instructor's full weekly grid),
    # mirroring the Rooms sheet. Reuses the single-workbook renderer.
    from core.services.timetable_export import _render_instructors_sheet

    _render_instructors_sheet(wb, scenario, _course_fill)

    # Save
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


# ── per-term sheet renderer ──────────────────────────────────────────────────


def _render_plan_rooms_sheet(
    *,
    wb,
    placements: list[SectionPlacement],
    lecture_slots: list[dict],
    lab_slots: list[dict],
    course_fill,
) -> None:
    """Add a room-centric schedule sheet to a per-plan workbook."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from core.models import Room

    roomed_placements = [p for p in placements if _norm(p.room) and _upper(p.room) != "UNASSIGNED"]
    if not roomed_placements:
        return

    used_room_codes = sorted({_norm(p.room) for p in roomed_placements})
    room_info = {
        _norm(room.room_code): room
        for room in Room.objects.filter(room_code__in=used_room_codes).order_by(
            "room_type", "room_code", "section"
        )
    }

    def _room_type(room_code: str) -> str:
        room = room_info.get(room_code)
        return _norm(getattr(room, "room_type", "")) or "lecture"

    lecture_room_codes = [code for code in used_room_codes if _room_type(code).lower() != "lab"]
    lab_room_codes = [code for code in used_room_codes if _room_type(code).lower() == "lab"]

    room_grid: dict[str, dict[tuple[str, str], list[str]]] = defaultdict(lambda: defaultdict(list))
    for p in roomed_placements:
        ts = p.term_section
        label = _section_cell_label(ts, set()) if ts else "Section"
        room_grid[_norm(p.room)][(_upper(p.day)[:3], _norm(p.start_time))].append(label)

    ws = wb.create_sheet(title="Rooms")
    ws.sheet_properties.tabColor = "2E86C1"

    room_hdr_fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
    room_hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    room_cell_font = Font(name="Consolas", size=8.5, bold=True)
    room_border = Border(
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
    )
    section_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    midday_fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
    lab_room_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
    room_name_fill = PatternFill(start_color="0A8E6E", end_color="0A8E6E", fill_type="solid")
    room_name_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    def _slot_rows(slots: list[dict]) -> list[dict]:
        return [{"start": _norm(s.get("start")), "end": _norm(s.get("end"))} for s in slots]

    lecture_slot_list = _slot_rows(lecture_slots or DEFAULT_SLOTS)
    lab_slot_list = _slot_rows(lab_slots or lecture_slots or DEFAULT_SLOTS)
    widest_slots = max(len(lecture_slot_list), len(lab_slot_list), 1)

    def _section_header(row: int, title: str, slot_count: int) -> int:
        end_col = 1 + max(slot_count, 1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=14)
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(2, end_col + 1):
            ws.cell(row=row, column=col).fill = section_fill
        return row + 2

    def _write_room_table(row_start: int, room_code: str, slots: list[dict]) -> int:
        room = room_info.get(room_code)
        capacity = getattr(room, "capacity", "?") if room else "?"
        is_lab = _room_type(room_code).lower() == "lab"
        row = row_start
        end_col = 1 + max(len(slots), 1)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=1, value=f"{room_code} ({capacity})")
        cell.font = room_name_font
        cell.fill = lab_room_fill if is_lab else room_name_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, end_col + 1):
            ws.cell(row=row, column=col).fill = lab_room_fill if is_lab else room_name_fill
        row += 1

        ws.cell(row=row, column=1).border = room_border
        for idx, slot in enumerate(slots):
            cell = ws.cell(row=row, column=2 + idx, value=f"{slot['start']}-{slot['end']}")
            cell.font = room_hdr_font
            cell.fill = room_hdr_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = room_border
        row += 1

        for day_code, day_name in zip(DAY_LABELS, DAY_NAMES, strict=False):
            day_cell = ws.cell(row=row, column=1, value=day_name)
            day_cell.font = Font(bold=True, size=9)
            day_cell.border = room_border
            for idx, slot in enumerate(slots):
                cell = ws.cell(row=row, column=2 + idx)
                cell.border = room_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if "11:35" <= slot["start"] <= "12:59":
                    cell.fill = midday_fill
                texts = room_grid.get(room_code, {}).get((day_code, slot["start"]), [])
                if texts:
                    cell.value = "\n".join(texts)
                    cell.font = room_cell_font
                    cell.fill = course_fill(texts[0].split()[0])
                    if len(texts) > 1:
                        cell.fill = PatternFill(
                            start_color="FADBD8", end_color="FADBD8", fill_type="solid"
                        )
            row += 1
        return row + 1

    current_row = _section_header(1, "Lectures", len(lecture_slot_list))
    for room_code in lecture_room_codes:
        current_row = _write_room_table(current_row, room_code, lecture_slot_list)

    if lab_room_codes:
        current_row += 1
        current_row = _section_header(current_row, "Labs", len(lab_slot_list))
        for room_code in lab_room_codes:
            current_row = _write_room_table(current_row, room_code, lab_slot_list)

    ws.column_dimensions["A"].width = 14
    for idx in range(widest_slots):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 15


def _render_plan_term_sheet(
    *,
    ws,
    scenario: TimetableScenario,
    program: str,
    term_num: int,
    plan_map: dict[str, int],
    placements: list[SectionPlacement],
    slot_config: list[dict],
    plan_lens: dict[str, Any],
    budget_by_key: dict[str, ScenarioSectionBudget],
    course_names: dict[str, str],
    course_status: dict[str, str],
    course_board_terms: dict[str, set[int]],
    board_term_by_id: dict[int, int],
    duplicate_course_codes: set[str],
    export_filter: ExportFilter,
    styles: dict,
    course_fill,
    apply_outer_border,
) -> None:
    """Render the day×slot grid + course info sidebar for one plan term."""
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Alignment, Font

    hdr_fill = styles["hdr_fill"]
    hdr_font = styles["hdr_font"]
    hdr_align = styles["hdr_align"]
    thin_border = styles["thin_border"]
    bold_font = styles["bold_font"]
    normal_font = styles["normal_font"]
    center_align = styles["center_align"]
    conflict_fill = styles["conflict_fill"]
    board_hdr_fill = styles["board_hdr_fill"]
    board_hdr_font = styles["board_hdr_font"]
    slot_fill = styles["slot_fill"]
    info_hdr_fill = styles["info_hdr_fill"]
    info_hdr_font = styles["info_hdr_font"]
    misplaced_fill = styles["misplaced_fill"]

    num_slots = len(slot_config)
    INFO_START_COL = 2 + num_slots + 1  # after slots + 1 gap

    # ── Header row ────────────────────────────────────────────────────────
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + num_slots)
    cell = ws.cell(row=row, column=1)
    cell.value = f"{program} — Term {term_num} ({len(placements)} placements)"
    cell.fill = board_hdr_fill
    cell.font = board_hdr_font
    cell.alignment = center_align
    for c in range(1, 2 + num_slots):
        ws.cell(row=row, column=c).fill = board_hdr_fill
    row += 1

    # ── If empty term, render an empty grid + plan-list sidebar ──────────
    # (still useful: shows planned courses with status badges)
    # Build sections-per-course map from this term's placements
    section_all_placements: dict[int, list[SectionPlacement]] = defaultdict(list)
    for p in placements:
        section_all_placements[p.term_section_id].append(p)

    # course_key -> {section_label: [SectionPlacement]} (representative per section)
    seen_sections: dict[str, dict[str, list[SectionPlacement]]] = defaultdict(dict)
    for p in placements:
        code = _term_section_course_key(p.term_section) or "?"
        sec = (p.term_section.section if p.term_section else "?") or "?"
        seen_sections[code].setdefault(sec, []).append(p)

    course_sections: dict[str, list[SectionPlacement]] = {
        code: [secs[s][0] for s in sorted(secs.keys())] for code, secs in seen_sections.items()
    }
    max_groups = max((len(secs) for secs in course_sections.values()), default=1)

    # Bitmask-based smart group split (clone of timetable_export logic) ──
    from core.services.timetable_workspace import _time_mask

    def _section_bitmask(ts_id: int) -> int:
        mask = 0
        for pp in section_all_placements.get(ts_id, []):  # noqa: B023
            mask |= _time_mask(pp.day, pp.start_time, pp.end_time)
        return mask

    course_options: dict[str, list[tuple]] = {
        code: [(s, _section_bitmask(s.term_section_id)) for s in secs]
        for code, secs in course_sections.items()
    }

    group_assignments: list[list[SectionPlacement]] = []
    used_sections: dict[str, set[int]] = defaultdict(set)
    for _g in range(max_groups):
        group_mask = 0
        group: list[SectionPlacement] = []
        sorted_codes = sorted(course_options.keys(), key=lambda c: -len(course_options[c]))
        for code in sorted_codes:
            options = course_options[code]
            best_idx = None
            best_conflicts = float("inf")
            for idx, (_sp, sec_mask) in enumerate(options):
                if idx in used_sections[code]:
                    continue
                overlap = bin(group_mask & sec_mask).count("1")
                if overlap < best_conflicts:
                    best_conflicts = overlap
                    best_idx = idx
            if best_idx is not None:
                sec_p, sec_mask = options[best_idx]
                group.append(sec_p)
                group_mask |= sec_mask
                used_sections[code].add(best_idx)
        group_assignments.append(group)

    # ── Render each group's grid ─────────────────────────────────────────
    def _tm(t: str) -> int:
        return int(t.split(":")[0]) * 60 + int(t.split(":")[1])

    def _is_lab_p(pp) -> bool:
        return (_tm(pp.end_time) - _tm(pp.start_time)) > 80

    def _best_lecture_slot(pp) -> int:
        lab_s, lab_e = _tm(pp.start_time), _tm(pp.end_time)
        best_idx, best_overlap = 0, -1
        for i, slot in enumerate(slot_config):  # noqa: B023
            s_s, s_e = _tm(slot["start"]), _tm(slot["end"])
            overlap = max(0, min(lab_e, s_e) - max(lab_s, s_s))
            if overlap > best_overlap:
                best_overlap = overlap
                best_idx = i
        return best_idx

    # If no placements, we still draw the empty grid header so the sheet
    # isn't blank. Skip group rendering loop.
    if not placements:
        # Day × slot empty grid
        grid_start_row = row
        ws.cell(row=row, column=1).border = thin_border
        for s_idx, _slot in enumerate(slot_config):
            cell = ws.cell(row=row, column=2 + s_idx)
            cell.value = s_idx + 1
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = thin_border
        row += 1
        ws.cell(row=row, column=1).border = thin_border
        for s_idx, slot in enumerate(slot_config):
            cell = ws.cell(row=row, column=2 + s_idx)
            cell.value = f"{slot['start']}-{slot['end']}"
            cell.fill = hdr_fill
            cell.font = Font(name="Consolas", size=8, bold=True, color="FFFFFF")
            cell.alignment = hdr_align
            cell.border = thin_border
        row += 1
        for _day_code, day_name in zip(DAY_LABELS, DAY_NAMES, strict=False):
            cell = ws.cell(row=row, column=1)
            cell.value = day_name
            cell.font = bold_font
            cell.fill = slot_fill
            cell.border = thin_border
            cell.alignment = center_align
            for s_idx in range(num_slots):
                cc = ws.cell(row=row, column=2 + s_idx)
                cc.border = thin_border
            row += 1
        apply_outer_border(ws, grid_start_row, 1, row - 1, 1 + num_slots)
        # Empty-plan note
        row += 1
        ws.cell(
            row=row, column=1, value="No placements for this plan term in the current scenario."
        ).font = Font(italic=True, color="888888", size=9)
    else:
        for group_idx, group_placements in enumerate(group_assignments):
            if not group_placements:
                continue
            if group_idx > 0:
                row += 1

            if max_groups > 1:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + num_slots)
                gcell = ws.cell(row=row, column=1)
                gcell.value = f"Group {group_idx + 1} — {len(group_placements)} courses"
                gcell.font = Font(bold=True, size=9, color="4056E3")
                gcell.alignment = center_align
                row += 1

            grid_start_row = row
            ws.cell(row=row, column=1).border = thin_border
            for s_idx, _slot in enumerate(slot_config):
                cell = ws.cell(row=row, column=2 + s_idx)
                cell.value = s_idx + 1
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align
                cell.border = thin_border
            row += 1
            ws.cell(row=row, column=1).border = thin_border
            for s_idx, slot in enumerate(slot_config):
                cell = ws.cell(row=row, column=2 + s_idx)
                cell.value = f"{slot['start']}-{slot['end']}"
                cell.fill = hdr_fill
                cell.font = Font(name="Consolas", size=8, bold=True, color="FFFFFF")
                cell.alignment = hdr_align
                cell.border = thin_border
            row += 1

            grid: dict[str, dict[str, dict]] = {}
            for day in DAY_LABELS:
                grid[day] = {}
                for slot in slot_config:
                    sk = f"{slot['start']}-{slot['end']}"
                    grid[day][sk] = {"text": "", "courses": [], "lab_time": "", "misplaced": False}

            for p in group_placements:
                ts_placements = section_all_placements.get(p.term_section_id, [])
                for pp in ts_placements:
                    day = pp.day.upper()[:3]
                    if day not in grid:
                        continue

                    is_lab = _is_lab_p(pp)
                    if is_lab:
                        idx = _best_lecture_slot(pp)
                        slot_key = f"{slot_config[idx]['start']}-{slot_config[idx]['end']}"
                    else:
                        slot_key = f"{pp.start_time}-{pp.end_time}"
                        if slot_key not in grid.get(day, {}):
                            for slot in slot_config:
                                if slot["start"] == pp.start_time:
                                    slot_key = f"{slot['start']}-{slot['end']}"
                                    break

                    ts = pp.term_section
                    code = _term_section_course_key(ts) or "?"
                    is_misplaced = course_status.get(code) == "misplaced"
                    meeting = TermSectionMeeting.objects.filter(term_section=ts, day=day).first()
                    instructor = meeting.instructor if meeting else ""
                    room = pp.room or (meeting.room if meeting else "")
                    text = _section_cell_label(ts, duplicate_course_codes)
                    if is_misplaced:
                        # Show the actual board term beside the course code so
                        # the registrar sees AT A GLANCE that the placement is
                        # off-plan — e.g. "AI113 S1 (on T3)".
                        actual = sorted(course_board_terms.get(code, set()))
                        if actual:
                            text += f" (on T{actual[0]})"
                    if instructor:
                        text += f"\n{instructor}"
                    if room:
                        text += f"\n{room}"

                    cd = grid[day][slot_key]
                    if cd["text"]:
                        cd["text"] += f"\n---\n{text}"
                    else:
                        cd["text"] = text
                    cd["courses"].append(code)
                    if is_lab:
                        cd["lab_time"] = f"{pp.start_time}-{pp.end_time}"
                    if is_misplaced:
                        cd["misplaced"] = True

            for day_code, day_name in zip(DAY_LABELS, DAY_NAMES, strict=False):
                cell = ws.cell(row=row, column=1)
                cell.value = day_name
                cell.font = bold_font
                cell.fill = slot_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                for s_idx, slot in enumerate(slot_config):
                    sk = f"{slot['start']}-{slot['end']}"
                    cell = ws.cell(row=row, column=2 + s_idx)
                    cd = grid.get(day_code, {}).get(
                        sk, {"text": "", "courses": [], "lab_time": "", "misplaced": False}
                    )

                    if cd["lab_time"]:
                        normal_if = InlineFont(sz=8.5, rFont="Consolas")
                        red_if = InlineFont(sz=8, rFont="Consolas", b=True, color="C03030")
                        cell.value = CellRichText(
                            TextBlock(normal_if, cd["text"]),
                            TextBlock(red_if, f"\n{cd['lab_time']}"),
                        )
                    else:
                        cell.value = cd["text"]
                        cell.font = normal_font

                    cell.border = thin_border
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                    if len(cd["courses"]) > 1:
                        cell.fill = conflict_fill
                    elif len(cd["courses"]) == 1:
                        if cd["misplaced"]:
                            cell.fill = misplaced_fill
                        else:
                            cell.fill = course_fill(cd["courses"][0])

                has_lab_in_row = any(
                    grid.get(day_code, {}).get(f"{s['start']}-{s['end']}", {}).get("lab_time")
                    for s in slot_config
                )
                if has_lab_in_row:
                    ws.row_dimensions[row].height = 42
                row += 1

            apply_outer_border(ws, grid_start_row, 1, row - 1, 1 + num_slots)

    # ── Course info sidebar (right side) ─────────────────────────────────
    plan_codes_for_term = sorted(code for code, t in plan_map.items() if t == term_num)
    if plan_codes_for_term:
        info_row = 1
        info_headers = ["Course", "Name", "Cr", "Sec.", "Students", "Plan", "Status"]
        for ci, hdr in enumerate(info_headers):
            cell = ws.cell(row=info_row, column=INFO_START_COL + ci)
            cell.value = hdr
            cell.fill = info_hdr_fill
            cell.font = info_hdr_font
            cell.alignment = hdr_align
            cell.border = thin_border
        info_row += 1

        total_sections = 0
        total_students = 0
        for code in plan_codes_for_term:
            budget = budget_by_key.get(code)
            cr = (budget.credit_hours if budget else None) or 0
            global_sections = (budget.planned_sections if budget else 0) or 0
            global_demand = (budget.total_demand if budget else 0) or 0
            planned_sections = _plan_lens_visible_section_count(
                plan_lens=plan_lens,
                course_key=code,
                program=program,
                export_filter=export_filter,
                fallback=global_sections,
            )
            total_demand = _plan_lens_program_demand(
                plan_lens=plan_lens,
                course_key=code,
                program=program,
                fallback=global_demand,
            )
            status = course_status.get(code, "missing")  # not placed at all → missing
            actual_terms = sorted(course_board_terms.get(code, set()))
            status_text = {
                "matched": "OK",
                "misplaced": f"On T{actual_terms[0]}" if actual_terms else "Off-plan",
                "missing": "Missing",
            }.get(status, status)

            c_cell = ws.cell(row=info_row, column=INFO_START_COL, value=_course_code_from_key(code))
            c_cell.font = Font(bold=True, size=9)
            c_cell.border = thin_border
            if status == "missing":
                c_cell.fill = styles["missing_fill"]
            elif status == "misplaced":
                c_cell.fill = styles["misplaced_fill"]
            else:
                c_cell.fill = course_fill(code)

            name_cell = ws.cell(
                row=info_row,
                column=INFO_START_COL + 1,
                value=_course_display_name(code, course_names),
            )
            name_cell.font = normal_font
            name_cell.border = thin_border

            cr_cell = ws.cell(row=info_row, column=INFO_START_COL + 2, value=cr)
            cr_cell.font = normal_font
            cr_cell.border = thin_border
            cr_cell.alignment = center_align

            sec_cell = ws.cell(row=info_row, column=INFO_START_COL + 3, value=planned_sections)
            sec_cell.font = normal_font
            sec_cell.border = thin_border
            sec_cell.alignment = center_align

            students_cell = ws.cell(row=info_row, column=INFO_START_COL + 4, value=total_demand)
            students_cell.font = normal_font
            students_cell.border = thin_border
            students_cell.alignment = center_align

            plan_cell = ws.cell(row=info_row, column=INFO_START_COL + 5, value=f"T{term_num}")
            plan_cell.font = normal_font
            plan_cell.border = thin_border
            plan_cell.alignment = center_align

            st_cell = ws.cell(row=info_row, column=INFO_START_COL + 6, value=status_text)
            st_cell.font = normal_font
            st_cell.border = thin_border
            st_cell.alignment = center_align
            if status == "missing":
                st_cell.fill = styles["missing_fill"]
            elif status == "misplaced":
                st_cell.fill = styles["misplaced_fill"]

            total_sections += planned_sections
            total_students += total_demand
            info_row += 1

        total_cell = ws.cell(row=info_row, column=INFO_START_COL, value="TOTAL")
        total_cell.font = Font(bold=True, size=9)
        total_cell.border = thin_border
        for ci in range(1, len(info_headers)):
            ws.cell(row=info_row, column=INFO_START_COL + ci).border = thin_border
        ws.cell(row=info_row, column=INFO_START_COL + 3, value=total_sections).font = Font(
            bold=True, size=9
        )
        ws.cell(row=info_row, column=INFO_START_COL + 3).border = thin_border
        ws.cell(row=info_row, column=INFO_START_COL + 3).alignment = center_align
        ws.cell(row=info_row, column=INFO_START_COL + 4, value=total_students).font = Font(
            bold=True, size=9
        )
        ws.cell(row=info_row, column=INFO_START_COL + 4).border = thin_border
        ws.cell(row=info_row, column=INFO_START_COL + 4).alignment = center_align

        apply_outer_border(ws, 1, INFO_START_COL, info_row, INFO_START_COL + 6)

        _write_plan_term_conflict_matrix(
            ws=ws,
            start_row=1,
            start_col=INFO_START_COL + 8,
            scenario=scenario,
            program=program,
            term_num=term_num,
            plan_course_keys=set(plan_codes_for_term),
            course_names=course_names,
            duplicate_course_codes=duplicate_course_codes,
            export_filter=export_filter,
            hdr_fill=hdr_fill,
            thin_border=thin_border,
            center_align=center_align,
        )

    # ── Column widths ────────────────────────────────────────────────────
    from openpyxl.utils import get_column_letter

    ws.column_dimensions["A"].width = 13
    for s_idx in range(num_slots):
        ws.column_dimensions[get_column_letter(2 + s_idx)].width = 18
    # Sidebar: gap, course, name, cr, sections, students, plan, status
    for ci, w in enumerate([3, 12, 30, 6, 6, 10, 6, 12]):
        ws.column_dimensions[get_column_letter(INFO_START_COL + ci - 1)].width = w


# ── plan-coverage / audit sheet ──────────────────────────────────────────────


def _matrix_course_label(
    course_key: str,
    course_names: dict[str, str],
    duplicate_course_codes: set[str],
) -> str:
    code = _course_code_from_key(course_key)
    if code not in duplicate_course_codes:
        return code
    name = _course_display_name(course_key, course_names)
    short = re.sub(r"\s+", " ", name).strip()
    return f"{code}\n{short[:18]}" if short and short != code else code


def _write_plan_term_conflict_matrix(
    *,
    ws,
    start_row: int,
    start_col: int,
    scenario: TimetableScenario,
    program: str,
    term_num: int,
    plan_course_keys: set[str],
    course_names: dict[str, str],
    duplicate_course_codes: set[str],
    export_filter: ExportFilter,
    hdr_fill,
    thin_border,
    center_align,
) -> None:
    """Render a student conflict matrix for this program/plan term."""
    from openpyxl.styles import Border, Font, PatternFill, Side

    demands = [
        row for row in load_scenario_course_demands(scenario.id) if row.primary_term == term_num
    ]
    student_ids = sorted({int(row.student_id) for row in demands})
    program_by_student = {
        int(student_id): _upper(program_name)
        for student_id, program_name in Student.objects.filter(
            student_id__in=student_ids
        ).values_list("student_id", "program")
    }
    target_program = _upper(program)
    rows = [row for row in demands if program_by_student.get(int(row.student_id)) == target_program]

    course_students: dict[str, set[int]] = defaultdict(set)
    for row in rows:
        course_key = _norm(row.course_key)
        if not course_key:
            continue
        if export_filter.text_tokens and not _matches_text_tokens(
            export_filter.text_tokens,
            _course_search_text(course_key, course_names),
        ):
            continue
        course_students[course_key].add(int(row.student_id))

    courses = sorted(
        course_students.keys(),
        key=lambda key: (
            key not in plan_course_keys,
            _course_code_from_key(key),
            _course_display_name(key, course_names),
        ),
    )
    if not courses:
        ws.cell(
            row=start_row,
            column=start_col,
            value=f"Conflict Matrix - {program} Term {term_num}",
        ).font = Font(bold=True, size=10)
        ws.cell(row=start_row + 1, column=start_col, value="No matching student demand.")
        return

    cross_term_courses = {key for key in courses if key not in plan_course_keys}
    native_count = len(courses) - len(cross_term_courses)

    r = start_row
    ws.cell(
        row=r, column=start_col, value=f"Conflict Matrix - {program} Term {term_num}"
    ).font = Font(bold=True, size=10)
    subtitle = f"{len(rows)} students, {native_count} courses"
    if cross_term_courses:
        subtitle += f" + {len(cross_term_courses)} cross-term (*)"
    ws.cell(row=r + 1, column=start_col, value=subtitle).font = Font(size=8, color="666666")
    r += 3

    sets = [course_students[key] for key in courses]
    matrix = [[0] * len(courses) for _ in courses]
    max_val = 0
    for i in range(len(courses)):
        matrix[i][i] = len(sets[i])
        for j in range(i + 1, len(courses)):
            shared = len(sets[i] & sets[j])
            matrix[i][j] = shared
            matrix[j][i] = shared
            max_val = max(max_val, shared)

    cross_hdr_fill = PatternFill(start_color="4056E3", end_color="4056E3", fill_type="solid")
    diag_fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
    zero_font = Font(name="Consolas", size=8, color="CCCCCC")
    header_font = Font(name="Consolas", size=8, bold=True, color="FFFFFF")

    ws.cell(row=r, column=start_col).border = thin_border
    for j, key in enumerate(courses):
        cell = ws.cell(row=r, column=start_col + 1 + j)
        label = _matrix_course_label(key, course_names, duplicate_course_codes)
        cell.value = f"*{label}" if key in cross_term_courses else label
        cell.fill = cross_hdr_fill if key in cross_term_courses else hdr_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    r += 1

    for i, key in enumerate(courses):
        cell = ws.cell(row=r, column=start_col)
        label = _matrix_course_label(key, course_names, duplicate_course_codes)
        cell.value = f"*{label}" if key in cross_term_courses else label
        cell.fill = cross_hdr_fill if key in cross_term_courses else hdr_fill
        cell.font = header_font
        cell.border = thin_border

        for j in range(len(courses)):
            val = matrix[i][j]
            data_cell = ws.cell(row=r, column=start_col + 1 + j)
            data_cell.value = val
            data_cell.border = thin_border
            data_cell.alignment = center_align
            if i == j:
                data_cell.fill = diag_fill
                data_cell.font = Font(name="Consolas", size=8, bold=True, color="0A8E6E")
            elif val == 0:
                data_cell.font = zero_font
            elif max_val > 0:
                t = min(val / max_val, 1.0)
                rc = int(200 - 150 * t)
                gc = int(230 - 140 * t)
                bc = int(220 - 10 * t)
                data_cell.fill = PatternFill(
                    start_color=f"{rc:02X}{gc:02X}{bc:02X}",
                    end_color=f"{rc:02X}{gc:02X}{bc:02X}",
                    fill_type="solid",
                )
                data_cell.font = Font(name="Consolas", size=8, bold=True)
        r += 1

    matrix_hdr_row = start_row + 3
    matrix_end_row = r - 1
    matrix_end_col = start_col + len(courses)
    thick = Side(style="medium", color="333333")
    for mr in range(matrix_hdr_row, matrix_end_row + 1):
        for mc in range(start_col, matrix_end_col + 1):
            cl = ws.cell(row=mr, column=mc)
            ex = cl.border
            cl.border = Border(
                left=thick if mc == start_col else ex.left,
                right=thick if mc == matrix_end_col else ex.right,
                top=thick if mr == matrix_hdr_row else ex.top,
                bottom=thick if mr == matrix_end_row else ex.bottom,
            )


def _write_plan_coverage_sheet(
    *,
    wb,
    scenario: TimetableScenario,
    program: str,
    plan_map: dict[str, int],
    course_status: dict[str, str],
    course_board_terms: dict[str, set[int]],
    course_names: dict[str, str],
    budget_by_key: dict[str, ScenarioSectionBudget],
    out_of_plan_codes: set[str],
    styles: dict,
    apply_outer_border,
) -> None:
    """Render the leading audit sheet showing plan-vs-actual coverage."""
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(title="Plan Coverage")

    hdr_fill = styles["hdr_fill"]
    hdr_font = styles["hdr_font"]
    hdr_align = styles["hdr_align"]
    thin_border = styles["thin_border"]
    bold_font = styles["bold_font"]
    normal_font = styles["normal_font"]
    center_align = styles["center_align"]
    matched_fill = styles["matched_fill"]
    misplaced_fill = styles["misplaced_fill"]
    missing_fill = styles["missing_fill"]
    board_hdr_fill = styles["board_hdr_fill"]
    board_hdr_font = styles["board_hdr_font"]

    # Title block
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws.cell(row=1, column=1)
    c.value = f"Plan Coverage — {program}  ·  Scenario: {scenario.name}"
    c.fill = board_hdr_fill
    c.font = board_hdr_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 7):
        ws.cell(row=1, column=col).fill = board_hdr_fill

    matched_count = sum(1 for code in plan_map if course_status.get(code) == "matched")
    misplaced_count = sum(1 for code in plan_map if course_status.get(code) == "misplaced")
    missing_count = sum(1 for code in plan_map if code not in course_status)
    ws.cell(
        row=2,
        column=1,
        value=(
            f"{matched_count} matched  ·  "
            f"{misplaced_count} misplaced  ·  "
            f"{missing_count} missing  ·  "
            f"{len(out_of_plan_codes)} courses placed but not in this plan"
        ),
    ).font = Font(size=9, color="666666", italic=True)

    # Plan table
    row = 4
    for col_idx, hdr in enumerate(
        ["Plan Term", "Course", "Name", "Credits", "Status", "Placed On"], start=1
    ):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = hdr
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border
    row += 1

    table_start_row = row
    plan_pairs = sorted(plan_map.items(), key=lambda kv: (kv[1], kv[0]))
    for code, plan_term in plan_pairs:
        if code in course_status:
            status = course_status[code]
        else:
            status = "missing"
        actual = sorted(course_board_terms.get(code, set()))
        if status == "matched":
            placed_on = f"T{plan_term} (matches plan)"
            row_fill = matched_fill
        elif status == "misplaced":
            placed_on = ", ".join(f"T{t}" for t in actual) or "?"
            row_fill = misplaced_fill
        else:
            placed_on = "—"
            row_fill = missing_fill

        ws.cell(row=row, column=1, value=f"T{plan_term}").alignment = center_align
        ws.cell(row=row, column=2, value=_course_code_from_key(code)).font = bold_font
        ws.cell(
            row=row, column=3, value=_course_display_name(code, course_names)
        ).font = normal_font
        budget = budget_by_key.get(code)
        budget_cred = budget.credit_hours if budget else 0
        ws.cell(row=row, column=4, value=budget_cred or 0).alignment = center_align
        ws.cell(row=row, column=5, value=status).alignment = center_align
        ws.cell(row=row, column=6, value=placed_on).alignment = center_align
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = row_fill
        row += 1

    apply_outer_border(ws, table_start_row, 1, row - 1, 6)

    # Out-of-plan section (courses placed in scenario but not part of this plan)
    if out_of_plan_codes:
        row += 2
        ws.cell(
            row=row,
            column=1,
            value=f"Other courses placed in this scenario but not in {program}'s plan",
        ).font = bold_font
        row += 1
        for code in sorted(out_of_plan_codes):
            ws.cell(row=row, column=2, value=_course_code_from_key(code)).font = bold_font
            ws.cell(
                row=row, column=3, value=_course_display_name(code, course_names)
            ).font = normal_font
            row += 1

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 22
    ws.freeze_panes = "A5"
