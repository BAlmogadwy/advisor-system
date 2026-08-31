"""Professional XLSX export for one advisor graduation-planning scenario.

The workbook is intentionally built from the same server-side report and
presentation used by the dedicated graduation page.  It is a read-only
planning aid: registered, formal-recommended, and optimized baselines are
exported separately so their provenance cannot be confused.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from core.services.advisor_graduation_optimization import OPTIMIZED_CURRENT_OFFERINGS
from core.services.student_graduation import RECOMMENDED_CURRENT_TERM

_SHEET_NAMES = (
    "Overview",
    "Term Plan",
    "Baseline Courses",
    "Prerequisite Map",
    "Blockers & Assumptions",
)


def _safe_excel_text(value: object) -> str:
    """Return literal spreadsheet text, neutralising formula prefixes."""
    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _student_value(student: object, field: str, default: object = "") -> object:
    if isinstance(student, dict):
        return student.get(field, default)
    return getattr(student, field, default)


def build_graduation_xlsx(
    *,
    student: object,
    academic_year: int,
    term: int,
    baseline_kind: str,
    report: dict[str, Any],
    presentation: dict[str, Any],
    language_code: str = "en",
) -> bytes:
    """Return a styled XLSX workbook for the selected graduation scenario."""
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover - dependency is pinned in production
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    navy = "111144"
    teal = "0A8E6E"
    royal = "4056E3"
    amber = "D97706"
    white = "FFFFFF"
    ink = "202735"
    slate = "5B6475"
    line = "CFD6DF"
    pale_navy = "E9EEF5"
    pale_teal = "E8F5F0"
    pale_royal = "E9ECFF"
    pale_amber = "FFF3CD"
    pale_red = "F8D7DA"
    pale_slate = "EEF1F5"

    thin = Side(style="thin", color=line)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor=navy)
    header_fill = PatternFill("solid", fgColor=teal)
    section_fill = PatternFill("solid", fgColor=pale_navy)
    title_font = Font(name="Aptos Display", size=16, bold=True, color=white)
    header_font = Font(name="Aptos", size=10, bold=True, color=white)
    label_font = Font(name="Aptos", size=9, bold=True, color=navy)
    body_font = Font(name="Aptos", size=9, color=ink)
    strong_font = Font(name="Aptos", size=10, bold=True, color=navy)
    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    is_arabic = str(language_code or "").lower().startswith("ar")
    baseline_is_recommended = baseline_kind == RECOMMENDED_CURRENT_TERM
    baseline_is_optimized = baseline_kind == OPTIMIZED_CURRENT_OFFERINGS
    if baseline_is_optimized:
        baseline_source = "Optimized current offerings"
        baseline_source_note = (
            "Exact hypothetical selection from the recorded current-section snapshot; "
            "these are not actual registrations or seat guarantees."
        )
        baseline_stage = "Optimized starting term"
    elif baseline_is_recommended:
        baseline_source = "Recommended starting-term courses"
        baseline_source_note = "System recommendations; these are not actual registrations."
        baseline_stage = "Recommended starting term"
    else:
        baseline_source = "Registered timetable"
        baseline_source_note = "Actual university-portal registrations."
        baseline_stage = "Registered"
    student_id = _student_value(student, "student_id", "")
    student_name = _student_value(student, "name", "")
    program = report.get("program") or _student_value(student, "program", "")
    section = _student_value(student, "section", "")
    registration_no = _student_value(student, "registration_no", "")
    baseline_year = int(report.get("planning_baseline_academic_year") or academic_year)
    baseline_term = int(report.get("planning_baseline_term") or term)

    workbook = Workbook()
    workbook.properties.creator = "Advisor Portal"
    workbook.properties.title = f"Graduation plan — {student_id}"
    workbook.properties.subject = f"Read-only {baseline_source.lower()} graduation scenario"
    workbook.properties.description = (
        "Read-only planning estimate. It is not an official graduation decision or registration."
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    def set_text(cell: Any, value: object) -> None:
        cell.value = _safe_excel_text(value)
        cell.data_type = "s"

    def style_cell(
        cell: Any,
        *,
        font: Any = body_font,
        alignment: Any = left,
        fill: Any | None = None,
    ) -> None:
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        if fill is not None:
            cell.fill = fill

    def merged_banner(
        sheet: Any,
        row: int,
        text: object,
        last_column: int,
        *,
        title: bool = False,
        warning: bool = False,
    ) -> None:
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=last_column,
        )
        fill = title_fill
        font = title_font
        height = 29
        if not title:
            fill = PatternFill("solid", fgColor=pale_amber if warning else "F5F7FA")
            font = Font(
                name="Aptos",
                size=9 if warning else 10,
                bold=warning,
                color=amber if warning else slate,
            )
            height = 25
        for column in range(1, last_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = fill
            cell.border = border
        first = sheet.cell(row=row, column=1)
        set_text(first, text)
        first.font = font
        first.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sheet.row_dimensions[row].height = height

    def section_banner(sheet: Any, row: int, text: object, last_column: int) -> None:
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=last_column,
        )
        for column in range(1, last_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = section_fill
            cell.border = border
        cell = sheet.cell(row=row, column=1)
        set_text(cell, text)
        cell.font = strong_font
        cell.alignment = left
        sheet.row_dimensions[row].height = 22

    def write_headers(sheet: Any, row: int, headers: tuple[str, ...]) -> None:
        for column, heading in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=column)
            set_text(cell, heading)
            style_cell(cell, font=header_font, alignment=center, fill=header_fill)
        sheet.row_dimensions[row].height = 28

    def add_table(sheet: Any, reference: str, name: str) -> None:
        table = Table(displayName=name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    def configure_sheet(
        sheet: Any,
        *,
        orientation: str,
        freeze: str,
        tab_color: str,
        print_rows: str = "1:4",
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.rightToLeft = is_arabic
        sheet.freeze_panes = freeze
        sheet.sheet_properties.tabColor = tab_color
        sheet.page_setup.orientation = orientation
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_title_rows = print_rows
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.45
        sheet.page_margins.bottom = 0.45
        sheet.oddFooter.center.text = "Advisor Portal · Read-only planning estimate"
        sheet.oddFooter.right.text = "Page &P of &N"
        sheet.oddFooter.center.size = 8
        sheet.oddFooter.center.color = slate
        sheet.oddFooter.right.size = 8
        sheet.oddFooter.right.color = slate

    def status_label(status: object, band: object) -> str:
        key = str(status or "").strip().lower()
        try:
            band_number = int(band)
        except (TypeError, ValueError):
            band_number = -1
        if key == "passed":
            return "Completed"
        if band_number == 1 and baseline_is_recommended:
            return "Recommended starting course"
        if band_number == 1 and baseline_is_optimized:
            return "Optimized proposed course"
        if key == "studying" or band_number == 1:
            return "Registered"
        if key == "open":
            return "Projected"
        return key.replace("_", " ").title() or "Unresolved"

    def exported_course_code(course: dict[str, Any]) -> object:
        return course.get("offered_course_code") or course.get("code", "")

    def exported_course_name(course: dict[str, Any]) -> str:
        name = str(course.get("offered_course_name") or course.get("name") or "")
        plan_code = str(course.get("fulfills_plan_code") or "").strip()
        return f"{name} (fulfills {plan_code})" if plan_code else name

    # ── Overview ──────────────────────────────────────────────
    overview = workbook.active
    overview.title = _SHEET_NAMES[0]
    merged_banner(overview, 1, "Graduation Planning Workbook", 8, title=True)
    merged_banner(
        overview,
        2,
        f"{student_name or 'Student'} · ID {student_id} · {program or 'Program not recorded'}",
        8,
    )
    merged_banner(
        overview,
        3,
        "Read-only advisory estimate — not an official graduation decision, registration, or guarantee of course availability.",
        8,
        warning=True,
    )
    section_banner(overview, 5, "Student and scenario", 8)

    detail_rows = (
        ("Student ID", student_id, "Student name", student_name),
        ("Registration number", registration_no, "Program", program),
        ("Section", section, "Planning term", f"{baseline_year}/{baseline_term}"),
        (
            "Starting-course source",
            baseline_source,
            "Generated",
            datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z"),
        ),
    )
    for row, values in enumerate(detail_rows, start=6):
        for offset in (0, 2):
            label_cell = overview.cell(row=row, column=1 + offset * 2)
            value_cell = overview.cell(row=row, column=2 + offset * 2)
            set_text(label_cell, values[offset])
            set_text(value_cell, values[offset + 1])
            style_cell(label_cell, font=label_font, fill=section_fill)
            style_cell(value_cell, alignment=wrap)
            overview.merge_cells(
                start_row=row,
                start_column=2 + offset * 2,
                end_row=row,
                end_column=4 + offset * 2,
            )
            for column in range(3 + offset * 2, 5 + offset * 2):
                overview.cell(row=row, column=column).border = border
                overview.cell(row=row, column=column).fill = PatternFill("solid", fgColor=white)

    section_banner(overview, 11, "Planning outlook", 8)
    metric_rows = (
        (
            "Plan courses completed",
            report.get("plan_courses_passed"),
            "Plan courses total",
            report.get("plan_courses_total"),
        ),
        (
            "Courses remaining",
            report.get("remaining_courses"),
            "Credits remaining",
            report.get("remaining_credits"),
        ),
        (
            "Registrar earned credits",
            report.get("earned_credits_registrar"),
            "Passed plan credits",
            report.get("passed_credits_in_plan"),
        ),
        (
            "Starting-term credits",
            report.get("planning_baseline_credits"),
            "Maximum credits per term",
            report.get("max_credits_per_term"),
        ),
        (
            "Productive projected terms",
            report.get("productive_terms_planned"),
            "Terms examined",
            report.get("simulated_terms_examined"),
        ),
    )
    for row, values in enumerate(metric_rows, start=12):
        for pair in range(2):
            label_col = 1 + pair * 4
            value_col = label_col + 2
            label = values[pair * 2]
            value = values[pair * 2 + 1]
            overview.merge_cells(
                start_row=row, start_column=label_col, end_row=row, end_column=label_col + 1
            )
            label_cell = overview.cell(row=row, column=label_col)
            set_text(label_cell, label)
            style_cell(label_cell, font=label_font, fill=section_fill)
            overview.cell(row=row, column=label_col + 1).border = border
            overview.merge_cells(
                start_row=row, start_column=value_col, end_row=row, end_column=value_col + 1
            )
            value_cell = overview.cell(row=row, column=value_col)
            if isinstance(value, int | float) and not isinstance(value, bool):
                value_cell.value = value
                value_cell.number_format = "#,##0.##"
            else:
                set_text(value_cell, "Not available" if value is None else value)
            style_cell(value_cell, font=strong_font, alignment=center)
            overview.cell(row=row, column=value_col + 1).border = border

    section_banner(overview, 18, "Projection status and verification", 8)
    overview_metrics = (
        (
            "Simulation status",
            "Complete" if report.get("simulation_completed") else "Incomplete — lower bound only",
        ),
        (
            "Projected terms including baseline",
            report.get("estimated_terms_including_planning_baseline")
            if report.get("simulation_completed")
            else "Not calculated while blockers remain",
        ),
        (
            "Verified lower bound including baseline",
            report.get("lower_bound_terms_including_planning_baseline"),
        ),
        ("GPA", report.get("gpa")),
    )
    for index, (label, value) in enumerate(overview_metrics):
        row = 19 + index // 2
        start_col = 1 + (index % 2) * 4
        label_cell = overview.cell(row=row, column=start_col)
        value_cell = overview.cell(row=row, column=start_col + 2)
        overview.merge_cells(
            start_row=row, start_column=start_col, end_row=row, end_column=start_col + 1
        )
        overview.merge_cells(
            start_row=row, start_column=start_col + 2, end_row=row, end_column=start_col + 3
        )
        set_text(label_cell, label)
        style_cell(label_cell, font=label_font, fill=section_fill)
        overview.cell(row=row, column=start_col + 1).border = border
        if isinstance(value, int | float) and not isinstance(value, bool):
            value_cell.value = value
            value_cell.number_format = "0.00" if label == "GPA" else "#,##0"
        else:
            set_text(value_cell, "Not available" if value is None else value)
        style_cell(value_cell, font=strong_font, alignment=center)
        overview.cell(row=row, column=start_col + 3).border = border

    overview["A22"] = "Completion %"
    style_cell(overview["A22"], font=label_font, fill=section_fill)
    overview.merge_cells("A22:B22")
    overview["C22"] = "=IFERROR(C12/G12,0)"
    overview["C22"].number_format = "0%"
    style_cell(overview["C22"], font=strong_font, alignment=center)
    overview.merge_cells("C22:D22")
    overview["E22"] = "Plan-course reconciliation"
    style_cell(overview["E22"], font=label_font, fill=section_fill)
    overview.merge_cells("E22:F22")
    overview["G22"] = "=C12+C13-G12"
    overview["G22"].number_format = "0;[Red]-0"
    style_cell(overview["G22"], font=strong_font, alignment=center)
    overview.merge_cells("G22:H22")
    overview["A23"] = (
        "A reconciliation value of 0 means completed plus remaining courses matches the recorded plan total."
    )
    overview.merge_cells("A23:H23")
    style_cell(overview["A23"], alignment=wrap, fill=PatternFill("solid", fgColor="F5F7FA"))
    overview.row_dimensions[23].height = 30

    for column, width in {
        "A": 18,
        "B": 12,
        "C": 18,
        "D": 18,
        "E": 22,
        "F": 14,
        "G": 22,
        "H": 18,
    }.items():
        overview.column_dimensions[column].width = width
    configure_sheet(overview, orientation="portrait", freeze="A5", tab_color=navy)
    overview.print_area = "A1:H23"

    # ── Term Plan ─────────────────────────────────────────────
    term_sheet = workbook.create_sheet(_SHEET_NAMES[1])
    term_headers = (
        "Step",
        "Academic Year",
        "Term",
        "Academic Term",
        "State",
        "Course Code",
        "Course Name",
        "Credits",
        "Requirement Type",
        "Elective Slot",
        "Term Credits",
        "Load %",
    )
    merged_banner(term_sheet, 1, "Term-by-Term Graduation Plan", len(term_headers), title=True)
    merged_banner(
        term_sheet,
        2,
        (
            f"Starting source: {baseline_source} · {baseline_source_note} "
            f"· Planning term {baseline_year}/{baseline_term}"
        ),
        len(term_headers),
    )
    write_headers(term_sheet, 4, term_headers)
    term_rows: list[dict[str, Any]] = []
    baseline_courses = list(report.get("planning_baseline_courses_assumed_passed") or [])
    baseline_total = int(report.get("planning_baseline_credits") or 0)
    for course in baseline_courses:
        term_rows.append(
            {
                "step": 0,
                "year": baseline_year,
                "term": baseline_term,
                "stage": baseline_stage,
                "course": course,
                "term_credits": baseline_total,
                "waiting": False,
            }
        )
    if not baseline_courses:
        term_rows.append(
            {
                "step": 0,
                "year": baseline_year,
                "term": baseline_term,
                "stage": baseline_stage,
                "course": {},
                "term_credits": baseline_total,
                "waiting": True,
                "message": "No starting-term courses were available for this scenario.",
            }
        )
    for planned in list(report.get("term_plan") or []):
        step = int(planned.get("sequence") or 0)
        courses = list(planned.get("courses") or [])
        if courses:
            for course in courses:
                term_rows.append(
                    {
                        "step": step,
                        "year": int(planned.get("academic_year") or 0),
                        "term": int(planned.get("term") or 0),
                        "stage": "Projected",
                        "course": course,
                        "term_credits": int(planned.get("credits") or 0),
                        "waiting": False,
                    }
                )
        else:
            term_rows.append(
                {
                    "step": step,
                    "year": int(planned.get("academic_year") or 0),
                    "term": int(planned.get("term") or 0),
                    "stage": "Waiting",
                    "course": {},
                    "term_credits": int(planned.get("credits") or 0),
                    "waiting": True,
                    "message": "Waiting for prerequisites",
                }
            )

    for row, item in enumerate(term_rows, start=5):
        course = item["course"]
        values: tuple[object, ...] = (
            item["step"],
            item["year"],
            item["term"],
            None,
            item["stage"],
            exported_course_code(course),
            exported_course_name(course) or item.get("message", ""),
            int(course.get("credits") or 0),
            course.get("requirement_type", ""),
            "Yes" if course.get("elective_slot") else "No",
            item["term_credits"],
            None,
        )
        for column, value in enumerate(values, start=1):
            cell = term_sheet.cell(row=row, column=column)
            if column in (1, 2, 3, 8, 11):
                cell.value = int(value or 0)
                cell.number_format = "#,##0"
                style_cell(cell, alignment=center)
            elif column == 4:
                cell.value = f'=B{row}&"/"&C{row}'
                style_cell(cell, alignment=center)
            elif column == 12:
                cell.value = f"=IFERROR(K{row}/'Overview'!G15,0)"
                cell.number_format = "0%"
                style_cell(cell, alignment=center)
            else:
                set_text(cell, value)
                style_cell(cell, alignment=wrap if column == 7 else center)
        stage_fill = (
            PatternFill("solid", fgColor=pale_amber)
            if item["waiting"]
            else PatternFill("solid", fgColor=pale_royal if item["step"] == 0 else pale_teal)
        )
        for column in range(1, len(term_headers) + 1):
            term_sheet.cell(row=row, column=column).fill = stage_fill
        term_sheet.row_dimensions[row].height = 28 if item["waiting"] else 23

    term_last_row = 4 + len(term_rows)
    if term_rows:
        term_sheet.auto_filter.ref = f"A4:L{term_last_row}"
        add_table(term_sheet, f"A4:L{term_last_row}", "GraduationTermPlan")
    for column, width in enumerate((8, 14, 9, 16, 19, 15, 42, 10, 20, 14, 13, 11), start=1):
        term_sheet.column_dimensions[
            term_sheet.cell(row=4, column=column).column_letter
        ].width = width
    configure_sheet(term_sheet, orientation="landscape", freeze="F5", tab_color=teal)
    term_sheet.print_area = f"A1:L{max(4, term_last_row)}"
    term_sheet.conditional_formatting.add(
        f"L5:L{max(5, term_last_row)}",
        FormulaRule(formula=["L5>1"], fill=PatternFill("solid", fgColor=pale_red)),
    )

    # ── Baseline Courses ──────────────────────────────────────
    baseline_sheet = workbook.create_sheet(_SHEET_NAMES[2])
    baseline_headers = (
        "Course Code",
        "Course Name",
        "Plan Requirement",
        "Section Evidence",
        "Credits",
        "Source",
    )
    merged_banner(baseline_sheet, 1, "Starting-Term Courses", len(baseline_headers), title=True)
    merged_banner(
        baseline_sheet,
        2,
        f"{baseline_source} · {baseline_source_note} · {baseline_year}/{baseline_term}",
        len(baseline_headers),
    )
    write_headers(baseline_sheet, 4, baseline_headers)
    baseline_data_rows = 0
    if baseline_courses:
        for row, course in enumerate(baseline_courses, start=5):
            values = (
                exported_course_code(course),
                exported_course_name(course),
                course.get("fulfills_plan_code") or course.get("code", ""),
                (
                    ", ".join(str(value) for value in course.get("recorded_sections") or [])
                    if baseline_is_optimized
                    else course.get("section", "")
                ),
                int(course.get("credits") or 0),
                baseline_source,
            )
            for column, value in enumerate(values, start=1):
                cell = baseline_sheet.cell(row=row, column=column)
                if column == 5:
                    cell.value = value
                    cell.number_format = "#,##0"
                    style_cell(cell, alignment=center)
                else:
                    set_text(cell, value)
                    style_cell(cell, alignment=wrap if column in (2, 4, 6) else center)
            baseline_data_rows += 1
    else:
        baseline_sheet.merge_cells("A5:F5")
        set_text(baseline_sheet["A5"], "No starting-term courses are available for this scenario.")
        style_cell(
            baseline_sheet["A5"], alignment=wrap, fill=PatternFill("solid", fgColor=pale_slate)
        )
        for column in range(2, 7):
            baseline_sheet.cell(row=5, column=column).border = border
    if baseline_data_rows:
        baseline_sheet.auto_filter.ref = f"A4:F{4 + baseline_data_rows}"
        add_table(baseline_sheet, f"A4:F{4 + baseline_data_rows}", "GraduationBaselineCourses")
    for column, width in enumerate((17, 43, 18, 30, 11, 48), start=1):
        baseline_sheet.column_dimensions[
            baseline_sheet.cell(row=4, column=column).column_letter
        ].width = width
    configure_sheet(baseline_sheet, orientation="landscape", freeze="A5", tab_color=royal)
    baseline_sheet.print_area = f"A1:F{max(5, 4 + baseline_data_rows)}"

    # ── Prerequisite Map ──────────────────────────────────────
    prereq_sheet = workbook.create_sheet(_SHEET_NAMES[3])
    prereq_headers = (
        "Relation",
        "Prerequisite Code",
        "Prerequisite Name",
        "Prerequisite Status",
        "Prerequisite Band",
        "Prerequisite Term",
        "Course Code",
        "Course Name",
        "Course Status",
        "Course Band",
        "Course Term",
        "Term Gap",
        "Order Check",
    )
    merged_banner(prereq_sheet, 1, "Prerequisite Map", len(prereq_headers), title=True)
    merged_banner(
        prereq_sheet,
        2,
        "Tabular version of the on-screen tree. Standalone courses are retained even when they have no visible prerequisite link.",
        len(prereq_headers),
    )
    write_headers(prereq_sheet, 4, prereq_headers)
    graph = presentation.get("graph") if isinstance(presentation, dict) else {}
    graph = graph if isinstance(graph, dict) else {}
    edges = [row for row in list(graph.get("items") or []) if isinstance(row, dict)]
    term_of = graph.get("termOf") if isinstance(graph.get("termOf"), dict) else {}
    name_of = graph.get("nameOf") if isinstance(graph.get("nameOf"), dict) else {}
    status_of = graph.get("statusOf") if isinstance(graph.get("statusOf"), dict) else {}
    band_labels = (
        presentation.get("band_labels") if isinstance(presentation.get("band_labels"), dict) else {}
    )
    nodes = {str(code) for code in list(graph.get("extraNodes") or []) if str(code).strip()}
    linked_nodes: set[str] = set()
    map_rows: list[dict[str, Any]] = []
    for edge in edges:
        course_code = str(edge.get("course_code") or "").strip()
        prereq_code = str(edge.get("prerequisite_course_code") or "").strip()
        if not course_code or not prereq_code:
            continue
        nodes.update((course_code, prereq_code))
        linked_nodes.update((course_code, prereq_code))
        map_rows.append(
            {
                "relation": "Prerequisite link",
                "prereq": prereq_code,
                "course": course_code,
            }
        )
    for code in sorted(nodes - linked_nodes):
        map_rows.append({"relation": "No prerequisite link", "prereq": "", "course": code})

    for row, item in enumerate(map_rows, start=5):
        prereq_code = item["prereq"]
        course_code = item["course"]
        prereq_band = term_of.get(prereq_code, "") if prereq_code else ""
        course_band = term_of.get(course_code, "")
        values: tuple[object, ...] = (
            item["relation"],
            prereq_code,
            name_of.get(prereq_code, "") if prereq_code else "",
            status_label(status_of.get(prereq_code), prereq_band) if prereq_code else "",
            prereq_band,
            band_labels.get(str(prereq_band), "") if prereq_code else "",
            course_code,
            name_of.get(course_code, ""),
            status_label(status_of.get(course_code), course_band),
            course_band,
            band_labels.get(str(course_band), ""),
            None,
            None,
        )
        for column, value in enumerate(values, start=1):
            cell = prereq_sheet.cell(row=row, column=column)
            if column in (5, 10) and value not in (None, ""):
                try:
                    cell.value = int(value)
                    cell.number_format = "0"
                except (TypeError, ValueError):
                    set_text(cell, value)
                style_cell(cell, alignment=center)
            elif column == 12:
                # COUNT distinguishes a real band 0 (completed before the
                # scenario) from an absent band on a standalone-node row.
                cell.value = f'=IF(COUNT(E{row},J{row})<2,"",J{row}-E{row})'
                cell.number_format = "0"
                style_cell(cell, alignment=center)
            elif column == 13:
                cell.value = (
                    f'=IF(COUNT(E{row},J{row})<2,"No link",IF(J{row}-E{row}>0,"OK",'
                    f'IF(J{row}-E{row}=0,"Same term","Prerequisite later")))'
                )
                style_cell(cell, alignment=center)
            else:
                set_text(cell, value)
                style_cell(cell, alignment=wrap if column in (3, 6, 8, 11) else center)

    if not map_rows:
        prereq_sheet.merge_cells("A5:M5")
        set_text(prereq_sheet["A5"], "No prerequisite-map rows are available for this scenario.")
        style_cell(
            prereq_sheet["A5"], alignment=wrap, fill=PatternFill("solid", fgColor=pale_slate)
        )
        for column in range(2, 14):
            prereq_sheet.cell(row=5, column=column).border = border
    prereq_last_row = 4 + len(map_rows)
    if map_rows:
        prereq_sheet.auto_filter.ref = f"A4:M{prereq_last_row}"
        add_table(prereq_sheet, f"A4:M{prereq_last_row}", "GraduationPrerequisiteMap")
        prereq_sheet.conditional_formatting.add(
            f"M5:M{prereq_last_row}",
            FormulaRule(
                formula=['M5="Prerequisite later"'],
                fill=PatternFill("solid", fgColor=pale_red),
            ),
        )
        prereq_sheet.conditional_formatting.add(
            f"M5:M{prereq_last_row}",
            FormulaRule(
                formula=['M5="Same term"'],
                fill=PatternFill("solid", fgColor=pale_amber),
            ),
        )
    for column, width in enumerate(
        (20, 18, 35, 24, 14, 31, 18, 35, 24, 14, 31, 11, 19),
        start=1,
    ):
        prereq_sheet.column_dimensions[
            prereq_sheet.cell(row=4, column=column).column_letter
        ].width = width
    configure_sheet(prereq_sheet, orientation="landscape", freeze="G5", tab_color=royal)
    prereq_sheet.print_area = f"A1:M{max(5, prereq_last_row)}"

    # ── Blockers and assumptions ──────────────────────────────
    blocker_sheet = workbook.create_sheet(_SHEET_NAMES[4])
    blocker_headers = (
        "Course Code",
        "Course Name",
        "Credits",
        "Requirement Type",
        "Elective Slot",
        "Missing Course Prerequisites",
        "Outside-plan Prerequisites",
        "Required Credits",
        "Effective Credits",
        "Credit Shortfall",
    )
    merged_banner(
        blocker_sheet, 1, "Unresolved Blockers and Assumptions", len(blocker_headers), title=True
    )
    merged_banner(
        blocker_sheet,
        2,
        "Blockers explain an incomplete simulation. They do not remove other valid planned terms from the export.",
        len(blocker_headers),
    )
    write_headers(blocker_sheet, 4, blocker_headers)
    blockers = [
        row for row in list(report.get("unresolved_requirements") or []) if isinstance(row, dict)
    ]
    blocker_data_rows = 0
    if blockers:
        for row, blocker in enumerate(blockers, start=5):
            gate = blocker.get("credit_hour_gate")
            gate = gate if isinstance(gate, dict) else {}
            missing = ", ".join(
                str(value) for value in blocker.get("missing_course_prerequisites") or []
            )
            outside = ", ".join(
                str(value) for value in blocker.get("missing_prerequisites_outside_plan") or []
            )
            required = gate.get("required")
            effective = gate.get("effective_in_scenario", gate.get("effective"))
            remaining = gate.get("remaining")
            values: tuple[object, ...] = (
                blocker.get("code", ""),
                blocker.get("name", ""),
                int(blocker.get("credits") or 0),
                blocker.get("requirement_type", ""),
                "Yes" if blocker.get("elective_slot") else "No",
                missing,
                outside,
                required,
                effective,
                remaining,
            )
            for column, value in enumerate(values, start=1):
                cell = blocker_sheet.cell(row=row, column=column)
                if column in (3, 8, 9, 10) and value is not None:
                    cell.value = int(value or 0)
                    cell.number_format = "#,##0"
                    style_cell(cell, alignment=center)
                else:
                    set_text(cell, value)
                    style_cell(cell, alignment=wrap if column in (2, 6, 7) else center)
            blocker_data_rows += 1
    else:
        blocker_sheet.merge_cells("A5:J5")
        set_text(blocker_sheet["A5"], "No unresolved blockers.")
        style_cell(
            blocker_sheet["A5"],
            font=Font(name="Aptos", size=10, bold=True, color=teal),
            alignment=left,
            fill=PatternFill("solid", fgColor=pale_teal),
        )
        for column in range(2, 11):
            blocker_sheet.cell(row=5, column=column).border = border
    if blocker_data_rows:
        blocker_sheet.auto_filter.ref = f"A4:J{4 + blocker_data_rows}"
        add_table(
            blocker_sheet,
            f"A4:J{4 + blocker_data_rows}",
            "GraduationBlockers",
        )

    assumptions_start = max(7, 6 + blocker_data_rows)
    section_banner(blocker_sheet, assumptions_start, "Simulation assumptions", len(blocker_headers))
    assumptions = list(report.get("simulation_assumptions") or [])
    if not assumptions:
        assumptions = [
            "No simulation assumptions were supplied with this report.",
            "This export remains a read-only planning aid and is not an official graduation decision.",
        ]
    for index, assumption in enumerate(assumptions, start=1):
        row = assumptions_start + index
        blocker_sheet.cell(row=row, column=1).value = index
        blocker_sheet.cell(row=row, column=1).number_format = "0"
        style_cell(blocker_sheet.cell(row=row, column=1), alignment=center)
        blocker_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        set_text(blocker_sheet.cell(row=row, column=2), assumption)
        style_cell(blocker_sheet.cell(row=row, column=2), alignment=wrap)
        for column in range(3, 11):
            blocker_sheet.cell(row=row, column=column).border = border
        blocker_sheet.row_dimensions[row].height = 31

    metadata_row = assumptions_start + len(assumptions) + 2
    section_banner(blocker_sheet, metadata_row, "Scenario metadata", len(blocker_headers))
    metadata = (
        ("Baseline kind", baseline_kind),
        ("Starting-course source", baseline_source),
        ("Planning term", f"{baseline_year}/{baseline_term}"),
        ("Simulation completed", "Yes" if report.get("simulation_completed") else "No"),
    )
    for offset, (label, value) in enumerate(metadata, start=metadata_row + 1):
        blocker_sheet.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=2)
        set_text(blocker_sheet.cell(row=offset, column=1), label)
        style_cell(blocker_sheet.cell(row=offset, column=1), font=label_font, fill=section_fill)
        blocker_sheet.cell(row=offset, column=2).border = border
        blocker_sheet.merge_cells(start_row=offset, start_column=3, end_row=offset, end_column=10)
        set_text(blocker_sheet.cell(row=offset, column=3), value)
        style_cell(blocker_sheet.cell(row=offset, column=3), alignment=wrap)
        for column in range(4, 11):
            blocker_sheet.cell(row=offset, column=column).border = border

    for column, width in enumerate((17, 38, 10, 20, 13, 34, 32, 15, 15, 15), start=1):
        blocker_sheet.column_dimensions[
            blocker_sheet.cell(row=4, column=column).column_letter
        ].width = width
    configure_sheet(blocker_sheet, orientation="landscape", freeze="F5", tab_color=amber)
    blocker_sheet.print_area = f"A1:J{metadata_row + len(metadata)}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


__all__ = ["build_graduation_xlsx"]
