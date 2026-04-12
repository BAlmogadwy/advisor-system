"""
core/services/eligibility_export.py
Styled XLSX export for the course eligibility report.

Generates a multi-sheet workbook:
  - Summary: overall stats, per-program breakdown with bar chart
  - Eligible Students: one row per eligible student with details
  - Blocked Analysis: top missing prerequisites and blocked samples
"""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import Student

# ── Styles ────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B4F72")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2C3E50")
_TEAL_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
_RED_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
_AMBER_FILL = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
_BORDER = Border(
    bottom=Side(style="thin", color="BDC3C7"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ── Main export ──────────────────────────────────────────────────


def export_eligibility_xlsx(payload: dict) -> Path:
    """Build a styled XLSX from the eligibility report payload.

    Returns the path to the temporary XLSX file.
    """
    wb = Workbook()
    course_code = payload.get("course_code", "?")
    mode = "Strict" if payload.get("strict_passed_only") else "Relaxed"
    filters = payload.get("filters", {})
    per_program = payload.get("per_program", [])

    # Fetch student details for eligible students (bulk) — no name/GPA for privacy
    all_eligible_ids: list[int] = []
    for row in per_program:
        all_eligible_ids.extend(row.get("eligible_student_ids", []))

    student_details: dict[int, dict] = {}
    if all_eligible_ids:
        for s in Student.objects.filter(student_id__in=all_eligible_ids).values(
            "student_id",
            "program",
            "total_earned_credits",
            "current_registered_credits",
            "section",
        ):
            student_details[s["student_id"]] = s

    # ── Sheet 1: Summary ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1B4F72"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Course Eligibility Report: {course_code}"
    ws["A1"].font = _TITLE_FONT

    ws["A3"] = "Course Code"
    ws["B3"] = course_code
    ws["A4"] = "Mode"
    ws["B4"] = mode
    ws["A5"] = "Section"
    ws["B5"] = filters.get("section") or "All"
    ws["A6"] = "Program"
    ws["B6"] = filters.get("program") or "All"

    for r in range(3, 7):
        ws.cell(row=r, column=1).font = _SUBTITLE_FONT

    ws["A8"] = "Total Students"
    ws["B8"] = payload.get("total_students", 0)
    ws["A9"] = "Total Eligible"
    ws["B9"] = payload.get("total_eligible", 0)
    ws["A10"] = "Eligibility Rate"
    total = payload.get("total_students", 0) or 1
    ws["B10"] = f"{payload.get('total_eligible', 0) / total * 100:.1f}%"

    for r in range(8, 11):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).font = Font(bold=True, color="1B4F72")

    # Per-program table
    row = 12
    ws.cell(row=row, column=1, value="Program")
    ws.cell(row=row, column=2, value="Students")
    ws.cell(row=row, column=3, value="Eligible")
    ws.cell(row=row, column=4, value="Blocked")
    ws.cell(row=row, column=5, value="Rate")
    ws.cell(row=row, column=6, value="Prerequisites")
    _style_header_row(ws, row, 6)

    for prog_data in per_program:
        row += 1
        ws.cell(row=row, column=1, value=prog_data["program"])
        ws.cell(row=row, column=2, value=prog_data["students"])
        ws.cell(row=row, column=3, value=prog_data["eligible_count"])
        ws.cell(row=row, column=4, value=prog_data["blocked_count"])
        rate = (
            prog_data["eligible_count"] / prog_data["students"] * 100
            if prog_data["students"]
            else 0
        )
        ws.cell(row=row, column=5, value=f"{rate:.0f}%")
        ws.cell(row=row, column=6, value=", ".join(prog_data.get("prerequisites", [])))

        # Color code rate
        rate_cell = ws.cell(row=row, column=5)
        if rate >= 50:
            rate_cell.fill = _TEAL_FILL
        elif rate >= 20:
            rate_cell.fill = _AMBER_FILL
        else:
            rate_cell.fill = _RED_FILL

    # Bar chart
    if len(per_program) > 1:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"Eligible Students per Program — {course_code}"
        chart.y_axis.title = "Students"
        chart.x_axis.title = "Program"
        chart.style = 10

        data_ref = Reference(ws, min_col=3, min_row=12, max_row=12 + len(per_program))
        cats_ref = Reference(ws, min_col=1, min_row=13, max_row=12 + len(per_program))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        ws.add_chart(chart, "H3")

    _auto_width(ws)

    # ── Sheet 2: Eligible Students ───────────────────────────────
    ws2 = wb.create_sheet("Eligible Students")
    ws2.sheet_properties.tabColor = "27AE60"

    headers = ["#", "Student ID", "Program", "Section", "Earned Credits", "Current Credits"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header_row(ws2, 1, len(headers))

    row = 1
    for prog_data in per_program:
        for sid in sorted(prog_data.get("eligible_student_ids", [])):
            row += 1
            details = student_details.get(sid, {})
            ws2.cell(row=row, column=1, value=row - 1)
            ws2.cell(row=row, column=2, value=sid)
            ws2.cell(row=row, column=3, value=details.get("program", prog_data["program"]))
            ws2.cell(row=row, column=4, value=details.get("section", ""))
            ws2.cell(row=row, column=5, value=details.get("total_earned_credits"))
            ws2.cell(row=row, column=6, value=details.get("current_registered_credits"))

            # Alternate row shading
            if (row - 1) % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws2.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )

    _auto_width(ws2)

    # ── Sheet 3: Blocked Analysis ────────────────────────────────
    ws3 = wb.create_sheet("Blocked Analysis")
    ws3.sheet_properties.tabColor = "E74C3C"

    # Top missing prerequisites across all programs
    ws3.merge_cells("A1:D1")
    ws3["A1"] = f"Blocked Analysis — {course_code}"
    ws3["A1"].font = _TITLE_FONT

    ws3.cell(row=3, column=1, value="Missing Prerequisite")
    ws3.cell(row=3, column=2, value="Program")
    ws3.cell(row=3, column=3, value="Students Blocked")
    ws3.cell(row=3, column=4, value="% of Program")
    _style_header_row(ws3, 3, 4)

    row = 3
    for prog_data in per_program:
        for mp in prog_data.get("top_missing_prerequisites", []):
            row += 1
            ws3.cell(row=row, column=1, value=mp["course_code"])
            ws3.cell(row=row, column=2, value=prog_data["program"])
            ws3.cell(row=row, column=3, value=mp["count"])
            pct = mp["count"] / prog_data["students"] * 100 if prog_data["students"] else 0
            ws3.cell(row=row, column=4, value=f"{pct:.0f}%")

    # Blocked samples
    row += 2
    ws3.cell(row=row, column=1, value="Blocked Student Samples")
    ws3.cell(row=row, column=1).font = _SUBTITLE_FONT

    row += 1
    ws3.cell(row=row, column=1, value="Student ID")
    ws3.cell(row=row, column=2, value="Program")
    ws3.cell(row=row, column=3, value="Reason")
    ws3.cell(row=row, column=4, value="Missing Prerequisites")
    _style_header_row(ws3, row, 4)

    for prog_data in per_program:
        for sample in prog_data.get("blocked_samples", []):
            if sample.get("reason") == "missing_prerequisites":
                row += 1
                ws3.cell(row=row, column=1, value=sample["student_id"])
                ws3.cell(row=row, column=2, value=prog_data["program"])
                ws3.cell(row=row, column=3, value=sample["reason"])
                ws3.cell(
                    row=row,
                    column=4,
                    value=", ".join(sample.get("missing_prerequisites", [])),
                )

    _auto_width(ws3)

    # ── Save ─────────────────────────────────────────────────────
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)
