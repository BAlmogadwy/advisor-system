"""
core/services/batch_export.py
XLSX export for the Batch Recommender (aggregate course demand).

Layout: each course gets a bold header row with demand count.
Prerequisites are listed as indented sub-rows below each course,
showing how many students are currently studying each prerequisite.
"""

from __future__ import annotations

import tempfile
from collections import defaultdict
from pathlib import Path


def export_batch_recommender_xlsx(
    year: int,
    semester: int,
    program: str | None,
    section: str | None,
    student_count: int,
    aggregate: dict[str, int],
    *,
    course_rows: list[dict[str, object]] | None = None,
) -> Path:
    """Export batch recommender results as styled XLSX."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required") from exc

    from core.models import Course, Prerequisite, Student, StudentCourse
    from core.services.student_helpers import normalize_code

    # ── Styles ───────────────────────────────────────────────────
    hdr_fill = PatternFill(start_color="0A8E6E", end_color="0A8E6E", fill_type="solid")
    hdr_font = Font(name="Consolas", size=9, bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center")
    title_fill = PatternFill(start_color="111144", end_color="111144", fill_type="solid")
    title_font = Font(name="Consolas", size=11, bold=True, color="FFFFFF")
    thick_side = Side(style="medium", color="333333")
    thin_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    course_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    course_font = Font(name="Consolas", size=10, bold=True, color="111144")
    count_font = Font(name="Consolas", size=11, bold=True, color="0A8E6E")
    prereq_indent_font = Font(size=9, color="666666")
    studying_font = Font(size=9, bold=True, color="C03030")
    studying_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    passed_font = Font(size=9, color="0A8E6E")
    passed_fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
    no_prereq_font = Font(size=8, italic=True, color="AAAAAA")
    center = Alignment(horizontal="center", vertical="center")

    # ── Pre-load data ────────────────────────────────────────────
    if course_rows is None:
        sorted_rows: list[dict[str, object]] = [
            {
                "course_code": normalize_code(code),
                "course_name": "",
                "count": int(count),
                "programs": [],
                "show_programs": False,
            }
            for code, count in aggregate.items()
        ]
    else:
        sorted_rows = [
            {
                "course_code": normalize_code(str(row.get("course_code", ""))),
                "course_name": str(row.get("course_name", "") or ""),
                "count": int(row.get("count", 0) or 0),
                "programs": list(row.get("programs", []))
                if isinstance(row.get("programs"), list)
                else [],
                "show_programs": bool(row.get("show_programs", False)),
            }
            for row in course_rows
            if row.get("course_code")
        ]
    sorted_rows.sort(
        key=lambda row: (
            -int(row.get("count", 0)),
            str(row.get("course_code", "")),
            str(row.get("course_name", "")),
        )
    )
    course_codes = sorted({str(row["course_code"]) for row in sorted_rows})

    # Course names
    course_names: dict[str, str] = {}
    for c in Course.objects.filter(course_code__in=course_codes):
        course_names[c.course_code.upper()] = c.description or ""

    def _display_course_name(row: dict[str, object]) -> str:
        code = normalize_code(str(row.get("course_code", "")))
        name = str(row.get("course_name") or course_names.get(code, "") or "")
        programs = row.get("programs", [])
        program_list = [str(p) for p in programs] if isinstance(programs, list) else []
        if row.get("show_programs") and program_list:
            prefix = ", ".join(program_list)
            return f"{prefix} - {name}" if name else prefix
        return name

    # Prerequisites
    programs_to_check = [p.strip() for p in (program or "").split(",") if p.strip()]
    prereq_map: dict[str, list[str]] = defaultdict(list)
    if programs_to_check:
        for prog in programs_to_check:
            for cc, prereq_cc in Prerequisite.objects.filter(
                program=prog, course_code__in=course_codes
            ).values_list("course_code", "prerequisite_course_code"):
                cc_n = normalize_code(cc)
                for part in str(prereq_cc).split(","):
                    p = normalize_code(part)
                    if p and p not in prereq_map[cc_n]:
                        prereq_map[cc_n].append(p)

    # ── Per-course prerequisite studying count ──────────────────
    # For each recommended course, count how many of the students who
    # NEED that course are currently studying each of its prerequisites.
    #
    # Step 1: Get per-student recommendations to know WHO needs WHAT
    from core.services.recommender_batch import batch_recommend, batch_recommend_multi_program
    from core.services.reporting import get_student_ids

    scoped_sids = get_student_ids(program=program, section=section)
    if program and "," not in (program or ""):
        all_recs = batch_recommend(scoped_sids, program, year, semester)
    else:
        all_recs = batch_recommend_multi_program(scoped_sids, year, semester)

    # Build: course_code -> set of student_ids who need it
    course_to_students: dict[str, set[int]] = defaultdict(set)
    student_programs: dict[int, str] = {
        sid: str(prog or "").strip()
        for sid, prog in Student.objects.filter(student_id__in=scoped_sids).values_list(
            "student_id", "program"
        )
    }
    course_to_students_by_program: dict[tuple[str, str], set[int]] = defaultdict(set)
    for sid, recs in all_recs.items():
        for code in recs:
            code_n = normalize_code(code)
            course_to_students[code_n].add(sid)
            course_to_students_by_program[(code_n, student_programs.get(sid, ""))].add(sid)

    # Step 2: For each prereq, get which students are studying it
    all_prereq_codes = set()
    for prereqs in prereq_map.values():
        all_prereq_codes.update(prereqs)

    # Build: student_id -> set of courses they're currently studying
    studying_by_student: dict[int, set[str]] = defaultdict(set)
    if all_prereq_codes:
        for sid_val, code in StudentCourse.objects.filter(
            course__course_code__in=all_prereq_codes,
            status="studying",
            student_id__in=scoped_sids,
        ).values_list("student_id", "course__course_code"):
            studying_by_student[sid_val].add(normalize_code(code))

    # Prereq course names
    prereq_names: dict[str, str] = {}
    if all_prereq_codes:
        for c in Course.objects.filter(course_code__in=all_prereq_codes):
            prereq_names[c.course_code.upper()] = c.description or ""

    wb = Workbook()
    wb.remove(wb.active)

    # ── Demand Sheet ─────────────────────────────────────────────
    ws = wb.create_sheet(title="Course Demand")

    # Title
    prog_label = program or "All Programs"
    sec_label = f"Section {section}" if section else "All Sections"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    tc = ws.cell(row=1, column=1)
    tc.value = f"Batch Recommender — {prog_label} | {sec_label} | {year}/T{semester} | {student_count} students"
    tc.fill = title_fill
    tc.font = title_font
    tc.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 6):
        ws.cell(row=1, column=c).fill = title_fill

    # Headers
    row = 2
    for col, (hdr, _w) in enumerate(
        [
            ("#", 5),
            ("Course", 14),
            ("Name", 32),
            ("Students", 12),
            ("Prerequisite Status", 40),
        ],
        start=1,
    ):
        cell = ws.cell(row=row, column=col)
        cell.value = hdr
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border

    # Data rows
    row = 3
    for rank, row_data in enumerate(sorted_rows, start=1):
        code = str(row_data["course_code"])
        count = int(row_data["count"])
        code_n = normalize_code(code)
        prereqs = prereq_map.get(code_n, [])
        name = _display_course_name(row_data)
        row_programs_raw = row_data.get("programs", [])
        row_programs = (
            [str(p) for p in row_programs_raw] if isinstance(row_programs_raw, list) else []
        )
        if row_programs:
            students_needing: set[int] = set()
            for prog in row_programs:
                students_needing.update(course_to_students_by_program.get((code_n, prog), set()))
        else:
            students_needing = set(course_to_students.get(code_n, set()))

        # ── Course header row ────────────────────────────────────
        # Rank
        ws.cell(row=row, column=1, value=rank).font = Font(size=9, bold=True, color="999999")
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = center

        # Course code (bold, colored background)
        cell = ws.cell(row=row, column=2, value=code)
        cell.font = course_font
        cell.fill = course_fill
        cell.border = thin_border

        # Name
        cell = ws.cell(row=row, column=3, value=name)
        cell.font = Font(size=8, color="444444")
        cell.fill = course_fill
        cell.border = thin_border

        # Student count
        cell = ws.cell(row=row, column=4, value=count)
        cell.font = count_font
        cell.alignment = center
        cell.fill = course_fill
        cell.border = thin_border
        # Demand gradient
        pct = round(count / max(student_count, 1) * 100)
        if pct >= 50:
            cell.fill = PatternFill(start_color="A9DFBF", end_color="A9DFBF", fill_type="solid")
        elif pct >= 30:
            cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

        # Prereq summary in course row
        if not prereqs:
            cell = ws.cell(row=row, column=5, value="No prerequisites")
            cell.font = no_prereq_font
            cell.fill = course_fill
            cell.border = thin_border
        else:
            prereq_studying_counts = {
                p: sum(1 for sid in students_needing if p in studying_by_student.get(sid, set()))
                for p in prereqs
            }
            studying_total = sum(prereq_studying_counts.values())
            if studying_total > 0:
                cell = ws.cell(
                    row=row,
                    column=5,
                    value=f"{len(prereqs)} prerequisites ({studying_total} of {count} still studying)",
                )
                cell.font = Font(size=8, bold=True, color="C03030")
                cell.fill = studying_fill
            else:
                cell = ws.cell(
                    row=row, column=5, value=f"{len(prereqs)} prerequisites (all passed)"
                )
                cell.font = Font(size=8, color="0A8E6E")
                cell.fill = passed_fill
            cell.border = thin_border

        row += 1

        # ── Prerequisite sub-rows ────────────────────────────────
        for pr in prereqs:
            studying_cnt = sum(
                1 for sid in students_needing if pr in studying_by_student.get(sid, set())
            )
            pr_name = prereq_names.get(pr.upper(), "")

            ws.cell(row=row, column=1).border = thin_border
            # Indented prereq code
            cell = ws.cell(row=row, column=2, value=f"  └ {pr}")
            cell.font = prereq_indent_font
            cell.border = thin_border

            # Prereq name
            cell = ws.cell(row=row, column=3, value=pr_name)
            cell.font = Font(size=8, color="888888")
            cell.border = thin_border

            ws.cell(row=row, column=4).border = thin_border

            # Studying count — among students who need THIS course
            if studying_cnt > 0:
                cell = ws.cell(
                    row=row, column=5, value=f"{studying_cnt} of {count} students still studying"
                )
                cell.font = studying_font
                cell.fill = studying_fill
            else:
                cell = ws.cell(row=row, column=5, value="All passed")
                cell.font = passed_font
                cell.fill = passed_fill
            cell.border = thin_border

            row += 1

    # Totals
    total_demand = sum(int(row.get("count", 0)) for row in sorted_rows)
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = Border(
            left=thin_side,
            right=thin_side,
            top=thick_side,
            bottom=thick_side,
        )
    ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True, size=10)
    ws.cell(row=row, column=3, value=f"{len(sorted_rows)} courses").font = Font(
        size=9, color="666666"
    )
    ws.cell(row=row, column=4, value=total_demand).font = Font(bold=True, size=10, color="0A8E6E")
    ws.cell(row=row, column=4).alignment = center

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 40
    ws.freeze_panes = "A3"

    # ── Summary Sheet ────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="Summary")

    ws_sum.cell(row=1, column=1, value="Batch Recommender Report").font = Font(bold=True, size=12)
    ws_sum.cell(row=3, column=1, value="Year").font = Font(bold=True, size=9)
    ws_sum.cell(row=3, column=2, value=year)
    ws_sum.cell(row=3, column=3, value="Semester").font = Font(bold=True, size=9)
    ws_sum.cell(row=3, column=4, value=semester)
    ws_sum.cell(row=4, column=1, value="Program").font = Font(bold=True, size=9)
    ws_sum.cell(row=4, column=2, value=program or "All")
    ws_sum.cell(row=4, column=3, value="Section").font = Font(bold=True, size=9)
    ws_sum.cell(row=4, column=4, value=section or "All")

    ws_sum.cell(row=6, column=1, value="Students Scanned").font = Font(bold=True, size=9)
    ws_sum.cell(row=6, column=2, value=student_count).font = Font(
        bold=True, size=14, color="0A8E6E"
    )
    ws_sum.cell(row=7, column=1, value="Courses Found").font = Font(bold=True, size=9)
    ws_sum.cell(row=7, column=2, value=len(sorted_rows))
    ws_sum.cell(row=8, column=1, value="Total Demand").font = Font(bold=True, size=9)
    ws_sum.cell(row=8, column=2, value=total_demand)

    # Top 10
    ws_sum.cell(row=10, column=1, value="Top 10 Courses").font = Font(bold=True, size=11)
    for col, hdr in enumerate(["Course", "Name", "Count"], start=1):
        cell = ws_sum.cell(row=11, column=col)
        cell.value = hdr
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border

    for i, row_data in enumerate(sorted_rows[:10]):
        code = str(row_data["course_code"])
        count = int(row_data["count"])
        r = 12 + i
        ws_sum.cell(row=r, column=1, value=code).font = Font(bold=True, size=9)
        ws_sum.cell(row=r, column=1).border = thin_border
        ws_sum.cell(row=r, column=2, value=_display_course_name(row_data)).font = Font(size=8)
        ws_sum.cell(row=r, column=2).border = thin_border
        ws_sum.cell(row=r, column=3, value=count).border = thin_border
        ws_sum.cell(row=r, column=3).alignment = center

    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 32
    ws_sum.column_dimensions["C"].width = 14

    # Save
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)
