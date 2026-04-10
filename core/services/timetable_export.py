"""
core/services/timetable_export.py
XLSX export for the Timetable Workspace.

Exports all boards in a scenario to a styled, print-ready workbook using
openpyxl.  The workbook structure is:

Sheets
------
* **Term N** (one per term level) -- contains:
    - A day-vs-slot timetable grid for each delivery board, with pastel
      colour-coding per course and red highlighting for time conflicts.
    - A course info sidebar (right of the grid) showing course name,
      credits, planned sections, and student demand.
    - A per-term conflict matrix showing how many students need each pair
      of courses simultaneously.
* **Conflicts (All)** -- a global NxN student-overlap heatmap across every
  course in the scenario.
* **Summary** -- scenario metadata, per-board stats (primary / visitor
  counts, conflict tallies), and the full section budget table.

Styling
-------
* Teal header fills (``#0A8E6E``), navy board headers (``#111144``),
  royal-blue info headers (``#4056E3``).
* 20 distinct pastel fills for course cells, conflict cells highlighted
  in salmon (``#F4CCCC``).
* Thick outer borders around every table block.

Called from ``timetable_workspace_views.tw_scenario_export_view``.
"""

from __future__ import annotations

import tempfile
from collections import defaultdict
from pathlib import Path

from core.models import (
    BoardStudentLink,
    DeliveryBoard,
    ScenarioSectionBudget,
    ScenarioStudentMap,
    SectionPlacement,
    TermSectionMeeting,
    TimetableScenario,
)
from core.services.timetable_autoplace import DEFAULT_SLOTS, get_meeting_pattern
from core.services.timetable_workspace import compute_scenario_budget, detect_board_conflicts


def export_scenario_xlsx(scenario_id: int) -> Path:
    """Export a scenario's timetable to a styled XLSX workbook.

    Parameters
    ----------
    scenario_id : int
        Primary key of the ``TimetableScenario`` to export.

    Returns
    -------
    Path
        Path to a temporary ``.xlsx`` file.  The caller is responsible for
        serving it (e.g. via ``FileResponse``) and cleaning up afterwards.

    Raises
    ------
    RuntimeError
        If ``openpyxl`` is not installed.
    TimetableScenario.DoesNotExist
        If the scenario ID is invalid.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    scenario = TimetableScenario.objects.get(id=scenario_id)
    boards = list(
        DeliveryBoard.objects.filter(scenario=scenario).order_by("display_order", "label")
    )

    # ── Styles ───────────────────────────────────────────────────
    hdr_fill = PatternFill(start_color="0A8E6E", end_color="0A8E6E", fill_type="solid")
    hdr_font = Font(name="Consolas", size=9, bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="CCCCCC")
    thick_side = Side(style="medium", color="333333")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def _apply_outer_border(ws, min_row, min_col, max_row, max_col):
        """Apply a thick (medium-weight) outer border around a rectangular cell range.

        Preserves existing inner borders; only overwrites the edge sides.
        """
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                existing = cell.border
                left = thick_side if c == min_col else existing.left
                right = thick_side if c == max_col else existing.right
                top = thick_side if r == min_row else existing.top
                bottom = thick_side if r == max_row else existing.bottom
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    bold_font = Font(bold=True, size=9)
    normal_font = Font(name="Consolas", size=9)
    center_align = Alignment(horizontal="center", vertical="center")
    conflict_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    board_hdr_fill = PatternFill(start_color="111144", end_color="111144", fill_type="solid")
    board_hdr_font = Font(name="Consolas", size=10, bold=True, color="FFFFFF")
    slot_fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
    info_hdr_fill = PatternFill(start_color="4056E3", end_color="4056E3", fill_type="solid")
    info_hdr_font = Font(name="Consolas", size=9, bold=True, color="FFFFFF")

    DAY_LABELS = ["SUN", "MON", "TUE", "WED", "THU"]
    # Course info sidebar: after slots (col B+) + 1 gap column
    # With 5 slots: B-F = slots, G = gap, H = info start (col 8)
    INFO_START_COL = 2 + len(DEFAULT_SLOTS) + 1  # dynamic based on slot count

    # ── Distinct pastel colors for courses ───────────────────────
    # 20 visually distinct pastels with good contrast for dark text
    COURSE_COLORS = [
        "D4E6F1",  # light blue
        "D5F5E3",  # light green
        "FADBD8",  # light pink
        "FCF3CF",  # light yellow
        "D7BDE2",  # light purple
        "A9DFBF",  # mint
        "F9E79F",  # gold
        "AED6F1",  # sky blue
        "F5CBA7",  # peach
        "A3E4D7",  # teal light
        "E8DAEF",  # lavender
        "FDEBD0",  # cream
        "ABB2B9",  # silver
        "A2D9CE",  # sea green
        "F5B7B1",  # salmon
        "D6DBDF",  # light gray
        "ABEBC6",  # pale green
        "FAD7A0",  # light orange
        "D2B4DE",  # orchid
        "AEB6BF",  # cool gray
    ]

    def _course_fill(course_code: str, color_map: dict[str, str]) -> PatternFill:
        """Get or assign a distinct pastel fill for a course.

        Colours are assigned in round-robin order from ``COURSE_COLORS``.
        The mapping is shared across all groups/boards so the same course
        always gets the same colour within a single term sheet.
        """
        if course_code not in color_map:
            idx = len(color_map) % len(COURSE_COLORS)
            color_map[course_code] = COURSE_COLORS[idx]
        hex_c = color_map[course_code]
        return PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")

    wb = Workbook()
    wb.remove(wb.active)

    # Pre-load budget for all terms
    budget_all = list(ScenarioSectionBudget.objects.filter(scenario=scenario))
    budget_by_term: dict[int, list] = defaultdict(list)
    for b in budget_all:
        budget_by_term[b.programme_term or 0].append(b)

    # Pre-load course names from Course table
    from core.models import Course

    course_names: dict[str, str] = {}
    for c in Course.objects.all():
        course_names[c.course_code.upper()] = c.description or c.course_code

    # ── Group boards by nominal_term ─────────────────────────────
    boards_by_term: dict[int, list[DeliveryBoard]] = defaultdict(list)
    for b in boards:
        boards_by_term[b.nominal_term or 0].append(b)

    # ── One sheet per term level ─────────────────────────────────
    for term_num in sorted(boards_by_term.keys()):
        term_boards = boards_by_term[term_num]
        ws = wb.create_sheet(title=f"Term {term_num}")

        row = 1

        for board_idx, board in enumerate(term_boards):
            if board_idx > 0:
                row += 2

            # Board header
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1)
            primary_count = BoardStudentLink.objects.filter(
                board=board, link_type="primary"
            ).count()
            visitor_count = BoardStudentLink.objects.filter(
                board=board, link_type="visitor"
            ).count()
            cell.value = f"{board.label} — {primary_count} primary + {visitor_count} visitors"
            cell.fill = board_hdr_fill
            cell.font = board_hdr_font
            cell.alignment = center_align
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = board_hdr_fill
            row += 1

            # Slot config
            slot_config = scenario.slot_config or []
            if not slot_config:
                slot_config = DEFAULT_SLOTS
            from core.services.timetable_autoplace import DEFAULT_LAB_SLOTS

            lab_slot_config = scenario.lab_slot_config or DEFAULT_LAB_SLOTS  # noqa: F841

            DAY_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]

            # Load placements
            placements = list(
                SectionPlacement.objects.filter(board=board)
                .select_related("term_section")
                .order_by("day", "start_time")
            )

            # ── Split placements into student groups ─────────────
            # A single section may have multiple placement rows (one per
            # weekly meeting, e.g. a 4-credit course meets 3 times/week).
            # We group by unique (course_code, section) so each section
            # appears once in the timetable grid, not once per meeting.
            from collections import defaultdict as _dd

            seen_sections: dict[str, dict[str, list]] = _dd(
                dict
            )  # code -> {sec_label -> [placements]}
            for p in placements:
                code = p.term_section.course_code
                sec = p.term_section.section
                if sec not in seen_sections[code]:
                    seen_sections[code][sec] = []
                seen_sections[code][sec].append(p)

            # course_sections: code -> [representative placement per unique section], sorted by label
            course_sections: dict[str, list] = {}
            for code, sec_dict in seen_sections.items():
                course_sections[code] = [
                    sec_dict[sec][0]  # pick first placement as representative
                    for sec in sorted(sec_dict.keys())
                ]

            max_groups = max((len(secs) for secs in course_sections.values()), default=1)
            # Also keep all placements per section for grid building
            section_all_placements: dict[int, list] = _dd(
                list
            )  # term_section_id -> [all placements]
            for p in placements:
                section_all_placements[p.term_section_id].append(p)

            # ── Smart group assignment: minimise time conflicts per group ──
            # When a course has multiple sections (e.g. S1, S2, S3), we need
            # to assign one section per group.  Instead of a naive round-robin
            # (S1->G1, S2->G2), we use bitmask overlap counting to greedily
            # pick the section that clashes least with sections already in
            # that group.  This produces cleaner, more printable timetables.
            from core.services.timetable_workspace import _time_mask

            def _section_bitmask(ts_id: int) -> int:
                """Compute a combined time-slot bitmask for all meetings of a section.

                Each bit represents a unique (day, time-slot) combination.
                Two sections overlap iff their bitmasks have common set bits.
                """
                mask = 0
                for pp in section_all_placements.get(ts_id, []):  # noqa: B023
                    mask |= _time_mask(pp.day, pp.start_time, pp.end_time)
                return mask

            # Build: code -> [(representative_placement, bitmask)] sorted by section label
            course_options: dict[str, list[tuple]] = {}
            for code, secs in course_sections.items():
                course_options[code] = [(s, _section_bitmask(s.term_section_id)) for s in secs]

            # Greedy assignment: for each group, pick the section per course
            # that has the least overlap with already-assigned sections in this group
            group_assignments: list[list] = []  # group_idx -> [placement representatives]
            used_sections: dict[str, set[int]] = _dd(set)  # code -> set of used indices

            for _g in range(max_groups):
                group_mask = 0  # combined bitmask of all sections assigned to this group
                group = []
                # Sort courses by demand descending (most constrained first)
                sorted_codes = sorted(course_options.keys(), key=lambda c: -len(course_options[c]))
                for code in sorted_codes:
                    options = course_options[code]
                    best_idx = None
                    best_conflicts = float("inf")
                    for idx, (_sec_p, sec_mask) in enumerate(options):
                        if idx in used_sections[code]:
                            continue  # already assigned to a previous group
                        # Count overlap bits with current group
                        overlap = bin(group_mask & sec_mask).count("1")
                        if overlap < best_conflicts:
                            best_conflicts = overlap
                            best_idx = idx
                    if best_idx is not None:
                        sec_p, sec_mask = options[best_idx]
                        group.append(sec_p)
                        group_mask |= sec_mask
                        used_sections[code].add(best_idx)

                group_assignments.append(group)

            # Assign distinct pastel color per course (shared across all groups)
            course_color_map: dict[str, str] = {}

            num_slots = len(slot_config)

            for group_idx, group_placements in enumerate(group_assignments):
                if not group_placements:
                    continue
                if group_idx > 0:
                    row += 1

                # Group header
                if max_groups > 1:
                    ws.merge_cells(
                        start_row=row, start_column=1, end_row=row, end_column=1 + num_slots
                    )
                    gcell = ws.cell(row=row, column=1)
                    gcell.value = f"Group {group_idx + 1} — {len(group_placements)} courses"
                    gcell.font = Font(bold=True, size=9, color="4056E3")
                    gcell.alignment = center_align
                    row += 1

                # ── Transposed layout: columns=slots, rows=days ──
                grid_start_row = row
                ws.cell(row=row, column=1).border = thin_border
                for s_idx, _slot in enumerate(slot_config):
                    cell = ws.cell(row=row, column=2 + s_idx)
                    cell.value = s_idx + 1
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = hdr_align
                    cell.border = thin_border
                row += 1

                ws.cell(row=row, column=1).border = thin_border
                for s_idx, slot in enumerate(slot_config):
                    cell = ws.cell(row=row, column=2 + s_idx)
                    cell.value = f"{slot['start']}-{slot['end']}"
                    cell.fill = hdr_fill
                    cell.font = Font(name="Consolas", size=8, bold=True, color="FFFFFF")
                    cell.alignment = hdr_align
                    cell.border = thin_border
                row += 1

                # Helper: convert "HH:MM" to minutes
                def _tm(t):
                    return int(t.split(":")[0]) * 60 + int(t.split(":")[1])

                # Helper: classify placement as lab (>80 min duration)
                def _is_lab_p(pp):
                    return (_tm(pp.end_time) - _tm(pp.start_time)) > 80

                # Helper: find the best-overlapping lecture slot for a lab placement
                def _best_lecture_slot(pp):
                    """Return the slot_config index whose time range overlaps
                    most with the lab placement, or the closest by start time."""
                    lab_s, lab_e = _tm(pp.start_time), _tm(pp.end_time)
                    best_idx, best_overlap = 0, -1
                    for i, slot in enumerate(slot_config):  # noqa: B023
                        s_s, s_e = _tm(slot["start"]), _tm(slot["end"])
                        overlap = max(0, min(lab_e, s_e) - max(lab_s, s_s))
                        if overlap > best_overlap:
                            best_overlap = overlap
                            best_idx = i
                    return best_idx

                # Build unified grid: day → slot_key → {text, courses, lab_times}
                # Labs go into the same grid as lectures, matched to best-overlapping slot
                grid: dict[str, dict[str, dict]] = {}
                for day in DAY_LABELS:
                    grid[day] = {}
                    for slot in slot_config:
                        sk = f"{slot['start']}-{slot['end']}"
                        grid[day][sk] = {"text": "", "courses": [], "lab_time": ""}

                for p in group_placements:
                    ts_placements = section_all_placements.get(p.term_section_id, [])
                    for pp in ts_placements:
                        day = pp.day.upper()[:3]
                        if day not in grid:
                            continue

                        is_lab = _is_lab_p(pp)

                        if is_lab:
                            # Find best-overlapping lecture slot column
                            best_idx = _best_lecture_slot(pp)
                            slot_key = (
                                f"{slot_config[best_idx]['start']}-{slot_config[best_idx]['end']}"
                            )
                        else:
                            slot_key = f"{pp.start_time}-{pp.end_time}"
                            if slot_key not in grid.get(day, {}):
                                # Match by start time
                                for slot in slot_config:
                                    if slot["start"] == pp.start_time:
                                        slot_key = f"{slot['start']}-{slot['end']}"
                                        break

                        ts = pp.term_section
                        meeting = TermSectionMeeting.objects.filter(
                            term_section=ts, day=day
                        ).first()
                        instructor = meeting.instructor if meeting else ""
                        room = pp.room or (meeting.room if meeting else "")
                        text = f"{ts.course_code} {ts.section}"
                        if instructor:
                            text += f"\n{instructor}"
                        if room:
                            text += f"\n{room}"

                        if day in grid and slot_key in grid[day]:
                            cd = grid[day][slot_key]
                            if cd["text"]:
                                cd["text"] += f"\n---\n{text}"
                            else:
                                cd["text"] = text
                            cd["courses"].append(ts.course_code)
                            if is_lab:
                                cd["lab_time"] = f"{pp.start_time}-{pp.end_time}"

                # ── Write day rows (unified: lectures + labs in same grid) ──
                from openpyxl.cell.rich_text import CellRichText, TextBlock
                from openpyxl.cell.text import InlineFont

                for _day_idx, (day_code, day_name) in enumerate(
                    zip(DAY_LABELS, DAY_NAMES, strict=False)
                ):
                    cell = ws.cell(row=row, column=1)
                    cell.value = day_name
                    cell.font = bold_font
                    cell.fill = slot_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    for s_idx, slot in enumerate(slot_config):
                        sk = f"{slot['start']}-{slot['end']}"
                        cell = ws.cell(row=row, column=2 + s_idx)
                        cd = grid.get(day_code, {}).get(
                            sk, {"text": "", "courses": [], "lab_time": ""}
                        )

                        if cd["lab_time"]:
                            # Rich text: course name in normal + lab time in red bold
                            # NOTE: do NOT set cell.font after CellRichText — it overrides rich text
                            normal_if = InlineFont(sz=8.5, rFont="Consolas")
                            red_if = InlineFont(sz=8, rFont="Consolas", b=True, color="C03030")
                            cell.value = CellRichText(
                                TextBlock(normal_if, cd["text"]),
                                TextBlock(red_if, f"\n{cd['lab_time']}"),
                            )
                        else:
                            cell.value = cd["text"]
                            cell.font = normal_font

                        cell.border = thin_border
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )

                        if len(cd["courses"]) > 1:
                            cell.fill = conflict_fill
                        elif len(cd["courses"]) == 1:
                            cell.fill = _course_fill(cd["courses"][0], course_color_map)

                    # Increase row height if any cell has lab time (needs 2 lines)
                    has_lab_in_row = any(
                        grid.get(day_code, {}).get(f"{s['start']}-{s['end']}", {}).get("lab_time")
                        for s in slot_config
                    )
                    if has_lab_in_row:
                        ws.row_dimensions[row].height = 42
                    row += 1

                # Outer border for this group's grid
                _apply_outer_border(ws, grid_start_row, 1, row - 1, 1 + num_slots)

            # Conflicts below grid
            conflicts = detect_board_conflicts(board.id)
            summary = conflicts["summary"]
            if summary["critical"] > 0 or summary["warning"] > 0:
                row += 1
                ws.cell(row=row, column=1, value="Conflicts:").font = Font(
                    bold=True, color="C03030", size=9
                )
                row += 1
                for o in conflicts.get("overlaps", []):
                    ws.cell(
                        row=row, column=1, value=f"OVERLAP: {' vs '.join(o.get('sections', []))}"
                    ).font = Font(color="C03030", size=8)
                    ws.cell(row=row, column=3, value=o.get("detail", ""))
                    row += 1
                for c in conflicts.get("instructor_clashes", []):
                    ws.cell(
                        row=row, column=1, value=f"INSTRUCTOR: {c.get('instructor', '')}"
                    ).font = Font(color="C03030", size=8)
                    ws.cell(row=row, column=3, value=" vs ".join(c.get("sections", [])))
                    row += 1
                for c in conflicts.get("room_clashes", []):
                    ws.cell(row=row, column=1, value=f"ROOM: {c.get('room', '')}").font = Font(
                        color="D97706", size=8
                    )
                    ws.cell(row=row, column=3, value=" vs ".join(c.get("sections", [])))
                    row += 1

        # ── Course Info Sidebar (right side of each sheet) ───────
        term_budget = budget_by_term.get(term_num, [])
        if term_budget:
            info_row = 1
            # Header
            info_headers = ["Course", "Name", "Cr", "Sec.", "Students"]
            for ci, hdr in enumerate(info_headers):
                cell = ws.cell(row=info_row, column=INFO_START_COL + ci)
                cell.value = hdr
                cell.fill = info_hdr_fill
                cell.font = info_hdr_font
                cell.alignment = hdr_align
                cell.border = thin_border
            info_row += 1

            total_sections = 0
            total_students = 0
            for b in sorted(term_budget, key=lambda x: x.course_code):
                cr = b.credit_hours or 0
                c_cell = ws.cell(row=info_row, column=INFO_START_COL, value=b.course_code)
                c_cell.font = Font(bold=True, size=9)
                c_cell.border = thin_border
                # Match the timetable grid color for this course
                c_cell.fill = _course_fill(b.course_code, course_color_map)
                name = course_names.get(b.course_code.upper(), b.course_code)
                ws.cell(row=info_row, column=INFO_START_COL + 1, value=name).font = normal_font
                ws.cell(row=info_row, column=INFO_START_COL + 1).border = thin_border
                ws.cell(row=info_row, column=INFO_START_COL + 2, value=cr).font = normal_font
                ws.cell(row=info_row, column=INFO_START_COL + 2).border = thin_border
                ws.cell(row=info_row, column=INFO_START_COL + 2).alignment = center_align
                ws.cell(
                    row=info_row, column=INFO_START_COL + 3, value=b.planned_sections
                ).font = normal_font
                ws.cell(row=info_row, column=INFO_START_COL + 3).border = thin_border
                ws.cell(row=info_row, column=INFO_START_COL + 3).alignment = center_align
                ws.cell(
                    row=info_row, column=INFO_START_COL + 4, value=b.total_demand
                ).font = normal_font
                ws.cell(row=info_row, column=INFO_START_COL + 4).border = thin_border
                ws.cell(row=info_row, column=INFO_START_COL + 4).alignment = center_align
                total_sections += b.planned_sections
                total_students += b.total_demand
                info_row += 1

            # Totals row
            ws.cell(row=info_row, column=INFO_START_COL, value="TOTAL").font = Font(
                bold=True, size=9
            )
            ws.cell(row=info_row, column=INFO_START_COL).border = thin_border
            ws.cell(row=info_row, column=INFO_START_COL + 1).border = thin_border
            ws.cell(row=info_row, column=INFO_START_COL + 2).border = thin_border
            ws.cell(row=info_row, column=INFO_START_COL + 3, value=total_sections).font = Font(
                bold=True, size=9
            )
            ws.cell(row=info_row, column=INFO_START_COL + 3).border = thin_border
            ws.cell(row=info_row, column=INFO_START_COL + 3).alignment = center_align
            ws.cell(row=info_row, column=INFO_START_COL + 4, value=total_students).font = Font(
                bold=True, size=9
            )
            ws.cell(row=info_row, column=INFO_START_COL + 4).border = thin_border
            ws.cell(row=info_row, column=INFO_START_COL + 4).alignment = center_align

            # Outer border for course info table
            _apply_outer_border(ws, 1, INFO_START_COL, info_row, INFO_START_COL + 4)

            # ── Per-Term Conflict Matrix (right of course info table) ─
            # Course info ends at column L (INFO_START_COL + 4 = col 12)
            # Gap column M (13), matrix starts at column N (14)
            MATRIX_START_COL = INFO_START_COL + 6  # col 14 = N
            term_courses = {b.course_code for b in term_budget}
            _write_mini_conflict_matrix(
                ws,
                1,
                MATRIX_START_COL,
                scenario,
                term_num,
                term_courses,
                hdr_fill,
                hdr_font,
                hdr_align,
                thin_border,
                normal_font,
                center_align,
            )

        # Column widths: A = day name, B..F+ = time slots
        ws.column_dimensions["A"].width = 13
        for s_idx in range(num_slots):
            cl = chr(ord("B") + s_idx) if s_idx < 25 else None
            if cl:
                ws.column_dimensions[cl].width = 18
        # Gap + course info sidebar columns
        gap_col_idx = 1 + num_slots + 1  # after slots + 1 gap
        for ci, w in enumerate([3, 12, 30, 6, 6, 10]):  # gap, Course, Name, Cr, Sec, Students
            cl = chr(ord("A") + gap_col_idx - 1 + ci) if (gap_col_idx + ci - 1) < 26 else None
            if cl:
                ws.column_dimensions[cl].width = w

    # ── Conflict Matrix Sheet ────────────────────────────────────
    _build_conflict_matrix_sheet(
        wb,
        scenario,
        hdr_fill,
        hdr_font,
        hdr_align,
        thin_border,
        normal_font,
        bold_font,
        center_align,
    )

    # ── Summary Sheet ────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="Summary")

    ws_sum.cell(row=1, column=1, value="Scenario").font = bold_font
    ws_sum.cell(row=1, column=2, value=scenario.name)
    ws_sum.cell(row=2, column=1, value="Year").font = bold_font
    ws_sum.cell(row=2, column=2, value=scenario.academic_year)
    ws_sum.cell(row=2, column=3, value="Term").font = bold_font
    ws_sum.cell(row=2, column=4, value=scenario.term)
    ws_sum.cell(row=3, column=1, value="Status").font = bold_font
    ws_sum.cell(row=3, column=2, value=scenario.status)

    # Board summary
    row = 5
    ws_sum.cell(row=row, column=1, value="Boards").font = Font(bold=True, size=11)
    row += 1
    for col_idx, hdr in enumerate(
        ["Board", "Term", "Primary", "Visitors", "Sections", "Critical", "Warnings"], start=1
    ):
        cell = ws_sum.cell(row=row, column=col_idx)
        cell.value = hdr
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border
    row += 1

    for board in boards:
        primary = BoardStudentLink.objects.filter(board=board, link_type="primary").count()
        visitor = BoardStudentLink.objects.filter(board=board, link_type="visitor").count()
        placed = SectionPlacement.objects.filter(board=board).count()
        conflicts = detect_board_conflicts(board.id)
        ws_sum.cell(row=row, column=1, value=board.label).border = thin_border
        ws_sum.cell(row=row, column=2, value=board.nominal_term).border = thin_border
        ws_sum.cell(row=row, column=3, value=primary).border = thin_border
        ws_sum.cell(row=row, column=4, value=visitor).border = thin_border
        ws_sum.cell(row=row, column=5, value=placed).border = thin_border
        c_cell = ws_sum.cell(row=row, column=6, value=conflicts["summary"]["critical"])
        c_cell.border = thin_border
        if conflicts["summary"]["critical"] > 0:
            c_cell.fill = conflict_fill
        w_cell = ws_sum.cell(row=row, column=7, value=conflicts["summary"]["warning"])
        w_cell.border = thin_border
        row += 1

    # Outer border for boards table (header row to last data row)
    _apply_outer_border(ws_sum, 6, 1, row - 1, 7)

    # Section budget
    row += 2
    ws_sum.cell(row=row, column=1, value="Section Budget").font = Font(bold=True, size=11)
    row += 1
    for col_idx, hdr in enumerate(
        ["Course", "Term", "Credits", "Meetings/wk", "Planned", "Used", "Remaining", "Demand"],
        start=1,
    ):
        cell = ws_sum.cell(row=row, column=col_idx)
        cell.value = hdr
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border
    row += 1

    budget = compute_scenario_budget(scenario_id)
    for b in budget:
        cr = b.get("credit_hours", 3)
        meetings = len(get_meeting_pattern(cr))
        ws_sum.cell(row=row, column=1, value=b["course_code"]).border = thin_border
        ws_sum.cell(row=row, column=2, value=b.get("programme_term")).border = thin_border
        ws_sum.cell(row=row, column=3, value=cr).border = thin_border
        ws_sum.cell(row=row, column=4, value=f"{meetings}x").border = thin_border
        ws_sum.cell(row=row, column=5, value=b["planned_sections"]).border = thin_border
        ws_sum.cell(row=row, column=6, value=b["used_sections"]).border = thin_border
        rem_cell = ws_sum.cell(row=row, column=7, value=b["remaining_sections"])
        rem_cell.border = thin_border
        if b["remaining_sections"] > 0:
            rem_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        ws_sum.cell(row=row, column=8, value=b["total_demand"]).border = thin_border
        row += 1

    # Outer border for budget table
    budget_hdr_row = row - len(budget) - 1  # header row of budget table
    _apply_outer_border(ws_sum, budget_hdr_row, 1, row - 1, 8)

    ws_sum.column_dimensions["A"].width = 16
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws_sum.column_dimensions[col_letter].width = 14
    ws_sum.freeze_panes = "A2"

    # ── Room Schedule Sheets ───────────────────────────────────────
    from core.models import Room

    # Get all rooms actually used in this scenario's placements
    used_room_codes = set(
        SectionPlacement.objects.filter(board__scenario=scenario)
        .exclude(room="")
        .exclude(room="UNASSIGNED")
        .values_list("room", flat=True)
        .distinct()
    )
    all_rooms = list(
        Room.objects.filter(room_code__in=used_room_codes).order_by("room_type", "room_code")
    )

    if all_rooms:
        lecture_slots = scenario.slot_config or DEFAULT_SLOTS
        lab_slots = scenario.lab_slot_config or []

        room_hdr_fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
        room_hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        room_cell_font = Font(name="Consolas", size=8.5, bold=True)
        room_border = Border(
            top=Side(style="thin", color="D5D8DC"),
            bottom=Side(style="thin", color="D5D8DC"),
            left=Side(style="thin", color="D5D8DC"),
            right=Side(style="thin", color="D5D8DC"),
        )

        # Collect all placements across all boards for the room view
        all_board_placements = list(
            SectionPlacement.objects.filter(board__scenario=scenario)
            .exclude(room="")
            .exclude(room="UNASSIGNED")
            .select_related("term_section")
        )

        if all_board_placements:
            room_info = {r.room_code: r for r in all_rooms}
            used_rooms = sorted({p.room for p in all_board_placements})

            # Build room grid: room_code → {(day, start) → [course_texts]}
            room_grid: dict[str, dict[tuple[str, str], list[str]]] = defaultdict(
                lambda: defaultdict(list)
            )
            for p in all_board_placements:
                text = f"{p.term_section.course_code} {p.term_section.section}"
                room_grid[p.room][(p.day, p.start_time)].append(text)

            # Build slot list (lecture + lab, sorted, no duplicates)
            lecture_slots = scenario.slot_config or DEFAULT_SLOTS
            lab_slots = scenario.lab_slot_config or []
            all_slot_list: list[dict] = []
            for s in lecture_slots:
                all_slot_list.append({"start": s["start"], "end": s["end"], "type": "L"})
            for s in lab_slots:
                all_slot_list.append({"start": s["start"], "end": s["end"], "type": "Lab"})
            seen_starts: set[str] = set()
            unique_slots: list[dict] = []
            for s in sorted(all_slot_list, key=lambda x: x["start"]):
                if s["start"] not in seen_starts:
                    seen_starts.add(s["start"])
                    unique_slots.append(s)

            num_slot_cols = len(unique_slots)

            # Separate lecture rooms and lab rooms
            lecture_rooms = [
                rc
                for rc in used_rooms
                if room_info.get(rc) and room_info[rc].room_type == "lecture"
            ]
            lab_room_codes = [
                rc for rc in used_rooms if room_info.get(rc) and room_info[rc].room_type == "lab"
            ]

            # Create ONE room sheet with one table per room
            ws_room = wb.create_sheet(title="Rooms")
            ws_room.sheet_properties.tabColor = "2E86C1"

            prayer_fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
            lab_room_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
            room_name_fill = PatternFill(
                start_color="0A8E6E", end_color="0A8E6E", fill_type="solid"
            )
            room_name_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

            current_row = 1

            # Title
            ws_room.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=1 + num_slot_cols
            )
            tc = ws_room.cell(row=1, column=1, value="Lectures")
            tc.font = Font(bold=True, size=14)
            tc.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            tc.alignment = Alignment(horizontal="center", vertical="center")
            for c in range(2, 2 + num_slot_cols):
                ws_room.cell(row=1, column=c).fill = PatternFill(
                    start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"
                )
            current_row = 3

            def _write_room_table(ws, row_start, room_code, grid_data, slots, info):
                """Write one room table block. Returns next row after the block."""
                r_obj = info.get(room_code)
                cap = r_obj.capacity if r_obj else "?"
                rtype = r_obj.room_type if r_obj else "lecture"
                is_lab = rtype == "lab"

                row = row_start
                # Room header row
                ws.merge_cells(
                    start_row=row, start_column=1, end_row=row, end_column=1 + len(slots)
                )
                rc = ws.cell(row=row, column=1, value=f"{room_code} ({cap})")
                rc.font = room_name_font
                rc.fill = lab_room_fill if is_lab else room_name_fill
                rc.alignment = Alignment(horizontal="left", vertical="center")
                for c in range(2, 2 + len(slots)):
                    ws.cell(row=row, column=c).fill = lab_room_fill if is_lab else room_name_fill
                row += 1

                # Slot header row
                ws.cell(row=row, column=1).border = room_border
                for si, slot in enumerate(slots):
                    c = ws.cell(row=row, column=2 + si, value=f"{slot['start']}-{slot['end']}")
                    c.font = room_hdr_font
                    c.fill = room_hdr_fill
                    c.alignment = Alignment(horizontal="center")
                    c.border = room_border
                row += 1

                # Day rows
                _days = ["SUN", "MON", "TUE", "WED", "THU"]
                _day_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]
                for day_code, day_name in zip(_days, _day_names, strict=False):
                    dc = ws.cell(row=row, column=1, value=day_name)
                    dc.font = Font(bold=True, size=9)
                    dc.border = room_border

                    for si, slot in enumerate(slots):
                        cell = ws.cell(row=row, column=2 + si)
                        cell.border = room_border
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )

                        # Prayer break column (12:00-13:00 area)
                        if slot["start"] >= "11:35" and slot["start"] <= "12:59":
                            cell.fill = prayer_fill

                        texts = grid_data.get((day_code, slot["start"]), [])
                        if texts:
                            cell.value = "\n".join(texts)
                            cell.font = room_cell_font
                            cell.fill = _course_fill(texts[0].split()[0], course_color_map)
                            if len(texts) > 1:
                                # Multiple courses = conflict, red tint
                                cell.fill = PatternFill(
                                    start_color="FADBD8", end_color="FADBD8", fill_type="solid"
                                )
                    row += 1

                return row + 1  # gap between room tables

            # Write lecture room tables
            for room_code in lecture_rooms:
                current_row = _write_room_table(
                    ws_room,
                    current_row,
                    room_code,
                    room_grid.get(room_code, {}),
                    unique_slots,
                    room_info,
                )

            # Lab rooms section
            if lab_room_codes:
                current_row += 1
                ws_room.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=1 + num_slot_cols,
                )
                tc = ws_room.cell(row=current_row, column=1, value="Labs")
                tc.font = Font(bold=True, size=14)
                tc.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
                tc.alignment = Alignment(horizontal="center")
                for c in range(2, 2 + num_slot_cols):
                    ws_room.cell(row=current_row, column=c).fill = PatternFill(
                        start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"
                    )
                current_row += 2

                for room_code in lab_room_codes:
                    current_row = _write_room_table(
                        ws_room,
                        current_row,
                        room_code,
                        room_grid.get(room_code, {}),
                        unique_slots,
                        room_info,
                    )

            # Column widths
            ws_room.column_dimensions["A"].width = 14
            from openpyxl.utils import get_column_letter as _gcl

            for si in range(num_slot_cols):
                ws_room.column_dimensions[_gcl(2 + si)].width = 15

    # ── Save ─────────────────────────────────────────────────────
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


def _write_mini_conflict_matrix(
    ws,
    start_row,
    start_col,
    scenario,
    term_num,
    term_courses,
    hdr_fill,
    hdr_font,
    hdr_align,
    thin_border,
    normal_font,
    center_align,
):
    """Write a per-term conflict matrix on an existing worksheet.

    The matrix shows how many students need each pair of courses
    simultaneously.  It includes **all** courses that the board's students
    need -- not just courses from this term's curriculum -- because a
    Term 3 student repeating MATH105 (Term 1) must not have it overlap
    with their Term 3 courses.  Cross-term courses are prefixed with
    ``*`` in both row and column headers and given a distinct header colour.

    Parameters
    ----------
    ws : openpyxl.worksheet.Worksheet
        The worksheet to write into.
    start_row, start_col : int
        Top-left cell coordinates for the matrix block.
    scenario : TimetableScenario
        The parent scenario (used to look up student maps).
    term_num : int
        The nominal term level for this matrix.
    term_courses : set[str]
        Course codes that belong to this term's curriculum (used to
        distinguish native vs cross-term courses).
    hdr_fill, hdr_font, hdr_align, thin_border, normal_font, center_align
        Pre-built openpyxl style objects shared across the workbook.
    """
    from openpyxl.styles import Border, Font, PatternFill, Side

    # Get students relevant to this term (primary + visitors)
    from core.models import BoardStudentLink

    board_ids = list(
        DeliveryBoard.objects.filter(scenario=scenario, nominal_term=term_num).values_list(
            "id", flat=True
        )
    )
    student_ids = set(
        BoardStudentLink.objects.filter(board_id__in=board_ids).values_list("student_id", flat=True)
    )

    # Get ALL recommended courses for these students (not filtered by term)
    student_maps = ScenarioStudentMap.objects.filter(scenario=scenario, student_id__in=student_ids)
    course_students: dict[str, set[int]] = defaultdict(set)
    for sm in student_maps:
        for code in sm.recommended_courses:
            course_students[code].add(sm.student_id)

    courses = sorted(course_students.keys())
    n = len(courses)
    if n == 0:
        return

    # Identify which courses are cross-term (not in this term's curriculum)
    cross_term_courses = {c for c in courses if c not in term_courses}
    native_count = n - len(cross_term_courses)

    # Title
    r = start_row
    ws.cell(row=r, column=start_col, value=f"Conflict Matrix — Term {term_num}").font = Font(
        bold=True, size=10
    )
    subtitle = f"{len(student_ids)} students, {native_count} courses"
    if cross_term_courses:
        subtitle += f" + {len(cross_term_courses)} cross-term (*)"
    ws.cell(row=r + 1, column=start_col, value=subtitle).font = Font(size=8, color="666666")
    r += 3

    # Build matrix
    sets = [course_students[c] for c in courses]
    matrix = [[0] * n for _ in range(n)]
    max_val = 0
    for i in range(n):
        matrix[i][i] = len(sets[i])
        for j in range(i + 1, n):
            shared = len(sets[i] & sets[j])
            matrix[i][j] = shared
            matrix[j][i] = shared
            if shared > max_val:
                max_val = shared

    # Header row
    cross_hdr_fill = PatternFill(start_color="4056E3", end_color="4056E3", fill_type="solid")
    ws.cell(row=r, column=start_col).border = thin_border
    for j, code in enumerate(courses):
        cell = ws.cell(row=r, column=start_col + 1 + j)
        is_cross = code in cross_term_courses
        cell.value = f"*{code}" if is_cross else code
        cell.fill = cross_hdr_fill if is_cross else hdr_fill
        cell.font = Font(name="Consolas", size=8, bold=True, color="FFFFFF")
        cell.alignment = center_align
        cell.border = thin_border
    r += 1

    diag_fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
    zero_font = Font(name="Consolas", size=8, color="CCCCCC")

    for i, code in enumerate(courses):
        is_cross = code in cross_term_courses
        cell = ws.cell(row=r, column=start_col)
        cell.value = f"*{code}" if is_cross else code
        cell.fill = cross_hdr_fill if is_cross else hdr_fill
        cell.font = Font(name="Consolas", size=8, bold=True, color="FFFFFF")
        cell.border = thin_border

        for j in range(n):
            cell = ws.cell(row=r, column=start_col + 1 + j)
            val = matrix[i][j]
            cell.value = val
            cell.border = thin_border
            cell.alignment = center_align

            if i == j:
                cell.fill = diag_fill
                cell.font = Font(name="Consolas", size=8, bold=True, color="0A8E6E")
            elif val == 0:
                cell.font = zero_font
            elif max_val > 0:
                t = min(val / max_val, 1.0)
                rc = int(200 - 150 * t)
                gc = int(230 - 140 * t)
                bc = int(220 - 10 * t)
                cell.fill = PatternFill(
                    start_color=f"{rc:02X}{gc:02X}{bc:02X}",
                    end_color=f"{rc:02X}{gc:02X}{bc:02X}",
                    fill_type="solid",
                )
                cell.font = Font(name="Consolas", size=8, bold=True)

        r += 1

    # Outer border for mini conflict matrix (header row to last data row)
    matrix_hdr_row = start_row + 3
    matrix_end_row = r - 1
    matrix_end_col = start_col + n
    if matrix_end_row >= matrix_hdr_row:
        thick = Side(style="medium", color="333333")
        for mr in range(matrix_hdr_row, matrix_end_row + 1):
            for mc in range(start_col, matrix_end_col + 1):
                cl = ws.cell(row=mr, column=mc)
                ex = cl.border
                cl.border = Border(
                    left=thick if mc == start_col else ex.left,
                    right=thick if mc == matrix_end_col else ex.right,
                    top=thick if mr == matrix_hdr_row else ex.top,
                    bottom=thick if mr == matrix_end_row else ex.bottom,
                )


def _build_conflict_matrix_sheet(
    wb, scenario, hdr_fill, hdr_font, hdr_align, thin_border, normal_font, bold_font, center_align
):
    """Build the global "Conflicts (All)" sheet -- an NxN student overlap heatmap.

    Unlike the per-term mini matrices, this sheet includes *every* course in
    the scenario regardless of term level.  The diagonal cells show total
    student count per course; off-diagonal cells show how many students need
    both courses.  Cells are colour-coded with a teal-to-royal gradient
    proportional to the overlap count.

    Parameters
    ----------
    wb : openpyxl.Workbook
        The workbook to add the sheet to.
    scenario : TimetableScenario
        The scenario whose ``ScenarioStudentMap`` records provide the data.
    hdr_fill, hdr_font, hdr_align, thin_border, normal_font, bold_font, center_align
        Pre-built openpyxl style objects shared across the workbook.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws = wb.create_sheet(title="Conflicts (All)")

    # Load student-course data from ScenarioStudentMap (all students, all courses)
    student_maps = ScenarioStudentMap.objects.filter(scenario=scenario)
    course_students: dict[str, set[int]] = defaultdict(set)
    for sm in student_maps:
        for code in sm.recommended_courses:
            course_students[code].add(sm.student_id)

    courses = sorted(course_students.keys())
    n = len(courses)

    if n == 0:
        ws.cell(row=1, column=1, value="No student data available")
        return

    # Build NxN matrix
    sets = [course_students[c] for c in courses]
    matrix = [[0] * n for _ in range(n)]
    max_val = 0
    for i in range(n):
        matrix[i][i] = len(sets[i])
        for j in range(i + 1, n):
            shared = len(sets[i] & sets[j])
            matrix[i][j] = shared
            matrix[j][i] = shared
            if shared > max_val:
                max_val = shared

    # Title
    ws.cell(row=1, column=1, value="Student Conflict Matrix").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=f"{n} courses, {len(student_maps)} students").font = Font(
        size=9, color="666666"
    )

    start_row = 4

    # Header row (course codes rotated)
    ws.cell(row=start_row, column=1, value="").border = thin_border
    for j, code in enumerate(courses):
        cell = ws.cell(row=start_row, column=j + 2)
        cell.value = code
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=90)
        cell.border = thin_border

    # Data rows
    diag_fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
    zero_font = Font(name="Consolas", size=9, color="CCCCCC")

    for i, code in enumerate(courses):
        r = start_row + 1 + i
        # Row header
        cell = ws.cell(row=r, column=1)
        cell.value = code
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = thin_border

        for j in range(n):
            cell = ws.cell(row=r, column=j + 2)
            val = matrix[i][j]
            cell.value = val
            cell.border = thin_border
            cell.alignment = center_align

            if i == j:
                # Diagonal: total students for this course
                cell.fill = diag_fill
                cell.font = Font(name="Consolas", size=9, bold=True, color="0A8E6E")
            elif val == 0:
                cell.font = zero_font
            else:
                # Heat-map: teal → royal gradient based on overlap count
                if max_val > 0:
                    t = min(val / max_val, 1.0)
                    r_c = int(200 - 150 * t)
                    g_c = int(230 - 140 * t)
                    b_c = int(220 - 10 * t)
                    hex_color = f"{r_c:02X}{g_c:02X}{b_c:02X}"
                    cell.fill = PatternFill(
                        start_color=hex_color, end_color=hex_color, fill_type="solid"
                    )
                    cell.font = Font(name="Consolas", size=9, bold=True)

    # Outer border for global conflict matrix
    if n > 0:
        thick = Side(style="medium", color="333333")
        end_row = start_row + n
        end_col = 1 + n
        for mr in range(start_row, end_row + 1):
            for mc in range(1, end_col + 1):
                cl = ws.cell(row=mr, column=mc)
                ex = cl.border
                cl.border = Border(
                    left=thick if mc == 1 else ex.left,
                    right=thick if mc == end_col else ex.right,
                    top=thick if mr == start_row else ex.top,
                    bottom=thick if mr == end_row else ex.bottom,
                )

    # Column widths
    ws.column_dimensions["A"].width = 12
    for j in range(n):
        col_letter = chr(ord("B") + j) if j < 25 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = 7

    ws.freeze_panes = "B5"
