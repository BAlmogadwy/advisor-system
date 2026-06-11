"""
core/services/exam_timetable.py
In-memory exam-timetable pipeline.

Pipeline sections (section numbers match ``# ── N.`` markers below):

0. Credit helpers             – build_credit_map, _credit_pair_penalty
1. Enrolled sets              – build_enrolled_sets → {course_code: {student_ids}}
2. Conflict graph             – build_conflict_graph → adjacency dict + edge list
3. Programme-plan term buckets – build_plan_term_buckets, check_bucket_feasibility
4. Greedy scheduler           – schedule → course→slot assignments (graph-coloring)
5. QA report                  – _build_qa → validation + soft-constraint metrics
6. Orchestrator               – build_exam_timetable → runs 0→5, persists JSON
7. Excel export               – export_exam_timetable_xlsx → styled .xlsx workbook
"""

from __future__ import annotations

import itertools
import json
import random

# ── Invigilator calculation rules ──────────────────────────────
#
# Department courses (CS, IS, COE, CYB, AI, DS) need invigilators FROM
# our department for every exam-room: 1 invigilator if the room holds
# fewer than 30 students, 2 invigilators if 30 or more.
#
# External / general-requirements courses (GS, EDCT, GSE, ENV, MATH,
# STAT, PHYS) only need an invigilator from our department when the
# room holds MORE than 30 students (1 invigilator), otherwise 0
# (the providing college supplies its own staff).
import re as _re
from collections import defaultdict
from pathlib import Path
from typing import Any

from core.models import (
    Course,
    ExamTimetableRun,
    ProgrammeRequirement,
    Room,
    Student,
    StudentCourse,
    StudentTermSection,
    TermSectionMeeting,
)
from core.services.course_identity import planner_course_key
from core.services.exam_run_schema import (
    STATUS_DERIVATION_VERSION,
    compute_enrolment_snapshot,
    derive_building_footprint,
    derive_multi_sitting_details,
    derive_status_surface,
    load_normalised_run,
    stamp_schema_version,
)

_DEPARTMENT_PREFIXES: set[str] = {"CS", "IS", "COE", "CYB", "AI", "DS"}
_EXTERNAL_PREFIXES: set[str] = {"GS", "EDCT", "GSE", "ENV", "MATH", "STAT", "PHYS"}

_DEPT_LARGE_THRESHOLD = 30  # >= triggers a second invigilator (department)
_EXT_LARGE_THRESHOLD = 30  # > triggers one invigilator from us (external)


def _course_prefix(course_code: str) -> str:
    """Extract the alphabetic prefix from a course code (e.g. 'CS112' → 'CS')."""
    if not course_code:
        return ""
    m = _re.match(r"^[A-Za-z]+", str(course_code))
    return m.group(0).upper() if m else ""


def _invigilators_needed(course_code: str, students_in_room: int) -> int:
    """Return how many invigilators FROM OUR DEPARTMENT this room needs.

    Department courses:  >=30 students → 2, otherwise 1
    External courses:    > 30 students → 1, otherwise 0
    Unknown prefix is treated as department (safer side).
    """
    prefix = _course_prefix(course_code)
    if prefix in _EXTERNAL_PREFIXES:
        return 1 if students_in_room > _EXT_LARGE_THRESHOLD else 0
    # Department or unknown
    return 2 if students_in_room >= _DEPT_LARGE_THRESHOLD else 1


# ── 0. Credit helpers ───────────────────────────────────────────

_CREDIT_DEFAULT = 3  # fallback for NULL / 0 / missing credit_hours

# Penalty ladder for the top-2 heaviest exams on a single day.
# Key = (max_credits, min_credits); value = penalty weight.
_CREDIT_PAIR_WEIGHTS: dict[tuple[int, int], int] = {
    (4, 4): 100,  # worst – two heavy courses
    (4, 3): 30,  # acceptable
    (4, 2): 0,  # ideal pairing
}
_CREDIT_PAIR_FALLBACK = 5  # other combos (3+3, 3+2, 2+2, …)


def build_credit_map(course_codes: list[str] | set[str]) -> dict[str, int]:
    """Return {course_code: credit_hours} for the given courses.

    Missing / NULL / zero credit_hours default to ``_CREDIT_DEFAULT``
    so the penalty formula degrades gracefully.
    """
    display_to_source = {str(cc): _source_code_for_display(str(cc)) for cc in course_codes}
    rows = Course.objects.filter(
        course_code__in=sorted(set(display_to_source.values())),
    ).values_list("course_code", "credit_hours")
    source_cm: dict[str, int] = {cc: (ch if ch and ch > 0 else _CREDIT_DEFAULT) for cc, ch in rows}
    cm: dict[str, int] = {}
    for display, source in display_to_source.items():
        cm[display] = source_cm.get(source, _CREDIT_DEFAULT)
    return cm


def _source_code_for_display(display_code: str, explicit_source: object = None) -> str:
    source = str(explicit_source or "").strip()
    if source:
        return source
    if display_code.endswith(")") and " (" in display_code:
        return display_code.rsplit(" (", 1)[0]
    return display_code


def _credit_pair_penalty(credits_on_day: list[int]) -> int:
    """Penalty for the top-2 heaviest exams on a single day for one student.

    Returns 0 when there are fewer than 2 exams.
    """
    if len(credits_on_day) < 2:
        return 0
    top2 = sorted(credits_on_day, reverse=True)[:2]
    pair = (top2[0], top2[1])
    return _CREDIT_PAIR_WEIGHTS.get(pair, _CREDIT_PAIR_FALLBACK)


# ── 1. Enrolled sets ────────────────────────────────────────────


def build_enrolled_sets(
    programs: list[str] | None = None,
    sections: list[str] | None = None,
) -> dict[str, set[int]]:
    """Return {course_code: {student_id, …}} for current-term enrolments.

    Uses StudentCourse.status='studying' as the exam source of truth.
    Same-code courses with different programme names are split into
    display keys such as "CS112 (1)" and "CS112 (2)".

    Optional filters narrow the student population:
        programs – only include students whose program is in this list
        sections – only include students whose section is in this list
    When a filter is None or empty, it is ignored (all values pass).
    """
    # Try StudentTermSection first (more accurate — actual section enrollments).
    # Detect the latest academic year/term that has ANY data. The fallback
    # only triggers when the source itself is empty — filters that match no
    # students must not silently cross over to StudentCourse.
    enrolled, _meta = build_enrolled_sets_with_meta(programs=programs, sections=sections)
    return enrolled

    latest = (
        StudentTermSection.objects.order_by("-academic_year", "-term")
        .values_list("academic_year", "term")
        .first()
    )

    if latest is not None:
        ay, tm = latest
        qs = StudentTermSection.objects.filter(academic_year=ay, term=tm).select_related(
            "term_section"
        )

        if programs:
            student_ids = set(
                Student.objects.filter(program__in=programs).values_list("student_id", flat=True)
            )
            qs = qs.filter(student_id__in=student_ids)
        if sections:
            student_ids_sec = set(
                Student.objects.filter(section__in=sections).values_list("student_id", flat=True)
            )
            qs = qs.filter(student_id__in=student_ids_sec)

        # NOTE: TermSection.course_code is the department prefix (e.g. "CS");
        # the full course identifier matching Course.course_code is course_key
        # (e.g. "CS101"). Grouping by course_code would collapse every CS
        # course into one bucket and destroy the exam schedule.
        rows = qs.values_list("term_section__course_key", "student_id")
        enrolled: dict[str, set[int]] = defaultdict(set)
        for course_key, student_id in rows:
            enrolled[course_key].add(student_id)
        return dict(enrolled)

    # Fallback: StudentCourse with status="studying" (no StudentTermSection data).
    qs_sc = StudentCourse.objects.filter(status="studying").select_related("course", "student")
    if programs:
        qs_sc = qs_sc.filter(student__program__in=programs)
    if sections:
        qs_sc = qs_sc.filter(student__section__in=sections)

    rows_sc = qs_sc.values_list("course__course_code", "student_id")
    enrolled_sc: dict[str, set[int]] = defaultdict(set)
    for course_code, student_id in rows_sc:
        enrolled_sc[course_code].add(student_id)
    return dict(enrolled_sc)


def build_enrolled_sets_with_meta(
    programs: list[str] | None = None,
    sections: list[str] | None = None,
) -> tuple[dict[str, set[int]], dict[str, dict]]:
    """Return selected students' studying courses plus display/source metadata."""
    qs_sc = StudentCourse.objects.filter(status="studying").select_related("course", "student")
    if programs:
        qs_sc = qs_sc.filter(student__program__in=programs)
    if sections:
        qs_sc = qs_sc.filter(student__section__in=sections)

    rows = list(
        qs_sc.values_list(
            "course__course_code",
            "course__description",
            "student_id",
            "student__program",
        )
    )
    source_codes = {str(code) for code, _desc, _sid, _program in rows}
    if not source_codes:
        latest = (
            StudentTermSection.objects.order_by("-academic_year", "-term")
            .values_list("academic_year", "term")
            .first()
        )
        if latest is None:
            return {}, {}

        ay, tm = latest
        qs_sts = StudentTermSection.objects.filter(academic_year=ay, term=tm).select_related(
            "term_section"
        )
        if programs:
            student_ids = set(
                Student.objects.filter(program__in=programs).values_list("student_id", flat=True)
            )
            qs_sts = qs_sts.filter(student_id__in=student_ids)
        if sections:
            student_ids_sec = set(
                Student.objects.filter(section__in=sections).values_list("student_id", flat=True)
            )
            qs_sts = qs_sts.filter(student_id__in=student_ids_sec)

        enrolled_sts: dict[str, set[int]] = defaultdict(set)
        meta_sts: dict[str, dict] = {}
        for course_key, course_name, student_id in qs_sts.values_list(
            "term_section__course_key",
            "term_section__course_name",
            "student_id",
        ):
            source_code = str(course_key or "").strip()
            if not source_code:
                continue
            enrolled_sts[source_code].add(int(student_id))
            meta_sts.setdefault(
                source_code,
                {
                    "source_course_code": source_code,
                    "course_name": str(course_name or "").strip(),
                    "course_identity": planner_course_key(source_code, course_name),
                },
            )
        return dict(enrolled_sts), meta_sts

    pr_rows = list(
        ProgrammeRequirement.objects.filter(course_code__in=source_codes).values_list(
            "program",
            "course_code",
            "course_name",
            "programme_term",
        )
    )
    pr_name_by_program_code = {
        (str(program), str(code)): str(name or "").strip() for program, code, name, _term in pr_rows
    }

    enrolled_by_identity: dict[tuple[str, str], set[int]] = defaultdict(set)
    identity_name: dict[tuple[str, str], str] = {}
    for source, course_desc, student_id, program in rows:
        source_code = str(source)
        name = (
            pr_name_by_program_code.get((str(program), source_code))
            or str(course_desc or "").strip()
        )
        identity = planner_course_key(source_code, name)
        enrolled_by_identity[(source_code, identity)].add(int(student_id))
        identity_name.setdefault((source_code, identity), name)

    identity_term_rank: dict[tuple[str, str], int] = {}
    for _program, source, name, term in pr_rows:
        source_code = str(source)
        identity = planner_course_key(source_code, name)
        key = (source_code, identity)
        rank = int(term or 999)
        identity_term_rank[key] = min(identity_term_rank.get(key, rank), rank)
        identity_name.setdefault(key, str(name or "").strip())

    identities_by_source: dict[str, list[str]] = defaultdict(list)
    for source, identity in enrolled_by_identity:
        identities_by_source[source].append(identity)
    for source in identities_by_source:
        identities_by_source[source] = sorted(
            set(identities_by_source[source]),
            key=lambda identity: (
                identity_term_rank.get((source, identity), 999),
                identity,
            ),
        )

    display_by_identity: dict[tuple[str, str], str] = {}
    for source, identities in identities_by_source.items():
        if len(identities) == 1:
            display_by_identity[(source, identities[0])] = source
        else:
            for idx, identity in enumerate(identities, start=1):
                display_by_identity[(source, identity)] = f"{source} ({idx})"

    enrolled: dict[str, set[int]] = {}
    meta: dict[str, dict] = {}
    for (source, identity), student_ids in enrolled_by_identity.items():
        display = display_by_identity[(source, identity)]
        enrolled[display] = set(student_ids)
        meta[display] = {
            "source_course_code": source,
            "course_name": identity_name.get((source, identity), ""),
            "course_identity": identity,
        }
    return dict(enrolled), meta


# ── 2. Conflict graph ──────────────────────────────────────────


def build_conflict_graph(
    enrolled_sets: dict[str, set[int]],
) -> tuple[list[dict], dict[str, dict[str, int]]]:
    """
    Build conflict edges from enrolled sets.

    Returns:
        conflicts  – list of {course_a, course_b, shared} with course_a < course_b
        adj        – adjacency dict {course: {neighbour: weight, …}}
    """
    # Invert: student_id → [course_codes] so we can iterate per-student
    student_courses: dict[int, list[str]] = defaultdict(list)
    for course_code, students in enrolled_sets.items():
        for sid in students:
            student_courses[sid].append(course_code)

    # Count pairwise overlaps: for each student, every pair of their courses
    # shares that student.  sorted() ensures (a,b) key is deterministic.
    edge_counts: dict[tuple[str, str], int] = defaultdict(int)
    for courses in student_courses.values():
        for a, b in itertools.combinations(sorted(set(courses)), 2):
            edge_counts[(a, b)] += 1

    # Build adjacency dict (bidirectional) + flat edge list for the frontend
    conflicts: list[dict] = []
    adj: dict[str, dict[str, int]] = defaultdict(dict)
    for (a, b), cnt in edge_counts.items():
        conflicts.append({"course_a": a, "course_b": b, "shared": cnt})
        adj[a][b] = cnt
        adj[b][a] = cnt

    return conflicts, dict(adj)


# ── 3. Programme-plan term buckets ─────────────────────────────


def build_plan_term_buckets(
    running_courses: set[str],
    course_meta: dict[str, dict] | None = None,
) -> tuple[dict[tuple[str, int], set[str]], dict[str, list[tuple[str, int]]]]:
    """Map running courses to (program, programme_term) buckets.

    Returns:
        buckets       – {(program, programme_term): {course_codes}}
        course_buckets – {course_code: [(program, term), …]} reverse index
    """
    meta = course_meta or {}
    source_to_display: dict[str, list[str]] = defaultdict(list)
    identity_by_display: dict[str, str] = {}
    for display in running_courses:
        display_code = str(display)
        m = meta.get(display_code, {})
        source = _source_code_for_display(display_code, m.get("source_course_code"))
        identity = str(m.get("course_identity") or source)
        source_to_display[source].append(display_code)
        identity_by_display[display_code] = identity

    rows = ProgrammeRequirement.objects.filter(
        course_code__in=set(source_to_display),
        programme_term__isnull=False,
    ).values_list("program", "course_code", "course_name", "programme_term")

    # Forward index: (program, term) → {course_codes}
    buckets: dict[tuple[str, int], set[str]] = defaultdict(set)
    # Reverse index: course_code → [(program, term), …]  (a course can appear
    # in multiple programmes, e.g. service courses shared across AI & DS)
    course_buckets: dict[str, list[tuple[str, int]]] = defaultdict(list)

    for program, course_code, course_name, programme_term in rows:
        key = (program, int(programme_term or 0))
        row_identity = planner_course_key(course_code, course_name)
        displays = [
            display
            for display in source_to_display.get(str(course_code), [])
            if identity_by_display.get(display, display) == row_identity
        ]
        if not displays:
            displays = [
                display
                for display in source_to_display.get(str(course_code), [])
                if identity_by_display.get(display, display) == str(course_code)
            ]
        for display in displays:
            buckets[key].add(display)
            course_buckets[display].append(key)

    return dict(buckets), dict(course_buckets)


def check_bucket_feasibility(
    buckets: dict[tuple[str, int], set[str]],
    num_days: int,
) -> list[dict]:
    """Return list of violations where a bucket has more courses than days.

    Each violation: {program, programme_term, bucket_size, num_days, courses}
    Empty list means all buckets are feasible.
    """
    violations: list[dict] = []
    for (program, term), courses in sorted(buckets.items()):
        if len(courses) > num_days:
            violations.append(
                {
                    "program": program,
                    "programme_term": term,
                    "bucket_size": len(courses),
                    "num_days": num_days,
                    "courses": sorted(courses),
                }
            )
    return violations


# ── 4. Greedy scheduler ───────────────────────────────────────


def schedule(
    courses: list[str],
    adj: dict[str, dict[str, int]],
    slots: list[dict],
    enrolled_sets: dict[str, set[int]] | None = None,
    max_per_day: int = 2,
    plan_term_buckets: dict[tuple[str, int], set[str]] | None = None,
    course_buckets: dict[str, list[tuple[str, int]]] | None = None,
    pinned: list[dict] | None = None,
    credit_map: dict[str, int] | None = None,
    preferred_slots: dict[str, int] | None = None,
    seed: int | None = None,
) -> list[dict]:
    """
    Greedy graph-coloring with day-spread soft constraint.

    Hard constraints:
      A. No two conflicting courses in the same slot (student clash).
      B. No two courses from the same (program, programme_term) bucket
         on the same day.

    Soft constraints (in priority order):
      1.   Minimise students with >max_per_day exams on one day.
      1.5  Credit-pair penalty: when multi-exam days are unavoidable,
           prefer lighter pairings (4+2 < 4+3 < 4+4).
      2.   Maximise spacing within (program, term) buckets (penalise
           small day gaps between bucket-mates).
      3.   Balance load across slots (prefer less-loaded slots as tiebreaker).

    Randomised tie-breaking:
      When ``seed`` is provided, courses with the same constraint degree
      are shuffled randomly before scheduling.  This produces a different
      (but equally valid) timetable on each run, letting the user pick
      the best variant.  When ``seed`` is None the order is deterministic
      (alphabetical within each tier).

    Args:
        courses            – list of course codes to schedule
        adj                – adjacency dict from build_conflict_graph
        slots              – list of {index, day, period} dicts
        enrolled_sets      – {course_code: {student_ids}} for soft-constraint scoring
        max_per_day        – soft cap on exams per student per day (default 2)
        plan_term_buckets  – {(program, term): {course_codes}} hard day-rule buckets
        course_buckets     – {course_code: [(program, term), …]} reverse index
        pinned             – list of {course_code, day, period} to fix before scheduling
        credit_map         – {course_code: credit_hours} for credit-pair penalty
        seed               – RNG seed for tie-breaking; None = deterministic order

    Returns:
        list of {course_code, slot_index, day, period}
    """
    # ── Preparation: build lookup tables ──
    max_slot_idx = max((s["index"] for s in slots), default=-1)
    slot_by_index: dict[int, dict] = {s["index"]: s for s in slots}

    # Day-index mapping: convert day names ("Sun", "Mon") to integers (0, 1)
    # so we can compute numeric spacing gaps between bucket-mates.
    unique_days: list[str] = []
    day_set: set[str] = set()
    for s in slots:
        if s["day"] not in day_set:
            day_set.add(s["day"])
            unique_days.append(s["day"])
    day_to_idx: dict[str, int] = {d: i for i, d in enumerate(unique_days)}

    # Shorthand aliases for optional dicts (avoid repeated `or {}` everywhere)
    _ptb = plan_term_buckets or {}  # (program, term) → {course_codes}
    _cb = course_buckets or {}  # course_code → [(program, term), …]
    _cm = credit_map or {}  # course_code → credit_hours
    _pref = preferred_slots or {}  # course_code → preferred slot_index

    def _constraint_degree(c: str) -> int:
        """Heuristic: courses with more conflicts + more bucket-mates are harder
        to place, so we schedule them first (most-constrained-first ordering)."""
        adj_deg = len(adj.get(c, {}))
        bucket_deg = sum(len(_ptb.get(bk, set())) for bk in _cb.get(c, []))
        return adj_deg + bucket_deg

    # Sort most-constrained first.  When seed is provided, we bucket
    # courses into WIDE degree bands (each band is ~20% of the degree
    # range) and shuffle inside every band.  Banding is wider than
    # strict equal-degree tiers so courses with close-but-different
    # degrees also mix together — this gives each run noticeably more
    # variety without letting low-degree courses leapfrog high-degree
    # ones entirely.
    if seed is not None:
        rng = random.Random(seed)
        degrees = [(_constraint_degree(c), c) for c in courses]
        if degrees:
            max_deg = max(d for d, _ in degrees)
            # Band width: ~10% of the max degree, floor of 1.
            band_width = max(1, max_deg // 10)
            # Bucket by band index (descending bands placed first)
            band_map: dict[int, list[str]] = defaultdict(list)
            for d, c in degrees:
                band_map[d // band_width].append(c)
            courses_sorted = []
            for band in sorted(band_map.keys(), reverse=True):
                members = band_map[band]
                rng.shuffle(members)
                courses_sorted.extend(members)
        else:
            courses_sorted = []
    else:
        rng = None
        # Deterministic: alphabetical within each tier (Python sort is stable)
        courses_sorted = sorted(courses, key=_constraint_degree, reverse=True)

    # ── Mutable state: updated as each course is placed ──
    #
    # assignment:          final result — which slot each course lands in
    # student_day_count:   how many exams each student has per day (Level 1 scoring)
    # student_day_courses: which courses each student has per day (Level 1.5 credit scoring)
    # slot_load:           how many courses are in each slot (Level 3 load-balancing)
    # bucket_day_courses:  which courses are assigned to each day per bucket (hard constraint B)
    # course_assigned_day: which day each course is on (spacing calculation)

    assignment: dict[str, int] = {}  # course_code → slot_index

    student_day_count: dict[int, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    student_day_courses: dict[int, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))

    slot_load: dict[int, int] = defaultdict(int)

    # Per-day course count — used as the primary load-balancing tiebreaker so
    # courses without other constraints (degree-0 / no bucket) get spread
    # across days instead of clustering on whichever slot index is lowest.
    day_load: dict[str, int] = defaultdict(int)

    bucket_day_courses: dict[tuple[str, int], dict[str, set[str]]] = defaultdict(
        lambda: defaultdict(set)
    )

    course_assigned_day: dict[str, str] = {}

    # Pre-assign pinned courses (user overrides — bypass constraints)
    if pinned:
        dp_to_slot = {(s["day"], s["period"]): s["index"] for s in slots}
        for pin in pinned:
            cc = pin.get("course_code", "")
            p_day = pin.get("day", "")
            p_period = pin.get("period", "")
            si = dp_to_slot.get((p_day, p_period))
            if si is None or cc not in set(courses):
                continue
            assignment[cc] = si
            slot_load[si] += 1
            day_load[p_day] += 1
            course_assigned_day[cc] = p_day
            if enrolled_sets and cc in enrolled_sets:
                for sid in enrolled_sets[cc]:
                    student_day_count[sid][p_day] += 1
                    student_day_courses[sid][p_day].append(cc)
            for bk in _cb.get(cc, []):
                bucket_day_courses[bk][p_day].add(cc)

    # ── Main scheduling loop ──
    # Process courses in most-constrained-first order.  For each course:
    #   1. Eliminate slots blocked by hard constraints (A: clash, B: bucket-day)
    #   2. If no slot survives → create OVERFLOW virtual slot
    #   3. Otherwise score each surviving candidate on 4-level soft priority
    #   4. Pick the candidate with the lowest (best) score tuple
    for course in courses_sorted:
        if course in assignment:
            continue  # already pinned by user

        # Hard constraint A: no two conflicting courses in the same slot.
        # Find which slots are already used by this course's neighbours.
        neighbours = adj.get(course, {})
        used_slots = {assignment[n] for n in neighbours if n in assignment}

        # Conflict-free candidates: every slot NOT used by a neighbour
        candidates = [si for si in range(max_slot_idx + 1) if si not in used_slots]

        # Hard constraint B: no two courses from the same (program, term)
        # bucket on the same day.  Remove candidates whose day already
        # has a bucket-mate from ANY of this course's buckets.
        my_buckets = _cb.get(course, [])
        if my_buckets and candidates:
            blocked_days: set[str] = set()
            for bk in my_buckets:
                for day, assigned in bucket_day_courses[bk].items():
                    if assigned:  # day already has a course from this bucket
                        blocked_days.add(day)
            if blocked_days:
                candidates = [
                    si for si in candidates if slot_by_index[si]["day"] not in blocked_days
                ]

        if not candidates:
            # ── OVERFLOW: no feasible slot exists ──
            # All real slots are blocked by hard constraints (student clash
            # or bucket-mate same day).  Create a virtual OVERFLOW slot so
            # the course isn't silently dropped; the QA report will flag it
            # and the UI shows a red overflow row.
            max_slot_idx += 1
            chosen = max_slot_idx
            slot_by_index[chosen] = {
                "index": chosen,
                "day": "OVERFLOW",
                "period": f"Extra-{chosen}",
            }
            assignment[course] = chosen
            slot_load[chosen] += 1
            day_load["OVERFLOW"] += 1
            course_assigned_day[course] = "OVERFLOW"
            # Maintain student tracking structures for consistency
            if enrolled_sets and course in enrolled_sets:
                for sid in enrolled_sets[course]:
                    student_day_count[sid]["OVERFLOW"] += 1
                    student_day_courses[sid]["OVERFLOW"].append(course)
            # Maintain bucket-day tracking (use my_buckets, already computed above)
            for bk in my_buckets:
                bucket_day_courses[bk]["OVERFLOW"].add(course)
            continue

        if enrolled_sets and course in enrolled_sets:
            # ── Soft-constraint scoring ──
            # Evaluate ALL conflict-free candidates and pick the best by a
            # five-level priority tuple (lower is better):
            #   (overload, credit_pair, spacing, day_load, slot_load)
            # Python tuple comparison ensures level 1 always trumps level 2, etc.
            # day_load comes before slot_load so courses with no other
            # constraints spread across days first, then balance within day.
            course_students = enrolled_sets[course]
            best_slot = candidates[0]
            best_score = (
                float("inf"),
                float("inf"),
                float("inf"),
                float("inf"),
                float("inf"),
                float("inf"),
            )
            # Reservoir sampling counter for ties when seed is provided.
            ties_seen = 0

            for si in candidates:
                day = slot_by_index[si]["day"]

                # Level 1 — Day-overload: count students who would exceed the
                # per-day cap if this course is placed on this day.
                penalty = 0
                for sid in course_students:
                    if student_day_count[sid][day] >= max_per_day:
                        penalty += 1

                # Level 1.5 — Credit-pair penalty: when multi-exam days are
                # unavoidable, prefer lighter pairings (4+2 over 4+3 over 4+4).
                credit_penalty = 0
                if _cm:
                    this_cr = _cm.get(course, _CREDIT_DEFAULT)
                    for sid in course_students:
                        existing = student_day_courses[sid][day]
                        if existing:
                            day_credits = [_cm.get(ec, _CREDIT_DEFAULT) for ec in existing] + [
                                this_cr
                            ]
                            credit_penalty += _credit_pair_penalty(day_credits)

                # Level 2 — Spacing within programme-plan buckets:
                # penalise placing bucket-mates on adjacent days so students
                # get breathing room.  Weights: 1-day gap=100, 2-day=30, 3-day=10.
                spacing_penalty = 0
                if my_buckets and day in day_to_idx:
                    cand_di = day_to_idx[day]
                    for bk in my_buckets:
                        for mate in _ptb.get(bk, set()) - {course}:
                            if mate in course_assigned_day:
                                mate_day = course_assigned_day[mate]
                                if mate_day in day_to_idx:
                                    gap = abs(cand_di - day_to_idx[mate_day])
                                    if gap <= 1:
                                        spacing_penalty += 100
                                    elif gap <= 2:
                                        spacing_penalty += 30
                                    elif gap <= 3:
                                        spacing_penalty += 10

                # Level 3 — Day load: spread across days first
                # Level 4 — Slot load: balance within day
                score = (
                    penalty,
                    credit_penalty,
                    spacing_penalty,
                    0 if _pref.get(course) == si else 1,
                    day_load[day],
                    slot_load[si],
                )
                if score < best_score:
                    best_score = score
                    best_slot = si
                    ties_seen = 1
                elif score == best_score and rng is not None:
                    # Reservoir-sample tied slots so the chosen one
                    # varies across runs (1/ties_seen probability).
                    ties_seen += 1
                    if rng.random() < 1.0 / ties_seen:
                        best_slot = si

            chosen = best_slot
        else:
            # No enrolled data available — fall back to least-loaded slot
            chosen = min(
                candidates, key=lambda si: (0 if _pref.get(course) == si else 1, slot_load[si])
            )

        assignment[course] = chosen
        slot_load[chosen] += 1
        chosen_day = slot_by_index[chosen]["day"]
        day_load[chosen_day] += 1
        course_assigned_day[course] = chosen_day

        # Update student day counts + courses-per-day tracking
        if enrolled_sets and course in enrolled_sets:
            for sid in enrolled_sets[course]:
                student_day_count[sid][chosen_day] += 1
                student_day_courses[sid][chosen_day].append(course)

        # Update bucket-day tracking
        for bk in my_buckets:
            bucket_day_courses[bk][chosen_day].add(course)

    # ── Build result list sorted by slot index ──
    result: list[dict] = []
    for cc, si in assignment.items():
        slot = slot_by_index[si]
        result.append(
            {
                "course_code": cc,
                "slot_index": si,
                "day": slot["day"],
                "period": slot["period"],
            }
        )

    return sorted(result, key=lambda r: (r["slot_index"], r["course_code"]))


# ── 5. QA report ────────────────────────────────────────────────


def _compute_thin_clash_risk(
    enrolled_sets: dict[str, set[int]],
    schedule_entries: list[dict],
) -> list[dict]:
    """Walk the schedule and report any student whose exams collide in
    the same slot.

    Used after thin-conflict relaxation to surface the realised cost
    of dropping edges. Returns one entry per (student, slot) collision
    with the colliding course list.
    """
    slot_to_courses: dict[int, list[str]] = defaultdict(list)
    for e in schedule_entries:
        if e.get("day") == "OVERFLOW":
            continue
        slot_to_courses[e["slot_index"]].append(e["course_code"])

    clashes: list[dict] = []
    for si, course_codes in slot_to_courses.items():
        if len(course_codes) < 2:
            continue
        # Find students enrolled in 2+ of these courses
        student_courses_in_slot: dict[int, list[str]] = defaultdict(list)
        for cc in course_codes:
            for sid in enrolled_sets.get(cc, ()):
                student_courses_in_slot[sid].append(cc)
        for sid, ccs in student_courses_in_slot.items():
            if len(ccs) >= 2:
                clashes.append(
                    {
                        "student_id": sid,
                        "slot_index": si,
                        "courses": sorted(ccs),
                    }
                )
    return clashes


def _build_qa(
    enrolled_sets: dict[str, set[int]],
    schedule_entries: list[dict],
    max_per_day: int = 2,
    plan_term_buckets: dict[tuple[str, int], set[str]] | None = None,
    credit_map: dict[str, int] | None = None,
) -> dict:
    """Validate the schedule and produce a QA report.

    Checks hard-constraint violations (which should only happen with
    user-pinned overrides) and computes soft-constraint metrics:

    Hard constraints:
      - Same-slot conflicts:    two courses sharing students in the same slot
      - Bucket day violations:  two bucket-mates placed on the same day

    Soft-constraint metrics:
      - Day-overload count:     students exceeding max_per_day
      - Credit load metrics:    max credit-load per day, heavy-day student count

    Also collects per-student detail records for KPI drilldown in the UI.
    """
    # ── Preparation ──
    all_students: set[int] = set()
    for sids in enrolled_sets.values():
        all_students |= sids

    # Lookup maps: course → its assigned slot index / day
    course_slot: dict[str, int] = {e["course_code"]: e["slot_index"] for e in schedule_entries}
    course_day: dict[str, str] = {e["course_code"]: e["day"] for e in schedule_entries}
    course_identity: dict[str, str] = {
        e["course_code"]: str(
            e.get("course_identity") or e.get("source_course_code") or e["course_code"]
        )
        for e in schedule_entries
    }

    slots_used = len({e["slot_index"] for e in schedule_entries})

    # Invert enrolled_sets: student_id → [course_codes] so we can iterate
    # per-student and check their personal schedule for violations.
    student_courses: dict[int, list[str]] = defaultdict(list)
    for cc, sids in enrolled_sets.items():
        for sid in sids:
            student_courses[sid].append(cc)

    _cm = credit_map or {}

    # ── Accumulators ──
    same_slot_conflicts: list[dict] = []  # hard-constraint violations
    max_exams_per_day: int = 0  # worst-case day load globally
    students_over_limit_per_day: int = 0  # students exceeding soft cap
    max_credit_load_per_day: int = 0  # worst-case credit sum on a day
    heavy_day_students: int = 0  # students with heavy credit pair

    # Detail records for KPI drilldown panel in the UI
    overload_details: list[dict] = []  # per-student, per-day overload records
    heavy_day_details: list[dict] = []  # per-student, per-day heavy-credit records

    # ── Per-student validation ──
    for sid, courses in student_courses.items():
        # Group this student's courses by slot and by day
        slot_groups: dict[int, list[str]] = defaultdict(list)
        day_groups: dict[str, list[str]] = defaultdict(list)
        for cc in courses:
            si = course_slot.get(cc)
            if si is not None:
                slot_groups[si].append(cc)
            day = course_day.get(cc)
            if day is not None:
                day_groups[day].append(cc)

        # If ≥2 courses land in the same slot → conflict (pinned override)
        for si, ccs in slot_groups.items():
            if len(ccs) >= 2:
                same_slot_conflicts.append(
                    {
                        "student_id": sid,
                        "slot_index": si,
                        "courses": ccs,
                    }
                )

        # ── Soft-constraint metrics per day ──
        has_overload = False  # does this student exceed the per-day cap?
        has_heavy_day = False  # does this student have a heavy credit pairing?
        for _day, ccs in day_groups.items():
            if _day == "OVERFLOW":
                continue  # OVERFLOW is a virtual day — skip for metrics

            day_count = len(ccs)
            max_exams_per_day = max(max_exams_per_day, day_count)

            # Day-overload check: student exceeds the soft cap
            if day_count > max_per_day:
                has_overload = True
                overload_details.append(
                    {
                        "student_id": sid,
                        "day": _day,
                        "count": day_count,
                        "courses": [
                            {"code": c, "credits": _cm.get(c, _CREDIT_DEFAULT) if _cm else None}
                            for c in sorted(ccs)
                        ],
                    }
                )

            # Credit-load check: evaluate the top-2 heaviest exams on this day.
            # Only relevant when credit_map is available AND student has ≥2 exams.
            if _cm and len(ccs) >= 2:
                day_credits = [_cm.get(c, _CREDIT_DEFAULT) for c in ccs]
                total_credits = sum(day_credits)
                max_credit_load_per_day = max(max_credit_load_per_day, total_credits)
                pair_penalty = _credit_pair_penalty(day_credits)
                # "Heavy day" threshold: penalty ≥ 30 catches (4,4)→100 and
                # (4,3)→30, but NOT mild combos like (3,3)→5 or (3,2)→5.
                if pair_penalty >= 30:
                    has_heavy_day = True
                    heavy_day_details.append(
                        {
                            "student_id": sid,
                            "day": _day,
                            "penalty": pair_penalty,
                            "total_credits": total_credits,
                            "courses": [
                                {"code": c, "credits": _cm.get(c, _CREDIT_DEFAULT)}
                                for c in sorted(ccs)
                            ],
                        }
                    )

        if has_overload:
            students_over_limit_per_day += 1
        if has_heavy_day:
            heavy_day_students += 1

    # ── Bucket (programme-plan term) day-rule verification ──
    # Hard constraint B says no two courses from the same (program, term)
    # bucket should share a day.  This can only be violated when the user
    # pins courses that override the scheduler's hard-constraint logic.
    bucket_day_violations: list[dict] = []
    bucket_count = 0
    if plan_term_buckets:
        bucket_count = len(plan_term_buckets)
        for (program, term), bucket_courses in sorted(plan_term_buckets.items()):
            # Group this bucket's courses by their assigned day
            day_groups_b: dict[str, dict[str, str]] = defaultdict(dict)
            for cc in bucket_courses:
                day = course_day.get(cc)
                if day is not None and day != "OVERFLOW":
                    identity = course_identity.get(cc, cc)
                    day_groups_b[day].setdefault(identity, cc)
            # Any day with ≥2 bucket-mates is a violation
            for day, ccs in day_groups_b.items():
                if len(ccs) >= 2:
                    bucket_day_violations.append(
                        {
                            "program": program,
                            "programme_term": term,
                            "day": day,
                            "courses": sorted(ccs.values()),
                        }
                    )

    return {
        "total_courses": len(enrolled_sets),
        "total_students": len(all_students),
        "slots_used": slots_used,
        "max_per_day": max_per_day,
        "max_exams_per_day_per_student": max_exams_per_day,
        "students_over_limit_per_day": students_over_limit_per_day,
        "same_slot_conflicts": same_slot_conflicts,
        "conflict_count": len(same_slot_conflicts),
        "bucket_count": bucket_count,
        "bucket_day_violations": bucket_day_violations,
        "bucket_day_violations_count": len(bucket_day_violations),
        "max_credit_load_per_day": max_credit_load_per_day,
        "heavy_day_students": heavy_day_students,
        "overload_details": overload_details,
        "heavy_day_details": heavy_day_details,
    }


# ── 5b. Room assignment (exam rooms) ───────────────────────────
#
# Exam rooms are allocated AFTER the greedy slot scheduler has decided
# which (day, period) every course goes into.  The workflow per slot:
#
#   1. Collect every demand unit scheduled in this slot, split by gender
#      (M students use M rooms, F students use F rooms — separate
#       buildings, same exam time, no cross-gender room sharing).
#   2. Attempt same-course same-gender section merges (combined ≤ biggest
#      available room) so small sections of one course collapse into one
#      room rather than eating two rooms.
#   3. Sort demand units by student_count DESC (largest first — classic
#      best-fit-decreasing bin packing).
#   4. For each unit, pick the tightest-fit available room; prefer the
#      room the section normally uses during regular term meetings.
#   5. Units that still don't fit get room_code="UNASSIGNED" and are
#      reported by the QA layer.
#
# Constraints enforced:
#   • Each room hosts at most one course per slot (no cross-course share)
#   • Same-course same-gender sections MAY share a room (after merging)
#   • M sections → M rooms only; F sections → F rooms only
#   • Room.capacity is respected
#   • Room.department is IGNORED during exams (all departments share)

_SYNTHETIC_SECTION_LABEL = "ALL"


def _section_gender(section_label: str) -> str:
    """Derive gender ('M' or 'F') from a TermSection.section label.

    Labels are like "M7", "M128", "F3" — the first character is the
    gender tag.  Falls back to "M" for anything unexpected.
    """
    if not section_label:
        return "M"
    first = section_label[0].upper()
    if first in ("M", "F"):
        return first
    return "M"


def build_section_enrollment(
    course_codes: set[str] | list[str],
    programs: list[str] | None = None,
    sections: list[str] | None = None,
) -> dict[str, list[dict]]:
    """Return per-section enrolment data for room assignment.

    Result shape:
        {
          "CS101": [
            {"section": "M7", "student_count": 32,
             "preferred_room": "172FA003", "gender": "M"},
            ...
          ],
          ...
        }

    Source of truth is StudentTermSection (latest academic_year / term),
    grouped by the underlying TermSection.  Only course_codes passed in
    are returned.  The ``programs`` / ``sections`` filters narrow the
    student population the same way build_enrolled_sets does, so the
    per-section counts stay consistent with the scheduled enrolled_sets.

    Courses with no StudentTermSection data fall back to a synthetic
    single "ALL" section sized from StudentCourse.status='studying' so
    they still get a room assignment.
    """
    wanted: set[str] = {str(c) for c in course_codes}
    if not wanted:
        return {}

    # Latest (academic_year, term) that has any data — same approach as
    # build_enrolled_sets so we stay on the same dataset.
    latest = (
        StudentTermSection.objects.order_by("-academic_year", "-term")
        .values_list("academic_year", "term")
        .first()
    )

    result: dict[str, list[dict]] = {}
    sts_course_keys: set[str] = set()

    if latest is not None:
        ay, tm = latest
        qs = StudentTermSection.objects.filter(
            academic_year=ay,
            term=tm,
            term_section__course_key__in=list(wanted),
        ).select_related("term_section")

        if programs:
            student_ids = set(
                Student.objects.filter(program__in=programs).values_list("student_id", flat=True)
            )
            qs = qs.filter(student_id__in=student_ids)
        if sections:
            student_ids_sec = set(
                Student.objects.filter(section__in=sections).values_list("student_id", flat=True)
            )
            qs = qs.filter(student_id__in=student_ids_sec)

        # Group by (course_key, term_section_id) and count distinct students.
        # Also remember the section label so we can derive gender later.
        per_section: dict[tuple[str, int], dict[str, Any]] = {}
        for row in qs.values(
            "term_section_id",
            "term_section__course_key",
            "term_section__section",
            "student_id",
        ):
            ck = row["term_section__course_key"]
            tsid = row["term_section_id"]
            key = (ck, tsid)
            entry = per_section.setdefault(
                key,
                {
                    "section": row["term_section__section"] or "",
                    "student_ids": set(),
                    "term_section_id": tsid,
                },
            )
            entry["student_ids"].add(row["student_id"])

        # Preferred room: for each TermSection, find the most-used room code
        # across its TermSectionMeeting rows (classroom where the section
        # normally meets during the term).
        involved_ts_ids = {k[1] for k in per_section}
        preferred_room_by_ts: dict[int, str] = {}
        if involved_ts_ids:
            meetings = TermSectionMeeting.objects.filter(
                term_section_id__in=involved_ts_ids,
            ).values_list("term_section_id", "room")
            room_counts: dict[int, dict[str, int]] = defaultdict(lambda: defaultdict(int))
            for tsid, room_code in meetings:
                if room_code:
                    room_counts[tsid][room_code] += 1
            for tsid, counts in room_counts.items():
                # Most-frequent wins; alphabetical as stable tiebreaker
                preferred_room_by_ts[tsid] = max(
                    counts.items(), key=lambda it: (it[1], -ord(it[0][0]) if it[0] else 0)
                )[0]

        for (course_key, _tsid), entry in per_section.items():
            sts_course_keys.add(course_key)
            section_label = entry["section"] or _SYNTHETIC_SECTION_LABEL
            result.setdefault(course_key, []).append(
                {
                    "section": section_label,
                    "student_count": len(entry["student_ids"]),
                    "preferred_room": preferred_room_by_ts.get(entry["term_section_id"], ""),
                    "gender": _section_gender(section_label),
                }
            )

    # Fallback for courses with zero STS data — build one synthetic section
    # from StudentCourse.status='studying', taking gender from the first
    # enrolled student.  This guarantees every scheduled course gets room
    # assignment even when timetable data is missing.
    missing = wanted - sts_course_keys
    if missing:
        sc_qs = StudentCourse.objects.filter(
            course__course_code__in=list(missing),
            status="studying",
        ).select_related("course", "student")
        if programs:
            sc_qs = sc_qs.filter(student__program__in=programs)
        if sections:
            sc_qs = sc_qs.filter(student__section__in=sections)

        fallback: dict[str, dict[str, set[int] | str]] = {}
        for sc in sc_qs.values("course__course_code", "student_id", "student__section"):
            cc = sc["course__course_code"]
            entry_f = fallback.setdefault(
                cc, {"student_ids": set(), "gender": sc["student__section"] or "M"}
            )
            # student_ids is typed as set[int] in this shape
            entry_f["student_ids"].add(sc["student_id"])  # type: ignore[union-attr]

        for cc, entry_f in fallback.items():
            sids = entry_f["student_ids"]
            gender = str(entry_f["gender"]).strip().upper()[:1] or "M"
            if gender not in ("M", "F"):
                gender = "M"
            result.setdefault(cc, []).append(
                {
                    "section": _SYNTHETIC_SECTION_LABEL,
                    "student_count": len(sids),  # type: ignore[arg-type]
                    "preferred_room": "",
                    "gender": gender,
                }
            )

    return result


def check_room_feasibility(
    section_enrollment: dict[str, list[dict]],
    rooms: list[dict],
) -> list[dict]:
    """Return sections that cannot fit in any single same-gender room.

    A violation means even the largest room of the matching gender is
    smaller than the section — the scheduler cannot place it without
    splitting students, which we do not support.  Returns the list of
    violations so the caller can surface them as warnings (does not
    block the build — unassignable sections just get UNASSIGNED).
    """
    if not rooms:
        return []
    max_cap_by_gender: dict[str, int] = {"M": 0, "F": 0}
    for r in rooms:
        g = str(r.get("section", "M")).upper() or "M"
        max_cap_by_gender[g] = max(max_cap_by_gender.get(g, 0), int(r.get("capacity", 0) or 0))

    violations: list[dict] = []
    for course_code, sections_data in section_enrollment.items():
        for s in sections_data:
            g = s.get("gender", "M")
            max_cap = max_cap_by_gender.get(g, 0)
            if s["student_count"] > max_cap:
                violations.append(
                    {
                        "course_code": course_code,
                        "section": s["section"],
                        "gender": g,
                        "student_count": s["student_count"],
                        "max_room_capacity": max_cap,
                    }
                )
    return violations


def _split_oversized_sections(
    sections: list[dict],
    max_cap: int,
) -> list[dict]:
    """Split any section larger than ``max_cap`` into roughly-equal parts
    so it can be distributed across multiple rooms.

    Needed for giant lecture sections (e.g. 100-student GS courses)
    where no single same-gender room is big enough — the registrar
    would handle this by splitting the section into two half-sized
    sitting groups during the exam.  We simulate that here.

    A section with ``student_count > max_cap`` is replaced by
    ``ceil(count / max_cap)`` parts labelled ``"<orig>/1"``, ``"<orig>/2"``,
    etc.  Only the first part inherits the preferred_room so the packer
    doesn't try to place every part in the same normal classroom.
    """
    if max_cap <= 0:
        return list(sections)
    import math

    out: list[dict] = []
    for s in sections:
        cnt = int(s.get("student_count", 0) or 0)
        if cnt <= max_cap:
            out.append(s)
            continue
        n_parts = max(2, math.ceil(cnt / max_cap))
        base = cnt // n_parts
        remainder = cnt - base * n_parts
        for i in range(n_parts):
            part_size = base + (1 if i < remainder else 0)
            out.append(
                {
                    "section": f"{s['section']}/{i + 1}",
                    "student_count": part_size,
                    "preferred_room": s.get("preferred_room", "") if i == 0 else "",
                    "gender": s.get("gender", "M"),
                    "_split_from": s["section"],
                }
            )
    return out


def _merge_same_course_sections(
    sections_data: list[dict],
    max_room_capacity: int,
) -> list[dict]:
    """Merge same-course same-gender sections if their combined size
    fits in the largest available room.

    Returns a list of demand units where each unit is either:
      - a single section (not merged), or
      - a merged group with ``merged_from`` listing the source sections.

    Merging reduces the number of rooms needed for popular courses
    (e.g. two M sections of 15 students each fit in one 30-seat room).
    Only attempts simple two-way and three-way merges — more aggressive
    bin packing isn't worth the complexity for ≤4 sections per gender.
    """
    if len(sections_data) <= 1:
        return list(sections_data)

    # Sort DESC by student_count so we try to pair up largest first
    by_size = sorted(sections_data, key=lambda s: s["student_count"], reverse=True)
    units: list[dict] = []
    used_indices: set[int] = set()

    for i, s_i in enumerate(by_size):
        if i in used_indices:
            continue
        group = [s_i]
        running = int(s_i["student_count"])
        used_indices.add(i)
        # Try to add smaller sections while they fit
        for j, s_j in enumerate(by_size):
            if j <= i or j in used_indices:
                continue
            if running + int(s_j["student_count"]) <= max_room_capacity:
                group.append(s_j)
                running += int(s_j["student_count"])
                used_indices.add(j)

        if len(group) == 1:
            units.append(s_i)
        else:
            preferred = next((g["preferred_room"] for g in group if g["preferred_room"]), "")
            units.append(
                {
                    "section": "+".join(g["section"] for g in group),
                    "student_count": running,
                    "preferred_room": preferred,
                    "gender": s_i["gender"],
                    "merged_from": [g["section"] for g in group],
                    # Keep original constituents so the assigner can split
                    # the group back if the merged room candidate fails.
                    "_constituents": [dict(g) for g in group],
                }
            )
    return units


def assign_rooms_to_schedule(
    schedule_entries: list[dict],
    section_enrollment: dict[str, list[dict]],
    rooms: list[dict],
    seed: int | None = None,
) -> list[dict]:
    """Assign rooms to every scheduled course, mutating ``schedule_entries``
    in place and returning the list for chaining.

    For each slot:
      1. Group courses scheduled in that slot by their per-gender sections.
      2. Try to merge same-course same-gender sections.
      3. Best-fit-decreasing assignment: tightest-fitting room wins, with
         a preference bonus for the section's normal meeting room.
      4. Rooms already taken in this slot are unavailable.
      5. Unassignable demand units get room_code="UNASSIGNED".

    Each entry in schedule_entries gains a ``rooms`` key:
        [
          {"section": "M7", "room_code": "172FA003",
           "student_count": 32, "room_capacity": 35,
           "gender": "M", "merged_from": ["M7"]},
          ...
        ]

    OVERFLOW entries are skipped (they have no real slot).
    When ``rooms`` is empty, every entry gets an empty ``rooms`` list.
    """
    if not schedule_entries:
        return schedule_entries

    # Ensure every entry has a rooms slot even if we bail early
    for e in schedule_entries:
        e.setdefault("rooms", [])

    if not rooms:
        return schedule_entries

    # Randomiser used for tie-breaking when a seed is provided.  Tied
    # room candidates and tied demand groups are reservoir-sampled so
    # the chosen one varies across runs — the overall quality stays
    # identical because we only break ties, never override better picks.
    rng = random.Random(seed) if seed is not None else None

    # Pre-split rooms by gender; sort ASC by capacity so best-fit picks
    # the tightest room first when iterating.
    rooms_by_gender: dict[str, list[dict]] = {"M": [], "F": []}
    for r in rooms:
        g = str(r.get("section", "M")).upper() or "M"
        if g in rooms_by_gender:
            rooms_by_gender[g].append(r)
    for g in rooms_by_gender:
        rooms_by_gender[g].sort(key=lambda r: int(r.get("capacity", 0) or 0))

    # Group schedule entries by slot_index so we can assign per-slot
    entries_by_slot: dict[int, list[dict]] = defaultdict(list)
    for e in schedule_entries:
        if e.get("day") == "OVERFLOW":
            continue
        entries_by_slot[e["slot_index"]].append(e)

    def _pack_course_gender_group(
        entry: dict,
        sections: list[dict],
        gender: str,
        taken: set[tuple[str, str]],
    ) -> None:
        """Room-aware packing for one (course, gender) block.

        Algorithm per (course, gender):
          1. Sort remaining sections DESC by student_count.
          2. Top section is the current "anchor".  Find all free same-
             gender rooms whose capacity ≥ anchor.student_count.
          3. If none: split the anchor in half and retry.  Only when
             the anchor is already smaller than every remaining room
             is it stamped UNASSIGNED (true capacity exhaustion).
          4. Score candidates: preferred-room bonus, then tightest-fit.
          5. Pack additional sections into the chosen room until no
             more fit, largest-first.
          6. Stamp one ``rooms`` record describing the packed block.
          7. Remove packed sections from remaining; loop until empty.

        Recursive splitting during step 3 handles the common case where
        a big section would have fit earlier but all the biggest rooms
        have been consumed by other courses in the same slot — halving
        it puts it within reach of smaller rooms still available.
        """
        remaining = sorted(
            (dict(s) for s in sections),
            key=lambda s: int(s.get("student_count", 0) or 0),
            reverse=True,
        )

        def _free_max_cap() -> int:
            return max(
                (
                    int(r.get("capacity", 0) or 0)
                    for r in rooms_by_gender.get(gender, [])
                    if (str(r.get("room_code", "")), gender) not in taken
                ),
                default=0,
            )

        while remaining:
            anchor = remaining[0]
            anchor_count = int(anchor.get("student_count", 0) or 0)
            preferred = anchor.get("preferred_room", "")

            candidates = [
                r
                for r in rooms_by_gender.get(gender, [])
                if (str(r.get("room_code", "")), gender) not in taken
                and int(r.get("capacity", 0) or 0) >= anchor_count
            ]

            if not candidates:
                # Before giving up, try to split the anchor in half —
                # the other half could fit a smaller room.  We only
                # truly give up when the anchor is already smaller than
                # every free same-gender room (real capacity exhaustion)
                # or when halving would drop the part below 1 student.
                free_max = _free_max_cap()
                if anchor_count >= 2 and free_max > 0 and anchor_count > free_max:
                    # Split roughly in half and push both halves back.
                    half_a = anchor_count // 2
                    half_b = anchor_count - half_a
                    base_label = anchor.get("_split_from") or anchor["section"]
                    remaining.pop(0)
                    remaining = [
                        {
                            "section": f"{anchor['section']}a",
                            "student_count": half_a,
                            "preferred_room": preferred,
                            "gender": gender,
                            "_split_from": base_label,
                        },
                        {
                            "section": f"{anchor['section']}b",
                            "student_count": half_b,
                            "preferred_room": "",
                            "gender": gender,
                            "_split_from": base_label,
                        },
                    ] + remaining
                    # Keep sorted so the loop's DESC invariant holds
                    remaining.sort(
                        key=lambda s: int(s.get("student_count", 0) or 0),
                        reverse=True,
                    )
                    continue

                entry["rooms"].append(
                    {
                        "section": anchor["section"],
                        "room_code": "UNASSIGNED",
                        "student_count": anchor_count,
                        "room_capacity": 0,
                        "gender": gender,
                        "merged_from": [anchor["section"]],
                    }
                )
                remaining.pop(0)
                continue

            def _score(room: dict, pref: str = preferred, demand: int = anchor_count) -> tuple:
                code = str(room.get("room_code", ""))
                cap = int(room.get("capacity", 0) or 0)
                return (0 if code == pref else 1, cap - demand)

            # Pick the tightest-fit room.  When multiple rooms tie on
            # the score and a seed is provided, reservoir-sample among
            # them so the chosen room varies run-to-run.
            if rng is None:
                chosen = min(candidates, key=_score)
            else:
                best_room_score: tuple | None = None
                chosen = candidates[0]
                ties = 0
                for r in candidates:
                    rs = _score(r)
                    if best_room_score is None or rs < best_room_score:
                        best_room_score = rs
                        chosen = r
                        ties = 1
                    elif rs == best_room_score:
                        ties += 1
                        if rng.random() < 1.0 / ties:
                            chosen = r
            chosen_code = str(chosen.get("room_code", ""))
            chosen_cap = int(chosen.get("capacity", 0) or 0)
            taken.add((chosen_code, gender))

            # Pack more sections into the chosen room, largest-first,
            # while they still fit.  This forms the merged block as we
            # go so its shape matches the actual room we're using.
            packed = [anchor]
            running = anchor_count
            remaining = remaining[1:]
            still_left: list[dict] = []
            for s in remaining:
                sc = int(s.get("student_count", 0) or 0)
                if running + sc <= chosen_cap:
                    packed.append(s)
                    running += sc
                else:
                    still_left.append(s)
            remaining = still_left

            entry["rooms"].append(
                {
                    "section": "+".join(p["section"] for p in packed)
                    if len(packed) > 1
                    else packed[0]["section"],
                    "room_code": chosen_code,
                    "student_count": running,
                    "room_capacity": chosen_cap,
                    "gender": gender,
                    "merged_from": [p["section"] for p in packed],
                }
            )

    for _si, entries in entries_by_slot.items():
        # Rooms consumed in this slot (by room_code + gender)
        taken: set[tuple[str, str]] = set()

        # Build one "demand group" per (entry, gender) so we can sort
        # ACROSS courses by the largest indivisible section each course
        # needs.  This guarantees a 90-student single section beats a
        # 40-section group of 6-student blocks for first pick.
        groups: list[tuple[dict, str, list[dict], int]] = []
        for entry in entries:
            course_code = entry["course_code"]
            sections_data = section_enrollment.get(course_code, [])
            if not sections_data:
                continue
            for gender in ("M", "F"):
                same_gender = [s for s in sections_data if s.get("gender") == gender]
                if not same_gender:
                    continue
                # Pre-split any section that exceeds the biggest same-
                # gender room — the registrar normally splits such huge
                # sections into two sitting groups for exams.
                max_cap_gender = max(
                    (int(r.get("capacity", 0) or 0) for r in rooms_by_gender[gender]),
                    default=0,
                )
                same_gender = _split_oversized_sections(same_gender, max_cap_gender)
                max_section = max(int(s.get("student_count", 0) or 0) for s in same_gender)
                groups.append((entry, gender, same_gender, max_section))

        # Sort by the biggest indivisible section DESC: a group whose
        # largest section is 90 students is placed before a group whose
        # largest section is 7, regardless of total size.  When a seed
        # is provided, groups with the same max-section value are
        # shuffled so tied groups get varied priority each run.
        if rng is None:
            groups.sort(key=lambda g: -g[3])
        else:
            # Banded shuffle: groups are bucketed by max-section tier
            # (bands of 10 students) and shuffled within each band
            # before processing.  Wider than strict ties so adjacent
            # sizes also mix.
            groups.sort(key=lambda g: -g[3])
            banded: dict[int, list] = defaultdict(list)
            for g in groups:
                banded[g[3] // 10].append(g)
            groups = []
            for band in sorted(banded.keys(), reverse=True):
                members = banded[band]
                rng.shuffle(members)
                groups.extend(members)

        for entry, gender, same_gender, _max_section in groups:
            _pack_course_gender_group(entry, same_gender, gender, taken)

    return schedule_entries


def _rebalance_invigilators_pass(
    schedule_entries: list[dict],
    section_enrollment: dict[str, list[dict]],
    rooms_list: list[dict],
    slots: list[dict],
    adj: dict[str, dict[str, int]],
    plan_term_buckets: dict[tuple[str, int], set[str]] | None,
    course_buckets: dict[str, list[tuple[str, int]]] | None,
    max_iterations: int = 30,
) -> int:
    """Final post-pass that moves courses between days to flatten the
    per-day invigilator demand.

    Local search:
      1. Recompute per-day invigilator totals using the current packing.
      2. Identify the hottest day (highest demand) and the coldest day.
      3. For each course on the hottest day, try moving it to a free slot
         on the coldest day.  A move is valid only if it doesn't create
         a same-slot student conflict and doesn't violate any
         (programme, term) bucket day rule.
      4. After each tentative move, re-pack rooms for the WHOLE schedule
         (because room demand on both affected slots changes).  Compute
         the new standard deviation of per-day invigilator totals.
         Accept the move if it strictly improves stddev; otherwise revert.
      5. Repeat until no improving move is found or max_iterations hit.

    Returns the number of moves accepted.  Idempotent and reversible —
    schedule_entries is mutated in place but never made worse than its
    starting state (we always revert non-improving moves).
    """
    if not schedule_entries or not rooms_list or not slots:
        return 0

    course_buckets = course_buckets or {}
    plan_term_buckets = plan_term_buckets or {}

    # Slot lookup helpers
    slots_by_day: dict[str, list[dict]] = defaultdict(list)
    for s in slots:
        slots_by_day[s["day"]].append(s)

    def _repack_all() -> None:
        """Clear current room assignments and re-run the full packer."""
        for e in schedule_entries:
            e["rooms"] = []
        assign_rooms_to_schedule(schedule_entries, section_enrollment, rooms_list, seed=None)

    def _per_day_invigilators() -> dict[str, dict[str, int]]:
        """Return ``{day: {'M': int, 'F': int, 'total': int}}`` so the
        rebalance metric can score per-gender stddev rather than only
        the combined total — a move that flattens the total but
        worsens the M-only or F-only spread should not be accepted.
        """
        per_day: dict[str, dict[str, int]] = defaultdict(lambda: {"M": 0, "F": 0, "total": 0})
        for e in schedule_entries:
            if e.get("day") == "OVERFLOW":
                continue
            for a in e.get("rooms", []) or []:
                if a.get("room_code") == "UNASSIGNED":
                    continue
                stu = int(a.get("student_count", 0) or 0)
                invigs = _invigilators_needed(e["course_code"], stu)
                gender = a.get("gender", "M")
                per_day[e["day"]][gender] += invigs
                per_day[e["day"]]["total"] += invigs
        return {k: dict(v) for k, v in per_day.items()}

    def _balance_score(per_day: dict[str, dict[str, int]]) -> tuple[float, int, int]:
        """Score a per-day distribution.  Lower is flatter.

        Returns a tuple ``(combined_stddev, max_total, max_minus_min_total)``
        so we sort lexicographically: primary objective is the combined
        per-gender standard deviation (M-series and F-series concatenated),
        with the day-total max and spread as tiebreakers.  Using both
        gender series guarantees a move that helps the overall total but
        worsens one gender's distribution will be rejected.
        """
        if not per_day:
            return (0.0, 0, 0)
        m_vals = [v["M"] for v in per_day.values()]
        f_vals = [v["F"] for v in per_day.values()]
        t_vals = [v["total"] for v in per_day.values()]
        # stddev across M-series + F-series combined (so each gender is
        # weighted equally regardless of which one happens to be larger)
        combined = m_vals + f_vals
        n = len(combined)
        mean = sum(combined) / n
        var = sum((x - mean) ** 2 for x in combined) / n
        return (var**0.5, max(t_vals), max(t_vals) - min(t_vals))

    def _causes_conflict(
        course: str, target_slot_idx: int, current_slot_of: dict[str, int]
    ) -> bool:
        for other, si in current_slot_of.items():
            if other == course or si != target_slot_idx:
                continue
            if other in adj.get(course, {}) or course in adj.get(other, {}):
                return True
        return False

    def _causes_bucket_violation(
        course: str, target_day: str, current_day_of: dict[str, str]
    ) -> bool:
        for bk in course_buckets.get(course, []):
            for mate in plan_term_buckets.get(bk, set()):
                if mate != course and current_day_of.get(mate) == target_day:
                    return True
        return False

    # Make sure we start from a clean packing
    _repack_all()
    current = _per_day_invigilators()
    base_score = _balance_score(current)
    moves_accepted = 0

    for _iter in range(max_iterations):
        if not current:
            break
        # Sort days by total invigilator load — pick the worst hot/cold pair
        days_by_load = sorted(current.items(), key=lambda x: x[1]["total"])
        coldest_day, cold_counts = days_by_load[0]
        hottest_day, hot_counts = days_by_load[-1]
        if hot_counts["total"] - cold_counts["total"] <= 2:
            break  # already pretty flat

        improved = False
        # Snapshot current slot/day lookups
        current_slot_of = {e["course_code"]: e["slot_index"] for e in schedule_entries}
        current_day_of = {e["course_code"]: e["day"] for e in schedule_entries}

        # Try every course on the hottest day
        hot_entries = [e for e in schedule_entries if e.get("day") == hottest_day]
        # Process larger courses first — they shift more invigilator weight
        hot_entries.sort(
            key=lambda e: -sum(
                int(s.get("student_count", 0) or 0)
                for s in section_enrollment.get(e["course_code"], [])
            )
        )

        for entry in hot_entries:
            cc = entry["course_code"]
            old_slot_idx = entry["slot_index"]
            old_day = entry["day"]
            old_period = entry["period"]

            for target_slot in slots_by_day.get(coldest_day, []):
                tsi = target_slot["index"]
                if tsi == old_slot_idx:
                    continue
                if _causes_conflict(cc, tsi, current_slot_of):
                    continue
                if _causes_bucket_violation(cc, coldest_day, current_day_of):
                    continue

                # Tentative move
                entry["slot_index"] = tsi
                entry["day"] = target_slot["day"]
                entry["period"] = target_slot["period"]
                _repack_all()
                new_per_day = _per_day_invigilators()
                new_score = _balance_score(new_per_day)

                # Accept only if strictly improving the lexicographic
                # (combined-stddev, max-day, spread) score.  A 0.01
                # tolerance on the stddev component prevents oscillation
                # when several moves have indistinguishable impact.
                accept = new_score[0] + 0.01 < base_score[0] or (
                    abs(new_score[0] - base_score[0]) <= 0.01 and new_score[1:] < base_score[1:]
                )
                if accept:
                    base_score = new_score
                    current = new_per_day
                    moves_accepted += 1
                    improved = True
                    break

                # Revert: restore entry fields AND re-pack so the rooms
                # data also matches the reverted state.  Without the
                # second repack the entry would briefly carry rooms from
                # the failed move until the next successful move (or the
                # final pass return) re-packed everything.
                entry["slot_index"] = old_slot_idx
                entry["day"] = old_day
                entry["period"] = old_period
                _repack_all()

            if improved:
                break

        if not improved:
            break

    # Final re-pack so the returned schedule_entries always carry rooms
    # data that matches their final slot assignments — important when
    # the loop exits mid-iteration.
    _repack_all()
    return moves_accepted


def _build_room_qa(
    schedule_entries: list[dict],
    rooms: list[dict],
) -> dict:
    """QA metrics for room assignment — rooms used, utilisation, unassigned,
    and double-booking defensive check (should never trigger).
    """
    if not rooms:
        return {
            "rooms_available": 0,
            "rooms_used": 0,
            "total_demand": 0,
            "total_capacity_used": 0,
            "avg_utilization": 0.0,
            "unassigned_room_sections": [],
            "room_double_bookings": [],
        }

    rooms_used_keys: set[tuple[int, str]] = set()
    total_demand = 0
    total_capacity_used = 0
    unassigned: list[dict] = []

    # Double-booking defensive check — (slot_index, room_code) should map
    # to a single course.  Track (slot, room) → course mappings.
    slot_room_course: dict[tuple[int, str], str] = {}
    double_bookings: list[dict] = []

    # Per-day invigilator totals, split by gender.
    # invigilators_per_day[day]['M'/'F'/'total'] = int
    invigilators_per_day: dict[str, dict[str, int]] = defaultdict(
        lambda: {"M": 0, "F": 0, "total": 0}
    )

    for e in schedule_entries:
        if e.get("day") == "OVERFLOW":
            continue
        si = int(e.get("slot_index", -1))
        for a in e.get("rooms", []):
            code = a.get("room_code", "")
            if code == "UNASSIGNED":
                unassigned.append(
                    {
                        "course_code": e["course_code"],
                        "day": e["day"],
                        "period": e["period"],
                        "section": a.get("section", ""),
                        "student_count": int(a.get("student_count", 0) or 0),
                        "gender": a.get("gender", ""),
                    }
                )
                continue
            rooms_used_keys.add((si, code))
            stu = int(a.get("student_count", 0) or 0)
            total_demand += stu
            total_capacity_used += int(a.get("room_capacity", 0) or 0)
            # Invigilator tally for this room
            invigs = _invigilators_needed(e["course_code"], stu)
            gender = a.get("gender", "M")
            day_label = e["day"]
            invigilators_per_day[day_label][gender] += invigs
            invigilators_per_day[day_label]["total"] += invigs
            prev = slot_room_course.get((si, code))
            if prev is not None and prev != e["course_code"]:
                double_bookings.append(
                    {
                        "slot_index": si,
                        "room_code": code,
                        "courses": [prev, e["course_code"]],
                    }
                )
            else:
                slot_room_course[(si, code)] = e["course_code"]

    avg_util = (total_demand / total_capacity_used) if total_capacity_used else 0.0
    # Convert invigilator dict to a stable JSON-serialisable shape
    invig_summary = {day: dict(counts) for day, counts in invigilators_per_day.items()}
    invig_grand_total = sum(c["total"] for c in invig_summary.values())
    invig_grand_M = sum(c["M"] for c in invig_summary.values())
    invig_grand_F = sum(c["F"] for c in invig_summary.values())
    return {
        "rooms_available": len(rooms),
        "rooms_used": len(rooms_used_keys),
        "total_demand": total_demand,
        "total_capacity_used": total_capacity_used,
        "avg_utilization": round(avg_util, 4),
        "unassigned_room_sections": unassigned,
        "room_double_bookings": double_bookings,
        "invigilators_per_day": invig_summary,
        "invigilators_total": invig_grand_total,
        "invigilators_total_M": invig_grand_M,
        "invigilators_total_F": invig_grand_F,
    }


def _build_section_enrollment_from_enrolled_sets(
    enrolled_sets: dict[str, set[int]],
) -> dict[str, list[dict]]:
    """Build exam room demand from the same student sets used for scheduling."""
    student_ids: set[int] = set()
    for sids in enrolled_sets.values():
        student_ids.update(sids)
    section_by_student = {
        int(sid): str(section or "").strip()
        for sid, section in Student.objects.filter(student_id__in=student_ids).values_list(
            "student_id",
            "section",
        )
    }
    result: dict[str, list[dict]] = {}
    for course_code, sids in enrolled_sets.items():
        grouped: dict[str, set[int]] = defaultdict(set)
        for sid in sids:
            grouped[section_by_student.get(int(sid), "") or _SYNTHETIC_SECTION_LABEL].add(int(sid))
        result[course_code] = [
            {
                "section": section_label,
                "student_count": len(section_sids),
                "preferred_room": "",
                "gender": _section_gender(section_label),
            }
            for section_label, section_sids in sorted(
                grouped.items(),
                key=lambda item: (_section_gender(item[0]), item[0]),
            )
        ]
    return result


# ── 6. Orchestrator ────────────────────────────────────────────


def build_exam_timetable(
    label: str,
    days: list[str],
    periods: list[str],
    max_per_day: int = 2,
    programs: list[str] | None = None,
    sections: list[str] | None = None,
    selected_courses: list[str] | None = None,
    pinned: list[dict] | None = None,
    seed: int | None = None,
    assign_rooms: bool = True,
    rebalance_invigilators: bool = True,
    thin_conflict_threshold: int = 0,
    persist: bool = True,
) -> dict:
    """
    End-to-end pipeline: build enrolled sets → conflict graph →
    programme-plan term buckets → feasibility check → greedy schedule
    → QA → persist JSON.

    Returns the full result dict. When ``persist=True`` (the default,
    matching pre-existing single-run callers) a row is written to
    ``ExamTimetableRun`` and the row id appears in ``result["run_id"]``.
    When ``persist=False`` the row is NOT written — the multi-start
    runner uses this mode to evaluate many seeded builds before
    selecting which 4 to persist as Pareto candidates, so the history
    panel doesn't fill with throw-away exploratory runs.

    If a bucket feasibility violation is found, returns an error dict
    without scheduling (key "feasibility_error": True).

    Args:
        selected_courses – if provided, restrict scheduling to only
                           these course codes (user-curated list from
                           the preview step).
        seed             – RNG seed for randomised tie-breaking in the
                           scheduler; None = deterministic (alphabetical).
                           Different seeds produce different timetable
                           variants from the same input data.
        thin_conflict_threshold – if > 0, courses with total enrolled
                           students <= threshold are dropped from the
                           conflict graph (their tiny enrolment no
                           longer blocks other courses from a slot).
                           Default 0 = current behaviour (no relaxation).
                           Realised same-slot clashes for thin students
                           are reported in qa.thin_clash_risk so the
                           registrar can pin them manually if needed.
        persist          – when True (default), writes an
                           ``ExamTimetableRun`` row and stamps the
                           returned dict with ``run_id``. When False,
                           skips the DB write and returns the result
                           dict unstamped — used by the multi-start
                           runner to evaluate candidates before
                           persisting only the selected ones.
    """
    # 1. Enrolled sets
    enrolled_sets, course_meta = build_enrolled_sets_with_meta(
        programs=programs,
        sections=sections,
    )

    # 1b. Filter to user-selected courses (if provided from preview step)
    if selected_courses is not None:
        keep = set(selected_courses)
        enrolled_sets = {cc: sids for cc, sids in enrolled_sets.items() if cc in keep}
        course_meta = {cc: meta for cc, meta in course_meta.items() if cc in keep}

    course_list = sorted(enrolled_sets.keys())

    # 1c. Build credit map for credit-weighted scoring
    credit_map = build_credit_map(course_list)

    all_students: set[int] = set()
    for sids in enrolled_sets.values():
        all_students |= sids

    # 2. Conflict graph
    conflicts, adj = build_conflict_graph(enrolled_sets)

    # 2b. Thin-conflict relaxation — when threshold > 0, drop courses
    # with total enrolment <= threshold from the conflict graph entirely.
    # They become degree-0 (no neighbours, no back-edges from peers) so
    # the scheduler treats their tiny student conflicts as soft instead
    # of hard. The realised same-slot clash count is computed below for
    # transparency.
    thin_courses_report: list[dict] = []
    if thin_conflict_threshold > 0:
        thin_set = {
            cc for cc, sids in enrolled_sets.items() if len(sids) <= thin_conflict_threshold
        }
        # Snapshot full neighbour list for every thin course BEFORE any
        # mutation. Otherwise, when courses A and B are mutual thin
        # neighbours, processing A first pops B's back-edge to A, so
        # B's "dropped edges" report would be short by 1 and missing A
        # from its neighbours list. The report must be independent of
        # iteration order.
        thin_neighbours = {cc: sorted(adj.get(cc, {}).keys()) for cc in thin_set}
        for cc in sorted(thin_set):
            dropped = thin_neighbours[cc]
            thin_courses_report.append(
                {
                    "course_code": cc,
                    "total_students": len(enrolled_sets[cc]),
                    "dropped_edges": len(dropped),
                    "neighbours": dropped,
                }
            )
            adj[cc] = {}
            for n in dropped:
                if n in adj:
                    adj[n].pop(cc, None)

    # 3. Programme-plan term buckets
    ptb, cb = build_plan_term_buckets(set(course_list), course_meta=course_meta)

    # 4. Feasibility pre-check
    violations = check_bucket_feasibility(ptb, len(days))
    if violations:
        return stamp_schema_version(
            {
                "feasibility_error": True,
                "status": "feasibility_error",
                # v2 status surface: feasibility_error short-circuits
                # to "infeasible" with no further flags.
                "primary_status": "infeasible",
                "status_flags": [],
                "status_derivation_version": STATUS_DERIVATION_VERSION,
                "violations": violations,
                "courses_count": len(course_list),
                "students_count": len(all_students),
                "bucket_count": len(ptb),
            }
        )

    # 5. Slot pool (Cartesian product)
    slots: list[dict] = []
    idx = 0
    for day in days:
        for period in periods:
            slots.append({"index": idx, "day": day, "period": period})
            idx += 1

    # 6. Schedule (with day-spread + bucket + credit-pair constraints)
    schedule_entries = schedule(
        course_list,
        adj,
        slots,
        enrolled_sets=enrolled_sets,
        max_per_day=max_per_day,
        plan_term_buckets=ptb,
        course_buckets=cb,
        pinned=pinned,
        credit_map=credit_map,
        seed=seed,
    )
    for entry in schedule_entries:
        meta = course_meta.get(entry["course_code"], {})
        source = _source_code_for_display(entry["course_code"], meta.get("source_course_code"))
        entry["source_course_code"] = source
        entry["course_name"] = str(meta.get("course_name") or "")
        entry["course_identity"] = str(meta.get("course_identity") or source)

    # 7. QA report — validate hard constraints and compute quality metrics
    qa = _build_qa(
        enrolled_sets,
        schedule_entries,
        max_per_day=max_per_day,
        plan_term_buckets=ptb,
        credit_map=credit_map,
    )

    # 7b. Thin-clash risk — when thin relaxation was active, surface any
    # student who actually ended up with two same-slot exams as a result
    # of the dropped conflict edges. Empty list when threshold == 0 or
    # when no clashes materialised.
    qa["thin_threshold"] = thin_conflict_threshold
    qa["thin_courses"] = thin_courses_report
    qa["thin_clash_risk"] = (
        _compute_thin_clash_risk(enrolled_sets, schedule_entries)
        if thin_conflict_threshold > 0
        else []
    )

    # 7b. Room assignment (Phase 2) — attach rooms to each schedule entry.
    # Feasibility violations and double-bookings are surfaced in QA but
    # never block the build: unfittable sections simply land in
    # "UNASSIGNED" and the UI flags them.
    section_enrollment: dict[str, list[dict]] = {}
    rooms_list: list[dict] = []
    room_feasibility: list[dict] = []
    room_qa: dict = {}
    if assign_rooms:
        section_enrollment = _build_section_enrollment_from_enrolled_sets(enrolled_sets)
        rooms_list = list(
            Room.objects.all().values(
                "room_code", "capacity", "section", "department", "building", "floor"
            )
        )
        room_feasibility = check_room_feasibility(section_enrollment, rooms_list)
        assign_rooms_to_schedule(
            schedule_entries,
            section_enrollment,
            rooms_list,
            seed=seed,
        )

        # 7c. Final optimisation — flatten per-day invigilator load by
        # moving courses between days whenever the move strictly
        # improves the per-day invigilator-count standard deviation.
        # Skipped when the caller opts out, when there are no rooms, or
        # when there's nothing meaningful to balance (single day).
        rebalance_moves = 0
        if rebalance_invigilators and rooms_list and len(days) > 1:
            rebalance_moves = _rebalance_invigilators_pass(
                schedule_entries,
                section_enrollment,
                rooms_list,
                slots,
                adj,
                ptb,
                cb,
            )

        room_qa = _build_room_qa(schedule_entries, rooms_list)
        # Re-run main QA after rebalance so credit/conflict metrics
        # reflect any moved courses (cheap — no DB hits).
        qa = _build_qa(
            enrolled_sets,
            schedule_entries,
            max_per_day=max_per_day,
            plan_term_buckets=ptb,
            credit_map=credit_map,
        )
        # Merge room QA into main QA dict for a single source of truth
        qa["rooms"] = room_qa
        qa["room_feasibility_violations"] = room_feasibility
        qa["rebalance_moves"] = rebalance_moves
        # Re-attach thin-relaxation report (lost in the QA rebuild above)
        qa["thin_threshold"] = thin_conflict_threshold
        qa["thin_courses"] = thin_courses_report
        qa["thin_clash_risk"] = (
            _compute_thin_clash_risk(enrolled_sets, schedule_entries)
            if thin_conflict_threshold > 0
            else []
        )

    # Bucket summary for the result (frontend renders bucket info cards)
    buckets_summary: list[dict] = []
    for (program, term), bucket_courses in sorted(ptb.items()):
        buckets_summary.append(
            {
                "program": program,
                "programme_term": term,
                "course_count": len(bucket_courses),
                "courses": sorted(bucket_courses),
            }
        )

    # ── v2 + v3 telemetry authoring ──
    # Step 4 enriches each schedule entry's ``rooms`` sub-list with the
    # ``building`` field (looked up by room_code from rooms_list) so the
    # v3 building-footprint derivation has the data it needs. This also
    # gets persisted, which means historic v3 rows can re-derive the
    # footprint on read if we ever need to — though by default we
    # populate the footprint at write time and store it under
    # ``qa.building_footprint``.
    if rooms_list:
        room_meta_by_code: dict[str, dict] = {str(r.get("room_code", "")): r for r in rooms_list}
        for entry in schedule_entries:
            for r in entry.get("rooms") or []:
                if not isinstance(r, dict):
                    continue
                meta = room_meta_by_code.get(str(r.get("room_code", "")))
                if meta:
                    r.setdefault("building", str(meta.get("building", "") or ""))
                    r.setdefault("floor", str(meta.get("floor", "") or ""))

    # Multi-sitting details computed from schedule_entries (each entry's
    # ``rooms`` list carries the section labels we detect splits from).
    # The derivation is in the schema module so the v1->v2 migrator and
    # this build site share the same logic.
    multi_sitting_details = derive_multi_sitting_details(schedule_entries)
    qa["multi_sitting_sections"] = len(multi_sitting_details)
    qa["multi_sitting_details"] = multi_sitting_details

    # Manual-override signal: in this system, all same_slot_conflicts
    # come from registrar pinned overrides (the scheduler refuses to
    # produce them otherwise). Surface explicitly under the v2-named
    # keys so the status derivation doesn't need to fall back to
    # legacy_incomplete_qa for fresh builds.
    qa["manual_override_count"] = qa.get("conflict_count", 0)
    qa["manual_override_details"] = list(qa.get("same_slot_conflicts", []))

    # ── v3 telemetry blocks (display-only — no ranking/scheduler effect) ──
    qa["building_footprint"] = derive_building_footprint(schedule_entries)

    # Enrolment snapshot integrity: distinct sections counted from
    # section_enrollment (a dict[course -> list[section_dict]]). The
    # fallback flag is plumbed from build_enrolled_sets via a marker
    # in the section labels: when the path used StudentCourse, every
    # course gets a synthetic "ALL" section.
    _sections_total = sum(len(v) for v in section_enrollment.values())
    _synthetic_all = sum(
        1
        for course, sections in section_enrollment.items()
        for s in sections
        if str(s.get("section", "")).upper() == "ALL"
    )
    _fallback_used = _synthetic_all > 0 and _synthetic_all == len(section_enrollment)
    qa["enrolment_snapshot"] = compute_enrolment_snapshot(
        enrolled_sets,
        sections_count=_sections_total,
        fallback_used=_fallback_used,
        synthetic_all_sections_count=_synthetic_all,
    )

    # ── Assemble result dict ──
    # This dict is: (a) returned to the frontend as JSON, (b) persisted
    # in ExamTimetableRun.result_json for later viewing/export.
    # Both consumers go through ``load_normalised_run`` /
    # ``normalise_exam_run_payload`` (see core.services.exam_run_schema),
    # so ``stamp_schema_version`` here is the only write site that needs
    # to know about the schema version constant.
    _draft: dict = {
        "status": "ok",
        "students_count": len(all_students),
        "courses": course_list,
        "courses_count": len(course_list),
        "conflicts": conflicts,
        "conflicts_count": len(conflicts),
        "slots": slots,
        "schedule": schedule_entries,
        "qa": qa,
        "buckets_summary": buckets_summary,
        "bucket_count": len(ptb),
        "credit_map": credit_map,
        "seed": seed,
        "section_enrollment": section_enrollment,
        "rooms_count": len(rooms_list),
        "assign_rooms": assign_rooms,
    }
    # Compute the registrar status surface from the assembled draft
    # so the headline + flags reflect the real run state.
    primary_status, status_flags = derive_status_surface(_draft)
    _draft["primary_status"] = primary_status
    _draft["status_flags"] = status_flags
    _draft["status_derivation_version"] = STATUS_DERIVATION_VERSION
    result: dict = stamp_schema_version(_draft)

    # Persist (skipped in multi-start exploration mode where we evaluate
    # many candidates and only persist the selected Pareto few).
    if persist:
        run = ExamTimetableRun.objects.create(
            label=label,
            result_json=json.dumps(result, ensure_ascii=False),
        )
        result["run_id"] = run.id

    return result


# ── 7. Excel export ──────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parents[2]
RUNTIME_DIR = BASE_DIR / "runtime"


def export_exam_timetable_xlsx(run_id: int) -> Path:
    """Export a saved ExamTimetableRun to a styled multi-sheet .xlsx workbook.

    Sheets:
        Schedule        – day × period grid (mirrors the on-screen table)
        Schedule (M)    – same grid filtered to male sections + rooms only
        Schedule (F)    – same grid filtered to female sections + rooms only
        Courses         – flat list with course code, enrolled count, day, period
        Students (M)    – per-course male section counts + student totals
        Students (F)    – per-course female section counts + student totals
        QA Summary      – key metrics + any conflict / bucket-day warnings

    Returns the Path to the written file (in the runtime/ directory).
    """
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.cell.rich_text import CellRichText, TextBlock  # type: ignore[import-untyped]
    from openpyxl.cell.text import InlineFont  # type: ignore[import-untyped]
    from openpyxl.styles import (  # type: ignore[import-untyped]
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    run = ExamTimetableRun.objects.get(id=run_id)
    # Read through the normaliser so historic runs (pre-schema-versioning,
    # or rows missing keys this exporter expects) render cleanly with safe
    # defaults. Single read path: never ``json.loads(run.result_json)`` here.
    data = load_normalised_run(run)

    # Hard rule for sentinel payloads: do NOT silently export an empty
    # workbook from an unrenderable run. Registrar trust attaches to the
    # exported artefact more than the web UI; a blank XLSX is worse than
    # a controlled error because the registrar may distribute it without
    # noticing the data is gone. Fail loudly with a message the view
    # layer can surface as a 500 with body text.
    status = data.get("status")
    if status == "unrenderable":
        reason = data.get("error", "payload could not be rendered")
        raise RuntimeError(
            f"Exam timetable run #{run_id} cannot be exported: {reason}. "
            "The stored payload is missing, corrupt, or otherwise unreadable; "
            "rebuild the run before exporting."
        )
    if status == "future_version_unrenderable":
        raise RuntimeError(
            f"Exam timetable run #{run_id} was created by a newer build of "
            "the exam scheduler and cannot be exported by this version. "
            "Upgrade the application or rebuild the run."
        )

    schedule = data["schedule"]  # list of {course_code, slot_index, day, period}
    slots = data["slots"]  # list of {index, day, period}
    qa = data["qa"]  # QA metrics dict from _build_qa()

    # ── Styling constants ──
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="0A8E6E")  # teal
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    warn_fill = PatternFill("solid", fgColor="FFF3CD")  # light yellow
    danger_fill = PatternFill("solid", fgColor="F8D7DA")  # light red

    # ── Course colour by code hash (matches frontend colorForCourse) ──
    _color_cache: dict[str, PatternFill] = {}

    def _course_color_fill(code: str) -> PatternFill:
        if code in _color_cache:
            return _color_cache[code]
        # Same hash as JS: h = (h * 31 + charCode) % 360
        h = 0
        for ch in str(code):
            h = (h * 31 + ord(ch)) % 360
        # HSL to RGB: s=70%, l=92% (light pastel, same as frontend light mode)
        import colorsys

        r, g, b = colorsys.hls_to_rgb(h / 360.0, 0.92, 0.70)
        hex_color = f"{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}"
        fill = PatternFill("solid", fgColor=hex_color)
        _color_cache[code] = fill
        return fill

    def style_header_row(ws: Any, col_count: int) -> None:
        """Apply teal background + white bold font to the first row."""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

    # ── Extract ordered days and periods from slots ──
    day_order: list[str] = []
    period_order: list[str] = []
    day_set: set[str] = set()
    period_set: set[str] = set()
    for s in slots:
        if s["day"] not in day_set:
            day_set.add(s["day"])
            day_order.append(s["day"])
        if s["period"] not in period_set:
            period_set.add(s["period"])
            period_order.append(s["period"])

    wb = Workbook()

    # ────────────────────────────────────────────────────────────
    # Sheet 1 / 1b / 1c: Schedule grids — All / M only / F only
    # ────────────────────────────────────────────────────────────
    _credit_map_sched = data.get("credit_map", {})

    # Rich-text inline fonts: course code stands out (bold, dark teal)
    # against the lighter, smaller, gray room codes underneath.
    _course_inline_font = InlineFont(rFont="Consolas", b=True, sz=10, color="064E3B")
    _room_inline_font = InlineFont(rFont="Consolas", b=False, sz=8, color="6B7280")

    # Detect runs that pre-date the gender field on room dicts (Session 20+).
    # Old runs render M/F sheets as blank grids by default — emit a single
    # placeholder row instead so the registrar isn't confused by an empty sheet.
    _has_gender_data = any("gender" in a for e in schedule for a in e.get("rooms", []) or ())

    def _render_schedule_sheet(ws: Any, gender_filter: str | None) -> None:
        """Render a day×period grid. If gender_filter is 'M' or 'F', a course
        only appears in a cell when it has at least one room of that gender,
        and only rooms of that gender are listed under the course.

        For runs predating the gender field, a single placeholder row is
        emitted instead of a blank grid.
        """
        if gender_filter and not _has_gender_data:
            ws.append(
                [
                    f"This run pre-dates gender-aware room assignment. "
                    f"Re-build the schedule to populate the {gender_filter} view."
                ]
            )
            ws.column_dimensions["A"].width = 90
            return
        grid_local: dict[str, dict[str, list[str]]] = {}
        rooms_by_entry: dict[tuple[str, str, str], list[str]] = {}
        for e in schedule:
            if e.get("day") == "OVERFLOW":
                continue
            all_rooms = e.get("rooms", [])
            if gender_filter:
                matching = [
                    a
                    for a in all_rooms
                    if a.get("gender") == gender_filter
                    and a.get("room_code")
                    and a.get("room_code") != "UNASSIGNED"
                ]
                if not matching:
                    continue
                room_codes = [a["room_code"] for a in matching]
            else:
                room_codes = [
                    a.get("room_code", "")
                    for a in all_rooms
                    if a.get("room_code") and a.get("room_code") != "UNASSIGNED"
                ]
            grid_local.setdefault(e["day"], {}).setdefault(e["period"], []).append(e["course_code"])
            if room_codes:
                rooms_by_entry[(e["day"], e["period"], e["course_code"])] = room_codes

        ws.append(["Day \\ Period"] + period_order)
        style_header_row(ws, 1 + len(period_order))

        for day in day_order:
            row_idx = ws.max_row + 1
            day_cell = ws.cell(row=row_idx, column=1, value=day)
            day_cell.font = header_font
            day_cell.border = thin_border
            day_cell.alignment = left_align

            for pi, period in enumerate(period_order):
                courses = sorted(grid_local.get(day, {}).get(period, []))
                cell = ws.cell(row=row_idx, column=2 + pi)
                cell.border = thin_border
                cell.alignment = center

                def _label_blocks(
                    code: str,
                    p: str = period,
                    d: str = day,
                    leading_newline: bool = False,
                ) -> list[TextBlock]:
                    """Course header in dark bold teal, room list in light gray
                    so the eye finds the course code at a glance.

                    Any leading newline (used to separate stacked courses in a
                    cell) is folded into the course-header text rather than a
                    standalone "\\n" block — Excel rejects whitespace-only runs
                    that lack xml:space="preserve" and shows a corruption dialog.
                    """
                    cr = _credit_map_sched.get(code, "")
                    head = f"{code} {cr}cr" if cr else code
                    if leading_newline:
                        head = "\n" + head
                    blocks: list[TextBlock] = [TextBlock(_course_inline_font, head)]
                    rooms_line = rooms_by_entry.get((d, p, code), [])
                    if rooms_line:
                        blocks.append(TextBlock(_room_inline_font, "\n" + ", ".join(rooms_line)))
                    return blocks

                if not courses:
                    cell.value = ""
                elif len(courses) == 1:
                    c = courses[0]
                    cell.value = CellRichText(_label_blocks(c))
                    cell.fill = _course_color_fill(c)
                else:
                    blocks: list[TextBlock] = []
                    for i, c in enumerate(courses):
                        blocks.extend(_label_blocks(c, leading_newline=(i > 0)))
                    cell.value = CellRichText(blocks)
                    cell.fill = _course_color_fill(courses[0])

        _period_col_width = 24
        ws.column_dimensions["A"].width = 14
        for i, _p in enumerate(period_order, start=2):
            ws.column_dimensions[get_column_letter(i)].width = _period_col_width

        def _visual_lines(val: Any, col_width_chars: int) -> int:
            """How many wrapped lines a cell renders at in Excel.

            Flattens rich text to a single string and counts each segment
            (between \\n) plus its wrap overflow. Uses a size-10 chars-per-
            line — the size-8 room line actually fits more, so this
            slightly over-estimates wraps for rooms, which is the safe
            direction (row ends up a touch taller, never clipped).
            """
            if isinstance(val, CellRichText):
                text = "".join(str(b.text if isinstance(b, TextBlock) else b) for b in val)
            elif isinstance(val, str):
                text = val
            else:
                return 1
            if not text:
                return 1
            cpl = max(8, int(col_width_chars * 11 / 10))
            total = 0
            for seg in text.split("\n"):
                total += 1 + (max(0, len(seg) - 1) // cpl)
            return total

        for r in range(2, ws.max_row + 1):
            max_lines = 1
            for c in range(2, ws.max_column + 1):
                lines = _visual_lines(ws.cell(row=r, column=c).value, _period_col_width)
                if lines > max_lines:
                    max_lines = lines
            # ~15 pt per wrapped line at size-10 is safe; add a small pad so
            # descenders of the last line don't touch the border.
            if max_lines > 1:
                ws.row_dimensions[r].height = max_lines * 15 + 3

    ws1 = wb.active
    ws1.title = "Schedule"
    _render_schedule_sheet(ws1, None)
    _render_schedule_sheet(wb.create_sheet("Schedule (M)"), "M")
    _render_schedule_sheet(wb.create_sheet("Schedule (F)"), "F")

    # ────────────────────────────────────────────────────────────
    # Sheet 2: Courses (flat list)
    # ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Courses")
    _credit_map = data.get("credit_map", {})
    ws2.append(["Course Code", "Credits", "Day", "Period", "Slot Index"])
    style_header_row(ws2, 5)

    sorted_schedule = sorted(schedule, key=lambda e: (e.get("slot_index", 999), e["course_code"]))
    for e in sorted_schedule:
        cr = _credit_map.get(e["course_code"], "")
        ws2.append([e["course_code"], cr, e["day"], e["period"], e.get("slot_index", "")])

    for r in range(2, ws2.max_row + 1):
        code_val = ws2.cell(row=r, column=1).value
        for c in range(1, 6):
            cell = ws2.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = center
            if code_val:
                cell.fill = _course_color_fill(code_val)

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 12

    # ────────────────────────────────────────────────────────────
    # Sheet 2b / 2c: Students (M) / Students (F) — per-course gender totals
    # ────────────────────────────────────────────────────────────
    def _render_students_sheet(ws: Any, gender_filter: str) -> None:
        if not _has_gender_data:
            ws.append(
                [
                    f"This run pre-dates gender-aware room assignment. "
                    f"Re-build the schedule to populate the {gender_filter} view."
                ]
            )
            ws.column_dimensions["A"].width = 90
            return

        ws.append(["Course Code", "Credits", "Day", "Period", "Sections", "Students"])
        style_header_row(ws, 6)

        rows: list[tuple[str, Any, str, str, int, int]] = []
        for e in sorted_schedule:
            matching = [a for a in e.get("rooms", []) if a.get("gender") == gender_filter]
            if not matching:
                continue
            sections = len(matching)
            students = sum(int(a.get("student_count", 0) or 0) for a in matching)
            cr = _credit_map.get(e["course_code"], "")
            rows.append((e["course_code"], cr, e["day"], e["period"], sections, students))

        for row in rows:
            ws.append(list(row))

        for r in range(2, ws.max_row + 1):
            code_val = ws.cell(row=r, column=1).value
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = center
                if code_val:
                    cell.fill = _course_color_fill(code_val)

        if rows:
            total_sections = sum(r[4] for r in rows)
            total_students = sum(r[5] for r in rows)
            total_row = ws.max_row + 1
            ws.cell(row=total_row, column=1, value="TOTAL").font = header_font
            ws.cell(row=total_row, column=5, value=total_sections).font = header_font
            ws.cell(row=total_row, column=6, value=total_students).font = header_font
            for c in range(1, 7):
                cell = ws.cell(row=total_row, column=c)
                cell.border = thin_border
                cell.alignment = center
                cell.fill = PatternFill("solid", fgColor="E8F5E9")

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12

    _render_students_sheet(wb.create_sheet("Students (M)"), "M")
    _render_students_sheet(wb.create_sheet("Students (F)"), "F")

    # ────────────────────────────────────────────────────────────
    # Sheet 3: QA Summary
    # ────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("QA Summary")

    # Key metrics as label-value pairs
    metrics = [
        ("Label", run.label),
        ("Total Courses", qa.get("total_courses", data.get("courses_count", 0))),
        ("Total Students", qa.get("total_students", data.get("students_count", 0))),
        ("Slots Used", qa.get("slots_used", 0)),
        ("Max Exams/Day/Student", qa.get("max_exams_per_day_per_student", 0)),
        ("Max Per Day Cap", qa.get("max_per_day", 2)),
        ("Students Over Limit", qa.get("students_over_limit_per_day", 0)),
        ("Max Credit Load/Day", qa.get("max_credit_load_per_day", 0)),
        ("Heavy Day Students", qa.get("heavy_day_students", 0)),
        ("Same-Slot Conflicts", qa.get("conflict_count", 0)),
        ("Programme Buckets", qa.get("bucket_count", 0)),
        ("Bucket Day Violations", qa.get("bucket_day_violations_count", 0)),
    ]

    ws3.append(["Metric", "Value"])
    style_header_row(ws3, 2)

    for label, value in metrics:
        ws3.append([label, value])

    # Highlight warning rows
    for r in range(2, ws3.max_row + 1):
        for c in range(1, 3):
            cell = ws3.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = left_align
        metric_name = ws3.cell(row=r, column=1).value
        metric_val = ws3.cell(row=r, column=2).value
        # Colour warning rows yellow/red
        if metric_name == "Same-Slot Conflicts" and metric_val and int(metric_val) > 0:
            ws3.cell(row=r, column=1).fill = danger_fill
            ws3.cell(row=r, column=2).fill = danger_fill
        elif metric_name in ("Students Over Limit", "Bucket Day Violations", "Heavy Day Students"):
            if metric_val and int(metric_val) > 0:
                ws3.cell(row=r, column=1).fill = warn_fill
                ws3.cell(row=r, column=2).fill = warn_fill

    # Append same-slot conflict details (if any)
    same_slot = qa.get("same_slot_conflicts", [])
    if same_slot:
        ws3.append([])
        ws3.append(["Same-Slot Conflict Details"])
        ws3.cell(row=ws3.max_row, column=1).font = header_font
        ws3.append(["Student ID", "Slot Index", "Courses"])
        r = ws3.max_row
        for c in range(1, 4):
            cell = ws3.cell(row=r, column=c)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor="EEEEEE")
            cell.border = thin_border
        for conflict in same_slot:
            ws3.append(
                [
                    conflict.get("student_id", ""),
                    conflict.get("slot_index", ""),
                    ", ".join(conflict.get("courses", [])),
                ]
            )

    # Append bucket day violation details (if any)
    bucket_viols = qa.get("bucket_day_violations", [])
    if bucket_viols:
        ws3.append([])
        ws3.append(["Bucket Day Violation Details"])
        ws3.cell(row=ws3.max_row, column=1).font = header_font
        ws3.append(["Programme", "Term", "Day", "Courses"])
        r = ws3.max_row
        for c in range(1, 5):
            cell = ws3.cell(row=r, column=c)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor="EEEEEE")
            cell.border = thin_border
        for v in bucket_viols:
            ws3.append(
                [
                    v.get("program", ""),
                    v.get("programme_term", ""),
                    v.get("day", ""),
                    ", ".join(v.get("courses", [])),
                ]
            )

    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 30

    # ────────────────────────────────────────────────────────────
    # Sheet 4: Room Assignments (one row per assigned section)
    # ────────────────────────────────────────────────────────────
    room_qa = qa.get("rooms") if isinstance(qa, dict) else None
    has_room_data = any(e.get("rooms") for e in schedule if e.get("day") != "OVERFLOW")
    if has_room_data:
        ws4 = wb.create_sheet("Room Assignments")
        ws4.append(
            [
                "Day",
                "Period",
                "Course",
                "Section",
                "Gender",
                "Students",
                "Room",
                "Capacity",
                "Utilization",
            ]
        )
        style_header_row(ws4, 9)

        # Sort by slot then course for a stable, readable ordering
        sorted_for_rooms = sorted(
            (e for e in schedule if e.get("day") != "OVERFLOW"),
            key=lambda e: (e.get("slot_index", 999), e["course_code"]),
        )
        for e in sorted_for_rooms:
            for a in e.get("rooms", []) or []:
                cap = int(a.get("room_capacity", 0) or 0)
                cnt = int(a.get("student_count", 0) or 0)
                util = f"{(cnt / cap * 100):.0f}%" if cap else "-"
                ws4.append(
                    [
                        e["day"],
                        e["period"],
                        e["course_code"],
                        a.get("section", ""),
                        a.get("gender", ""),
                        cnt,
                        a.get("room_code", ""),
                        cap if cap else "",
                        util,
                    ]
                )

        for r in range(2, ws4.max_row + 1):
            code_val = ws4.cell(row=r, column=3).value
            for c in range(1, 10):
                cell = ws4.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = center
            if code_val:
                ws4.cell(row=r, column=3).fill = _course_color_fill(str(code_val))
            room_val = ws4.cell(row=r, column=7).value
            if room_val == "UNASSIGNED":
                for c in range(1, 10):
                    ws4.cell(row=r, column=c).fill = danger_fill

        ws4.column_dimensions["A"].width = 10
        ws4.column_dimensions["B"].width = 16
        ws4.column_dimensions["C"].width = 14
        ws4.column_dimensions["D"].width = 22
        ws4.column_dimensions["E"].width = 10
        ws4.column_dimensions["F"].width = 12
        ws4.column_dimensions["G"].width = 14
        ws4.column_dimensions["H"].width = 12
        ws4.column_dimensions["I"].width = 14

        # Append the room QA summary on the same sheet below the table
        if isinstance(room_qa, dict):
            ws4.append([])
            ws4.append(["Room QA Summary"])
            ws4.cell(row=ws4.max_row, column=1).font = header_font
            room_metrics: list[tuple[str, Any]] = [
                ("Rooms Available", room_qa.get("rooms_available", 0)),
                ("Rooms Used (slot × room)", room_qa.get("rooms_used", 0)),
                ("Total Demand", room_qa.get("total_demand", 0)),
                ("Capacity Used", room_qa.get("total_capacity_used", 0)),
                (
                    "Avg Utilization",
                    f"{(float(room_qa.get('avg_utilization', 0) or 0) * 100):.1f}%",
                ),
                (
                    "Unassigned Sections",
                    len(room_qa.get("unassigned_room_sections", []) or []),
                ),
                (
                    "Double Bookings",
                    len(room_qa.get("room_double_bookings", []) or []),
                ),
            ]
            for lbl, val in room_metrics:
                ws4.append([lbl, val])
                rr = ws4.max_row
                ws4.cell(row=rr, column=1).font = header_font
                for c in range(1, 3):
                    ws4.cell(row=rr, column=c).border = thin_border

    # ────────────────────────────────────────────────────────────
    # Sheet 5: Invigilators — per-day totals (M/F) + per-room detail
    # ────────────────────────────────────────────────────────────
    if has_room_data:
        ws5 = wb.create_sheet("Invigilators")

        # ── Section A: rules legend ──
        ws5.append(["Invigilator Rules"])
        ws5.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws5.append(
            [
                "Department courses (CS/IS/COE/CYB/AI/DS):  1 invigilator if room has <30 students, "
                "2 if 30+"
            ]
        )
        ws5.append(
            [
                "External courses (GS/EDCT/GSE/ENV/MATH/STAT/PHYS):  1 invigilator only if room "
                "has more than 30 students, 0 otherwise"
            ]
        )
        ws5.append([])

        # ── Section B: daily summary ──
        summary_header_row = ws5.max_row + 1
        ws5.append(["Day", "M Invigilators", "F Invigilators", "Total"])
        style_header_row_at = summary_header_row
        for col in range(1, 5):
            cell = ws5.cell(row=style_header_row_at, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # Iterate days in slot order so the table matches the schedule grid
        invig_per_day_data = (
            (room_qa or {}).get("invigilators_per_day", {}) if isinstance(room_qa, dict) else {}
        )
        # Re-derive on the fly from the schedule too in case room_qa wasn't computed
        if not invig_per_day_data:
            tally: dict[str, dict[str, int]] = {}
            for e in schedule:
                if e.get("day") == "OVERFLOW":
                    continue
                day = e["day"]
                day_t = tally.setdefault(day, {"M": 0, "F": 0, "total": 0})
                for a in e.get("rooms", []) or []:
                    if a.get("room_code") == "UNASSIGNED":
                        continue
                    invigs = _invigilators_needed(
                        e["course_code"], int(a.get("student_count", 0) or 0)
                    )
                    g = a.get("gender", "M")
                    day_t[g] = day_t.get(g, 0) + invigs
                    day_t["total"] += invigs
            invig_per_day_data = tally

        running_M = 0
        running_F = 0
        for day in day_order:
            counts = invig_per_day_data.get(day, {"M": 0, "F": 0, "total": 0})
            ws5.append([day, counts["M"], counts["F"], counts["total"]])
            running_M += counts["M"]
            running_F += counts["F"]
            r = ws5.max_row
            for col in range(1, 5):
                ws5.cell(row=r, column=col).border = thin_border
                ws5.cell(row=r, column=col).alignment = center

        # Grand total row
        ws5.append(["TOTAL", running_M, running_F, running_M + running_F])
        r = ws5.max_row
        for col in range(1, 5):
            cell = ws5.cell(row=r, column=col)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = center
            cell.fill = PatternFill("solid", fgColor="E0E0E0")

        ws5.column_dimensions["A"].width = 18
        ws5.column_dimensions["B"].width = 18
        ws5.column_dimensions["C"].width = 18
        ws5.column_dimensions["D"].width = 14

        # ── Section C: per-room detail ──
        ws5.append([])
        detail_header_row = ws5.max_row + 1
        ws5.append(
            [
                "Day",
                "Period",
                "Course",
                "Type",
                "Section",
                "Gender",
                "Students",
                "Room",
                "Invigilators",
            ]
        )
        for col in range(1, 10):
            cell = ws5.cell(row=detail_header_row, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        sorted_for_invig = sorted(
            (e for e in schedule if e.get("day") != "OVERFLOW"),
            key=lambda e: (e.get("slot_index", 999), e["course_code"]),
        )
        for e in sorted_for_invig:
            cc = e["course_code"]
            prefix = _course_prefix(cc)
            ctype = "Department" if prefix in _DEPARTMENT_PREFIXES else "External"
            for a in e.get("rooms", []) or []:
                if a.get("room_code") == "UNASSIGNED":
                    continue
                stu = int(a.get("student_count", 0) or 0)
                invigs = _invigilators_needed(cc, stu)
                ws5.append(
                    [
                        e["day"],
                        e["period"],
                        cc,
                        ctype,
                        a.get("section", ""),
                        a.get("gender", ""),
                        stu,
                        a.get("room_code", ""),
                        invigs,
                    ]
                )
                rr = ws5.max_row
                for col in range(1, 10):
                    ws5.cell(row=rr, column=col).border = thin_border
                    ws5.cell(row=rr, column=col).alignment = center
                # Highlight rows with 0 invigilators (no department staffing needed)
                if invigs == 0:
                    for col in range(1, 10):
                        ws5.cell(row=rr, column=col).fill = PatternFill("solid", fgColor="EDEDED")
                # Highlight rows with 2 invigilators (heavy room)
                elif invigs >= 2:
                    for col in range(1, 10):
                        ws5.cell(row=rr, column=col).fill = warn_fill

        ws5.column_dimensions["E"].width = 22
        ws5.column_dimensions["F"].width = 10
        ws5.column_dimensions["G"].width = 10
        ws5.column_dimensions["H"].width = 14
        ws5.column_dimensions["I"].width = 14

    # ── Write to disk ──
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    out = RUNTIME_DIR / f"exam_timetable_{run_id}.xlsx"
    wb.save(str(out))
    return out
