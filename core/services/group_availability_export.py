"""Styled XLSX export for the Group Availability workspace.

The exporter consumes the same server-computed payload used by the screen. It
keeps the exact busy counts, while making incomplete coverage explicit so a
zero is never presented as certainty when no registered timetable was loaded.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, cast

_DAY_LABELS = {
    "SUN": "Sunday",
    "MON": "Monday",
    "TUE": "Tuesday",
    "WED": "Wednesday",
    "THU": "Thursday",
}

_GRID_SPECS = (
    ("lecture", "Lecture 75m", "Lecture Availability — 75 minutes"),
    ("lab", "Lab 100m", "Lab Availability — 100 minutes"),
    ("timeline", "Full Day 10m", "Full-day Availability — 10-minute periods"),
)


def _safe_excel_text(value: object) -> str:
    """Return literal cell text, neutralising spreadsheet formula prefixes."""
    text = str(value or "")
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def build_group_availability_xlsx(result: dict[str, Any]) -> bytes:
    """Build a readable, in-memory XLSX workbook from an availability result."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - dependency is pinned in production
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    navy = "111144"
    teal = "0A8E6E"
    pale_teal = "E8F5F0"
    amber = "D97706"
    pale_amber = "FFF3CD"
    red = "C03030"
    pale_red = "F8D7DA"
    slate = "5B6475"
    pale_slate = "EEF1F5"
    white = "FFFFFF"
    line = "CFD6DF"

    title_fill = PatternFill("solid", fgColor=navy)
    header_fill = PatternFill("solid", fgColor=teal)
    free_fill = PatternFill("solid", fgColor=pale_teal)
    some_fill = PatternFill("solid", fgColor=pale_amber)
    most_fill = PatternFill("solid", fgColor=pale_red)
    unknown_fill = PatternFill("solid", fgColor=pale_slate)
    label_fill = PatternFill("solid", fgColor="E9EEF5")
    thin = Side(style="thin", color=line)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    title_font = Font(name="Aptos Display", size=16, bold=True, color=white)
    subtitle_font = Font(name="Aptos", size=10, color=slate)
    header_font = Font(name="Aptos", size=10, bold=True, color=white)
    label_font = Font(name="Aptos", size=9, bold=True, color=navy)
    body_font = Font(name="Aptos", size=9, color="202735")

    def set_text(cell: Any, value: object) -> None:
        cell.value = _safe_excel_text(value)
        cell.data_type = "s"

    def banner(ws: Any, row: int, text: object, last_col: int, *, title: bool) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
        cell = ws.cell(row=row, column=1)
        banner_fill = title_fill if title else PatternFill("solid", fgColor="F5F7FA")
        set_text(cell, text)
        cell.fill = banner_fill
        cell.font = title_font if title else subtitle_font
        cell.alignment = left
        cell.border = border
        ws.row_dimensions[row].height = 28 if title else 22
        for column in range(2, last_col + 1):
            ws.cell(row=row, column=column).fill = banner_fill
            ws.cell(row=row, column=column).border = border

    requested = int(result.get("requested_count", 0) or 0)
    resolved = int(result.get("resolved_count", 0) or 0)
    unresolved = int(result.get("unresolved_count", max(0, requested - resolved)) or 0)
    partial = int(result.get("partial_schedule_count", 0) or 0)
    academic_year = _safe_excel_text(result.get("academic_year", ""))
    term = _safe_excel_text(result.get("term", ""))
    coverage_complete = bool(result.get("coverage_complete", False))
    not_found = {int(value) for value in result.get("not_found", [])}
    no_schedule = {int(value) for value in result.get("no_schedule", [])}
    partial_ids = {int(value) for value in result.get("partial_schedule", [])}

    workbook = Workbook()
    workbook.properties.creator = "Advisor Portal"
    workbook.properties.title = f"Group Availability {academic_year}/{term}"
    workbook.properties.subject = "Registered-schedule group availability"

    # ── Summary ────────────────────────────────────────────────
    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    banner(summary, 1, "Group Availability", 6, title=True)
    banner(
        summary,
        2,
        f"Registered schedules only · Academic term {academic_year}/{term}",
        6,
        title=False,
    )

    metrics = (
        ("Requested students", requested),
        ("Registered schedules", resolved),
        ("Unresolved", unresolved),
        ("Partially timed", partial),
        ("Coverage", "Complete" if coverage_complete else "Incomplete"),
        ("Source", "Registered only"),
    )
    for index, (label, value) in enumerate(metrics):
        row = 4 + index // 3
        label_column = 1 + (index % 3) * 2
        value_column = label_column + 1
        label_cell = summary.cell(row=row, column=label_column)
        set_text(label_cell, label)
        label_cell.fill = label_fill
        label_cell.font = label_font
        label_cell.alignment = left
        label_cell.border = border
        value_cell = summary.cell(row=row, column=value_column)
        if isinstance(value, int):
            value_cell.value = value
            value_cell.number_format = "#,##0"
        else:
            set_text(value_cell, value)
        value_cell.font = Font(name="Aptos", size=10, bold=True, color=teal)
        value_cell.alignment = center
        value_cell.border = border

    banner(summary, 8, "Coverage warnings", 6, title=False)
    warnings = (
        ("Not found", sorted(not_found)),
        ("No registered schedule", sorted(no_schedule)),
        ("Partially timed schedules", sorted(partial_ids)),
    )
    for offset, (label, ids) in enumerate(warnings, start=9):
        label_cell = summary.cell(row=offset, column=1)
        set_text(label_cell, label)
        label_cell.fill = label_fill
        label_cell.font = label_font
        label_cell.alignment = left
        label_cell.border = border
        summary.merge_cells(start_row=offset, start_column=2, end_row=offset, end_column=6)
        value_cell = summary.cell(row=offset, column=2)
        set_text(value_cell, ", ".join(str(value) for value in ids) if ids else "None")
        value_cell.font = body_font
        value_cell.alignment = wrap
        value_cell.border = border
        for column in range(3, 7):
            summary.cell(row=offset, column=column).border = border

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 24
    summary.column_dimensions["D"].width = 18
    summary.column_dimensions["E"].width = 18
    summary.column_dimensions["F"].width = 20
    summary.freeze_panes = "A4"
    summary.page_setup.orientation = "landscape"
    summary.page_setup.fitToWidth = 1
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    summary.print_area = "A1:F11"

    # ── Student coverage ───────────────────────────────────────
    students_sheet = workbook.create_sheet("Students")
    students_sheet.sheet_view.showGridLines = False
    banner(students_sheet, 1, "Student Coverage", 6, title=True)
    banner(
        students_sheet,
        2,
        "Flagged students remain listed and never block calculations for loaded schedules.",
        6,
        title=False,
    )
    student_headers = (
        "Student ID",
        "Name",
        "Program",
        "Coverage status",
        "Timed meetings",
        "Untimed sections",
    )
    for column, heading in enumerate(student_headers, start=1):
        cell = students_sheet.cell(row=4, column=column)
        set_text(cell, heading)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    students = list(result.get("students", []) or [])
    for row, student in enumerate(students, start=5):
        student_id = int(student.get("student_id", 0) or 0)
        meeting_count = int(student.get("meeting_count", 0) or 0)
        untimed_count = int(student.get("unscheduled_section_count", 0) or 0)
        if student_id in not_found:
            status = "Not found"
            status_fill = most_fill
        elif student_id in no_schedule:
            status = "No registered schedule"
            status_fill = unknown_fill
        elif student_id in partial_ids or untimed_count:
            status = "Partially timed"
            status_fill = some_fill
        elif meeting_count:
            status = "Loaded"
            status_fill = free_fill
        else:
            status = "Unresolved"
            status_fill = unknown_fill

        values: tuple[object, ...] = (
            student_id,
            student.get("name", ""),
            student.get("program", ""),
            status,
            meeting_count,
            untimed_count,
        )
        for column, value in enumerate(values, start=1):
            cell = students_sheet.cell(row=row, column=column)
            if column in (1, 5, 6):
                cell.value = cast(int, value)
                cell.number_format = "0"
            else:
                set_text(cell, value)
            cell.font = body_font
            cell.alignment = left if column in (2, 3, 4) else center
            cell.border = border
        students_sheet.cell(row=row, column=4).fill = status_fill

    for column, width in enumerate((15, 34, 16, 25, 18, 18), start=1):
        students_sheet.column_dimensions[get_column_letter(column)].width = width
    students_sheet.freeze_panes = "A5"
    if students:
        students_sheet.auto_filter.ref = f"A4:F{4 + len(students)}"
        students_sheet.print_area = f"A1:F{4 + len(students)}"
    else:
        students_sheet.print_area = "A1:F4"
    students_sheet.page_setup.orientation = "landscape"
    students_sheet.page_setup.fitToWidth = 1
    students_sheet.sheet_properties.pageSetUpPr.fitToPage = True

    # ── Availability matrices ──────────────────────────────────
    grids = result.get("grids", {}) or {}
    days = [str(day) for day in (result.get("weekdays", []) or _DAY_LABELS)]
    for grid_key, sheet_name, title in _GRID_SPECS:
        grid = grids.get(grid_key, {}) or {}
        slots = list(grid.get("slots", []) or [])
        cells = grid.get("cells", {}) or {}
        last_column = 1 + len(days)
        last_letter = get_column_letter(last_column)
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.tabColor = teal
        banner(sheet, 1, title, last_column, title=True)
        banner(
            sheet,
            2,
            (
                f"Term {academic_year}/{term} · {resolved}/{requested} registered schedules · "
                f"{unresolved} unresolved · {int(grid.get('free_for_resolved_count', 0) or 0)} free cells"
            ),
            last_column,
            title=False,
        )

        legend = (
            ("0 · Free", free_fill, teal),
            ("Some busy", some_fill, amber),
            ("Most busy", most_fill, red),
            ("? · No schedule data", unknown_fill, slate),
        )
        for column, (label, fill, color) in enumerate(legend, start=1):
            if column > last_column:
                break
            cell = sheet.cell(row=3, column=column)
            set_text(cell, label)
            cell.fill = fill
            cell.font = Font(name="Aptos", size=8, bold=True, color=color)
            cell.alignment = center
            cell.border = border

        headers = ["Time", *(_DAY_LABELS.get(day, day) for day in days)]
        for column, heading in enumerate(headers, start=1):
            cell = sheet.cell(row=5, column=column)
            set_text(cell, heading)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for slot_index, slot in enumerate(slots):
            row = 6 + slot_index
            time_cell = sheet.cell(row=row, column=1)
            set_text(time_cell, f"{slot.get('start', '')}–{slot.get('end', '')}")
            time_cell.fill = label_fill
            time_cell.font = label_font
            time_cell.alignment = center
            time_cell.border = border

            for day_index, day in enumerate(days, start=2):
                day_cells = cells.get(day, []) or []
                cell_data = day_cells[slot_index] if slot_index < len(day_cells) else {}
                busy = int(cell_data.get("busy_count", 0) or 0)
                cell = sheet.cell(row=row, column=day_index)
                if resolved <= 0:
                    set_text(cell, "?")
                    cell.fill = unknown_fill
                    cell.font = Font(name="Aptos", size=10, bold=True, color=slate)
                else:
                    cell.value = busy
                    cell.number_format = "0"
                    ratio = busy / max(1, resolved)
                    if busy <= 0:
                        cell.fill = free_fill
                        color = teal
                    elif ratio >= 0.5:
                        cell.fill = most_fill
                        color = red
                    else:
                        cell.fill = some_fill
                        color = amber
                    cell.font = Font(name="Aptos", size=10, bold=True, color=color)
                cell.alignment = center
                cell.border = border

        sheet.column_dimensions["A"].width = 18
        for column in range(2, last_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 17
        sheet.freeze_panes = "B6"
        sheet.auto_filter.ref = f"A5:{last_letter}{5 + len(slots)}"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_title_rows = "1:5"
        sheet.print_area = f"A1:{last_letter}{5 + len(slots)}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
