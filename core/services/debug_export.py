"""
core/services/debug_export.py
XLSX export for the Recommendation Debug report.

Creates a styled workbook with:
  - Students sheet: one row per student with ID, program, term, passed/studying counts,
    recommended courses as colored pills, and prerequisite status
  - Summary sheet: metrics, course demand breakdown, student distribution by term
  - Course Detail sheet: for each recommended course, which students need it
"""

from __future__ import annotations

import tempfile
from collections import Counter, defaultdict
from pathlib import Path


def export_recommendation_debug_xlsx(payload: dict) -> Path:
    """Export recommendation debug report as styled XLSX.

    Args:
        payload: dict from build_recommendation_debug_report() with keys:
            count, filters, items (list of student dicts)

    Returns:
        Path to temporary .xlsx file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    items = payload.get("items", [])
    filters = payload.get("filters", {})

    # ── Styles ───────────────────────────────────────────────────
    hdr_fill = PatternFill(start_color="0A8E6E", end_color="0A8E6E", fill_type="solid")
    hdr_font = Font(name="Consolas", size=9, bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    bold_font = Font(bold=True, size=9)
    normal_font = Font(name="Consolas", size=9)
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    rec_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    no_rec_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    studying_fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
    passed_fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
    thick_side = Side(style="medium", color="333333")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Students Sheet ───────────────────────────────────────────
    ws = wb.create_sheet(title="Students")

    headers = [
        "Student ID", "Program", "Real Term", "Next Term",
        "Passed", "Studying", "Recommended Courses",
        "Prerequisite Details",
    ]
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = hdr
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border

    row = 2
    for item in items:
        sid = item.get("student_id", "")
        program = item.get("program", "")
        real_term = item.get("real_term", "")
        next_term = item.get("next_term", "")
        passed = item.get("passed", [])
        studying = item.get("studying", [])
        recs = item.get("recommended_courses", [])
        rec_details = item.get("recommendation_details", [])

        has_recs = len(recs) > 0

        # Student ID
        cell = ws.cell(row=row, column=1, value=sid)
        cell.font = bold_font
        cell.border = thin_border

        # Program
        cell = ws.cell(row=row, column=2, value=program)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center_align

        # Real Term
        cell = ws.cell(row=row, column=3, value=real_term)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center_align

        # Next Term
        cell = ws.cell(row=row, column=4, value=next_term)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center_align

        # Passed count
        cell = ws.cell(row=row, column=5, value=len(passed))
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center_align
        cell.fill = passed_fill

        # Studying count
        cell = ws.cell(row=row, column=6, value=len(studying))
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center_align
        if studying:
            cell.fill = studying_fill

        # Recommended courses
        recs_text = ", ".join(recs) if recs else "—"
        cell = ws.cell(row=row, column=7, value=recs_text)
        cell.font = Font(name="Consolas", size=9, bold=has_recs)
        cell.border = thin_border
        cell.alignment = wrap_align
        cell.fill = rec_fill if has_recs else no_rec_fill

        # Prerequisite details
        prereq_parts = []
        for rd in rec_details:
            code = rd.get("course_code", "")
            prereqs = rd.get("prerequisite_status", [])
            if prereqs:
                statuses = ", ".join(
                    f"{ps['prerequisite']}({ps['status'][0]})"
                    for ps in prereqs
                )
                prereq_parts.append(f"{code}: {statuses}")
            else:
                prereq_parts.append(f"{code}: no prereqs")
        cell = ws.cell(row=row, column=8, value="\n".join(prereq_parts) if prereq_parts else "—")
        cell.font = Font(size=8, color="666666")
        cell.border = thin_border
        cell.alignment = wrap_align

        row += 1

    # Outer border
    if row > 2:
        for r in range(1, row):
            for c in range(1, len(headers) + 1):
                cl = ws.cell(row=r, column=c)
                ex = cl.border
                cl.border = Border(
                    left=thick_side if c == 1 else ex.left,
                    right=thick_side if c == len(headers) else ex.right,
                    top=thick_side if r == 1 else ex.top,
                    bottom=thick_side if r == row - 1 else ex.bottom,
                )

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 35
    ws.column_dimensions["H"].width = 45
    ws.freeze_panes = "A2"

    # ── Summary Sheet ────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="Summary")

    # Filters
    ws_sum.cell(row=1, column=1, value="Recommendation Debug Report").font = Font(bold=True, size=12)
    ws_sum.cell(row=2, column=1, value="Year").font = bold_font
    ws_sum.cell(row=2, column=2, value=filters.get("year", ""))
    ws_sum.cell(row=2, column=3, value="Semester").font = bold_font
    ws_sum.cell(row=2, column=4, value=filters.get("semester", ""))
    ws_sum.cell(row=3, column=1, value="Program").font = bold_font
    ws_sum.cell(row=3, column=2, value=filters.get("program", "All"))
    ws_sum.cell(row=3, column=3, value="Section").font = bold_font
    ws_sum.cell(row=3, column=4, value=filters.get("section", "All"))

    # Metrics
    total = len(items)
    with_recs = sum(1 for i in items if i.get("recommended_courses"))
    without_recs = total - with_recs

    row = 5
    ws_sum.cell(row=row, column=1, value="Students").font = bold_font
    ws_sum.cell(row=row, column=2, value=total)
    ws_sum.cell(row=row, column=3, value="With Recs").font = bold_font
    ws_sum.cell(row=row, column=4, value=with_recs)
    ws_sum.cell(row=row, column=4).fill = rec_fill
    ws_sum.cell(row=row, column=5, value="Empty Recs").font = bold_font
    ws_sum.cell(row=row, column=6, value=without_recs)
    ws_sum.cell(row=row, column=6).fill = no_rec_fill

    # Course demand breakdown
    row = 7
    ws_sum.cell(row=row, column=1, value="Course Demand").font = Font(bold=True, size=11)
    row += 1
    for col, hdr in enumerate(["Course", "Students Needing", "% of Total"], start=1):
        cell = ws_sum.cell(row=row, column=col)
        cell.value = hdr
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border
    row += 1

    course_demand: Counter[str] = Counter()
    for item in items:
        for code in item.get("recommended_courses", []):
            course_demand[code] += 1

    for code, count in course_demand.most_common():
        ws_sum.cell(row=row, column=1, value=code).font = bold_font
        ws_sum.cell(row=row, column=1).border = thin_border
        ws_sum.cell(row=row, column=2, value=count).border = thin_border
        ws_sum.cell(row=row, column=2).alignment = center_align
        pct = round(count / max(total, 1) * 100)
        ws_sum.cell(row=row, column=3, value=f"{pct}%").border = thin_border
        ws_sum.cell(row=row, column=3).alignment = center_align
        row += 1

    # Student distribution by real_term
    row += 1
    ws_sum.cell(row=row, column=1, value="Students by Term").font = Font(bold=True, size=11)
    row += 1
    for col, hdr in enumerate(["Real Term", "Count", "With Recs", "Empty"], start=1):
        cell = ws_sum.cell(row=row, column=col)
        cell.value = hdr
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border
    row += 1

    term_stats: dict[int, dict] = defaultdict(lambda: {"total": 0, "with": 0, "without": 0})
    for item in items:
        rt = item.get("real_term", 0)
        term_stats[rt]["total"] += 1
        if item.get("recommended_courses"):
            term_stats[rt]["with"] += 1
        else:
            term_stats[rt]["without"] += 1

    for rt in sorted(term_stats.keys()):
        s = term_stats[rt]
        ws_sum.cell(row=row, column=1, value=rt).border = thin_border
        ws_sum.cell(row=row, column=1).alignment = center_align
        ws_sum.cell(row=row, column=2, value=s["total"]).border = thin_border
        ws_sum.cell(row=row, column=2).alignment = center_align
        ws_sum.cell(row=row, column=3, value=s["with"]).border = thin_border
        ws_sum.cell(row=row, column=3).alignment = center_align
        ws_sum.cell(row=row, column=3).fill = rec_fill if s["with"] > 0 else PatternFill()
        ws_sum.cell(row=row, column=4, value=s["without"]).border = thin_border
        ws_sum.cell(row=row, column=4).alignment = center_align
        if s["without"] > 0:
            ws_sum.cell(row=row, column=4).fill = no_rec_fill
        row += 1

    ws_sum.column_dimensions["A"].width = 16
    ws_sum.column_dimensions["B"].width = 16
    ws_sum.column_dimensions["C"].width = 14
    ws_sum.column_dimensions["D"].width = 14
    ws_sum.column_dimensions["E"].width = 14
    ws_sum.column_dimensions["F"].width = 14
    ws_sum.freeze_panes = "A2"

    # ── Course Detail Sheet ──────────────────────────────────────
    ws_cd = wb.create_sheet(title="Course Detail")

    ws_cd.cell(row=1, column=1, value="Course").font = hdr_font
    ws_cd.cell(row=1, column=1).fill = hdr_fill
    ws_cd.cell(row=1, column=1).border = thin_border
    ws_cd.cell(row=1, column=2, value="Students").font = hdr_font
    ws_cd.cell(row=1, column=2).fill = hdr_fill
    ws_cd.cell(row=1, column=2).border = thin_border
    ws_cd.cell(row=1, column=3, value="Student IDs").font = hdr_font
    ws_cd.cell(row=1, column=3).fill = hdr_fill
    ws_cd.cell(row=1, column=3).border = thin_border

    # Build course → student list
    course_students: dict[str, list[int]] = defaultdict(list)
    for item in items:
        sid = item.get("student_id", "")
        for code in item.get("recommended_courses", []):
            course_students[code].append(sid)

    row = 2
    for code in sorted(course_students.keys()):
        sids = course_students[code]
        ws_cd.cell(row=row, column=1, value=code).font = bold_font
        ws_cd.cell(row=row, column=1).border = thin_border
        ws_cd.cell(row=row, column=2, value=len(sids)).border = thin_border
        ws_cd.cell(row=row, column=2).alignment = center_align
        ws_cd.cell(row=row, column=3, value=", ".join(str(s) for s in sorted(sids))).font = Font(size=8)
        ws_cd.cell(row=row, column=3).border = thin_border
        ws_cd.cell(row=row, column=3).alignment = wrap_align
        row += 1

    ws_cd.column_dimensions["A"].width = 14
    ws_cd.column_dimensions["B"].width = 10
    ws_cd.column_dimensions["C"].width = 80
    ws_cd.freeze_panes = "A2"

    # ── Save ─────────────────────────────────────────────────────
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)
