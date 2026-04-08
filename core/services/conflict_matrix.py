from __future__ import annotations

from pathlib import Path

from core.models import Student
from core.services.debug_reporting import _build_students_query
from core.services.recommender import recommend_next_courses
from core.services.recommender_batch import batch_recommend

BASE_DIR = Path(__file__).resolve().parents[2]
RUNTIME_DIR = BASE_DIR / "runtime"


def build_conflict_matrix_report(
    current_academic_year: int,
    current_semester: int,
    section: str | None = None,
    program: str | None = None,
    join_year_prefixes: list[str] | None = None,
    limit: int = 150,
) -> dict:
    """Build an NxN conflict matrix from recommended courses.

    Returns a dict with:
      - courses: sorted list of course codes
      - matrix: NxN list-of-lists (matrix[i][j] = students sharing both courses)
      - student_count: total students processed
      - students: list of {student_id, program, recommended_courses}
    """
    student_ids = _build_students_query(section, program, join_year_prefixes)[:limit]

    # Build course→students mapping using batch recommender when possible
    course_to_students: dict[str, set[int]] = {}
    student_rows: list[dict] = []

    if program:
        all_recs = batch_recommend(student_ids, program, current_academic_year, current_semester)
        for sid, recs in all_recs.items():
            student_rows.append({"student_id": sid, "recommended_courses": recs})
            for code in recs:
                course_to_students.setdefault(code, set()).add(sid)
    else:
        # No program filter — need per-student program lookup
        for sid in student_ids:
            recs = recommend_next_courses(sid, current_academic_year, current_semester)
            if recs:
                student_rows.append({"student_id": sid, "recommended_courses": recs})
                for code in recs:
                    course_to_students.setdefault(code, set()).add(sid)

    # Build NxN matrix
    courses = sorted(course_to_students.keys())
    n = len(courses)
    sets = [course_to_students[c] for c in courses]

    matrix = [[0] * n for _ in range(n)]
    for i in range(n):
        matrix[i][i] = len(sets[i])
        for j in range(i + 1, n):
            shared = len(sets[i] & sets[j])
            matrix[i][j] = shared
            matrix[j][i] = shared

    return {
        "courses": courses,
        "matrix": matrix,
        "student_count": len(student_ids),
        "filters": {
            "section": section,
            "program": program,
            "join_year_prefixes": join_year_prefixes or [],
            "limit": limit,
            "year": current_academic_year,
            "semester": current_semester,
        },
    }


def export_conflict_matrix_xlsx(
    current_academic_year: int,
    current_semester: int,
    section: str | None = None,
    program: str | None = None,
    join_year_prefixes: list[str] | None = None,
    limit: int = 150,
) -> Path:
    """Export the conflict matrix as a styled XLSX workbook."""
    report = build_conflict_matrix_report(
        current_academic_year,
        current_semester,
        section,
        program,
        join_year_prefixes,
        limit,
    )

    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    out = RUNTIME_DIR / "conflict_matrix.xlsx"

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except Exception as exc:
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Conflict Matrix"

    courses = report["courses"]
    matrix = report["matrix"]
    n = len(courses)

    # Find max off-diagonal value for heat-map
    max_val = (
        max(
            (matrix[i][j] for i in range(n) for j in range(n) if i != j),
            default=1,
        )
        or 1
    )

    # Styles
    thin = Side(style="thin", color="CCCCCC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    hdr_fill = PatternFill(start_color="0A8E6E", end_color="0A8E6E", fill_type="solid")
    hdr_font = Font(name="Consolas", size=9, bold=True, color="FFFFFF")
    diag_fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
    diag_font = Font(name="Consolas", size=9, bold=True, color="0A8E6E")
    zero_font = Font(name="Consolas", size=9, color="AAAAAA")
    center = Alignment(horizontal="center", vertical="center")

    # Corner cell
    ws.cell(row=1, column=1, value="").border = border

    # Column headers (row 1)
    for j, code in enumerate(courses):
        cell = ws.cell(row=1, column=j + 2, value=code)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=90)
        cell.border = border

    # Data rows
    for i, row_code in enumerate(courses):
        # Row header
        rh = ws.cell(row=i + 2, column=1, value=row_code)
        rh.fill = hdr_fill
        rh.font = hdr_font
        rh.alignment = Alignment(horizontal="left", vertical="center")
        rh.border = border

        for j in range(n):
            val = matrix[i][j]
            cell = ws.cell(row=i + 2, column=j + 2, value=val)
            cell.alignment = center
            cell.border = border

            if i == j:
                cell.fill = diag_fill
                cell.font = diag_font
            elif val == 0:
                cell.font = zero_font
            else:
                t = min(val / max_val, 1.0)
                # Teal → royal gradient
                r = int(200 - 150 * t)
                g = int(230 - 140 * t)
                b = int(220 - 10 * t)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(
                    start_color=hex_color, end_color=hex_color, fill_type="solid"
                )
                cell.font = Font(
                    name="Consolas", size=9, bold=True, color="FFFFFF" if t > 0.6 else "333333"
                )

    # Column widths
    ws.column_dimensions["A"].width = 14
    for j in range(n):
        col_letter = chr(66 + j) if j < 25 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = 8

    # Freeze panes (top row + first column)
    ws.freeze_panes = "B2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2.append(["Students", report["student_count"]])
    ws2.append(["Courses", n])
    conflicts = sum(1 for i in range(n) for j in range(i + 1, n) if matrix[i][j] > 0)
    ws2.append(["Conflict Pairs", conflicts])
    ws2.append(["Section", section or "All"])
    ws2.append(["Program", program or "All"])
    ws2.append(["Year", current_academic_year])
    ws2.append(["Semester", current_semester])
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 20

    wb.save(str(out))
    return out
