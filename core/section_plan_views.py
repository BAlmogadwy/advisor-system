"""
core/section_plan_views.py
Next Semester Section Planning — page view + API endpoints.

Endpoints:
    GET  section_plan_page          – render the section planning page
    POST section_plan_generate_view – compute section demand from recommendations
    POST section_plan_export_view   – download section plan as styled .xlsx
"""

from __future__ import annotations

import json
import logging
import math
import tempfile
from pathlib import Path

from django.contrib.auth.decorators import login_required
from django.http import FileResponse, HttpRequest, HttpResponse, HttpResponseBase, JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_GET, require_POST

from core.authz import role_required, throttle
from core.models import ProgrammeRequirement
from core.services.rbac import ROLE_GENERAL_ADVISOR, get_user_role
from core.services.reporting import build_aggregate_counts
from core.services.section_planning import (
    DEFAULT_MAX_EXTERNAL,
    DEFAULT_MAX_LOCAL_4CR,
    DEFAULT_MAX_LOCAL_OTHER,
    compute_plan_summary,
    compute_section_plan,
    get_all_courses_with_defaults,
    load_programme_capacities,
)
from core.services.student_helpers import normalize_code
from core.settings_views import load_defaults
from core.sidebar_context import get_sidebar_context

logger = logging.getLogger(__name__)


def _apply_programme_course_names(plan: list[dict], program: str) -> list[dict]:
    """Overlay ProgrammeRequirement.course_name for display-only plan rows."""
    course_codes = [
        normalize_code(str(row.get("course_code", ""))) for row in plan if row.get("course_code")
    ]
    if not course_codes:
        return [dict(row) for row in plan]

    names = {
        normalize_code(code): course_name or ""
        for code, course_name in ProgrammeRequirement.objects.filter(
            program=program,
            course_code__in=course_codes,
        ).values_list("course_code", "course_name")
    }

    updated: list[dict] = []
    for row in plan:
        row_copy = dict(row)
        code = normalize_code(str(row_copy.get("course_code", "")))
        row_copy["course_name"] = names.get(code) or row_copy.get("course_name", "")
        updated.append(row_copy)
    return updated


def _merge_section_plan_rows_by_course_identity(
    program_plans: list[tuple[str, list[dict]]],
) -> list[dict]:
    """Merge multi-program rows by course code and plan-specific course name."""
    merged: dict[tuple[str, str], dict] = {}

    for program, plan in program_plans:
        for row in plan:
            code = normalize_code(str(row.get("course_code", "")))
            name = str(row.get("course_name", "") or "")
            key = (code, name)

            if key not in merged:
                row_copy = dict(row)
                row_copy["course_code"] = code
                row_copy["course_name"] = name
                row_copy["total_students"] = 0
                row_copy["programs"] = []
                merged[key] = row_copy

            target = merged[key]
            if program not in target["programs"]:
                target["programs"].append(program)
            target["total_students"] += int(row.get("total_students") or 0)
            target["max_per_section"] = min(
                int(target.get("max_per_section") or 1),
                int(row.get("max_per_section") or 1),
            )
            target["is_external"] = bool(target.get("is_external")) or bool(row.get("is_external"))

    result = []
    for row in merged.values():
        max_per_section = max(1, int(row.get("max_per_section") or 1))
        total_students = int(row.get("total_students") or 0)
        num_sections = max(1, math.ceil(total_students / max_per_section))
        avg_per_section = math.ceil(total_students / num_sections)
        fill_percent = round((avg_per_section / max_per_section) * 100)

        if avg_per_section >= max_per_section:
            status = "full"
        elif avg_per_section < 10:
            status = "underfilled"
        else:
            status = ""

        row["total_students"] = total_students
        row["num_sections"] = num_sections
        row["avg_per_section"] = avg_per_section
        row["fill_percent"] = fill_percent
        row["status"] = status
        row["programs"] = sorted(row.get("programs") or [])
        result.append(row)

    result.sort(
        key=lambda r: (
            r.get("department", ""),
            r.get("course_code", ""),
            r.get("course_name", ""),
        )
    )
    return result


def _format_export_course_name(row: dict, course_names: dict[str, str]) -> str:
    """Return the display name written into the XLSX export."""
    code = normalize_code(str(row.get("course_code", "")))
    course_name = str(row.get("course_name") or course_names.get(code, "") or "")
    programs = [str(program) for program in row.get("programs", []) if str(program).strip()]
    if programs:
        program_label = ", ".join(programs)
        return f"{program_label} - {course_name}" if course_name else program_label
    return course_name


def _require_general_advisor(request: HttpRequest) -> JsonResponse | None:
    """Guard: returns a 403 JsonResponse if user is below GENERAL_ADVISOR, else None."""
    from core.services.rbac import ROLE_SUPER_ADMIN

    role = get_user_role(request.user)
    if role not in {ROLE_GENERAL_ADVISOR, ROLE_SUPER_ADMIN}:
        return JsonResponse({"error": "General Advisor access required"}, status=403)
    return None


def _parse_payload(request: HttpRequest) -> tuple[dict | None, JsonResponse | None]:
    """Parse JSON body and extract validated parameters.

    Returns (params_dict, None) on success or (None, error_response) on failure.
    """
    try:
        body = json.loads(request.body.decode("utf-8")) if request.body else {}
    except (json.JSONDecodeError, UnicodeDecodeError):
        return None, JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    try:
        year = int(body.get("year", 0))
        semester = int(body.get("semester", 0))
    except (ValueError, TypeError):
        return None, JsonResponse(
            {"ok": False, "error": "year and semester must be integers"},
            status=400,
        )

    if not (1400 <= year <= 1600):
        return None, JsonResponse(
            {"ok": False, "error": "year must be between 1400 and 1600"},
            status=400,
        )
    if semester not in (1, 2, 3):
        return None, JsonResponse(
            {"ok": False, "error": "semester must be 1, 2, or 3"},
            status=400,
        )

    # Support comma-separated programs  e.g. "AI,DS" → ["AI", "DS"]
    program_raw = str(body.get("program", "")).strip()
    if program_raw and "," in program_raw:
        program: str | list[str] | None = [p.strip() for p in program_raw.split(",") if p.strip()]
        if not program:
            program = None
    else:
        program = program_raw or None

    section = str(body.get("section", "")).strip() or None

    try:
        max_local_4cr = int(body.get("max_local_4cr", DEFAULT_MAX_LOCAL_4CR))
        max_local_other = int(body.get("max_local_other", DEFAULT_MAX_LOCAL_OTHER))
        max_external = int(body.get("max_external", DEFAULT_MAX_EXTERNAL))
    except (ValueError, TypeError):
        return None, JsonResponse(
            {"ok": False, "error": "Capacity limits must be integers"},
            status=400,
        )

    # Clamp to reasonable range
    max_local_4cr = max(5, min(max_local_4cr, 200))
    max_local_other = max(5, min(max_local_other, 200))
    max_external = max(5, min(max_external, 200))

    # Per-course capacity overrides  { "CS101": 30, "AI201": 20, ... }
    raw_overrides = body.get("course_overrides") or {}
    course_overrides: dict[str, int] = {}
    if isinstance(raw_overrides, dict):
        for k, v in raw_overrides.items():
            try:
                val = int(v)
                if val >= 1:
                    course_overrides[normalize_code(str(k))] = min(val, 500)
            except (ValueError, TypeError):
                pass

    # Department prefix filter for export  e.g. "CS,AI"
    dept_filter_raw = str(body.get("dept_filter", "")).strip().upper()
    dept_prefixes = (
        [p.strip() for p in dept_filter_raw.split(",") if p.strip()] if dept_filter_raw else []
    )

    return {
        "year": year,
        "semester": semester,
        "program": program,
        "section": section,
        "max_local_4cr": max_local_4cr,
        "max_local_other": max_local_other,
        "max_external": max_external,
        "course_overrides": course_overrides,
        "dept_filter": dept_prefixes,
    }, None


# ── Page view ──────────────────────────────────────────────────


@login_required(login_url="login")
@require_GET
def section_plan_page(request: HttpRequest) -> HttpResponse:
    """Render the section planning page."""
    deny = _require_general_advisor(request)
    if deny:
        return deny
    defaults = load_defaults()
    ctx = {
        **get_sidebar_context(request),
        "default_year": defaults["academic_year"],
        "default_term": defaults["term"],
    }
    return render(request, "core/section_planning.html", ctx)


# ── Generate API ───────────────────────────────────────────────


@role_required(ROLE_GENERAL_ADVISOR)
@require_POST
@throttle(max_calls=3, window_seconds=120)
def section_plan_generate_view(request: HttpRequest) -> JsonResponse:
    """Compute section demand from batch recommendations."""
    params, err = _parse_payload(request)
    if err:
        return err
    assert params is not None

    program = params["program"]
    capacity_kwargs = {
        "max_local_4cr": params["max_local_4cr"],
        "max_local_other": params["max_local_other"],
        "max_external": params["max_external"],
        "course_overrides": params.get("course_overrides"),
    }

    try:
        if isinstance(program, list):
            # ── Multi-program mode ──
            result_programs: list[dict] = []
            total_student_count = 0
            for prog in program:
                sc, agg = build_aggregate_counts(
                    params["year"],
                    params["semester"],
                    program=prog,
                    section=params["section"],
                    resolve_electives=True,
                )
                pr_caps = load_programme_capacities(prog, list(agg.keys()))
                plan = compute_section_plan(
                    agg,
                    **capacity_kwargs,
                    programme_capacities=pr_caps,
                )
                plan = _apply_programme_course_names(plan, prog)
                summary = compute_plan_summary(plan)
                result_programs.append(
                    {
                        "program": prog,
                        "student_count": sc,
                        "plan": plan,
                        "summary": summary,
                    }
                )
                total_student_count += sc

            # Build combined plan by the same identity shown in the table.
            combined_plan = _merge_section_plan_rows_by_course_identity(
                [(prog_data["program"], prog_data["plan"]) for prog_data in result_programs]
            )
            combined_summary = compute_plan_summary(combined_plan)

            # Build additive department summary (sum per-program dept stats)
            # This reflects actual teaching load across programs, not recomputed
            additive_dept: dict[str, dict[str, int]] = {}
            for prog_data in result_programs:
                for dept in (prog_data.get("summary") or {}).get("departments", []):
                    d = dept["department"]
                    if d not in additive_dept:
                        additive_dept[d] = {
                            "courses": 0,
                            "sections": 0,
                            "students": 0,
                            "total_credits": 0,
                        }
                    additive_dept[d]["courses"] += dept["courses"]
                    additive_dept[d]["sections"] += dept["sections"]
                    additive_dept[d]["students"] += dept["students"]
                    additive_dept[d]["total_credits"] += dept["total_credits"]
            combined_summary["departments"] = [
                {"department": d, **v} for d, v in sorted(additive_dept.items())
            ]

            return JsonResponse(
                {
                    "ok": True,
                    "mode": "multi",
                    "year": params["year"],
                    "semester": params["semester"],
                    "student_count": total_student_count,
                    "combined_plan": combined_plan,
                    "combined_summary": combined_summary,
                    "programs": result_programs,
                }
            )

        elif isinstance(program, str):
            # ── Single-program mode ──
            student_count, aggregate = build_aggregate_counts(
                params["year"],
                params["semester"],
                program=program,
                section=params["section"],
                resolve_electives=True,
            )
            pr_caps = load_programme_capacities(program, list(aggregate.keys()))
            plan = compute_section_plan(
                aggregate,
                **capacity_kwargs,
                programme_capacities=pr_caps,
            )
            plan = _apply_programme_course_names(plan, program)
            summary = compute_plan_summary(plan)

            return JsonResponse(
                {
                    "ok": True,
                    "mode": "single",
                    "year": params["year"],
                    "semester": params["semester"],
                    "student_count": student_count,
                    "plan": plan,
                    "summary": summary,
                }
            )

        else:
            # ── Combined mode (no program filter) ──
            student_count, aggregate = build_aggregate_counts(
                params["year"],
                params["semester"],
                program=None,
                section=params["section"],
                resolve_electives=True,
            )
            plan = compute_section_plan(aggregate, **capacity_kwargs)
            summary = compute_plan_summary(plan)

            return JsonResponse(
                {
                    "ok": True,
                    "mode": "combined",
                    "year": params["year"],
                    "semester": params["semester"],
                    "student_count": student_count,
                    "plan": plan,
                    "summary": summary,
                }
            )

    except Exception as exc:
        logger.exception("section_plan_generate error")
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)


# ── Courses list API (for advanced per-course settings) ───────


@role_required(ROLE_GENERAL_ADVISOR)
@require_GET
def section_plan_courses_view(request: HttpRequest) -> JsonResponse:
    """Return all courses with their computed default capacities.

    Query params (optional): max_local_4cr, max_local_other, max_external
    — so the list reflects the current global settings.
    """
    try:
        max_local_4cr = int(request.GET.get("max_local_4cr", DEFAULT_MAX_LOCAL_4CR))
        max_local_other = int(request.GET.get("max_local_other", DEFAULT_MAX_LOCAL_OTHER))
        max_external = int(request.GET.get("max_external", DEFAULT_MAX_EXTERNAL))
    except (ValueError, TypeError):
        max_local_4cr = DEFAULT_MAX_LOCAL_4CR
        max_local_other = DEFAULT_MAX_LOCAL_OTHER
        max_external = DEFAULT_MAX_EXTERNAL

    program_raw = request.GET.get("program", "").strip()
    if program_raw and "," in program_raw:
        program: str | list[str] | None = [p.strip() for p in program_raw.split(",") if p.strip()]
    else:
        program = program_raw or None
    courses = get_all_courses_with_defaults(
        max_local_4cr, max_local_other, max_external, program=program
    )
    return JsonResponse({"ok": True, "courses": courses})


# ── Save per-course capacity API ──────────────────────────────


@role_required(ROLE_GENERAL_ADVISOR)
@require_POST
def section_plan_save_capacity_view(request: HttpRequest) -> JsonResponse:
    """Persist a per-course max_capacity override to ProgrammeRequirement rows.

    Accepts JSON body:
        {
            "programs": ["AI", "DS"],
            "course_code": "CS211",
            "max_capacity": 30        // or null to clear
        }

    Updates ProgrammeRequirement.max_capacity for every row matching
    (program IN programs) AND (course_code = course_code).
    Returns {"ok": True, "updated": <count>}.
    """
    try:
        body = json.loads(request.body.decode("utf-8")) if request.body else {}
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    # ── Validate programs ──
    programs = body.get("programs")
    if not programs or not isinstance(programs, list):
        return JsonResponse(
            {"ok": False, "error": "'programs' must be a non-empty list of program codes"},
            status=400,
        )
    programs = [str(p).strip() for p in programs if str(p).strip()]
    if not programs:
        return JsonResponse(
            {"ok": False, "error": "'programs' must be a non-empty list of program codes"},
            status=400,
        )

    # ── Validate course_code ──
    course_code = str(body.get("course_code", "")).strip()
    if not course_code:
        return JsonResponse(
            {"ok": False, "error": "'course_code' is required"},
            status=400,
        )
    course_code = normalize_code(course_code)

    # ── Validate max_capacity ──
    raw_cap = body.get("max_capacity")
    if raw_cap is None or raw_cap == "" or raw_cap == "null":
        max_capacity = None
    else:
        try:
            max_capacity = int(raw_cap)
            if max_capacity < 1:
                return JsonResponse(
                    {"ok": False, "error": "'max_capacity' must be a positive integer or null"},
                    status=400,
                )
            max_capacity = min(max_capacity, 500)  # reasonable upper bound
        except (ValueError, TypeError):
            return JsonResponse(
                {"ok": False, "error": "'max_capacity' must be a positive integer or null"},
                status=400,
            )

    # ── Perform update ──
    updated = ProgrammeRequirement.objects.filter(
        program__in=programs,
        course_code=course_code,
    ).update(max_capacity=max_capacity)

    logger.info(
        "save_capacity: user=%s programs=%s course=%s max_capacity=%s updated=%d",
        request.user.username,
        programs,
        course_code,
        max_capacity,
        updated,
    )

    return JsonResponse({"ok": True, "updated": updated})


@login_required(login_url="login")
@require_POST
def section_plan_save_overrides_bulk_view(request: HttpRequest) -> JsonResponse:
    """Persist all per-course overrides from the advanced panel to the DB.

    Accepts JSON: ``{"overrides": {"CS211": 30, "AI492": 5, ...}}``
    Updates ProgrammeRequirement.max_capacity for ALL programmes that have
    each course code.  Returns count of updated rows.
    """
    deny = _require_general_advisor(request)
    if deny:
        return deny

    try:
        body = json.loads(request.body.decode("utf-8")) if request.body else {}
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    overrides = body.get("overrides", {})
    if not isinstance(overrides, dict):
        return JsonResponse({"ok": False, "error": "'overrides' must be a dict"}, status=400)

    total_updated = 0
    for code, cap in overrides.items():
        code_n = normalize_code(str(code))
        if not code_n:
            continue
        try:
            cap_val = int(cap)
            if cap_val < 1:
                continue
            cap_val = min(cap_val, 500)
        except (ValueError, TypeError):
            continue

        updated = ProgrammeRequirement.objects.filter(
            course_code=code_n,
        ).update(max_capacity=cap_val)
        total_updated += updated

    logger.info(
        "save_overrides_bulk: user=%s courses=%d updated=%d",
        request.user.username,
        len(overrides),
        total_updated,
    )

    return JsonResponse({"ok": True, "updated": total_updated, "courses": len(overrides)})


# ── Export XLSX API ────────────────────────────────────────────


@role_required(ROLE_GENERAL_ADVISOR)
@require_POST
@throttle(max_calls=5, window_seconds=120)
def section_plan_export_view(request: HttpRequest) -> HttpResponseBase:
    """Generate and download section plan as styled XLSX."""
    params, err = _parse_payload(request)
    if err:
        return err
    assert params is not None

    program = params["program"]
    capacity_kwargs = {
        "max_local_4cr": params["max_local_4cr"],
        "max_local_other": params["max_local_other"],
        "max_external": params["max_external"],
        "course_overrides": params.get("course_overrides"),
    }

    dept_prefixes = params.get("dept_filter", [])

    def _filter_plan(plan: list[dict]) -> list[dict]:
        if not dept_prefixes:
            return plan
        return [
            entry
            for entry in plan
            if any(normalize_code(entry["course_code"]).startswith(p) for p in dept_prefixes)
        ]

    try:
        if isinstance(program, list):
            # ── Multi-program export ──
            programs_data: list[dict] = []
            for prog in program:
                _sc, agg = build_aggregate_counts(
                    params["year"],
                    params["semester"],
                    program=prog,
                    section=params["section"],
                    resolve_electives=True,
                )
                pr_caps = load_programme_capacities(prog, list(agg.keys()))
                plan = compute_section_plan(
                    agg,
                    **capacity_kwargs,
                    programme_capacities=pr_caps,
                )
                plan = _apply_programme_course_names(plan, prog)
                plan = _filter_plan(plan)
                summary = compute_plan_summary(plan)
                programs_data.append(
                    {
                        "program": prog,
                        "plan": plan,
                        "summary": summary,
                    }
                )
            path = _export_section_plan_xlsx(
                params=params,
                mode="multi",
                programs_data=programs_data,
            )
            filename = f"section_plan_{params['year']}_{params['semester']}_multi.xlsx"

        elif isinstance(program, str):
            # ── Single-program export ──
            _sc, aggregate = build_aggregate_counts(
                params["year"],
                params["semester"],
                program=program,
                section=params["section"],
                resolve_electives=True,
            )
            pr_caps = load_programme_capacities(program, list(aggregate.keys()))
            plan = compute_section_plan(
                aggregate,
                **capacity_kwargs,
                programme_capacities=pr_caps,
            )
            plan = _apply_programme_course_names(plan, program)
            plan = _filter_plan(plan)
            summary = compute_plan_summary(plan)
            path = _export_section_plan_xlsx(plan, summary, params, mode="single")
            filename = f"section_plan_{params['year']}_{params['semester']}_{program}.xlsx"

        else:
            # ── Combined export (no program filter) ──
            _sc, aggregate = build_aggregate_counts(
                params["year"],
                params["semester"],
                program=None,
                section=params["section"],
                resolve_electives=True,
            )
            plan = _filter_plan(compute_section_plan(aggregate, **capacity_kwargs))
            summary = compute_plan_summary(plan)
            path = _export_section_plan_xlsx(plan, summary, params, mode="combined")
            filename = f"section_plan_{params['year']}_{params['semester']}.xlsx"

        return FileResponse(
            path.open("rb"),
            as_attachment=True,
            filename=filename,
        )

    except Exception as exc:
        logger.exception("section_plan_export error")
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)


def _export_section_plan_xlsx(
    plan: list[dict] | None = None,
    summary: dict | None = None,
    params: dict | None = None,
    *,
    mode: str = "single",
    programs_data: list[dict] | None = None,
) -> Path:
    """Build a styled XLSX workbook with Sections + Summary sheets.

    For mode="single" or "combined": uses plan/summary directly (one pair of sheets).
    For mode="multi": iterates programs_data, creating per-program sheet pairs.
    """
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import (  # type: ignore[import-untyped]
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    wb = Workbook()

    headers = [
        "#",
        "Department",
        "Course",
        "Name",
        "Credits",
        "External",
        "Students",
        "Sections",
        "Max/Section",
        "Avg/Section",
        "Fill %",
        "Status",
    ]

    # Build course name lookup from Course + ProgrammeRequirement tables
    from core.models import Course as CourseModel

    _course_names: dict[str, str] = {}
    for code, desc in CourseModel.objects.values_list("course_code", "description"):
        _course_names[normalize_code(code)] = desc or ""
    # Also check ElectiveCourse for elective names
    from core.models import ElectiveCourse as EC

    for code, name in EC.objects.values_list("course_code", "course_name"):
        nc = normalize_code(code)
        if nc not in _course_names or not _course_names[nc]:
            _course_names[nc] = name or ""

    # ── Styles ──
    thin = Side(style="thin", color="D5D8DC")
    cell_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0A8E6E", end_color="0A8E6E", fill_type="solid")
    title_fill = PatternFill(start_color="1B2631", end_color="1B2631", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    center = Alignment(horizontal="center", vertical="center")
    left_center = Alignment(horizontal="left", vertical="center")
    alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    full_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    under_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    ext_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    bold = Font(bold=True)
    mono = Font(name="Consolas", bold=True, size=10)
    dept_font = Font(name="Calibri", bold=True, size=10, color="2E4053")
    num_font = Font(name="Consolas", size=10)
    status_full_font = Font(bold=True, color="1E8449")
    status_under_font = Font(bold=True, color="C0392B")

    # Department colour map (hash-based pastel, same idea as exam export)
    _dept_colors: dict[str, PatternFill] = {}

    def _dept_fill(dept: str) -> PatternFill:
        if dept in _dept_colors:
            return _dept_colors[dept]
        import colorsys

        h = 0
        for ch in str(dept):
            h = (h * 31 + ord(ch)) % 360
        r, g, b = colorsys.hls_to_rgb(h / 360.0, 0.94, 0.50)
        hex_c = f"{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}"
        fill = PatternFill("solid", fgColor=hex_c)
        _dept_colors[dept] = fill
        return fill

    col_widths = [5, 12, 14, 30, 8, 9, 10, 10, 12, 12, 9, 12]

    def _write_sections_sheet(ws, plan_data: list[dict]) -> None:
        """Write the sections data rows with styled formatting."""
        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        tc = ws.cell(row=1, column=1, value="Section Planning")
        tc.font = title_font
        tc.fill = title_fill
        tc.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(2, len(headers) + 1):
            ws.cell(row=1, column=c).fill = title_fill

        # Header row
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = cell_border

        # Data rows
        for i, row in enumerate(plan_data, 1):
            r = i + 2
            is_alt = i % 2 == 0
            is_ext = row.get("is_external", False)
            dept = row.get("department", "")

            # Row number
            c = ws.cell(row=r, column=1, value=i)
            c.alignment = center
            c.font = num_font
            c.border = cell_border

            # Department — with colour band
            c = ws.cell(row=r, column=2, value=dept)
            c.font = dept_font
            c.alignment = left_center
            c.border = cell_border
            c.fill = _dept_fill(dept)

            # Course code
            c = ws.cell(row=r, column=3, value=row["course_code"])
            c.font = mono
            c.alignment = left_center
            c.border = cell_border

            # Course name
            course_name = _format_export_course_name(row, _course_names)
            c = ws.cell(row=r, column=4, value=course_name)
            c.font = Font(size=9, color="566573")
            c.alignment = left_center
            c.border = cell_border

            # Credits
            c = ws.cell(row=r, column=5, value=row["credit_hours"])
            c.alignment = center
            c.font = num_font
            c.border = cell_border

            # External
            c = ws.cell(row=r, column=6, value="Yes" if is_ext else "")
            c.alignment = center
            c.border = cell_border
            if is_ext:
                c.font = Font(color="2980B9", bold=True)

            # Students
            c = ws.cell(row=r, column=7, value=row["total_students"])
            c.alignment = center
            c.font = Font(bold=True, size=10)
            c.border = cell_border

            # Sections = CEILING(Students / Max, 1)
            c = ws.cell(row=r, column=8)
            c.value = f"=CEILING(G{r}/I{r},1)"
            c.alignment = center
            c.font = num_font
            c.border = cell_border

            # Max/Section (editable — user can change this)
            c = ws.cell(row=r, column=9, value=row["max_per_section"])
            c.alignment = center
            c.font = num_font
            c.border = cell_border

            # Avg/Section = ROUND(Students / Sections, 0)
            c = ws.cell(row=r, column=10)
            c.value = f"=IF(H{r}>0,ROUND(G{r}/H{r},0),0)"
            c.alignment = center
            c.font = num_font
            c.border = cell_border

            # Fill % = Avg / Max * 100
            c = ws.cell(row=r, column=11)
            c.value = f'=IF(I{r}>0,ROUND(G{r}/(H{r}*I{r})*100,0)&"%","")'
            c.alignment = center
            c.border = cell_border
            # Conditional formatting via static check (formulas recalculate in Excel)
            fill_pct = row.get("fill_percent", 0)
            if fill_pct >= 90:
                c.font = Font(bold=True, color="1E8449")
            elif fill_pct < 50:
                c.font = Font(bold=True, color="C0392B")
            else:
                c.font = num_font

            # Status
            status = row.get("status", "")
            c = ws.cell(row=r, column=12, value=status.title() if status else "")
            c.alignment = center
            c.border = cell_border
            if status == "full":
                c.fill = full_fill
                c.font = status_full_font
            elif status == "underfilled":
                c.fill = under_fill
                c.font = status_under_font

            # Row fill: external gets blue tint, alternating gets grey
            row_fill = ext_fill if is_ext else (alt_fill if is_alt else None)
            if row_fill:
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=r, column=col)
                    # Don't override dept colour (col 2) or status colour (col 12)
                    if (
                        col not in (2, 12)
                        and not cell.fill.fgColor
                        or cell.fill.fgColor.rgb == "00000000"
                    ):
                        cell.fill = row_fill

        # Column widths
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Freeze panes below headers
        ws.freeze_panes = "A3"

    from openpyxl.styles import Border, Side  # type: ignore[import-untyped]

    thin_border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )
    title_fill = PatternFill(start_color="1B2631", end_color="1B2631", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=12)
    alt_fill = PatternFill(start_color="F4F6F7", end_color="F4F6F7", fill_type="solid")

    def _write_summary_sheet(
        ws, summary_data: dict, p: dict, sec_sheet: str = "Sections", data_rows: int = 0
    ) -> None:
        """Write summary with formulas referencing the Sections sheet."""
        # Title row
        ws.merge_cells("A1:F1")
        title_val = f"Section Plan Summary — {p.get('year', '')}/{p.get('semester', '')}"
        dept_filter = p.get("dept_filter", [])
        if dept_filter:
            title_val += f" (filtered: {', '.join(dept_filter)})"
        tc = ws.cell(row=1, column=1, value=title_val)
        tc.font = title_font
        tc.fill = title_fill
        tc.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(2, 7):
            ws.cell(row=1, column=c).fill = title_fill

        # Sections sheet reference (quote if has spaces/hyphens)
        sq = f"'{sec_sheet}'" if "-" in sec_sheet or " " in sec_sheet else sec_sheet
        # Data range in sections sheet: row 3 to 2+data_rows
        last_row = 2 + max(data_rows, 1)

        # KPI row — formulas referencing Sections sheet
        ws.cell(row=3, column=1, value="Total Courses").font = bold
        ws.cell(row=3, column=2).value = f"=COUNTA({sq}!C3:C{last_row})"
        ws.cell(row=3, column=2).font = Font(bold=True, size=11, color="0A8E6E")

        ws.cell(row=3, column=3, value="Total Sections").font = bold
        ws.cell(row=3, column=4).value = f"=SUM({sq}!H3:H{last_row})"
        ws.cell(row=3, column=4).font = Font(bold=True, size=11, color="0A8E6E")

        ws.cell(row=3, column=5, value="Total Students").font = bold
        ws.cell(row=3, column=6).value = f"=SUM({sq}!G3:G{last_row})"
        ws.cell(row=3, column=6).font = Font(bold=True, size=11, color="0A8E6E")

        # Department Summary table
        ws.cell(row=5, column=1, value="Department Summary").font = Font(bold=True, size=11)
        dept_headers = ["Department", "Courses", "Sections", "Students", "Total Credits"]
        for col_idx, h in enumerate(dept_headers, 1):
            cell = ws.cell(row=6, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        depts = summary_data.get("departments", [])
        for i, dept in enumerate(depts):
            r = 7 + i
            dept_name = dept["department"]

            # Department name with colour
            c = ws.cell(row=r, column=1, value=dept_name)
            c.font = bold
            c.border = thin_border
            c.fill = _dept_fill(dept_name)

            # Courses = COUNTIF(Sections!B:B, dept_name)
            c = ws.cell(row=r, column=2)
            c.value = f'=COUNTIF({sq}!B3:B{last_row},"{dept_name}")'
            c.alignment = center
            c.border = thin_border

            # Sections = SUMPRODUCT for matching dept
            c = ws.cell(row=r, column=3)
            c.value = f'=SUMPRODUCT(({sq}!B3:B{last_row}="{dept_name}")*{sq}!H3:H{last_row})'
            c.alignment = center
            c.border = thin_border

            # Students = SUMPRODUCT for matching dept
            c = ws.cell(row=r, column=4)
            c.value = f'=SUMPRODUCT(({sq}!B3:B{last_row}="{dept_name}")*{sq}!G3:G{last_row})'
            c.alignment = center
            c.border = thin_border

            # Total Credits = SUMPRODUCT(sections * credits) for matching dept
            c = ws.cell(row=r, column=5)
            c.value = f'=SUMPRODUCT(({sq}!B3:B{last_row}="{dept_name}")*{sq}!H3:H{last_row}*{sq}!E3:E{last_row})'
            c.alignment = center
            c.border = thin_border

            if i % 2 == 1:
                for col in range(2, 6):
                    ws.cell(row=r, column=col).fill = alt_fill

        # Totals row — SUM of the formula rows above
        if depts:
            tr = 7 + len(depts)
            ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, size=10)
            ws.cell(row=tr, column=1).border = thin_border
            first_dept_row = 7
            last_dept_row = 6 + len(depts)
            for col in range(2, 6):
                c = ws.cell(row=tr, column=col)
                col_letter = chr(64 + col)
                c.value = f"=SUM({col_letter}{first_dept_row}:{col_letter}{last_dept_row})"
                c.font = Font(bold=True, size=10)
                c.alignment = center
                c.border = thin_border

        for col_idx in range(1, 7):
            ws.column_dimensions[chr(64 + col_idx)].width = 18

    if mode == "multi" and programs_data:
        # ── Combined sheet first (all programs merged) ──
        combined_plan = _merge_section_plan_rows_by_course_identity(
            [(prog_entry["program"], prog_entry["plan"]) for prog_entry in programs_data]
        )
        combined_summary = compute_plan_summary(combined_plan)

        default_sheet = wb.active
        default_sheet.title = "Sections-All"
        _write_sections_sheet(default_sheet, combined_plan)
        sum_all_ws = wb.create_sheet("Summary-All")
        _write_summary_sheet(
            sum_all_ws,
            combined_summary,
            params or {},
            sec_sheet="Sections-All",
            data_rows=len(combined_plan),
        )

        # ── Per-program sheet pairs ──
        for prog_entry in programs_data:
            prog_name = prog_entry["program"]
            sec_name = f"Sections-{prog_name}"
            sec_ws = wb.create_sheet(sec_name)
            _write_sections_sheet(sec_ws, prog_entry["plan"])
            sum_ws = wb.create_sheet(f"Summary-{prog_name}")
            _write_summary_sheet(
                sum_ws,
                prog_entry["summary"],
                params or {},
                sec_sheet=sec_name,
                data_rows=len(prog_entry["plan"]),
            )
    else:
        # Single or combined — keep existing behaviour exactly
        ws = wb.active
        ws.title = "Sections"
        _write_sections_sheet(ws, plan or [])
        ws2 = wb.create_sheet("Summary")
        _write_summary_sheet(
            ws2,
            summary or {},
            params or {},
            sec_sheet="Sections",
            data_rows=len(plan or []),
        )

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)
