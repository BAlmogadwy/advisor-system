"""Seed the female-cohort expected timetable from the two optimiser workbooks.

Dry-run by default, like every destructive command here. Nothing is written
without `--apply`, and the run that reports and the run that writes compute the
same plan from the same code.

    python manage.py import_f_sections \
        --timetable "F_Course_Sections_Timetable_1448_S1.xlsx" \
        --registration "F_Registration_Optimised_1448_T1.xlsx" \
        --year 1448 --term 1
"""

from __future__ import annotations

import hashlib
import json
import pathlib
from typing import Any

from django.core.management.base import BaseCommand, CommandError

from core.services.f_section_import import (
    ROSTER_SHEET,
    SCHEDULE_SHEET,
    apply_plan,
    build_plan,
    check_students,
    check_time_disagreements,
)


def _rows(path: pathlib.Path, sheet: str) -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise CommandError(f"{path.name} has no sheet {sheet!r} (has {wb.sheetnames})")
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    # Headers are bilingual and newline-separated ("الوحدات\nUnits"); the English
    # half is the stable key.
    header = [str(h).split("\n")[-1].strip() if h is not None else "" for h in next(it)]
    out = [dict(zip(header, row, strict=False)) for row in it if any(c is not None for c in row)]
    wb.close()
    return out


class Command(BaseCommand):
    help = "Seed the female-cohort EXPECTED timetable from the optimiser workbooks."

    def add_arguments(self, parser: Any) -> None:
        parser.add_argument("--timetable", required=True, help="F_Course_Sections_Timetable xlsx")
        parser.add_argument("--registration", required=True, help="F_Registration_Optimised xlsx")
        parser.add_argument("--year", required=True)
        parser.add_argument("--term", required=True, choices=["1", "2", "3"])
        parser.add_argument(
            "--source",
            default="",
            help="StudentTermSection.source; defaults to registration_plan_<year>_t<term>.",
        )
        parser.add_argument("--apply", action="store_true", help="Actually write.")
        parser.add_argument("--report", default="", help="Write the whole plan as JSON.")

    def handle(self, *args: Any, **options: Any) -> None:
        timetable = pathlib.Path(options["timetable"])
        registration = pathlib.Path(options["registration"])
        for path in (timetable, registration):
            if not path.exists():
                raise CommandError(f"{path} does not exist")

        year, term = str(options["year"]), str(options["term"])
        source = options["source"] or f"registration_plan_{year}_t{term}"

        schedule_rows = _rows(timetable, SCHEDULE_SHEET)
        roster_rows = _rows(registration, ROSTER_SHEET)
        plan = build_plan(schedule_rows, roster_rows)

        for path in (timetable, registration):
            digest = hashlib.sha256(path.read_bytes()).hexdigest()
            self.stdout.write(f"{path.name}  sha256 {digest[:16]}…")
        self.stdout.write(f"term {year}/{term}  source {source}")
        self.stdout.write(plan.summary())

        missing, wrong_cohort = check_students(plan)
        if missing:
            self.stdout.write(
                self.style.ERROR(f"\n{len(missing)} roster student(s) absent from Student:")
            )
            self.stdout.write(f"  {missing[:20]}")
        if wrong_cohort:
            self.stdout.write(
                self.style.ERROR(
                    f"\n{len(wrong_cohort)} roster student(s) are NOT cohort F — seating them "
                    "in female sections is refused:"
                )
            )
            self.stdout.write(f"  {wrong_cohort[:20]}")

        # Reported by BOTH the dry run and the apply, from the same function, so
        # what an operator reads is what the write will do.
        disagreements = check_time_disagreements(plan)
        if disagreements:
            self.stdout.write(
                self.style.WARNING(
                    f"\n{len(disagreements)} section(s) whose stored times differ from the "
                    "workbook. Times are NOT changed by this command:"
                )
            )
            for item in disagreements[:20]:
                self.stdout.write(f"  {item['section']}")
                self.stdout.write(f"    stored  : {item['stored']}")
                self.stdout.write(f"    workbook: {item['workbook']}")

        if plan.notices:
            self.stdout.write(self.style.WARNING(f"\n{len(plan.notices)} notice(s):"))
            for notice in plan.notices[:20]:
                self.stdout.write(f"  {notice}")

        if options["report"]:
            report_path = pathlib.Path(options["report"])
            report_path.parent.mkdir(parents=True, exist_ok=True)
            report_path.write_text(
                json.dumps(
                    {
                        "academic_year": year,
                        "term": term,
                        "source": source,
                        "summary": plan.summary(),
                        "sections": {
                            key: {
                                "course_key": s.course_key,
                                "section": s.label,
                                "credits": s.credits,
                                "meetings": s.meetings,
                            }
                            for key, s in sorted(plan.sections.items())
                        },
                        "links": [[sid, key] for sid, key in plan.links],
                        "time_disagreements": disagreements,
                        "notices": [str(n) for n in plan.notices],
                        "problems": [str(p) for p in plan.problems],
                        "students_missing": missing,
                        "students_wrong_cohort": wrong_cohort,
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            self.stdout.write(f"\nPlan written to {report_path}")

        if not plan.ok:
            self.stdout.write(
                self.style.ERROR(f"\n{len(plan.problems)} problem(s); nothing written.")
            )
            for problem in plan.problems[:40]:
                self.stdout.write(f"  {problem}")
            raise CommandError("the workbooks failed validation")

        if not options["apply"]:
            self.stdout.write(
                self.style.WARNING("\nDRY RUN — nothing written. Re-run with --apply.")
            )
            return

        result = apply_plan(plan, year, term, source)
        self.stdout.write(
            self.style.SUCCESS(
                f"\nWrote {result['written']} links for {result['students']} students; "
                f"{result['sections_created']} new section(s) of {result['sections_total']}; "
                f"{result['meetings_written']} meeting(s) written where there were none; "
                f"replaced {result['removed']} existing row(s) of the same source."
            )
        )
