"""Seed expected next-term timetables from an approved registration-plan workbook.

Dry-run by default, like every destructive command here. The report and the write
compute the SAME plan from the same code, so what the dry run prints is what an
apply would do — including the number of rows it would DELETE, which the first
version never showed.

    python manage.py import_registration_plan plan.xlsx --year 1448 --term 1
    python manage.py import_registration_plan plan.xlsx --year 1448 --term 1 --apply

`--expect-links` and `--expect-students` are drift guards: a workbook that no
longer produces the numbers you checked stops rather than writing something else.

`--accept-moved-times` links a section whose times match nothing on file, where
the course has exactly one section. Off by default: it is how a student ends up
seated in a section matching nothing they were told. Read the disagreement report
first.

`--report` writes the whole plan — including the coverage gaps and the workbook's
own SHA-256 — to a JSON file. THE GAPS ARE NOT PERSISTED ANYWHERE ELSE. Nothing in
the database distinguishes "no section exists for this course" from "not
registered", so without this file the only record of which students have an
incomplete week is a terminal that has since scrolled.
"""

from __future__ import annotations

import hashlib
import json
import pathlib
from typing import Any

from django.core.management.base import BaseCommand, CommandError

from core.services.registration_plan_import import (
    apply_plan,
    build_plan,
    check_students_exist,
)

ROSTERS_SHEET = "Section Rosters"
DETAIL_SHEET = "Student Courses (detail)"

#: Every column this importer reads is positional, so a reordered sheet would be
#: interpreted through positional coincidence rather than rejected. For an approved
#: fixed-format file the headers are part of the contract, and checking them is the
#: difference between "this workbook is not the one we agreed" and seating students
#: from whichever column happened to land in slot 4.
EXPECTED_HEADERS = {
    ROSTERS_SHEET: (
        "course",
        "section",
        "lectures",
        "lab period",
        "cap",
        "students",
        "student_ids",
    ),
    DETAIL_SHEET: (
        "student_id",
        "program",
        "course",
        "kind",
        "section",
        "lectures",
        "lab period",
        "room(s)",
        "instructor",
    ),
}


class Command(BaseCommand):
    help = "Link students to expected next-term sections from a registration-plan workbook."

    def add_arguments(self, parser: Any) -> None:
        parser.add_argument("path")
        parser.add_argument("--year", required=True)
        parser.add_argument("--term", required=True)
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Actually write. Without it this validates and reports only.",
        )
        parser.add_argument(
            "--accept-moved-times",
            action="store_true",
            help=(
                "Link a single-section course whose workbook times match nothing on "
                "file. Off by default — read the disagreement report first."
            ),
        )
        parser.add_argument(
            "--report",
            default=None,
            help="Write the full plan, gaps and workbook hash to this JSON path.",
        )
        parser.add_argument("--expect-links", type=int, default=None)
        parser.add_argument("--expect-students", type=int, default=None)

    def handle(self, *args: Any, **options: Any) -> None:
        import openpyxl

        path = pathlib.Path(options["path"])
        if not path.exists():
            raise CommandError(f"no such file: {path}")
        digest = hashlib.sha256(path.read_bytes()).hexdigest()

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in (ROSTERS_SHEET, DETAIL_SHEET):
            if sheet not in workbook.sheetnames:
                raise CommandError(f"the workbook has no {sheet!r} sheet")

        def rows(name: str) -> list[tuple]:
            all_rows = list(workbook[name].iter_rows(values_only=True))
            if not all_rows:
                raise CommandError(f"{name!r} is empty")
            expected = EXPECTED_HEADERS[name]
            found = tuple(str(h or "").strip().lower() for h in all_rows[0][: len(expected)])
            if found != expected:
                raise CommandError(
                    f"UNEXPECTED_HEADER in {name!r}: expected {expected}, found {found}. "
                    "Every column this importer reads is positional, so a reordered "
                    "sheet must fail rather than be read through positional coincidence."
                )
            return all_rows[1:]

        year, term = str(options["year"]).strip(), str(options["term"]).strip()
        plan = build_plan(
            rows(ROSTERS_SHEET),
            rows(DETAIL_SHEET),
            year,
            term,
            accept_moved_times=options["accept_moved_times"],
        )

        self.stdout.write(f"term {year}/{term}")
        self.stdout.write(f"workbook {path.name}  sha256 {digest[:16]}…")
        self.stdout.write(plan.summary())

        # Reported, never repaired. The owner's decision is link-only: a section
        # whose time moved in the plan keeps its database time, and a student would
        # otherwise be shown a slot that contradicts the plan that seated them.
        if plan.time_disagreements:
            self.stdout.write(
                self.style.WARNING(
                    f"\n{len(plan.time_disagreements)} section(s) differ from the database. "
                    "Times are NOT changed by this command:"
                )
            )
            for item in plan.time_disagreements:
                self.stdout.write(f"  {item['course']} {item['file_section']}")
                self.stdout.write(f"     workbook: {item['file_times']}")
                for section_id, times in item["database_times"].items():
                    self.stdout.write(f"     database #{section_id}: {times}")

        # THE GAP. Reported prominently because the seeded timetable is incomplete
        # by exactly this much, and a student looking at a blank Sunday evening
        # cannot tell "nothing scheduled" from "we do not hold it".
        if plan.uncovered:
            total = sum(len(v) for v in plan.uncovered.values())
            self.stdout.write(
                self.style.WARNING(
                    f"\nGAP — {total} registration(s) cannot be seeded because no section "
                    f"exists for these {len(plan.uncovered)} course(s):"
                )
            )
            for course, entries in sorted(plan.uncovered.items(), key=lambda kv: -len(kv[1])):
                slots = sorted({str(r["times"]) for r in entries})
                self.stdout.write(
                    f"  {course:8} {len(entries):4} registration(s), "
                    f"{len({r['student_id'] for r in entries})} student(s), "
                    f"{len(slots)} distinct slot(s) in the plan"
                )
                for slot in slots[:3]:
                    self.stdout.write(f"       {slot}")
            self.stdout.write(
                "  These students will have an INCOMPLETE week until those sections exist."
            )
            if not options["report"]:
                self.stdout.write(
                    self.style.WARNING(
                        "  Nothing in the database records this. Pass --report to keep it."
                    )
                )

        if options["report"]:
            report_path = pathlib.Path(options["report"])
            report_path.write_text(
                json.dumps(
                    {
                        "workbook": str(path),
                        "sha256": digest,
                        "academic_year": year,
                        "term": term,
                        "accept_moved_times": options["accept_moved_times"],
                        "summary": plan.summary(),
                        "counts": {
                            "links": len(plan.links),
                            "students": len(plan.students),
                            "replaces": plan.replaces,
                            "detail_rows_read": plan.detail_rows_read,
                            "blank_rows": plan.blank_rows,
                            "duplicate_rows": plan.duplicate_rows,
                            "skipped_unplaceable": plan.skipped_unplaceable,
                        },
                        "uncovered": plan.uncovered,
                        "time_disagreements": plan.time_disagreements,
                        "problems": [str(p) for p in plan.problems],
                        "links": plan.links,
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            self.stdout.write(f"\nPlan written to {report_path}")

        if not plan.ok:
            self.stdout.write(
                self.style.ERROR(f"\n{len(plan.problems)} problem(s); nothing written.")
            )
            for problem in plan.problems[:40]:
                self.stdout.write(f"  {problem}")
            raise CommandError("the workbook failed validation")

        missing = check_students_exist(plan)
        if missing:
            # Never invent a Student row from a registration file.
            raise CommandError(
                f"{len(missing)} student(s) in the workbook have no Student record: {missing[:10]}"
            )

        for flag, actual in (
            ("expect_links", len(plan.links)),
            ("expect_students", len(plan.students)),
        ):
            expected = options.get(flag)
            if expected is not None and expected != actual:
                raise CommandError(
                    f"--{flag.replace('_', '-')} says {expected}, the workbook produces {actual}"
                )

        # A warning scrolls away, which is the exact failure the report was added to
        # close. If the import leaves students with an incomplete week, the record of
        # WHICH students is a precondition of writing, not a suggestion. No default
        # path is chosen here: the file carries student identifiers, so the operator
        # names somewhere restricted rather than having one picked for them.
        if options["apply"] and plan.uncovered and not options["report"]:
            raise CommandError(
                f"--report is required when applying a plan with "
                f"{sum(len(v) for v in plan.uncovered.values())} uncovered registration(s); "
                "nothing else records which students are left with an incomplete week"
            )

        if not options["apply"]:
            self.stdout.write(
                f"\nDry run — nothing written. An apply would DELETE {plan.replaces} "
                f"existing row(s) and write {len(plan.links)}. Re-run with --apply."
            )
            return

        result = apply_plan(plan, year, term)
        self.stdout.write(
            self.style.SUCCESS(
                f"\nSeeded: {result['written']} links for {result['students']} students "
                f"({result['removed']} pre-existing rows for those students replaced)."
            )
        )
