from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet

from core.services.advisor_graduation_optimization import OPTIMIZED_CURRENT_OFFERINGS
from core.services.graduation_export import build_graduation_xlsx
from core.services.student_graduation import (
    RECOMMENDED_CURRENT_TERM,
    REGISTERED_TIMETABLE,
)

SHEET_NAMES = [
    "Overview",
    "Term Plan",
    "Baseline Courses",
    "Prerequisite Map",
    "Blockers & Assumptions",
]


def _student(*, name: str = "Portfolio Student") -> SimpleNamespace:
    return SimpleNamespace(
        student_id=4_801_234,
        registration_no="4801234",
        name=name,
        program="AI",
        section="M",
        advisor_id="ADV-PORTFOLIO",
    )


def _report(baseline_kind: str) -> dict[str, Any]:
    return {
        "program": "AI",
        "plan_courses_total": 10,
        "plan_courses_passed": 5,
        "percent_courses": 50,
        "remaining_courses": 5,
        "remaining_credits": 15,
        "passed_credits_in_plan": 45,
        "earned_credits_registrar": 60,
        "planning_baseline_academic_year": 1448,
        "planning_baseline_term": 2,
        "planning_baseline_kind": baseline_kind,
        "planning_baseline_credits": 3,
        "planning_baseline_courses_assumed_passed": [
            {
                "code": "AI201",
                "name": "Introduction to AI",
                "credits": 3,
                "requirement_type": "Core",
            }
        ],
        "chain_floor_terms": 2,
        "capacity_floor_terms_after_planning_baseline": 1,
        "lower_bound_additional_terms": 2,
        "lower_bound_terms_including_planning_baseline": 3,
        "estimated_additional_terms": None,
        "estimated_terms_including_planning_baseline": None,
        "simulation_completed": False,
        "simulated_terms_examined": 2,
        "productive_terms_planned": 1,
        "max_credits_per_term": 18,
        "term_plan": [
            {
                "sequence": 1,
                "academic_year": 1449,
                "term": 1,
                "waiting_term": True,
                "credits": 0,
                "course_codes": [],
                "courses": [],
            },
            {
                "sequence": 2,
                "academic_year": 1449,
                "term": 2,
                "waiting_term": False,
                "credits": 6,
                "course_codes": ["AI301", "AI302"],
                "courses": [
                    {
                        "code": "AI301",
                        "name": "Machine Learning",
                        "credits": 3,
                        "requirement_type": "Core",
                    },
                    {
                        "code": "AI302",
                        "name": "Natural Language Processing",
                        "credits": 3,
                        "requirement_type": "Core",
                    },
                ],
            },
        ],
        "unresolved_requirements": [
            {
                "code": "AI499",
                "name": "Graduation Project",
                "credits": 3,
                "requirement_type": "Core",
                "missing_course_prerequisites": ["AI398"],
                "missing_prerequisites_outside_plan": ["AI398"],
                "credit_hour_gate": {
                    "required": 100,
                    "effective_in_scenario": 90,
                    "remaining": 10,
                },
            }
        ],
        "hour_gates": [
            {
                "code": "AI499",
                "name": "Graduation Project",
                "required": 100,
                "effective": 90,
                "remaining": 10,
            }
        ],
        "simulation_assumptions": [
            "All planning-baseline and simulated courses are passed on the first attempt.",
            "Every simulated main term uses a maximum of 18 credits.",
            "The scenario is read-only and does not update the student record.",
        ],
    }


def _presentation(baseline_kind: str) -> dict[str, Any]:
    return {
        "kind": "graduation_scenario",
        "planning_baseline_kind": baseline_kind,
        "band_labels": {
            "1": {
                REGISTERED_TIMETABLE: "Registered timetable 1448/2",
                RECOMMENDED_CURRENT_TERM: "Recommended starting term 1448/2",
                OPTIMIZED_CURRENT_OFFERINGS: "Optimized current offerings 1448/2",
            }[baseline_kind],
            "2": "Projected 1449/1",
            "3": "Projected 1449/2",
        },
        "graph": {
            "items": [
                {
                    "course_code": "AI301",
                    "prerequisite_course_code": "AI201",
                },
                {
                    "course_code": "AI499",
                    "prerequisite_course_code": "AI301",
                },
            ],
            "termOf": {"AI201": 1, "AI301": 3, "AI499": 4, "IS101": 2},
            "nameOf": {
                "AI201": "Introduction to AI",
                "AI301": "Machine Learning",
                "AI499": "Graduation Project",
                "IS101": "Islamic Studies",
            },
            "statusOf": {
                "AI201": ("studying" if baseline_kind == REGISTERED_TIMETABLE else "open"),
                "AI301": "open",
                "AI499": "locked",
                "IS101": "open",
            },
            # IS101 is deliberately isolated. It must not disappear from a tabular
            # representation merely because it has no prerequisite edge.
            "extraNodes": ["AI201", "AI301", "AI499", "IS101"],
        },
    }


def _workbook(
    baseline_kind: str = REGISTERED_TIMETABLE,
    *,
    student: SimpleNamespace | None = None,
    report: dict[str, Any] | None = None,
    presentation: dict[str, Any] | None = None,
):
    content = build_graduation_xlsx(
        student=student or _student(),
        academic_year=1448,
        term=2,
        baseline_kind=baseline_kind,
        report=report if report is not None else _report(baseline_kind),
        presentation=(presentation if presentation is not None else _presentation(baseline_kind)),
        language_code="en",
    )
    assert isinstance(content, bytes)
    assert content.startswith(b"PK")
    return load_workbook(BytesIO(content))


def _sheet_text(sheet: Worksheet) -> list[str]:
    return [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None]


def _all_cells(workbook) -> list[Cell]:
    return [cell for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row]


def _row_containing(sheet: Worksheet, value: object) -> tuple[object, ...]:
    for row in sheet.iter_rows(values_only=True):
        if value in row:
            return row
    raise AssertionError(f"{value!r} was not exported on {sheet.title!r}")


def _header_row(sheet: Worksheet, required: set[str]) -> tuple[int, list[Cell]]:
    for row_number, row in enumerate(sheet.iter_rows(), start=1):
        values = {str(cell.value) for cell in row if cell.value is not None}
        if required <= values:
            return row_number, list(row)
    raise AssertionError(f"Missing headers {sorted(required)!r} on {sheet.title!r}")


def _cell_after_label(sheet: Worksheet, label: str) -> Cell:
    """Return the next populated cell on the row containing ``label``."""
    for row in sheet.iter_rows():
        for index, cell in enumerate(row):
            if cell.value != label:
                continue
            for candidate in row[index + 1 :]:
                if candidate.value is not None:
                    return candidate
            raise AssertionError(f"{label!r} has no value on {sheet.title!r}")
    raise AssertionError(f"{label!r} was not exported on {sheet.title!r}")


@pytest.mark.parametrize(
    ("baseline_kind", "truthful_label", "forbidden_label"),
    [
        (
            REGISTERED_TIMETABLE,
            "Registered timetable",
            "Recommended starting-term courses",
        ),
        (
            RECOMMENDED_CURRENT_TERM,
            "Recommended starting-term courses",
            "Registered timetable",
        ),
        (
            OPTIMIZED_CURRENT_OFFERINGS,
            "Optimized current offerings",
            "Registered timetable",
        ),
    ],
)
def test_export_has_five_complete_sheets_and_truthful_baseline_labels(
    baseline_kind: str,
    truthful_label: str,
    forbidden_label: str,
) -> None:
    workbook = _workbook(baseline_kind)
    try:
        assert workbook.sheetnames == SHEET_NAMES
        assert workbook.properties.creator == "Advisor Portal"

        overview_text = _sheet_text(workbook["Overview"])
        baseline_text = _sheet_text(workbook["Baseline Courses"])
        assert truthful_label in overview_text
        assert truthful_label in baseline_text
        assert forbidden_label not in overview_text
        assert forbidden_label not in baseline_text

        overview = workbook["Overview"]
        assert _cell_after_label(overview, "Student ID").value == "4801234"
        assert _cell_after_label(overview, "Student name").value == "Portfolio Student"
        assert _cell_after_label(overview, "Program").value == "AI"
        assert _cell_after_label(overview, "Planning term").value == "1448/2"
        assert _cell_after_label(overview, "Plan courses completed").value == 5
        assert _cell_after_label(overview, "Plan courses total").value == 10
        assert _cell_after_label(overview, "Courses remaining").value == 5
        assert _cell_after_label(overview, "Credits remaining").value == 15
        assert _cell_after_label(overview, "Maximum credits per term").value == 18
        assert _cell_after_label(overview, "Starting-course source").value == truthful_label

        baseline_row = _row_containing(workbook["Baseline Courses"], "AI201")
        assert "Introduction to AI" in baseline_row
        assert 3 in baseline_row
    finally:
        workbook.close()


def test_optimized_export_separates_real_mapped_elective_from_plan_placeholder() -> None:
    report = _report(OPTIMIZED_CURRENT_OFFERINGS)
    report["planning_baseline_courses_assumed_passed"] = [
        {
            "code": "IS1",
            "name": "Information Retrieval",
            "credits": 3,
            "offered_course_code": "IS481",
            "offered_course_name": "Information Retrieval",
            "fulfills_plan_code": "IS1",
            "recorded_sections": ["M2", "M3"],
            "elective_slot": True,
        }
    ]
    workbook = _workbook(OPTIMIZED_CURRENT_OFFERINGS, report=report)
    try:
        baseline_row = _row_containing(workbook["Baseline Courses"], "IS481")
        assert "IS1" in baseline_row
        assert "M2, M3" in baseline_row
        assert "Optimized current offerings" in baseline_row
        assert "Registered timetable" not in baseline_row

        term_row = _row_containing(workbook["Term Plan"], "IS481")
        assert "Optimized starting term" in term_row
        assert "Information Retrieval (fulfills IS1)" in term_row
        assert "Registered" not in term_row
    finally:
        workbook.close()


def test_term_plan_preserves_waiting_terms_and_uses_typed_values() -> None:
    workbook = _workbook()
    try:
        sheet = workbook["Term Plan"]
        _header_row(
            sheet,
            {
                "Step",
                "Academic Year",
                "Term",
                "State",
                "Course Code",
                "Course Name",
                "Credits",
                "Requirement Type",
            },
        )

        waiting = _row_containing(sheet, "Waiting for prerequisites")
        assert 1 in waiting
        assert 1449 in waiting
        assert 0 in waiting
        assert "AI301" not in waiting

        ai301 = _row_containing(sheet, "AI301")
        ai302 = _row_containing(sheet, "AI302")
        for row, course_name in (
            (ai301, "Machine Learning"),
            (ai302, "Natural Language Processing"),
        ):
            assert 2 in row
            assert 1449 in row
            assert course_name in row
            assert 3 in row
            assert "Core" in row

        for value in (1, 2, 1449, 0, 3):
            matching = [
                cell
                for row in sheet.iter_rows()
                for cell in row
                if cell.value == value and not isinstance(cell.value, bool)
            ]
            assert matching
            assert all(cell.data_type == "n" for cell in matching)
    finally:
        workbook.close()


def test_prerequisite_map_exports_edges_and_isolated_courses() -> None:
    workbook = _workbook()
    try:
        sheet = workbook["Prerequisite Map"]
        values = list(sheet.iter_rows(values_only=True))

        ai301_rows = [row for row in values if "AI301" in row]
        assert any("AI201" in row and "Introduction to AI" in row for row in ai301_rows)
        assert any("AI499" in row and "Graduation Project" in row for row in ai301_rows)

        isolated = _row_containing(sheet, "IS101")
        assert "Islamic Studies" in isolated
        # An isolated node must be explicit rather than being mistaken for a
        # prerequisite relationship to another course.
        course_codes = {"AI201", "AI301", "AI499"}
        assert not any(code in isolated for code in course_codes)
    finally:
        workbook.close()


def test_blockers_hour_gates_and_assumptions_are_auditable() -> None:
    workbook = _workbook()
    try:
        sheet = workbook["Blockers & Assumptions"]
        blocker = _row_containing(sheet, "AI499")
        assert "Graduation Project" in blocker
        assert any("AI398" in str(value) for value in blocker if value is not None)
        assert {100, 90, 10} <= {value for value in blocker if isinstance(value, int)}

        text = _sheet_text(sheet)
        assert any("passed on the first attempt" in value for value in text)
        assert any("maximum of 18 credits" in value for value in text)
        assert any("read-only" in value for value in text)
    finally:
        workbook.close()


def test_workbook_has_professional_screen_and_print_formatting() -> None:
    workbook = _workbook()
    try:
        assert workbook.active.title == "Overview"
        for sheet in workbook.worksheets:
            assert sheet.sheet_view.showGridLines is False
            assert any("A1" in merged for merged in map(str, sheet.merged_cells.ranges))

            title = sheet["A1"]
            assert title.value
            assert title.font.bold is True
            assert float(title.font.sz or 0) >= 14
            assert title.fill.fill_type == "solid"
            assert sheet.row_dimensions[1].height is not None
            assert sheet.row_dimensions[1].height >= 24

            assert sheet.page_setup.fitToWidth == 1
            assert sheet.sheet_properties.pageSetUpPr.fitToPage is True
            assert sheet.print_area

        for sheet_name, required_headers in {
            "Term Plan": {"Step", "Academic Year", "Course Code", "Course Name"},
            "Baseline Courses": {"Course Code", "Course Name", "Credits"},
            "Prerequisite Map": {"Course Code", "Course Name"},
        }.items():
            sheet = workbook[sheet_name]
            header_number, header_cells = _header_row(sheet, required_headers)
            assert sheet.freeze_panes is not None
            assert sheet.auto_filter.ref
            assert all(
                cell.font.bold and cell.fill.fill_type == "solid"
                for cell in header_cells
                if cell.value is not None
            )
            assert all(
                (sheet.column_dimensions[cell.column_letter].width or 0) >= 8
                for cell in header_cells
                if cell.value is not None and cell.column_letter != "A"
            )
            freeze_row, _freeze_column = coordinate_to_tuple(str(sheet.freeze_panes))
            assert freeze_row > header_number

        term_plan = workbook["Term Plan"]
        assert term_plan.page_setup.orientation == "landscape"
        name_cells = [cell for cell in _all_cells(workbook) if cell.value == "Machine Learning"]
        assert name_cells
        assert name_cells[0].alignment.wrap_text is True
    finally:
        workbook.close()


def test_identifiers_are_text_and_formula_prefixes_are_exported_as_literal_text() -> None:
    report = _report(REGISTERED_TIMETABLE)
    report["planning_baseline_courses_assumed_passed"][0]["code"] = "+AI201"
    report["planning_baseline_courses_assumed_passed"][0]["name"] = "+SUM(1,1)"
    report["term_plan"][1]["courses"][0]["code"] = "-AI301"
    report["term_plan"][1]["courses"][0]["name"] = "-1+1"
    report["unresolved_requirements"][0]["code"] = "@AI499"
    report["unresolved_requirements"][0]["name"] = "@SUM(1,1)"
    report["simulation_assumptions"][0] = '=HYPERLINK("https://example.invalid")'

    student = _student(name='=WEBSERVICE("https://example.invalid")')
    student.registration_no = "=1+1"
    workbook = _workbook(
        student=student,
        report=report,
    )
    try:
        exported = {cell.value: cell for cell in _all_cells(workbook) if cell.value is not None}
        for literal in (
            "4801234",
            "'=1+1",
            '\'=WEBSERVICE("https://example.invalid")',
            "'+AI201",
            "'+SUM(1,1)",
            "'-AI301",
            "'-1+1",
            "'@AI499",
            "'@SUM(1,1)",
            '\'=HYPERLINK("https://example.invalid")',
        ):
            assert literal in exported
            assert exported[literal].data_type == "s"
    finally:
        workbook.close()


def test_trusted_reconciliation_and_planning_formulas_are_retained() -> None:
    workbook = _workbook()
    try:
        overview = workbook["Overview"]
        assert _cell_after_label(overview, "Completion %").data_type == "f"
        assert _cell_after_label(overview, "Plan-course reconciliation").data_type == "f"

        term_plan = workbook["Term Plan"]
        academic_term_header, academic_term_column = next(
            (row_number, cell.column)
            for row_number, row in enumerate(term_plan.iter_rows(), start=1)
            for cell in row
            if cell.value == "Academic Term"
        )
        load_header, load_column = next(
            (row_number, cell.column)
            for row_number, row in enumerate(term_plan.iter_rows(), start=1)
            for cell in row
            if cell.value == "Load %"
        )
        assert academic_term_header == load_header
        assert term_plan.cell(academic_term_header + 1, academic_term_column).data_type == "f"
        assert term_plan.cell(load_header + 1, load_column).data_type == "f"

        prereq = workbook["Prerequisite Map"]
        term_gap_header, term_gap_column = next(
            (row_number, cell.column)
            for row_number, row in enumerate(prereq.iter_rows(), start=1)
            for cell in row
            if cell.value == "Term Gap"
        )
        order_header, order_column = next(
            (row_number, cell.column)
            for row_number, row in enumerate(prereq.iter_rows(), start=1)
            for cell in row
            if cell.value == "Order Check"
        )
        assert term_gap_header == order_header
        assert prereq.cell(term_gap_header + 1, term_gap_column).data_type == "f"
        assert prereq.cell(order_header + 1, order_column).data_type == "f"
    finally:
        workbook.close()


def test_empty_sections_still_produce_valid_explanatory_sheets() -> None:
    report = _report(REGISTERED_TIMETABLE)
    report.update(
        {
            "planning_baseline_courses_assumed_passed": [],
            "term_plan": [],
            "unresolved_requirements": [],
            "hour_gates": [],
            "simulation_assumptions": [],
            "simulation_completed": True,
            "estimated_additional_terms": 0,
            "estimated_terms_including_planning_baseline": 0,
        }
    )
    presentation = _presentation(REGISTERED_TIMETABLE)
    presentation["graph"] = {
        "items": [],
        "termOf": {},
        "nameOf": {},
        "statusOf": {},
        "extraNodes": [],
    }

    workbook = _workbook(report=report, presentation=presentation)
    try:
        assert workbook.sheetnames == SHEET_NAMES
        for sheet_name in SHEET_NAMES[1:]:
            text = _sheet_text(workbook[sheet_name])
            assert text, f"{sheet_name} must not be a blank worksheet"
            assert any(value.strip().lower().startswith(("no ", "none")) for value in text), (
                f"{sheet_name} must explain that its section has no records"
            )
    finally:
        workbook.close()
