from __future__ import annotations

from io import BytesIO
from unittest.mock import Mock

import pytest
from django.contrib.auth.models import Group, User
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook

from core.models import Student
from core.services.advisor_graduation_optimization import OPTIMIZED_CURRENT_OFFERINGS
from core.services.rbac import ROLE_ADVISOR, ROLE_STUDENT, set_user_scope
from core.services.student_graduation import (
    RECOMMENDED_CURRENT_TERM,
    REGISTERED_TIMETABLE,
)

pytestmark = pytest.mark.django_db

STUDENT_ID = 4_801_234
ADVISOR_ID = "ADV-PORTFOLIO"


@pytest.fixture
def portfolio_student() -> Student:
    return Student.objects.create(
        student_id=STUDENT_ID,
        registration_no=str(STUDENT_ID),
        name="Portfolio Student",
        program="AI",
        section="M",
        advisor_id=ADVISOR_ID,
        status="ACTIVE",
    )


def _client(
    username: str,
    *,
    role: str = ROLE_ADVISOR,
    advisor_id: str = ADVISOR_ID,
    student_id: int | None = None,
) -> Client:
    user = User.objects.create_user(username=username, password="x")
    user.groups.add(Group.objects.get_or_create(name=role)[0])
    set_user_scope(user.id, advisor_id=advisor_id, student_id=student_id)
    client = Client()
    client.force_login(user)
    return client


def _url() -> str:
    return reverse(
        "advisor_portfolio_student_graduation",
        kwargs={"student_id": STUDENT_ID},
    )


def _page_url() -> str:
    return reverse(
        "advisor_portfolio_student_graduation_page",
        kwargs={"student_id": STUDENT_ID},
    )


def _export_url() -> str:
    return reverse(
        "advisor_portfolio_student_graduation_export_xlsx",
        kwargs={"student_id": STUDENT_ID},
    )


def _report(baseline_kind: str) -> dict:
    return {
        "program": "AI",
        "remaining_courses": 2,
        "remaining_credits": 6,
        "estimated_additional_terms": 2,
        "estimated_terms_including_planning_baseline": 3,
        "planning_baseline_kind": baseline_kind,
        "planning_baseline_academic_year": 1448,
        "planning_baseline_term": 2,
        "planning_baseline_credits": 3,
        "planning_baseline_courses_assumed_passed": [
            {"code": "AI201", "name": "Baseline", "credits": 3}
        ],
        "term_plan": [
            {
                "sequence": 1,
                "academic_year": 1449,
                "term": 1,
                "waiting_term": True,
                "credits": 0,
                "course_codes": [],
                "courses": [],
            },
            {
                "sequence": 2,
                "academic_year": 1449,
                "term": 2,
                "waiting_term": False,
                "credits": 6,
                "course_codes": ["AI301", "AI302"],
                "courses": [
                    {"code": "AI301", "name": "One", "credits": 3},
                    {"code": "AI302", "name": "Two", "credits": 3},
                ],
            },
        ],
        "unresolved_requirements": [
            {"code": "AI499", "name": "Project", "missing_course_prerequisites": ["AI398"]}
        ],
        "hour_gates": [{"code": "AI499", "required": 100, "effective": 90, "remaining": 10}],
        "scenario_graph": {"items": [], "statusOf": {}, "nameOf": {}},
    }


def _install_successful_services(monkeypatch, baseline_kind: str = REGISTERED_TIMETABLE):
    build = Mock(return_value=_report(baseline_kind))
    presentation = Mock(return_value={"graph": {"extraNodes": ["AI201", "AI301"]}})
    monkeypatch.setattr("core.portfolio_views.build_graduation_report", build)
    monkeypatch.setattr(
        "core.portfolio_views.graduation_presentation_from_tool_results", presentation
    )
    monkeypatch.setattr(
        "core.portfolio_views.load_defaults",
        lambda: {
            "currentYear": "1448",
            "currentTerm": "2",
            # These legacy values deliberately disagree. The endpoint must not use them.
            "academic_year": "9999",
            "term": "3",
        },
    )
    return build, presentation


def test_endpoint_requires_advisor_role_and_student_scope(portfolio_student, monkeypatch):
    build, _presentation = _install_successful_services(monkeypatch)

    assert Client().get(_url()).status_code == 401

    student_client = _client(
        "student-role",
        role=ROLE_STUDENT,
        advisor_id="",
        student_id=STUDENT_ID,
    )
    assert student_client.get(_url()).status_code == 403

    other_advisor = _client("other-advisor", advisor_id="SOMEONE-ELSE")
    assert other_advisor.get(_url()).status_code == 403
    build.assert_not_called()

    response = _client("assigned-advisor").get(_url())
    assert response.status_code == 200
    assert response.json()["student_id"] == STUDENT_ID
    assert build.call_count == 1


def test_dedicated_page_requires_the_same_advisor_scope(portfolio_student, monkeypatch):
    build, _presentation = _install_successful_services(monkeypatch)

    assert Client().get(_page_url()).status_code == 401
    assert (
        _client("page-other-advisor", advisor_id="SOMEONE-ELSE").get(_page_url()).status_code == 403
    )
    build.assert_not_called()

    response = _client("page-assigned-advisor").get(_page_url())
    assert response.status_code == 200
    assert response.context["student_id"] == STUDENT_ID
    assert response.context["baseline_kind"] == REGISTERED_TIMETABLE
    assert response.context["grad"] == _report(REGISTERED_TIMETABLE)
    assert response.templates[0].name == "core/advisor_student_graduation.html"


def test_dedicated_page_switches_to_recommended_without_an_embedded_drawer(
    portfolio_student, monkeypatch
):
    build, _presentation = _install_successful_services(monkeypatch, RECOMMENDED_CURRENT_TERM)

    response = _client("page-recommended").get(_page_url(), {"baseline": RECOMMENDED_CURRENT_TERM})

    assert response.status_code == 200
    body = response.content.decode()
    assert 'id="advisorGraduationResults"' in body
    assert 'id="agRecommendedTab"' in body
    assert 'id="apGraduationPanel"' not in body
    assert "AI301" in body
    build.assert_called_once_with(
        STUDENT_ID,
        1448,
        2,
        planning_baseline_kind=RECOMMENDED_CURRENT_TERM,
    )


def test_dedicated_page_exports_only_the_active_scenario(portfolio_student, monkeypatch):
    _install_successful_services(monkeypatch, RECOMMENDED_CURRENT_TERM)

    response = _client("page-export-link").get(_page_url(), {"baseline": RECOMMENDED_CURRENT_TERM})

    assert response.status_code == 200
    body = response.content.decode()
    assert f"{_export_url()}?baseline={RECOMMENDED_CURRENT_TERM}" in body
    assert "Export this scenario (XLSX)" in body


def test_export_endpoint_requires_advisor_scope_before_building(portfolio_student, monkeypatch):
    build, _presentation = _install_successful_services(monkeypatch)
    export = Mock(return_value=b"PK-test")
    monkeypatch.setattr("core.portfolio_views.build_graduation_xlsx", export)

    assert Client().get(_export_url()).status_code == 401
    assert (
        _client(
            "export-student-role",
            role=ROLE_STUDENT,
            advisor_id="",
            student_id=STUDENT_ID,
        )
        .get(_export_url())
        .status_code
        == 403
    )
    assert (
        _client("export-other-advisor", advisor_id="SOMEONE-ELSE").get(_export_url()).status_code
        == 403
    )
    missing_url = reverse(
        "advisor_portfolio_student_graduation_export_xlsx",
        kwargs={"student_id": STUDENT_ID + 1},
    )
    assert _client("export-missing-student").get(missing_url).status_code == 404
    build.assert_not_called()
    export.assert_not_called()


@pytest.mark.parametrize(
    ("baseline_kind", "filename_part"),
    [
        (REGISTERED_TIMETABLE, "registered"),
        (RECOMMENDED_CURRENT_TERM, "recommended"),
    ],
)
def test_export_endpoint_returns_loadable_selected_scenario_workbook(
    portfolio_student, monkeypatch, baseline_kind, filename_part
):
    build, _presentation = _install_successful_services(monkeypatch, baseline_kind)

    response = _client(f"export-{filename_part}").get(_export_url(), {"baseline": baseline_kind})

    assert response.status_code == 200
    assert response.headers["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["Content-Disposition"] == (
        f'attachment; filename="graduation_plan_{STUDENT_ID}_{filename_part}_1448_T2.xlsx"'
    )
    assert response.headers["X-Content-Type-Options"] == "nosniff"
    assert "no-store" in response.headers["Cache-Control"]
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    try:
        assert workbook.sheetnames == [
            "Overview",
            "Term Plan",
            "Baseline Courses",
            "Prerequisite Map",
            "Blockers & Assumptions",
        ]
    finally:
        workbook.close()
    build.assert_called_once_with(
        STUDENT_ID,
        1448,
        2,
        planning_baseline_kind=baseline_kind,
    )


@pytest.mark.parametrize("invalid_baseline", ["ignore_registration", ""])
def test_export_endpoint_rejects_unknown_baseline(portfolio_student, monkeypatch, invalid_baseline):
    report_builder = Mock()
    exporter = Mock()
    monkeypatch.setattr("core.portfolio_views.build_graduation_report", report_builder)
    monkeypatch.setattr("core.portfolio_views.build_graduation_xlsx", exporter)

    response = _client(f"bad-export-{invalid_baseline or 'blank'}").get(
        _export_url(), {"baseline": invalid_baseline}
    )

    assert response.status_code == 400
    assert response.json()["code"] == "INVALID_GRADUATION_BASELINE"
    report_builder.assert_not_called()
    exporter.assert_not_called()


@pytest.mark.parametrize("invalid_baseline", ["ignore_registration", ""])
def test_endpoint_rejects_unknown_baseline(portfolio_student, monkeypatch, invalid_baseline):
    build = Mock()
    monkeypatch.setattr("core.portfolio_views.build_graduation_report", build)

    response = _client(f"bad-baseline-{invalid_baseline or 'blank'}").get(
        _url(), {"baseline": invalid_baseline}
    )

    assert response.status_code == 400
    assert response.json()["code"] == "INVALID_GRADUATION_BASELINE"
    assert set(response.json()["allowed_baselines"]) == {
        REGISTERED_TIMETABLE,
        RECOMMENDED_CURRENT_TERM,
        OPTIMIZED_CURRENT_OFFERINGS,
    }
    build.assert_not_called()


def test_endpoint_rejects_non_main_global_current_term(portfolio_student, monkeypatch):
    build = Mock()
    monkeypatch.setattr("core.portfolio_views.build_graduation_report", build)
    monkeypatch.setattr(
        "core.portfolio_views.load_defaults",
        lambda: {
            "currentYear": "1448",
            "currentTerm": "3",
            # A valid legacy term must not hide the unsupported global current term.
            "academic_year": "1448",
            "term": "1",
        },
    )

    response = _client("summer-term").get(_url(), {"baseline": RECOMMENDED_CURRENT_TERM})

    assert response.status_code == 400
    assert response.json() == {
        "error": "Graduation planning supports global current terms 1 and 2 only.",
        "code": "UNSUPPORTED_GRADUATION_TERM",
        "academic_year": 1448,
        "term": 3,
        "supported_terms": [1, 2],
    }
    build.assert_not_called()


@pytest.mark.parametrize(
    ("query", "expected_baseline"),
    [
        ({}, REGISTERED_TIMETABLE),
        ({"baseline": REGISTERED_TIMETABLE}, REGISTERED_TIMETABLE),
        ({"baseline": RECOMMENDED_CURRENT_TERM}, RECOMMENDED_CURRENT_TERM),
    ],
)
def test_endpoint_forwards_mode_and_preserves_full_report(
    portfolio_student,
    monkeypatch,
    query,
    expected_baseline,
):
    build, presentation = _install_successful_services(monkeypatch, expected_baseline)

    response = _client(f"mode-{expected_baseline}-{len(query)}").get(_url(), query)

    assert response.status_code == 200
    assert "no-store" in response.headers["Cache-Control"]
    payload = response.json()
    assert payload["academic_year"] == 1448
    assert payload["term"] == 2
    assert payload["baseline"] == expected_baseline
    assert payload["report"] == _report(expected_baseline)
    assert payload["report"]["term_plan"][0]["waiting_term"] is True
    assert payload["presentation"] == {"graph": {"extraNodes": ["AI201", "AI301"]}}
    build.assert_called_once_with(
        STUDENT_ID,
        1448,
        2,
        planning_baseline_kind=expected_baseline,
    )
    tool_result = presentation.call_args.args[0][0]
    assert tool_result["tool"] == "graduation_progress"
    assert tool_result["ok"] is True
    assert tool_result["scenario_academic_year"] == 1448
    assert tool_result["scenario_term"] == 2
    assert tool_result["term_plan"][0]["waiting_term"] is True


def test_advisor_only_optimized_mode_uses_verified_section_snapshot_and_preserves_provenance(
    portfolio_student, monkeypatch
):
    report = _report(OPTIMIZED_CURRENT_OFFERINGS)
    report["planning_baseline"] = {"kind": OPTIMIZED_CURRENT_OFFERINGS}
    report["offering_optimization"] = {
        "candidate_count": 4,
        "evaluated_maximal_subset_count": 2,
    }
    optimized = Mock(return_value=report)
    normal = Mock()
    presentation = Mock(
        return_value={
            "planning_baseline_kind": OPTIMIZED_CURRENT_OFFERINGS,
            "graph": {"extraNodes": ["AI201"]},
        }
    )
    monkeypatch.setattr("core.portfolio_views.build_optimized_current_offerings_report", optimized)
    monkeypatch.setattr("core.portfolio_views.build_graduation_report", normal)
    monkeypatch.setattr(
        "core.portfolio_views.graduation_presentation_from_tool_results", presentation
    )
    monkeypatch.setattr(
        "core.portfolio_views.load_defaults",
        lambda: {
            "currentYear": "1448",
            "currentTerm": "2",
            "academic_year": "1448",
            "term": "2",
        },
    )

    response = _client("optimized-api").get(_url(), {"baseline": OPTIMIZED_CURRENT_OFFERINGS})

    assert response.status_code == 200
    payload = response.json()
    assert payload["baseline"] == OPTIMIZED_CURRENT_OFFERINGS
    assert payload["report"]["planning_baseline_kind"] == OPTIMIZED_CURRENT_OFFERINGS
    assert payload["report"]["planning_baseline"]["kind"] == OPTIMIZED_CURRENT_OFFERINGS
    assert payload["presentation"]["planning_baseline_kind"] == OPTIMIZED_CURRENT_OFFERINGS
    optimized.assert_called_once_with(
        STUDENT_ID,
        1448,
        2,
        section_snapshot_academic_year=1448,
        section_snapshot_term=2,
    )
    normal.assert_not_called()


def test_optimized_export_uses_distinct_filename_and_source_label(portfolio_student, monkeypatch):
    report = _report(OPTIMIZED_CURRENT_OFFERINGS)
    report["planning_baseline"] = {"kind": OPTIMIZED_CURRENT_OFFERINGS}
    monkeypatch.setattr(
        "core.portfolio_views.build_optimized_current_offerings_report",
        Mock(return_value=report),
    )
    monkeypatch.setattr(
        "core.portfolio_views.graduation_presentation_from_tool_results",
        Mock(return_value={"graph": {"extraNodes": []}}),
    )
    monkeypatch.setattr(
        "core.portfolio_views.load_defaults",
        lambda: {
            "currentYear": "1448",
            "currentTerm": "2",
            "academic_year": "1448",
            "term": "2",
        },
    )

    response = _client("optimized-export").get(
        _export_url(), {"baseline": OPTIMIZED_CURRENT_OFFERINGS}
    )

    assert response.status_code == 200
    assert response.headers["Content-Disposition"] == (
        f'attachment; filename="graduation_plan_{STUDENT_ID}_optimized_offerings_1448_T2.xlsx"'
    )
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    try:
        values = {
            str(cell.value)
            for row in workbook["Overview"].iter_rows()
            for cell in row
            if cell.value is not None
        }
        assert "Optimized current offerings" in values
        assert "Registered timetable" not in values
    finally:
        workbook.close()
