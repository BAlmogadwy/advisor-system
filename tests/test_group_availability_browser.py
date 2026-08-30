"""Real-browser regressions for the Group Availability workspace.

The service tests prove the aggregation math.  These tests cover the parts
assembled by JavaScript: partial-coverage wording, registered-only sourcing, the continuous
ten-minute view, stale-response protection, keyboard semantics, and compact
viewport scrolling.
"""

from __future__ import annotations

import os
import threading
from io import BytesIO
from unittest import mock

# Playwright's synchronous API runs through a greenlet.  Fixture creation is
# still synchronous ORM work, while Django serves the page on its own thread.
os.environ.setdefault("DJANGO_ALLOW_ASYNC_UNSAFE", "1")

import pytest
from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.staticfiles.testing import StaticLiveServerTestCase
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook

from core.models import Student, StudentTermSection, TermSection, TermSectionMeeting
from core.services.rbac import ROLE_ADVISOR

playwright_api = pytest.importorskip("playwright.sync_api")
expect = playwright_api.expect
sync_playwright = playwright_api.sync_playwright

YEAR = "1448"
TERM = "1"

REGISTERED_ID = 8801001
REGISTERED_ID_2 = 8801002
NO_SCHEDULE_ID = 8801003
UNKNOWN_ID = 8899999


class GroupAvailabilityBrowserTests(StaticLiveServerTestCase):
    """Exercise the production template, JavaScript, endpoint, and service."""

    @classmethod
    def setUpClass(cls) -> None:
        super().setUpClass()
        cls._pw = sync_playwright().start()
        cls.browser = cls._pw.chromium.launch()

    @classmethod
    def tearDownClass(cls) -> None:
        cls.browser.close()
        cls._pw.stop()
        super().tearDownClass()

    # ── Browser and deterministic timetable fixtures ──────────
    def _session_cookie(self) -> dict[str, str]:
        user = get_user_model().objects.create_user(
            username="group-availability-browser-advisor",
            password="unused",
        )
        advisor_group, _ = Group.objects.get_or_create(name=ROLE_ADVISOR)
        user.groups.add(advisor_group)
        client = Client()
        client.force_login(user)
        return {
            "name": settings.SESSION_COOKIE_NAME,
            "value": client.cookies[settings.SESSION_COOKIE_NAME].value,
            "url": self.live_server_url,
        }

    def _page(self, *, viewport: dict[str, int] | None = None):
        context = self.browser.new_context(
            locale="en-US",
            extra_http_headers={"Accept-Language": "en"},
            viewport=viewport or {"width": 1280, "height": 900},
        )
        context.add_cookies(
            [
                self._session_cookie(),
                {
                    "name": settings.LANGUAGE_COOKIE_NAME,
                    "value": "en",
                    "url": self.live_server_url,
                },
            ]
        )
        self.addCleanup(context.close)
        page = context.new_page()
        page.goto(f"{self.live_server_url}{reverse('group_availability_page')}")
        page.wait_for_load_state("networkidle")
        return page

    @staticmethod
    def _enrol(
        student_id: int,
        *,
        name: str,
        course_key: str,
        section: str,
        source: str,
        day: str,
        start: str,
        end: str,
    ) -> None:
        Student.objects.create(
            student_id=student_id,
            name=name,
            program="DS",
            section="M",
        )
        term_section = TermSection.objects.create(
            course_code=course_key[:2],
            course_number=course_key[2:],
            course_key=course_key,
            course_name=course_key,
            section=section,
            source_tag="browser-test",
        )
        TermSectionMeeting.objects.create(
            term_section=term_section,
            day=day,
            start_time=start,
            end_time=end,
        )
        StudentTermSection.objects.create(
            student_id=student_id,
            academic_year=YEAR,
            term=TERM,
            term_section=term_section,
            source=source,
        )

    def _seed_partial_group(self) -> list[int]:
        # The off-grid registrar meeting overlaps exactly the 12:00 and 12:10
        # timeline periods.  A second registered schedule supplies a standard
        # lecture-grid conflict.
        self._enrol(
            REGISTERED_ID,
            name="Registered Student",
            course_key="DS431",
            section="M1",
            source="scraper_timetable",
            day="TUE",
            start="12:05",
            end="12:15",
        )
        self._enrol(
            REGISTERED_ID_2,
            name="Registered Student 2",
            course_key="DS432",
            section="M2",
            source="scraper_timetable",
            day="WED",
            start="09:00",
            end="10:15",
        )
        Student.objects.create(
            student_id=NO_SCHEDULE_ID,
            name="No Timetable",
            program="DS",
            section="M",
        )
        return [REGISTERED_ID, REGISTERED_ID_2, NO_SCHEDULE_ID, UNKNOWN_ID]

    def _seed_one_registered_student(self) -> None:
        self._enrol(
            REGISTERED_ID,
            name="Registered Student",
            course_key="DS431",
            section="M1",
            source="scraper_timetable",
            day="MON",
            start="09:00",
            end="10:15",
        )

    @staticmethod
    def _compute(page, ids: list[int]) -> None:
        page.locator("#gaIds").fill("\n".join(str(student_id) for student_id in ids))
        page.locator("#gaCompute").click()
        expect(page.locator("#gaSummary")).to_be_visible(timeout=10_000)

    # ── Partial coverage + registered source + full-day timeline ──
    def test_partial_coverage_keeps_results_warnings_source_and_full_timeline(self):
        ids = self._seed_partial_group()
        page = self._page()
        self._compute(page, ids)

        loaded = page.locator("#gaSummaryStats .ga-stat").filter(has_text="Registered schedules")
        unresolved = page.locator("#gaSummaryStats .ga-stat").filter(has_text="Unresolved")
        assert loaded.locator(".ga-stat-num").inner_text() == "2/4"
        assert unresolved.locator(".ga-stat-num").inner_text() == "2"

        summary_cards = page.locator("#gaSummaryStats .ga-stat")
        assert summary_cards.count() == 5
        card_boxes = summary_cards.evaluate_all(
            "nodes => nodes.map(node => ({ x: node.offsetLeft, y: node.offsetTop, width: node.offsetWidth }))"
        )
        assert card_boxes[0]["y"] == card_boxes[1]["y"]
        assert card_boxes[2]["y"] > card_boxes[0]["y"]
        assert card_boxes[4]["width"] > card_boxes[0]["width"] * 1.8
        for index in range(summary_cards.count()):
            overflow = (
                summary_cards.nth(index)
                .locator(".ga-stat-lbl")
                .evaluate(
                    "node => ({ clientWidth: node.clientWidth, scrollWidth: node.scrollWidth })"
                )
            )
            assert overflow["scrollWidth"] <= overflow["clientWidth"] + 1

        status = page.locator("#gaStatus").inner_text()
        assert "calculated from 2 of 4 registered schedules" in status
        assert "2 unresolved flagged" in status

        flags = page.locator("#gaFlags").inner_text()
        assert "Coverage is incomplete" in flags
        assert f"Not found: 1 ({UNKNOWN_ID})" in flags
        assert f"No registered schedule: 1 ({NO_SCHEDULE_ID})" in flags
        # The old phrase overstated partial results as certainty for the whole
        # pasted group.  It must not return anywhere in the rendered result.
        assert "Free for all" not in page.locator(".ga-shell").inner_text()

        provenance = page.locator("#gaProvenance").inner_text()
        assert "Registered schedules:" in provenance
        assert provenance.endswith("2")
        assert "Expected" not in page.locator(".ga-shell").inner_text()
        assert "Working" not in page.locator(".ga-shell").inner_text()
        assert page.locator("#gaLegendFree").inner_text() == "Free"

        lecture_rows = page.locator("#gaPanelLecture tbody tr")
        assert lecture_rows.count() == 8
        assert lecture_rows.nth(7).locator(".ga-grid-slotlabel").inner_text().splitlines() == [
            "17:15",
            "18:30",
        ]

        page.locator("#gaTabLab").click()
        lab_rows = page.locator("#gaPanelLab tbody tr")
        assert lab_rows.count() == 6
        assert lab_rows.nth(5).locator(".ga-grid-slotlabel").inner_text().splitlines() == [
            "16:50",
            "18:30",
        ]

        page.locator("#gaTabTimeline").click()
        timeline = page.locator("#gaPanelTimeline")
        expect(timeline).to_be_visible()
        rows = timeline.locator("tbody tr")
        assert rows.count() == 57
        assert rows.nth(0).locator(".ga-grid-slotlabel").inner_text().splitlines() == [
            "09:00",
            "09:10",
        ]
        assert rows.nth(56).locator(".ga-grid-slotlabel").inner_text().splitlines() == [
            "18:20",
            "18:30",
        ]

        noon = rows.nth(18)  # 12:00–12:10; noon is never omitted.
        after_noon = rows.nth(19)  # 12:10–12:20
        known_free = rows.nth(20)  # 12:20–12:30
        assert noon.locator(".ga-grid-slotlabel").inner_text().splitlines() == ["12:00", "12:10"]
        assert after_noon.locator(".ga-grid-slotlabel").inner_text().splitlines() == [
            "12:10",
            "12:20",
        ]
        # td order is Sunday, Monday, Tuesday, Wednesday, Thursday.
        assert noon.locator("td").nth(2).locator(".ga-cell-num").inner_text() == "1"
        assert after_noon.locator("td").nth(2).locator(".ga-cell-num").inner_text() == "1"
        assert known_free.locator("td").nth(2).locator(".ga-cell-num").inner_text() == "0"
        assert known_free.locator("td").nth(2).locator(".ga-cell-sub").inner_text() == "Free"
        noon_button_label = (
            noon.locator("td").nth(2).locator(".ga-cell-button").get_attribute("aria-label")
        )
        assert "Tuesday 12:00–12:10" in noon_button_label
        assert "1 busy among 2 registered schedules" in noon_button_label
        assert "0 busy" in known_free.locator("td").nth(2).get_attribute("aria-label")

        # The source stays explicit and registrar-only at both summary and
        # individual-conflict level.
        page.locator("#gaTabLecture").click()
        registered_conflict = page.locator(
            '#gaPanelLecture .ga-cell-button[aria-label^="Wednesday 09:00"]'
        )
        registered_conflict.click()
        expect(page.locator("#gaDetail")).to_be_visible()
        detail = page.locator("#gaDetail").inner_text()
        assert "DS432 · M2" in detail
        assert "Registered" in detail
        assert "Expected" not in detail
        assert "Working" not in detail

    # ── Input changes invalidate every representation of old data ──
    def test_editing_ids_immediately_hides_old_results_and_details(self):
        self._seed_one_registered_student()
        page = self._page()
        export_button = page.locator("#gaExport")
        expect(export_button).to_be_disabled()
        self._compute(page, [REGISTERED_ID])
        expect(export_button).to_be_enabled()

        conflict = page.locator('#gaPanelLecture .ga-cell-button[aria-label^="Monday 09:00"]')
        conflict.click()
        expect(page.locator("#gaDetail")).to_be_visible()

        page.locator("#gaIds").fill(f"{REGISTERED_ID}\n{UNKNOWN_ID}")

        expect(export_button).to_be_disabled()
        expect(page.locator("#gaSummary")).to_be_hidden()
        expect(page.locator("#gaDetail")).to_be_hidden()
        assert page.locator(".ga-grid-wrap table.ga-grid").count() == 0
        assert page.locator(".ga-grid-wrap .ga-empty").count() == 3
        # Do not couple this regression to transient status copy: Playwright's
        # fill may emit an input for the clear and another for the replacement.
        # The durable contract is that no result representation survives.
        assert "calculated from" not in page.locator("#gaStatus").inner_text()

    # ── A cancelled request may finish, but may not repaint ──────
    def test_clear_during_delayed_compute_cannot_repopulate_stale_results(self):
        self._seed_one_registered_student()
        page = self._page()

        from core.services.group_availability import compute_group_availability

        started = threading.Event()
        release = threading.Event()
        finished = threading.Event()

        def delayed_compute(student_ids):
            started.set()
            try:
                if not release.wait(timeout=8):
                    raise TimeoutError("browser test did not release delayed availability request")
                return compute_group_availability(student_ids)
            finally:
                finished.set()

        try:
            with mock.patch(
                "core.group_availability_views.compute_group_availability",
                side_effect=delayed_compute,
            ):
                page.locator("#gaIds").fill(str(REGISTERED_ID))
                page.locator("#gaCompute").click()
                expect(page.locator("#gaExport")).to_be_disabled()
                assert started.wait(timeout=5), (
                    "the live endpoint never entered the delayed service"
                )

                page.locator("#gaClear").click()
                assert page.locator("#gaIds").input_value() == ""
                expect(page.locator("#gaExport")).to_be_disabled()
                expect(page.locator("#gaSummary")).to_be_hidden()

                release.set()
                assert finished.wait(timeout=5), "the delayed service did not finish"
                # Let the fetch rejection/late response handlers run in Chromium.
                page.wait_for_timeout(250)
        finally:
            release.set()

        assert page.locator("#gaIds").input_value() == ""
        assert page.locator("#gaStatus").inner_text() == ""
        assert page.locator(".ga-grid-wrap table.ga-grid").count() == 0
        expect(page.locator("#gaSummary")).to_be_hidden()
        expect(page.locator("#gaCompute")).to_be_enabled()

    # ── XLSX export always includes every grid, independent of active tab ──
    def test_xlsx_export_downloads_all_grids_and_invalidates_on_edit(self):
        self._seed_one_registered_student()
        page = self._page()
        export_button = page.locator("#gaExport")
        expect(export_button).to_be_disabled()
        self._compute(page, [REGISTERED_ID])
        expect(export_button).to_be_enabled()

        page.locator("#gaTabTimeline").click()
        with page.expect_download(timeout=15_000) as download_info:
            export_button.click()
        download = download_info.value

        assert download.suggested_filename.startswith("group_availability_")
        assert download.suggested_filename.endswith(".xlsx")
        workbook = load_workbook(BytesIO(download.path().read_bytes()), read_only=True)
        assert workbook.sheetnames == [
            "Summary",
            "Students",
            "Lecture 75m",
            "Lab 100m",
            "Full Day 10m",
        ]
        assert (
            workbook["Lecture 75m"].cell(row=workbook["Lecture 75m"].max_row, column=1).value
            == "17:15–18:30"
        )
        assert (
            workbook["Lab 100m"].cell(row=workbook["Lab 100m"].max_row, column=1).value
            == "16:50–18:30"
        )
        assert (
            workbook["Full Day 10m"].cell(row=workbook["Full Day 10m"].max_row, column=1).value
            == "18:20–18:30"
        )
        workbook.close()

        expect(export_button).to_be_enabled()
        assert export_button.get_attribute("aria-busy") == "false"
        page.locator("#gaIds").fill(f"{REGISTERED_ID}\n{UNKNOWN_ID}")
        expect(export_button).to_be_disabled()

    def test_editing_ids_during_xlsx_export_prevents_stale_download(self):
        self._seed_one_registered_student()
        page = self._page()
        self._compute(page, [REGISTERED_ID])
        export_button = page.locator("#gaExport")
        downloads = []
        page.on("download", lambda download: downloads.append(download))

        from core.services.group_availability_export import build_group_availability_xlsx

        started = threading.Event()
        release = threading.Event()
        finished = threading.Event()

        def delayed_export(result):
            started.set()
            try:
                if not release.wait(timeout=8):
                    raise TimeoutError("browser test did not release delayed XLSX export")
                return build_group_availability_xlsx(result)
            finally:
                finished.set()

        try:
            with mock.patch(
                "core.services.group_availability_export.build_group_availability_xlsx",
                side_effect=delayed_export,
            ):
                export_button.click()
                assert started.wait(timeout=5), "the live endpoint never entered XLSX export"
                expect(export_button).to_be_disabled()
                assert export_button.get_attribute("aria-busy") == "true"
                assert export_button.inner_text() == "Exporting…"

                page.locator("#gaIds").fill(f"{REGISTERED_ID}\n{UNKNOWN_ID}")
                expect(export_button).to_be_disabled()
                expect(page.locator("#gaSummary")).to_be_hidden()

                release.set()
                assert finished.wait(timeout=5), "the delayed XLSX export did not finish"
                page.wait_for_timeout(350)
        finally:
            release.set()

        assert downloads == []
        expect(export_button).to_be_disabled()
        assert export_button.get_attribute("aria-busy") == "false"

    # ── Keyboard tabs, labelled conflicts, and focus restoration ──
    def test_tabs_and_busy_details_are_keyboard_and_screen_reader_operable(self):
        self._seed_one_registered_student()
        page = self._page()
        self._compute(page, [REGISTERED_ID])

        tabs = page.locator('.ga-tabs [role="tab"]')
        assert tabs.count() == 3
        lecture = page.locator("#gaTabLecture")
        lab = page.locator("#gaTabLab")
        timeline = page.locator("#gaTabTimeline")
        assert lecture.get_attribute("aria-controls") == "gaPanelLecture"
        assert page.locator("#gaPanelLecture").get_attribute("aria-labelledby") == "gaTabLecture"
        assert lecture.get_attribute("aria-selected") == "true"
        assert lecture.get_attribute("tabindex") == "0"

        lecture.focus()
        page.keyboard.press("ArrowRight")
        assert page.evaluate("document.activeElement.id") == "gaTabLab"
        assert lab.get_attribute("aria-selected") == "true"
        assert lecture.get_attribute("tabindex") == "-1"
        expect(page.locator("#gaPanelLecture")).to_be_hidden()
        expect(page.locator("#gaPanelLab")).to_be_visible()

        page.keyboard.press("End")
        assert page.evaluate("document.activeElement.id") == "gaTabTimeline"
        assert timeline.get_attribute("aria-selected") == "true"
        page.keyboard.press("Home")
        assert page.evaluate("document.activeElement.id") == "gaTabLecture"
        assert lecture.get_attribute("aria-selected") == "true"

        busy = page.locator('#gaPanelLecture .ga-cell-button[aria-label^="Monday 09:00–10:15"]')
        assert busy.get_attribute("aria-controls") == "gaDetail"
        assert busy.get_attribute("aria-expanded") == "false"
        assert "1 busy among 1 registered schedules" in busy.get_attribute("aria-label")

        busy.focus()
        page.keyboard.press("Enter")
        expect(page.locator("#gaDetail")).to_be_visible()
        assert page.evaluate("document.activeElement.id") == "gaDetailClose"
        assert busy.get_attribute("aria-expanded") == "true"
        assert "Registered" in page.locator("#gaDetail").inner_text()

        page.keyboard.press("Escape")
        expect(page.locator("#gaDetail")).to_be_hidden()
        assert page.evaluate("document.activeElement.classList.contains('ga-cell-button')") is True
        assert busy.get_attribute("aria-expanded") == "false"

    # ── Compact screens keep the table usable through local scrolling ──
    def test_narrow_viewport_grid_scrolls_horizontally_inside_its_panel(self):
        self._seed_one_registered_student()
        page = self._page(viewport={"width": 375, "height": 812})
        self._compute(page, [REGISTERED_ID])

        panel = page.locator("#gaPanelLecture")
        metrics = panel.evaluate(
            """node => ({
                clientWidth: node.clientWidth,
                scrollWidth: node.scrollWidth,
                overflowX: getComputedStyle(node).overflowX
            })"""
        )
        assert metrics["overflowX"] in {"auto", "scroll"}
        assert metrics["scrollWidth"] > metrics["clientWidth"], metrics
        moved = panel.evaluate("node => { node.scrollLeft = 120; return node.scrollLeft; }")
        assert moved > 0
