from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace
from typing import Any

from openpyxl import load_workbook

from core.services.advisor_presentations import graduation_presentation_from_tool_results
from core.services.graduation_export import build_graduation_xlsx
from core.services.student_graduation import REGISTERED_TIMETABLE


def _course(
    code: str,
    *,
    name: str,
    role: str,
    source: str,
    must_have: bool = False,
    auto_prerequisite: bool = False,
) -> dict[str, Any]:
    return {
        "code": code,
        "name": name,
        "credits": 3,
        "requirement_type": "Core",
        "scenario_role": role,
        "source": source,
        "must_have": must_have,
        "auto_added_prerequisite": auto_prerequisite,
    }


def _report(*, with_what_if: bool = True) -> dict[str, Any]:
    retained = _course(
        "AI201",
        name="Introduction to AI",
        role="retained_baseline",
        source="registered_timetable",
    )
    must_have = _course(
        "AI401",
        name="Advanced AI",
        role="must_have",
        source="admin_override",
        must_have=True,
    )
    prerequisite = _course(
        "AI301",
        name="Machine Learning",
        role="auto_prerequisite",
        source="same_term_direct_prerequisite",
        auto_prerequisite=True,
    )
    current = [retained, must_have, prerequisite] if with_what_if else [retained]
    report: dict[str, Any] = {
        "tool": "graduation_progress",
        "ok": True,
        "program": "AI",
        "planning_baseline_academic_year": 1448,
        "planning_baseline_term": 2,
        "planning_baseline_kind": REGISTERED_TIMETABLE,
        "planning_baseline_credits": sum(row["credits"] for row in current),
        "planning_baseline_courses_assumed_passed": current,
        "plan_courses_total": 4,
        "plan_courses_passed": 0,
        "remaining_courses": 1,
        "remaining_credits": 3,
        "passed_credits_in_plan": 0,
        "earned_credits_registrar": 70,
        "lower_bound_terms_including_planning_baseline": 2,
        "estimated_terms_including_planning_baseline": 2,
        "simulation_completed": True,
        "simulated_terms_examined": 1,
        "productive_terms_planned": 1,
        "max_credits_per_term": 18,
        "term_plan": [
            {
                "sequence": 1,
                "academic_year": 1449,
                "term": 1,
                "credits": 3,
                "courses": [
                    {
                        "code": "AI499",
                        "name": "Graduation Project",
                        "credits": 3,
                        "requirement_type": "Core",
                    }
                ],
            }
        ],
        "unresolved_requirements": [],
        "simulation_assumptions": ["All scenario courses are assumed passed on the first attempt."],
        "scenario_graph": {
            "items": [
                {
                    "course_code": "AI401",
                    "prerequisite_course_code": "AI301",
                },
                {
                    "course_code": "AI499",
                    "prerequisite_course_code": "AI401",
                },
            ],
            # Deliberately claim every current course is studying. The
            # presentation boundary must correct the two hypothetical rows.
            "statusOf": {
                "AI201": "studying",
                "AI301": "studying",
                "AI401": "studying",
                "AI499": "open",
            },
            "nameOf": {
                "AI201": "Introduction to AI",
                "AI301": "Machine Learning",
                "AI401": "Advanced AI",
                "AI499": "Graduation Project",
            },
        },
    }
    if with_what_if:
        report["what_if"] = {
            "mode": "must_have_current_term",
            "valid": True,
            "allow_same_term_direct_prerequisites": True,
            "same_term_direct_prerequisite_approval": True,
            "requested_must_have_course_codes": ["AI201", "AI401"],
            "must_have_courses": [retained, must_have],
            "already_in_baseline_courses": [retained],
            "added_must_have_courses": [must_have],
            "auto_added_prerequisites": [prerequisite],
            "same_term_direct_prerequisite_edges": [
                {
                    "course_code": "AI401",
                    "prerequisite_code": "AI301",
                    "exception": "DIRECT_PREREQUISITE_SAME_TERM",
                }
            ],
            "displaced_baseline_courses": [
                {
                    "code": "IS101",
                    "name": "Islamic Studies",
                    "credits": 2,
                    "source": "registered_timetable",
                }
            ],
            "validation_errors": [],
            "baseline": {
                "planning_baseline_kind": REGISTERED_TIMETABLE,
                "planning_baseline_credits": 3,
                "simulation_completed": True,
                "estimated_terms_including_planning_baseline": 3,
                "lower_bound_terms_including_planning_baseline": 3,
                "planning_baseline_courses_assumed_passed": [retained],
            },
            "scenario": {
                "planning_baseline_kind": REGISTERED_TIMETABLE,
                "planning_baseline_credits": 9,
                "simulation_completed": True,
                "estimated_terms_including_planning_baseline": 2,
                "lower_bound_terms_including_planning_baseline": 2,
                "planning_baseline_courses_assumed_passed": current,
            },
            "comparison": {
                "timing_effect": "EARLIER",
                "term_difference": -1,
                "terms_saved": 1,
                "exact_timing_comparison_available": True,
                "proven_improvement": True,
                "complete_forecast_improved": True,
                "improvement_basis": "COMPLETE_FORECAST",
                "plan_changed": True,
                "baseline_planning_credits": 3,
                "scenario_planning_credits": 9,
                "planning_credit_change": 6,
            },
            "timetable_check_required": True,
            "note": "Read-only administrative simulation.",
        }
    return report


def _presentation(report: dict[str, Any]) -> dict[str, Any]:
    presentation = graduation_presentation_from_tool_results([report])
    assert presentation
    return presentation


def _workbook(report: dict[str, Any]):
    content = build_graduation_xlsx(
        student=SimpleNamespace(
            student_id=4_501_234,
            registration_no="4501234",
            name="Scenario Student",
            program="AI",
            section="M",
        ),
        academic_year=1448,
        term=2,
        baseline_kind=REGISTERED_TIMETABLE,
        report=report,
        presentation=_presentation(report),
        language_code="en",
    )
    return load_workbook(BytesIO(content), data_only=False)


def _all_text(workbook_or_sheet) -> list[str]:
    sheets = (
        workbook_or_sheet.worksheets
        if hasattr(workbook_or_sheet, "worksheets")
        else [workbook_or_sheet]
    )
    return [
        str(cell.value)
        for sheet in sheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]


def _row_with(sheet, value: object) -> tuple[object, ...]:
    for row in sheet.iter_rows(values_only=True):
        if value in row:
            return row
    raise AssertionError(f"{value!r} is missing from {sheet.title!r}")


def test_must_have_presentation_preserves_evidence_and_never_claims_hypothetical_registration() -> (
    None
):
    presentation = _presentation(_report())

    assert presentation["graph"]["statusOf"]["AI201"] == "studying"
    assert presentation["graph"]["statusOf"]["AI401"] == "open"
    assert presentation["graph"]["statusOf"]["AI301"] == "open"

    what_if = presentation["what_if"]
    assert what_if["mode"] == "must_have_current_term"
    assert what_if["valid"] is True
    assert what_if["same_term_direct_prerequisite_approval"] is True
    assert what_if["special_approval_required"] is True
    assert what_if["no_eligibility_or_registration_approval"] is True
    assert what_if["requested_must_have_course_codes"] == ["AI201", "AI401"]
    assert what_if["added_must_have_courses"][0]["source"] == "admin_override"
    assert what_if["added_must_have_courses"][0]["scenario_role"] == "must_have"
    assert what_if["auto_added_prerequisites"][0]["source"] == ("same_term_direct_prerequisite")
    assert {(row["code"], row["source"]) for row in presentation["added_current_courses"]} == {
        ("AI401", "admin_override"),
        ("AI301", "same_term_direct_prerequisite"),
    }
    assert what_if["same_term_direct_prerequisite_edges"] == [
        {
            "course_code": "AI401",
            "prerequisite_code": "AI301",
            "exception": "DIRECT_PREREQUISITE_SAME_TERM",
        }
    ]
    assert what_if["displaced_baseline_courses"][0]["code"] == "IS101"
    assert what_if["baseline"]["estimated_terms_including_planning_baseline"] == 3
    assert what_if["scenario"]["estimated_terms_including_planning_baseline"] == 2
    assert what_if["comparison"]["term_difference"] == -1
    assert what_if["comparison"]["timing_effect"] == "EARLIER"


def test_invalid_must_have_presentation_keeps_validation_error_kind() -> None:
    report = _report(with_what_if=False)
    report["what_if"] = {
        "mode": "must_have_current_term",
        "valid": False,
        "requested_must_have_course_codes": ["ZZ999"],
        "must_have_courses": [],
        "validation_errors": [{"kind": "MUST_HAVE_COURSE_NOT_IN_PLAN", "course_code": "ZZ999"}],
        "note": "Read-only administrative simulation.",
    }

    presentation = _presentation(report)

    assert presentation["what_if"]["valid"] is False
    assert presentation["what_if"]["validation_errors"] == [
        {
            "code": "MUST_HAVE_COURSE_NOT_IN_PLAN",
            "message": "",
            "course_code": "ZZ999",
        }
    ]


def test_must_have_workbook_labels_hypothetical_rows_exceptions_and_decision_boundary() -> None:
    workbook = _workbook(_report())
    try:
        assert workbook.properties.subject == (
            "Read-only hypothetical admin must-have graduation scenario"
        )
        assert "does not approve registration" in str(workbook.properties.description)

        overview_text = _all_text(workbook)
        assert any("Hypothetical admin override" in value for value in overview_text)
        assert any("no eligibility" in value.lower() for value in overview_text)

        term_sheet = workbook["Term Plan"]
        assert "Registered" in _row_with(term_sheet, "AI201")
        assert "Hypothetical must-have" in _row_with(term_sheet, "AI401")
        assert "Hypothetical same-term prerequisite" in _row_with(term_sheet, "AI301")

        baseline_sheet = workbook["Baseline Courses"]
        retained = _row_with(baseline_sheet, "AI201")
        forced = _row_with(baseline_sheet, "AI401")
        prerequisite = _row_with(baseline_sheet, "AI301")
        assert "Registered timetable" in retained
        assert any("Hypothetical admin must-have" in str(value) for value in forced)
        assert any("No registration evidence" in str(value) for value in forced)
        assert any("Hypothetical auto-added" in str(value) for value in prerequisite)
        assert "Registered timetable" not in forced
        assert "Registered timetable" not in prerequisite

        exception = _row_with(
            workbook["Prerequisite Map"],
            "Same-term prerequisite exception",
        )
        assert "AI301" in exception and "AI401" in exception
        assert "Approved hypothetical same-term exception" in exception

        blocker_text = _all_text(workbook["Blockers & Assumptions"])
        for expected in (
            "Added hypothetical must-have courses",
            "Auto-added same-term prerequisites",
            "Displaced baseline courses",
            "Relaxed prerequisite edges",
            "Special academic approval",
            "Timetable conflict check",
            "Comparison",
            "Decision boundary",
        ):
            assert expected in blocker_text
        assert "AI401" in blocker_text
        assert "AI301" in blocker_text
        assert "IS101" in blocker_text
        assert "Required" in blocker_text
        assert any("no eligibility" in value.lower() for value in blocker_text)
    finally:
        workbook.close()


def test_ordinary_registered_export_keeps_existing_labels_and_has_no_override_metadata() -> None:
    report = _report(with_what_if=False)
    workbook = _workbook(report)
    try:
        text = _all_text(workbook)
        assert not any("Hypothetical admin" in value for value in text)
        assert not any("must-have override" in value.lower() for value in text)
        assert "Registered" in _row_with(workbook["Term Plan"], "AI201")
        assert "Registered timetable" in _row_with(
            workbook["Baseline Courses"],
            "AI201",
        )
    finally:
        workbook.close()
