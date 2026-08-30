"""Tests for the group-availability (common free-slot) finder.

Covers the aggregation service (busy/free cells, off-grid overlap, the lab
grid, not-found / no-schedule reporting, ID normalisation) and the page +
compute view wiring end-to-end via the Django test client.
"""

from __future__ import annotations

import json
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook

from core.models import (
    Student,
    StudentTermSection,
    TermSection,
    TermSectionMeeting,
    TimetableScenario,
)
from core.services.group_availability import (
    _section_course_identity,
    compute_group_availability,
    normalise_student_ids,
)
from core.services.rbac import ROLE_STUDENT

pytestmark = pytest.mark.django_db

YEAR = "1448"
TERM = "1"

# Lecture slot indices (DEFAULT_SLOTS): 0=09:00-10:15, 1=10:30-11:45,
# 2=10:50-12:05, 3=13:00-14:15, 4=14:30-15:45, 5=14:45-16:00,
# 6=16:00-17:15, plus the Group Availability extension 7=17:15-18:30.


def _make_global_section(
    course_code: str, section: str, meetings: list[tuple[str, str, str]]
) -> TermSection:
    """Create a GLOBAL (scenario=NULL) section with the given meetings.

    ``get_student_term_baseline`` only reads scenario-NULL sections, matching
    the system-of-record registrations the finder is built on.
    """
    ts = TermSection.objects.create(
        course_code=course_code,
        course_number=course_code,
        course_key=course_code,
        section=section,
        course_name=course_code,
        source_tag="test",
    )
    for day, start, end in meetings:
        TermSectionMeeting.objects.create(term_section=ts, day=day, start_time=start, end_time=end)
    return ts


def _enrol(
    student_id: int,
    ts: TermSection,
    *,
    program: str = "CS",
    cohort: str = "",
    source: str = "scraper_timetable",
) -> None:
    Student.objects.get_or_create(
        student_id=student_id,
        defaults={"name": f"Student {student_id}", "program": program, "section": cohort},
    )
    StudentTermSection.objects.create(
        student_id=student_id,
        academic_year=YEAR,
        term=TERM,
        term_section=ts,
        source=source,
    )


def _cell(result: dict, grid: str, day: str, slot_index: int) -> dict:
    return result["grids"][grid]["cells"][day][slot_index]


def test_shared_busy_slot_is_not_free_and_other_slot_is_free():
    sec = _make_global_section("CS101", "S1", [("MON", "09:00", "10:15")])
    _enrol(1001, sec)
    _enrol(1002, sec)

    result = compute_group_availability([1001, 1002], YEAR, TERM)

    assert result["requested_count"] == 2
    assert result["resolved_count"] == 2

    busy = _cell(result, "lecture", "MON", 0)
    assert busy["busy_count"] == 2
    assert busy["free"] is False

    free = _cell(result, "lecture", "TUE", 0)
    assert free["busy_count"] == 0
    assert free["free"] is True


def test_partial_overlap_counts_only_busy_students():
    sec = _make_global_section("CS201", "S1", [("MON", "10:30", "11:45")])
    _enrol(2001, sec)
    Student.objects.create(student_id=2002, name="Free", program="CS")
    # 2002 has no registered section → free everywhere.

    result = compute_group_availability([2001, 2002], YEAR, TERM)

    partial = _cell(result, "lecture", "MON", 1)
    assert partial["busy_count"] == 1
    assert partial["free"] is False
    # The busy student's other slots stay free.
    assert _cell(result, "lecture", "MON", 0)["free"] is True


def test_offgrid_meeting_marks_every_overlapping_slot():
    # 09:30–10:45 straddles lecture slot 0 (09:00-10:15) and slot 1 (10:30-11:45).
    sec = _make_global_section("CS301", "S1", [("TUE", "09:30", "10:45")])
    _enrol(3001, sec)

    result = compute_group_availability([3001], YEAR, TERM)

    assert _cell(result, "lecture", "TUE", 0)["busy_count"] == 1
    assert _cell(result, "lecture", "TUE", 1)["busy_count"] == 1
    # A non-overlapping slot is still free.
    assert _cell(result, "lecture", "TUE", 2)["free"] is True


def test_timeline_uses_all_continuous_ten_minute_intervals():
    result = compute_group_availability([], YEAR, TERM)

    slots = result["grids"]["timeline"]["slots"]
    assert len(slots) == 57
    assert slots[0] == {"label": "09:00-09:10", "start": "09:00", "end": "09:10"}
    assert slots[-1] == {"label": "18:20-18:30", "start": "18:20", "end": "18:30"}
    assert [(slot["start"], slot["end"]) for slot in slots[17:25]] == [
        ("11:50", "12:00"),
        ("12:00", "12:10"),
        ("12:10", "12:20"),
        ("12:20", "12:30"),
        ("12:30", "12:40"),
        ("12:40", "12:50"),
        ("12:50", "13:00"),
        ("13:00", "13:10"),
    ]
    assert [(slot["start"], slot["end"]) for slot in slots[48:]] == [
        ("17:00", "17:10"),
        ("17:10", "17:20"),
        ("17:20", "17:30"),
        ("17:30", "17:40"),
        ("17:40", "17:50"),
        ("17:50", "18:00"),
        ("18:00", "18:10"),
        ("18:10", "18:20"),
        ("18:20", "18:30"),
    ]
    assert result["grids"]["timeline"]["free_for_all_count"] == 285
    assert result["grids"]["timeline"]["free_for_resolved_count"] == 285


def test_timeline_marks_exactly_the_ten_minute_intervals_a_meeting_overlaps():
    section = _make_global_section("CS302", "M1", [("TUE", "12:05", "12:15")])
    _enrol(3051, section, cohort="M", source="scraper_timetable")

    result = compute_group_availability([3051], YEAR, TERM)

    assert _cell(result, "timeline", "TUE", 17)["busy_count"] == 0  # 11:50-12:00
    assert _cell(result, "timeline", "TUE", 18)["busy_count"] == 1  # 12:00-12:10
    assert _cell(result, "timeline", "TUE", 19)["busy_count"] == 1  # 12:10-12:20
    assert _cell(result, "timeline", "TUE", 20)["busy_count"] == 0  # 12:20-12:30
    assert _cell(result, "timeline", "MON", 18)["busy_count"] == 0


def test_all_grids_mark_a_meeting_in_the_new_final_interval():
    section = _make_global_section("CS303", "M2", [("THU", "18:20", "18:30")])
    _enrol(3052, section, cohort="M", source="scraper_timetable")

    result = compute_group_availability([3052], YEAR, TERM)

    assert _cell(result, "lecture", "THU", 6)["busy_count"] == 0  # 16:00-17:15
    assert _cell(result, "lecture", "THU", 7)["busy_count"] == 1  # 17:15-18:30
    assert _cell(result, "lab", "THU", 4)["busy_count"] == 0  # 16:30-18:10
    assert _cell(result, "lab", "THU", 5)["busy_count"] == 1  # 16:50-18:30
    assert _cell(result, "timeline", "THU", 55)["busy_count"] == 0  # 18:10-18:20
    assert _cell(result, "timeline", "THU", 56)["busy_count"] == 1  # 18:20-18:30


def test_lab_grid_reflects_lab_length_meeting():
    # Lab slot 0 is 09:00-10:40.
    sec = _make_global_section("CS401", "S1", [("WED", "09:00", "10:40")])
    _enrol(4001, sec)

    result = compute_group_availability([4001], YEAR, TERM)

    assert _cell(result, "lab", "WED", 0)["busy_count"] == 1
    assert _cell(result, "lab", "WED", 0)["free"] is False
    # free_for_all_count drops below the full 30 cells when something is busy.
    assert result["grids"]["lab"]["free_for_all_count"] == 29


def test_scenario_scoped_section_is_included():
    """In this system, schedules live under a planning scenario — the finder
    must read scenario-owned sections, not only global (scenario-NULL) ones."""
    scenario = TimetableScenario.objects.create(academic_year=YEAR, term=TERM, name="S")
    ts = TermSection.objects.create(
        scenario=scenario,
        course_code="CS900",
        course_number="CS900",
        course_key="CS900",
        section="S1",
        course_name="CS900",
        source_tag="test",
    )
    TermSectionMeeting.objects.create(
        term_section=ts, day="MON", start_time="09:00", end_time="10:15"
    )
    Student.objects.create(student_id=9001, name="Scoped", program="CS")
    StudentTermSection.objects.create(
        student_id=9001,
        academic_year=YEAR,
        term=TERM,
        term_section=ts,
        source="scraper_timetable",
    )

    result = compute_group_availability([9001])  # term auto-detected
    assert result["resolved_count"] == 1
    assert _cell(result, "lecture", "MON", 0)["busy_count"] == 1


def test_auto_detects_current_term_without_explicit_term():
    sec = _make_global_section("CS950", "S1", [("TUE", "13:00", "14:15")])
    _enrol(9501, sec)

    result = compute_group_availability([9501])  # no year/term passed
    assert result["academic_year"] == YEAR
    assert result["term"] == TERM
    # 13:00-14:15 is lecture index 3: the 10:50-12:05 post-lab slot sits at
    # index 2 in the default grid, so the afternoon slots shift down by one.
    assert _cell(result, "lecture", "TUE", 3)["busy_count"] == 1


def test_auto_detected_term_ignores_a_later_expected_only_term():
    registered = _make_global_section("CS951", "M1", [("MON", "09:00", "10:15")])
    future_plan = _make_global_section("CS952", "M2", [("TUE", "09:00", "10:15")])
    _enrol(9511, registered, cohort="M", source="scraper_timetable")
    StudentTermSection.objects.create(
        student_id=9511,
        academic_year="1449",
        term="1",
        term_section=future_plan,
        source="registration_plan_1449_t1",
    )

    result = compute_group_availability([9511])

    assert (result["academic_year"], result["term"]) == (YEAR, TERM)
    assert _cell(result, "lecture", "MON", 0)["busy_count"] == 1
    assert _cell(result, "lecture", "TUE", 0)["busy_count"] == 0


def test_occupants_carry_course_identity():
    sec = _make_global_section("CS501", "S2", [("SUN", "13:00", "14:15")])
    _enrol(5001, sec)

    result = compute_group_availability([5001], YEAR, TERM)
    # 13:00-14:15 is lecture index 3 (10:50-12:05 post-lab slot is index 2).
    cell = _cell(result, "lecture", "SUN", 3)
    assert cell["busy_count"] == 1
    occ = cell["occupants"]
    assert len(occ) == 1
    assert occ[0]["student_id"] == 5001
    assert occ[0]["course_code"] == "CS501"
    assert occ[0]["section"] == "S2"


def test_known_student_cohort_excludes_opposite_local_section_before_resolution():
    male = _make_global_section("DS432", "M3", [("MON", "09:00", "10:15")])
    female = _make_global_section("DS432", "F5", [("TUE", "09:00", "10:15")])
    expected = _make_global_section("DS433", "M4", [("WED", "09:00", "10:15")])
    # Cohort filtering happens before the registered-only source selection. The
    # valid registrar row remains, while both the opposite F row and M plan go.
    _enrol(5101, male, cohort="M", source="scraper_timetable")
    _enrol(5101, female, cohort="M", source="scraper_timetable")
    _enrol(5101, expected, cohort="M", source="registration_plan_1448_t1")

    result = compute_group_availability([5101], YEAR, TERM)

    assert _cell(result, "lecture", "MON", 0)["busy_count"] == 1
    assert _cell(result, "lecture", "TUE", 0)["busy_count"] == 0
    assert _cell(result, "lecture", "WED", 0)["busy_count"] == 0
    assert {
        occupant["section"] for occupant in _cell(result, "lecture", "MON", 0)["occupants"]
    } == {"M3"}
    assert result["students"][0]["snapshot_class"] == "registrar"


def test_female_cohort_excludes_male_sections_but_keeps_shared_sections():
    female = _make_global_section("AI431", "F2", [("MON", "09:00", "10:15")])
    male = _make_global_section("AI432", "M2", [("TUE", "09:00", "10:15")])
    shared = _make_global_section("AI433", "S2", [("WED", "09:00", "10:15")])
    _enrol(5151, female, cohort="F", source="scraper_timetable")
    _enrol(5151, male, cohort="F", source="scraper_timetable")
    _enrol(5151, shared, cohort="F", source="scraper_timetable")

    result = compute_group_availability([5151], YEAR, TERM)

    assert _cell(result, "lecture", "MON", 0)["busy_count"] == 1
    assert _cell(result, "lecture", "TUE", 0)["busy_count"] == 0
    assert _cell(result, "lecture", "WED", 0)["busy_count"] == 1


def test_blank_student_cohort_keeps_local_sections_but_rejects_other_branch():
    male = _make_global_section("DS433", "M1", [("MON", "09:00", "10:15")])
    female = _make_global_section("DS434", "F1", [("TUE", "09:00", "10:15")])
    other_branch = _make_global_section("DS435", "YM1", [("WED", "09:00", "10:15")])
    _enrol(5201, male, source="scraper_timetable")
    _enrol(5201, female, source="scraper_timetable")
    _enrol(5201, other_branch, source="scraper_timetable")

    result = compute_group_availability([5201], YEAR, TERM)

    assert _cell(result, "lecture", "MON", 0)["busy_count"] == 1
    assert _cell(result, "lecture", "TUE", 0)["busy_count"] == 1
    assert _cell(result, "lecture", "WED", 0)["busy_count"] == 0


def test_conflict_details_prefer_full_course_key_over_prefix():
    section = TermSection.objects.create(
        course_code="DS",
        course_number="432",
        course_key="DS432",
        section="M3",
        course_name="Data Mining",
        source_tag="test",
    )
    TermSectionMeeting.objects.create(
        term_section=section, day="TUE", start_time="13:00", end_time="14:15"
    )
    _enrol(5301, section, cohort="M", source="scraper_timetable")

    result = compute_group_availability([5301], YEAR, TERM)

    occupant = _cell(result, "lecture", "TUE", 3)["occupants"][0]
    assert occupant["course_code"] == "DS432"


@pytest.mark.parametrize(
    ("course_code", "course_number", "expected"),
    [("DS", "432", "DS432"), ("GS101", "GS101", "GS101")],
)
def test_course_identity_safely_composes_legacy_rows_without_a_key(
    course_code: str, course_number: str, expected: str
):
    section = TermSection(course_code=course_code, course_number=course_number, course_key="")

    assert _section_course_identity(section) == expected


def test_only_registrar_rows_contribute_to_availability():
    expected_superseded = _make_global_section("CS541", "M1", [("MON", "09:00", "10:15")])
    registered = _make_global_section("CS542", "M2", [("TUE", "09:00", "10:15")])
    expected_only = _make_global_section("CS543", "F1", [("WED", "09:00", "10:15")])
    working_only = _make_global_section("CS544", "M3", [("THU", "09:00", "10:15")])

    _enrol(5401, expected_superseded, cohort="M", source="registration_plan_1448_t1")
    _enrol(5401, registered, cohort="M", source="scraper_timetable")
    _enrol(5402, expected_only, cohort="F", source="registration_plan_1448_t1")
    _enrol(5403, working_only, cohort="M", source="planner")

    result = compute_group_availability([5401, 5402, 5403], YEAR, TERM)

    assert _cell(result, "lecture", "MON", 0)["busy_count"] == 0
    assert _cell(result, "lecture", "TUE", 0)["occupants"][0]["snapshot_class"] == "registrar"
    assert _cell(result, "lecture", "WED", 0)["busy_count"] == 0
    assert _cell(result, "lecture", "THU", 0)["busy_count"] == 0
    assert {student["student_id"]: student["snapshot_class"] for student in result["students"]} == {
        5401: "registrar",
        5402: "",
        5403: "",
    }
    assert result["no_schedule"] == [5402, 5403]
    assert result["resolved_count"] == 1
    assert result["snapshot_class_counts"] == {
        "registrar": 1,
        "expected": 0,
        "working": 0,
    }


def test_expected_and_working_rows_never_fallback_when_registrar_is_absent():
    expected = _make_global_section("CS545", "M1", [("MON", "09:00", "10:15")])
    working = _make_global_section("CS546", "M2", [("TUE", "09:00", "10:15")])
    _enrol(5451, expected, cohort="M", source="registration_plan_1448_t1")
    _enrol(5452, working, cohort="M", source="planner")

    result = compute_group_availability([5451, 5452], YEAR, TERM)

    assert result["resolved_count"] == 0
    assert result["no_schedule"] == [5451, 5452]
    assert all(
        cell["busy_count"] == 0
        for cells in result["grids"]["timeline"]["cells"].values()
        for cell in cells
    )


def test_not_found_vs_no_schedule_reporting():
    sec = _make_global_section("CS601", "S1", [("MON", "09:00", "10:15")])
    _enrol(6001, sec)
    Student.objects.create(student_id=6002, name="No sections", program="CS")  # exists, unenrolled

    result = compute_group_availability([6001, 6002, 9999], YEAR, TERM)

    assert result["requested_count"] == 3
    assert result["resolved_count"] == 1
    assert result["no_schedule"] == [6002]
    assert result["not_found"] == [9999]


def test_partial_coverage_is_flagged_without_blocking_resolved_calculation():
    section = _make_global_section("CS602", "M1", [("MON", "09:00", "10:15")])
    _enrol(6101, section, cohort="M", source="scraper_timetable")
    Student.objects.create(student_id=6102, name="No timetable", program="CS", section="M")

    result = compute_group_availability([6101, 6102, 6199], YEAR, TERM)

    assert result["resolved_count"] == 1
    assert result["unresolved_count"] == 2
    assert result["coverage_complete"] is False
    assert result["no_schedule"] == [6102]
    assert result["not_found"] == [6199]
    assert result["snapshot_class_counts"] == {
        "registrar": 1,
        "expected": 0,
        "working": 0,
    }
    busy = _cell(result, "lecture", "MON", 0)
    assert busy["busy_count"] == 1
    assert busy["free"] is False
    assert busy["free_for_resolved"] is False
    free = _cell(result, "lecture", "TUE", 0)
    assert free["free"] is True
    assert free["free_for_resolved"] is True
    assert (
        result["grids"]["lecture"]["free_for_resolved_count"]
        == result["grids"]["lecture"]["free_for_all_count"]
    )


def test_partially_timed_schedule_is_flagged_without_blocking_timed_sections():
    timed = _make_global_section("CS603", "M1", [("MON", "09:00", "10:15")])
    untimed = _make_global_section("CS604", "M2", [])
    _enrol(6201, timed, cohort="M", source="scraper_timetable")
    _enrol(6201, untimed, cohort="M", source="scraper_timetable")

    result = compute_group_availability([6201], YEAR, TERM)

    assert result["resolved_count"] == 1
    assert result["unresolved_count"] == 0
    assert result["partial_schedule"] == [6201]
    assert result["partial_schedule_count"] == 1
    assert result["coverage_complete"] is False
    assert result["students"][0]["unscheduled_section_count"] == 1
    assert result["students"][0]["snapshot_class"] == "registrar"
    assert _cell(result, "lecture", "MON", 0)["busy_count"] == 1
    assert _cell(result, "timeline", "MON", 0)["busy_count"] == 1


def test_normalise_student_ids_dedupes_and_drops_nonnumeric():
    assert normalise_student_ids(["5", "5", 5, "x", 7, None]) == [5, 7]


def test_empty_group_returns_all_free():
    result = compute_group_availability([], YEAR, TERM)
    assert result["requested_count"] == 0
    # Group Availability includes one late option ending at 18:30 in both the
    # 75-minute and 100-minute views.
    assert result["grids"]["lecture"]["free_for_all_count"] == 40
    assert result["grids"]["lab"]["free_for_all_count"] == 30
    assert result["grids"]["timeline"]["free_for_all_count"] == 285


# ── View wiring ──────────────────────────────────────────────


def _login_client() -> Client:
    user = get_user_model().objects.create_user(username="ga_tester", password="x")
    client = Client()
    client.force_login(user)
    return client


def test_page_renders_with_config():
    client = _login_client()
    resp = client.get(reverse("group_availability_page"))
    assert resp.status_code == 200
    assert b"groupAvailabilityConfig" in resp.content
    assert reverse("group_availability_export_xlsx").encode() in resp.content
    assert b'id="gaExport"' in resp.content


def test_compute_endpoint_returns_grids():
    sec = _make_global_section("CS701", "S1", [("MON", "09:00", "10:15")])
    _enrol(7001, sec)
    client = _login_client()

    resp = client.post(
        reverse("group_availability_compute"),
        data=json.dumps({"student_ids": [7001]}),
        content_type="application/json",
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["resolved_count"] == 1
    assert body["grids"]["lecture"]["cells"]["MON"][0]["busy_count"] == 1


def test_compute_endpoint_rejects_empty_ids():
    client = _login_client()
    resp = client.post(
        reverse("group_availability_compute"),
        data=json.dumps({"student_ids": []}),
        content_type="application/json",
    )
    assert resp.status_code == 400


def test_compute_endpoint_parses_freetext_ids():
    sec = _make_global_section("CS801", "S1", [("TUE", "10:30", "11:45")])
    _enrol(8001, sec)
    _enrol(8002, sec)
    client = _login_client()

    resp = client.post(
        reverse("group_availability_compute"),
        data=json.dumps({"student_ids": "8001, 8002\n8001"}),
        content_type="application/json",
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["requested_count"] == 2  # deduped
    assert body["grids"]["lecture"]["cells"]["TUE"][1]["busy_count"] == 2


def test_export_endpoint_returns_all_styled_grids_and_truthful_coverage():
    Student.objects.create(
        student_id=9001,
        name="=2+2",
        program="CS",
        section="M",
    )
    section = _make_global_section("CS901", "M2", [("THU", "18:20", "18:30")])
    _enrol(9001, section, cohort="M", source="scraper_timetable")
    unknown_id = 9099
    client = _login_client()

    response = client.post(
        reverse("group_availability_export_xlsx"),
        data=json.dumps({"student_ids": [9001, unknown_id]}),
        content_type="application/json",
    )

    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "group_availability_" in response["Content-Disposition"]
    assert response["Content-Disposition"].endswith('.xlsx"')

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == [
        "Summary",
        "Students",
        "Lecture 75m",
        "Lab 100m",
        "Full Day 10m",
    ]

    summary = workbook["Summary"]
    assert summary["B4"].value == 2
    assert summary["D4"].value == 1
    assert summary["F4"].value == 1
    assert summary["D5"].value == "Incomplete"
    assert str(unknown_id) in summary["B9"].value

    students = workbook["Students"]
    assert students["B5"].value == "'=2+2"
    assert students["B5"].data_type == "s"
    assert students["D6"].value == "Not found"

    expected_final_slots = {
        "Lecture 75m": "17:15–18:30",
        "Lab 100m": "16:50–18:30",
        "Full Day 10m": "18:20–18:30",
    }
    for sheet_name, final_slot in expected_final_slots.items():
        sheet = workbook[sheet_name]
        assert sheet.cell(row=sheet.max_row, column=1).value == final_slot
        assert sheet.cell(row=sheet.max_row, column=6).value == 1  # Thursday

    timeline = workbook["Full Day 10m"]
    timeline_slots = [timeline.cell(row=row, column=1).value for row in range(6, 63)]
    assert len(timeline_slots) == 57
    assert timeline_slots[18:21] == ["12:00–12:10", "12:10–12:20", "12:20–12:30"]


def test_export_endpoint_marks_grid_unknown_when_no_schedule_resolves():
    client = _login_client()
    response = client.post(
        reverse("group_availability_export_xlsx"),
        data=json.dumps({"student_ids": [9199]}),
        content_type="application/json",
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["Summary"]["D5"].value == "Incomplete"
    assert workbook["Lecture 75m"]["B6"].value == "?"
    assert workbook["Lab 100m"]["B6"].value == "?"
    assert workbook["Full Day 10m"]["B6"].value == "?"


def test_export_endpoint_validates_method_ids_and_role():
    client = _login_client()
    assert client.get(reverse("group_availability_export_xlsx")).status_code == 405
    empty = client.post(
        reverse("group_availability_export_xlsx"),
        data=json.dumps({"student_ids": []}),
        content_type="application/json",
    )
    assert empty.status_code == 400

    student_user = get_user_model().objects.create_user(username="ga_student", password="x")
    student_group, _ = Group.objects.get_or_create(name=ROLE_STUDENT)
    student_user.groups.add(student_group)
    student_client = Client()
    student_client.force_login(student_user)
    denied = student_client.post(
        reverse("group_availability_export_xlsx"),
        data=json.dumps({"student_ids": [9001]}),
        content_type="application/json",
    )
    assert denied.status_code == 403
