"""The export's "Instructors" sheet — one weekly grid per assigned instructor,
showing only instructors actually on the current timetable.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from core.models import (
    DeliveryBoard,
    SectionPlacement,
    TermSection,
    TermSectionMeeting,
    TimetableScenario,
)
from core.services.timetable_export import _render_instructors_sheet

_NOFILL = PatternFill()


def _section(scenario, board, code, day, start, end, instructor):
    ts = TermSection.objects.create(
        scenario=scenario,
        course_code=code,
        course_number=code,
        course_key=code,
        course_name=code,
        section="S1",
        source_tag="test",
    )
    TermSectionMeeting.objects.create(
        term_section=ts, day=day, start_time=start, end_time=end, instructor=instructor
    )
    SectionPlacement.objects.create(
        board=board, term_section=ts, day=day, start_time=start, end_time=end, room="R1"
    )
    return ts


def _cells(ws) -> list[str]:
    return [str(c.value) for r in ws.iter_rows() for c in r if c.value is not None]


@pytest.mark.django_db
def test_instructors_sheet_lists_only_assigned() -> None:
    sc = TimetableScenario.objects.create(academic_year="1448", term="1", name="x")
    board = DeliveryBoard.objects.create(scenario=sc, label="T1", nominal_term=1, program="AI")

    # Dr Alpha teaches two courses; one on SUN, one on SUN later (a gap in their day).
    _section(sc, board, "AI301", "SUN", "09:00", "10:15", "Dr Alpha")
    _section(sc, board, "AI302", "MON", "09:00", "10:15", "Dr Alpha")
    # AI303 has NO instructor → must not produce a header.
    _section(sc, board, "AI303", "TUE", "09:00", "10:15", "")

    wb = Workbook()
    _render_instructors_sheet(wb, sc, lambda cc: _NOFILL)

    assert "Instructors" in wb.sheetnames
    ws = wb["Instructors"]
    values = _cells(ws)

    assert "Dr Alpha" in values  # assigned instructor has a table
    assert any("AI301 S1" in v for v in values)  # their course shows in the grid
    assert any("AI302 S1" in v for v in values)
    # The unassigned course's section must not appear (no instructor to bucket it).
    assert not any("AI303" in v for v in values)


@pytest.mark.django_db
def test_no_sheet_when_no_instructors_assigned() -> None:
    sc = TimetableScenario.objects.create(academic_year="1448", term="1", name="x")
    board = DeliveryBoard.objects.create(scenario=sc, label="T1", nominal_term=1, program="AI")
    _section(sc, board, "AI301", "SUN", "09:00", "10:15", "")  # no instructor

    wb = Workbook()
    _render_instructors_sheet(wb, sc, lambda cc: _NOFILL)

    assert "Instructors" not in wb.sheetnames  # nothing assigned → no sheet


@pytest.mark.django_db
def test_same_slot_clash_is_flagged() -> None:
    sc = TimetableScenario.objects.create(academic_year="1448", term="1", name="x")
    board = DeliveryBoard.objects.create(scenario=sc, label="T1", nominal_term=1, program="AI")
    # Same instructor, two courses, same day+slot → both land in one cell.
    _section(sc, board, "AI301", "SUN", "09:00", "10:15", "Dr Beta")
    _section(sc, board, "AI302", "SUN", "09:00", "10:15", "Dr Beta")

    wb = Workbook()
    _render_instructors_sheet(wb, sc, lambda cc: _NOFILL)
    ws = wb["Instructors"]
    # The clashed cell holds both course texts joined by a newline.
    assert any(
        c.value and "AI301 S1" in str(c.value) and "AI302 S1" in str(c.value)
        for r in ws.iter_rows()
        for c in r
    )
